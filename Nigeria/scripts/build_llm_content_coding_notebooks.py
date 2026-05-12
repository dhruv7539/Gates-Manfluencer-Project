"""
Build the LLM Coding Notebooks for Content Analysis (Nigeria + Kenya).

Mirrors the audience-side notebooks
(Nigeria/Notebooks/LLM Coding Notebook - Audience Analysis.ipynb,
 Kenya/Notebooks/LLM Coding Notebook  - Audience Analysis.ipynb)
but uses the content codebook (Q1-Q18b) instead of the audience one (Q1-Q21h).

Run:
    python Nigeria/scripts/build_llm_content_coding_notebooks.py

Outputs:
    Nigeria/Notebooks/LLM Coding Notebook - Content Analysis.ipynb
    Kenya/Notebooks/LLM Coding Notebook - Content Analysis.ipynb
"""
from __future__ import annotations

import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]


# ──────────────────────────────────────────────────────────────────────────────
# Closed vocabularies (shared with audience workbook for cross-comparability)
# ──────────────────────────────────────────────────────────────────────────────

THEMES = [
    "Authority and Submission", "Male Victimhood", "Gender Grievance",
    "Sexual Morality", "Relationship Tactics", "Provider and Status",
    "Male Accountability", "Egalitarian Partnership",
    "Gender-Based Violence and Consent", "Trauma and Mental Health",
    "Self-Discipline", "Marriage and Family", "Faith and Moral Repair",
    "Unclear",
]
SENTIMENTS = ["Positive", "Negative", "Neutral", "Unclear"]
EMOTIONS = ["Joy", "Happiness", "Surprise", "Anger", "Fear", "Contempt",
            "Sadness", "Hope", "Empathy", "None of these"]
TONES = ["Earnest", "Sarcastic", "Hostile", "Humorous",
         "Empathetic", "Authoritative", "Detached"]
NORMATIVE_ORIENTATIONS = ["Progressive", "Regressive", "Mixed", "Unclear"]
TARGETS = ["Men/boys", "Women/girls", "Feminists/modern women",
           "Children/family", "Institutions/law/society", "Creator/content",
           "Self/personal life", "Mixed", "Unclear"]


# ──────────────────────────────────────────────────────────────────────────────
# Closed vocabularies for Q1-Q18b (content codebook)
# ──────────────────────────────────────────────────────────────────────────────

Q1_OPTS = ["Yes", "No"]
Q1A_OPTS = [
    "Compelling question", "Use of all CAPS", "Humor or sarcasm",
    "Shares something violent or gross", "Shares something sexual",
    "Shares something surprising",
    "Uses a news headline or social media trend as opener",
    "Interesting visual or meme", "Other",
]
Q2_OPTS = [
    "Dating/marriage", "Friends/socializing", "Family/children",
    "Money/status", "Fitness/self-improvement", "Mental health",
    "Gender issues, e.g. equality", "Social issues, e.g. corruption",
    "Religion/morality", "Gaming/technology", "Other",
]
Q3_OPTS = [
    "Interview/conversational content", "Motivational/self-help content",
    "Commentary/reaction content", "Other",
]
Q4_OPTS = ["Yes, explicitly", "Yes, implicitly", "No"]
Q5_OPTS = [
    "More regressive/traditional/restrictive",
    "More progressive/equitable/expansive",
    "Mixed/unclear",
    "Does not address masculinity or gender norms",
]
Q6_OPTS = ["Yes", "No", "Unclear", "Not applicable"]
Q7_OPTS = [
    "Men need to dominate/lead", "Men need to provide/succeed",
    "Men are disadvantaged/victims", "Men need to improve themselves",
    "Men need to be fully self-reliant", "Men need to be emotionally open",
    "Men need to not show emotions", "Men need to be equal partners",
    "Mixed/unclear", "Other", "Not applicable",
]
Q8_OPTS = [
    "Kenyan or Nigerian political/social problems",
    "Global political/social/cultural problems",
    "Western political/social influence", "Women/feminism",
    "Men’s behavior", "Economic/status pressure",
    "Mental health/emotional struggle",
    "No clear problem is identified", "Other",
]
Q9_OPTS = [
    "Social or political change", "Assert dominance/control",
    "More wealth/status", "More self-discipline/fitness",
    "More emotional growth/healing", "More equality/respect for men",
    "More equality/respect for women", "Building community",
    "No clear solution", "Other",
]
Q10_OPTS = [
    "Advice/instruction", "Personal story", "Commentary/opinion",
    "Debate/argument", "Humor/satire", "Motivational speech",
    "News/telling facts", "Other",
]
Q11_OPTS = [
    "Entertainment/escapism", "Information seeking",
    "Connection/social interaction",
    "Self expression/identity construction", "Status seeking",
    "Documentation of events", "None of these apply",
]
Q12_OPTS = [
    "Generalizations about men/women", "Personal experience",
    "Stories about men/women", "Cultural/social observations",
    "Facts/statistics", "Moral/religious claims", "Mixed",
    "No support", "Other",
]
Q13_OPTS = [
    "No justification", "Anecdotal examples", "Presented as common sense",
    "References data", "References religion/tradition",
    "References external sources, such as other influencers", "Other",
]
Q14_OPTS = ["Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned"]
Q15_OPTS = Q14_OPTS
Q16_OPTS = Q14_OPTS
Q17_OPTS = ["Yes", "Somewhat", "No"]
Q18_OPTS = ["Yes", "No"]
Q18A_OPTS = [
    "Calls for audience to like the content",
    "Calls for audience to share the content",
    "Calls for audience to follow the speaker on social media",
    "Calls for men to follow more traditional gender norms",
    "Calls for men to follow more equitable gender norms",
    "Calls for women to follow more traditional gender norms",
    "Calls for women to follow more equitable gender norms",
    "Calls for politicians or social figures to do something",
    "Calls for audience to vote in a different way", "Other",
]


# ──────────────────────────────────────────────────────────────────────────────
# Prompt builder (shared between Nigeria + Kenya, country tag injected)
# ──────────────────────────────────────────────────────────────────────────────

def build_prompt(country: str) -> str:
    country_note = {
        "Nigeria": (
            "Note on Nigeria context: Pidgin/Yoruba/Igbo/Hausa terms appear "
            "(agba, oga, wahala, abeg, ashawo). Faith framing is common "
            "(Christian + Islamic). Pidgin code-switching is normal."
        ),
        "Kenya": (
            "Note on Kenya context: Sheng/Swahili terms appear "
            "(mubaba, malaya, sponsor, jasho). Faith framing is common "
            "(Christian + Islamic). Hustle/discipline language is part of the "
            "regressive register (Amerix-style #MasculinitySaturday)."
        ),
    }[country]

    return (
        f"You are a senior research-grade content annotator coding {country} masculinity-focused "
        f"social-media CREATOR CONTENT (videos, podcasts, tweets) for the Norman Lear Center / "
        f"Gates Foundation Manfluencer Project.\n"
        f"\n"
        f"You are coding ONE content unit at a time. Read the full text carefully before assigning any code.\n"
        f"\n"
        f"Treat each field as a defensible research judgment, not a guess. Apply the codebook strictly. "
        f"When in doubt, choose the more conservative option: Unclear, No, None of these apply, "
        f"Not mentioned, Not applicable, or empty string for conditional open-text fields.\n"
        f"\n"
        f"Do not invent labels. Use ONLY the exact options listed below.\n"
        f"\n"
        f"Open-text fields must be specific and quote-grounded. Never write generic placeholders like "
        f"\"explains the topic\", \"the creator argues something\", or \"the content is about gender\".\n"
        f"\n"
        f"If a field is conditional, leave it as an empty string when the gating answer is No or not applicable.\n"
        f"\n"
        f"{country_note}\n"
        f"\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"CRITICAL CODING RULES (read first — these prevent the most common errors)\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"\n"
        f"RULE 1 — Theme reflects what the CREATOR advocates in this segment.\n"
        f"Code the position the creator is putting forward. If the creator is critiquing a regressive "
        f"position to argue for accountability or equality, the theme is the progressive frame the "
        f"creator advances (e.g., Male Accountability, Egalitarian Partnership, GBV and Consent), "
        f"NOT the regressive code that describes what they are critiquing.\n"
        f"\n"
        f"RULE 2 — Q4 is INTERNALLY CONSISTENT with Q5/Q6/Q7.\n"
        f"If Q4 = \"No\", then Q5 MUST be \"Does not address masculinity or gender norms\", Q6 MUST be "
        f"\"Not applicable\", Q7 MUST be [\"Not applicable\"], and Q7a MUST be empty. If you assign any "
        f"non-default value to Q5/Q6/Q7, then Q4 MUST be \"Yes, explicitly\" or \"Yes, implicitly\".\n"
        f"\n"
        f"RULE 3 — Q6 ↔ Q7 consistency.\n"
        f"If Q6 = \"No\" / \"Unclear\" / \"Not applicable\", then Q7 MUST be [\"Not applicable\"] and Q7a empty. "
        f"Q7 prescriptions (e.g., \"Men need to dominate/lead\") are only valid when Q6 = \"Yes\".\n"
        f"\n"
        f"RULE 4 — \"Other\" gates are tight.\n"
        f"Q1b only when Q1a includes \"Other\". Q2a only when Q2 includes \"Other\". Q3a only when Q3 = \"Other\". "
        f"Q7a only when Q7 includes \"Other\". Q8a only when Q8 includes \"Other\". Q9a only when Q9 includes \"Other\". "
        f"Q10a only when Q10 includes \"Other\". Q12a only when Q12 = \"Other\". Q13a only when Q13 = \"Other\". "
        f"Q18b only when Q18a includes \"Other\". Outside those gates, the matching open-text field MUST be empty.\n"
        f"\n"
        f"RULE 5 — Q1 ↔ Q1a.\n"
        f"If Q1 = \"No\" then Q1a MUST be empty list and Q1b empty. If Q1 = \"Yes\", Q1a MUST contain at least one option.\n"
        f"\n"
        f"RULE 6 — Q18 ↔ Q18a/Q18b.\n"
        f"If Q18 = \"No\", Q18a MUST be empty list and Q18b empty. If Q18 = \"Yes\", Q18a MUST contain at least one option.\n"
        f"\n"
        f"RULE 7 — Q14/Q15/Q16 use the full 6-value vocabulary.\n"
        f"Use \"Not mentioned\" only when the content makes NO reference to that group / construct. \"Mixed\" is "
        f"reserved for content that genuinely contains both positive and negative framing.\n"
        f"\n"
        f"RULE 8 — Q17 is reserved for fear-/threat-based persuasion.\n"
        f"\"Yes\" only when the content uses warnings, dire consequences, scare framing, or threat language to "
        f"drive the message. \"Somewhat\" for mild fear cues. \"No\" otherwise. Sarcasm or anger alone is NOT fear/threat.\n"
        f"\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"PRIMARY AND SECONDARY THEMES\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"Pick exactly ONE primary_theme.\n"
        f"\n"
        f"Pick up to TWO secondary themes:\n"
        f"- secondary_theme_1 may be empty if there is no strong secondary theme.\n"
        f"- secondary_theme_2 may be empty if there is no second strong secondary theme.\n"
        f"- Do not duplicate the primary theme.\n"
        f"- Do not duplicate secondary themes.\n"
        f"- If primary_theme = \"Unclear\", both secondary fields must be empty.\n"
        f"\n"
        f"Allowed theme labels (use these exact strings):\n"
        f"\n"
        f"1. Authority and Submission — explicit hierarchy: headship, submission, control, surname, command.\n"
        f"2. Male Victimhood — men framed as exploited, disadvantaged, harmed by women, courts, false accusations.\n"
        f"3. Gender Grievance — generalized distrust of women / feminists / equality framed as scam or threat.\n"
        f"4. Sexual Morality — cheating, body count, abortion, body policing, sexual respectability, double standards.\n"
        f"5. Relationship Tactics — tactical dating advice (scarcity, pursuit, availability, masculine-frame instructions).\n"
        f"6. Provider and Status — money/income/career/status/respectability as proof of manhood or relationship worth.\n"
        f"7. Male Accountability — \"men must change\", hold men accountable, refusal of deflection.\n"
        f"8. Egalitarian Partnership — mutual respect, shared parenting, listening, allyship, healthy reciprocity.\n"
        f"9. Gender-Based Violence and Consent — rape, consent, abuse, victim stigma, false accusations, justice.\n"
        f"10. Trauma and Mental Health — trauma, depression, grief, healing, vulnerability, psychological harm.\n"
        f"11. Self-Discipline — personal responsibility, restraint, growth, learning. Constructive only.\n"
        f"12. Marriage and Family — marriage / divorce / family / fatherhood / motherhood as the OBJECT.\n"
        f"13. Faith and Moral Repair — explicit faith / scripture / God / prayer / sin / testimony.\n"
        f"14. Unclear — uncodable / off-topic / low-signal.\n"
        f"\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"FRONT-SUMMARY FIELDS\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"masculinity_identity: \"Yes\" / \"No\" — Yes if content directly discusses men, boys, manhood, masculinity, "
        f"male socialization, male identity, male behavior, or male roles. Stricter than Q4.\n"
        f"\n"
        f"normative_orientation: {' / '.join(NORMATIVE_ORIENTATIONS)}\n"
        f"  Progressive — challenges hierarchy, supports equality, male accountability, empathy, consent, healthy "
        f"vulnerability, or shared partnership.\n"
        f"  Regressive — reinforces hierarchy, misogyny, rigid gender roles, male dominance, female submission, "
        f"anti-feminist grievance, sexual double standards, or victim-blaming.\n"
        f"  Mixed — both progressive and regressive elements present.\n"
        f"  Unclear — no clear ideological direction.\n"
        f"\n"
        f"target_of_claim: {' / '.join(TARGETS)} — pick the main target being addressed, blamed, praised, advised, "
        f"defended, or evaluated.\n"
        f"\n"
        f"sentiment: {' / '.join(SENTIMENTS)}\n"
        f"emotion: {' / '.join(EMOTIONS)} — pick the single dominant emotion expressed.\n"
        f"tone: {' / '.join(TONES)}\n"
        f"  Earnest — sincere, direct, no irony.\n"
        f"  Sarcastic — ironic, mocking, opposite-meaning surface.\n"
        f"  Hostile — aggressive, attacking, confrontational, insulting.\n"
        f"  Humorous — playful, joking, light-hearted (genuine humor, not sarcasm).\n"
        f"  Empathetic — supportive, compassionate, validating.\n"
        f"  Authoritative — didactic, prescriptive, lecturing.\n"
        f"  Detached — neutral, observational, distant.\n"
        f"\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"HUMAN CODEBOOK QUESTIONS (Q1–Q18b)\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"\n"
        f"q1 — content starts with attention-getter ({' / '.join(Q1_OPTS)}). Verbal or visual hook (caps, exclamations, "
        f"shocking opener, headline-style title, meme).\n"
        f"q1a — if q1 = \"Yes\", multi-select from: {', '.join(Q1A_OPTS)}. Empty list if q1 = No.\n"
        f"q1b — if q1a includes \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q2 — primary topic(s), multi-select: {', '.join(Q2_OPTS)}.\n"
        f"q2a — if q2 includes \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q3 — content type, single-select: {', '.join(Q3_OPTS)}.\n"
        f"q3a — if q3 = \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q4 — addresses masculinity / gender norms, single-select: {', '.join(Q4_OPTS)}.\n"
        f"q5 — masculinity/gender-norm characterization (RULE 2): {', '.join(Q5_OPTS)}.\n"
        f"q6 — addresses what men should do/be (RULE 3): {', '.join(Q6_OPTS)}.\n"
        f"q7 — what content indicates men do/need to do, multi-select: {', '.join(Q7_OPTS)}. Use [\"Not applicable\"] "
        f"when q6 ≠ Yes.\n"
        f"q7a — if q7 includes \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q8 — problem identified, multi-select: {', '.join(Q8_OPTS)}.\n"
        f"q8a — if q8 includes \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q9 — solution proposed, multi-select: {', '.join(Q9_OPTS)}.\n"
        f"q9a — if q9 includes \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q10 — communication mode, multi-select: {', '.join(Q10_OPTS)}.\n"
        f"q10a — if q10 includes \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q11 — audience needs served, multi-select: {', '.join(Q11_OPTS)}.\n"
        f"\n"
        f"q12 — claim support, single-select: {', '.join(Q12_OPTS)}.\n"
        f"q12a — if q12 = \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q13 — claim justification, single-select: {', '.join(Q13_OPTS)}.\n"
        f"q13a — if q13 = \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"q14 — sentiment toward men: {', '.join(Q14_OPTS)}.\n"
        f"q15 — sentiment toward women: {', '.join(Q15_OPTS)}.\n"
        f"q16 — sentiment toward traditional gender norms: {', '.join(Q16_OPTS)}.\n"
        f"\n"
        f"q17 — uses fear/threat: {', '.join(Q17_OPTS)} (RULE 8).\n"
        f"\n"
        f"q18 — calls to action: {', '.join(Q18_OPTS)} (RULE 6).\n"
        f"q18a — if q18 = \"Yes\", multi-select: {', '.join(Q18A_OPTS)}. Empty list if q18 = No.\n"
        f"q18b — if q18a includes \"Other\", ONE phrase. Else empty.\n"
        f"\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"QUALITY CHECKLIST (silent, before returning)\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"1. primary_theme reflects what the CREATOR advocates (RULE 1).\n"
        f"2. q4 is internally consistent with q5/q6/q7 (RULE 2).\n"
        f"3. q6/q7 are consistent (RULE 3).\n"
        f"4. All \"Other\" open-text gates obey RULE 4.\n"
        f"5. q1/q1a/q1b consistent (RULE 5).\n"
        f"6. q18/q18a/q18b consistent (RULE 6).\n"
        f"7. q14/q15/q16 vocabularies obey RULE 7.\n"
        f"8. q17 reserved for fear/threat (RULE 8).\n"
        f"\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"OUTPUT — JSON ONLY, no markdown, no commentary\n"
        f"────────────────────────────────────────────────────────────────────\n"
        f"{{\n"
        f"  \"primary_theme\": \"\",\n"
        f"  \"secondary_theme_1\": \"\",\n"
        f"  \"secondary_theme_2\": \"\",\n"
        f"  \"masculinity_identity\": \"\",\n"
        f"  \"normative_orientation\": \"\",\n"
        f"  \"target_of_claim\": \"\",\n"
        f"  \"sentiment\": \"\",\n"
        f"  \"emotion\": \"\",\n"
        f"  \"tone\": \"\",\n"
        f"  \"q1\": \"\", \"q1a\": [], \"q1b\": \"\",\n"
        f"  \"q2\": [], \"q2a\": \"\",\n"
        f"  \"q3\": \"\", \"q3a\": \"\",\n"
        f"  \"q4\": \"\", \"q5\": \"\", \"q6\": \"\",\n"
        f"  \"q7\": [], \"q7a\": \"\",\n"
        f"  \"q8\": [], \"q8a\": \"\",\n"
        f"  \"q9\": [], \"q9a\": \"\",\n"
        f"  \"q10\": [], \"q10a\": \"\",\n"
        f"  \"q11\": [],\n"
        f"  \"q12\": \"\", \"q12a\": \"\",\n"
        f"  \"q13\": \"\", \"q13a\": \"\",\n"
        f"  \"q14\": \"\", \"q15\": \"\", \"q16\": \"\",\n"
        f"  \"q17\": \"\",\n"
        f"  \"q18\": \"\", \"q18a\": [], \"q18b\": \"\"\n"
        f"}}\n"
    )


# ──────────────────────────────────────────────────────────────────────────────
# Cell builders
# ──────────────────────────────────────────────────────────────────────────────

def md(text: str) -> dict:
    return {"cell_type": "markdown", "metadata": {}, "source": text.splitlines(keepends=True)}


def code(text: str) -> dict:
    return {
        "cell_type": "code",
        "execution_count": None,
        "metadata": {},
        "outputs": [],
        "source": text.splitlines(keepends=True),
    }


def build_notebook(country: str) -> dict:
    """Build full notebook JSON for one country."""

    is_nigeria = country == "Nigeria"
    sheet_color = "305496" if is_nigeria else "C65911"
    cache_filename = "llm_content_coding_cache.parquet" if is_nigeria else "llm_content_coding_cache_kenya.parquet"

    if is_nigeria:
        input_path_str = (
            "ROOT / 'Nigeria' / 'Content Analysis' / 'Translated' / 'content_final_translated.parquet'"
        )
        load_block = """\
df = pd.read_parquet(INPUT_PATH)
print(f'full content set: {len(df)} rows, columns: {list(df.columns)}')
print(df['creator'].value_counts().to_string())

# stratified sample (cap at min(N_PER_CREATOR, available per creator))
samples = []
for cr, sub in df.groupby('creator'):
    n = min(N_PER_CREATOR, len(sub))
    samples.append(sub.sample(n=n, random_state=SEED))
sample = pd.concat(samples).reset_index(drop=True)
print(f'\\nstratified sample: {len(sample)} rows')
print(sample.groupby('creator').size().to_string())

# canonical text column for prompting
sample['_text'] = sample['text_english'].astype(str)
sample['_id']   = sample['content_id'].astype(str)
sample['_context'] = sample['context'].astype(str)
"""
    else:
        input_path_str = (
            "ROOT / 'Kenya' / 'Content Analysis' / 'Kenya Content Analysis Final.xlsx'"
        )
        load_block = """\
import openpyxl
wb_in = openpyxl.load_workbook(INPUT_PATH, data_only=True)
all_dfs = []
for sn in wb_in.sheetnames:
    if sn.lower() == 'summary': continue
    ws = wb_in[sn]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: continue
    hdr = [h for h in rows[0] if h]   # drop trailing None columns
    data = []
    for r in rows[1:]:
        if not r or r[0] is None: continue
        data.append(dict(zip(hdr, r[:len(hdr)])))
    sub = pd.DataFrame(data)
    if not sub.empty:
        all_dfs.append(sub)
df = pd.concat(all_dfs, ignore_index=True)

# normalize column names: Eric/Amerix uses 'Tweet ID' instead of 'Segment ID'
if 'Tweet ID' in df.columns and 'Segment ID' in df.columns:
    df['Segment ID'] = df['Segment ID'].fillna(df['Tweet ID'])
elif 'Tweet ID' in df.columns:
    df = df.rename(columns={'Tweet ID': 'Segment ID'})

print(f'full Kenya content set: {len(df)} rows')
print(df['Influencer'].value_counts().to_string())

# stratified sample (cap at min(N_PER_CREATOR, available))
samples = []
for cr, sub in df.groupby('Influencer'):
    n = min(N_PER_CREATOR, len(sub))
    samples.append(sub.sample(n=n, random_state=SEED))
sample = pd.concat(samples).reset_index(drop=True)
print(f'\\nstratified sample: {len(sample)} rows')
print(sample.groupby('Influencer').size().to_string())

# canonical columns for prompting (Kenya text is largely English; Sheng/Swahili
# handled by the LLM directly, same as the Kenya audience notebook)
sample['_text'] = sample['Text'].astype(str)
sample['_id']   = sample['Segment ID'].astype(str)
sample['_context'] = sample['Context'].astype(str)
sample['creator'] = sample['Influencer']
"""

    cells = []

    # ── intro ────────────────────────────────────────────────────────────────
    cells.append(md(
f"""# LLM Coding Notebook — Content Analysis ({country})

**Norman Lear Center × Gates Foundation — Manfluencer Project**

Codes a stratified random sample of {country} **creator content** segments using
`gpt-4o-mini` against the content codebook (Q1–Q18b — same as the human content
codebook in `Codebooks/Human Codebooks/content/Human A - Content Analysis Codebook.xlsx`).

Mirrors the audience-side LLM coding notebook, but for content (creator-side) units.

## Inputs
{'- `Nigeria/Content Analysis/Translated/content_final_translated.parquet` — translated coding-unit segments × 6 creators.' if is_nigeria else '- `Kenya/Content Analysis/Kenya Content Analysis Final.xlsx` — coding-unit segments × 5 creators (Andrew Kibe, Eric Amerix, Eddy Kimani, Onyango Otieno/Rixpoet, Philip Karanja).'}

## Sampling
Stratified up to **50 per creator** (capped at availability), seed=42. Smaller creator pools
(e.g., Deyemi @ 24) take everything they have.

## Coding produced per row
Front summary columns (LLM-generated, same vocabularies as audience workbook for cross-comparability):
- **Themes** — up to 3 from the 13-theme list + Unclear
- **Masculinity Identity / Normative Orientation / Target of Claim**
- **Sentiment / Emotion Detection / Tone**

Plus Q1–Q18b matching the human content codebook exactly.

## Output
Sheet **`{country} - LLM Coding`** in `Codebooks/LLM Codebook/LLM Coding - Content Analysis.xlsx`.

## Cost
~$0.20–0.40 + 5–8 min per country wall time on a cold cache. Re-runs are free.
"""
    ))

    # ── 0 setup ─────────────────────────────────────────────────────────────
    cells.append(md("## 0 — Setup\n"))
    cells.append(code(
f"""from __future__ import annotations
import asyncio, hashlib, json, os, re
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openai import AsyncOpenAI
from tqdm.asyncio import tqdm as atqdm

ROOT = Path.cwd().resolve()
while ROOT.name != 'Gates-Manfluencer-Project' and ROOT.parent != ROOT:
    ROOT = ROOT.parent
assert ROOT.name == 'Gates-Manfluencer-Project'

load_dotenv(ROOT / '.env')
assert os.getenv('OPENAI_API_KEY'), 'OPENAI_API_KEY missing'

INPUT_PATH = {input_path_str}
OUT_DIR    = ROOT / 'Codebooks' / 'LLM Codebook'
OUT_XLSX   = OUT_DIR / 'LLM Coding - Content Analysis.xlsx'
CACHE      = ROOT / 'temp' / '{cache_filename}'
CACHE.parent.mkdir(parents=True, exist_ok=True)
OUT_DIR.mkdir(parents=True, exist_ok=True)

MODEL = 'gpt-4o-mini'
SEED  = 42
N_PER_CREATOR = 50      # capped at availability
CONCURRENCY = 8

print(f'ROOT = {{ROOT}}')
print(f'model = {{MODEL}}, seed = {{SEED}}, n per creator (cap) = {{N_PER_CREATOR}}')
"""
    ))

    # ── 1 load + sample ─────────────────────────────────────────────────────
    cells.append(md("## 1 — Load + stratified sample\n"))
    cells.append(code(load_block))

    # ── 2 prompt ────────────────────────────────────────────────────────────
    cells.append(md("## 2 — Coding prompt (single LLM call per row → full structured JSON)\n"))
    prompt_text = build_prompt(country)
    # embed PROMPT as a normal Python triple-quoted string (no f-string here).
    cells.append(code(
f"""THEMES = {THEMES!r}
SENTIMENTS = {SENTIMENTS!r}
EMOTIONS = {EMOTIONS!r}
TONES = {TONES!r}
NORMATIVE_ORIENTATIONS = {NORMATIVE_ORIENTATIONS!r}
TARGETS = {TARGETS!r}

Q1_OPTS  = {Q1_OPTS!r}
Q1A_OPTS = {Q1A_OPTS!r}
Q2_OPTS  = {Q2_OPTS!r}
Q3_OPTS  = {Q3_OPTS!r}
Q4_OPTS  = {Q4_OPTS!r}
Q5_OPTS  = {Q5_OPTS!r}
Q6_OPTS  = {Q6_OPTS!r}
Q7_OPTS  = {Q7_OPTS!r}
Q8_OPTS  = {Q8_OPTS!r}
Q9_OPTS  = {Q9_OPTS!r}
Q10_OPTS = {Q10_OPTS!r}
Q11_OPTS = {Q11_OPTS!r}
Q12_OPTS = {Q12_OPTS!r}
Q13_OPTS = {Q13_OPTS!r}
Q14_OPTS = {Q14_OPTS!r}
Q15_OPTS = {Q15_OPTS!r}
Q16_OPTS = {Q16_OPTS!r}
Q17_OPTS = {Q17_OPTS!r}
Q18_OPTS = {Q18_OPTS!r}
Q18A_OPTS = {Q18A_OPTS!r}

PROMPT = {prompt_text!r}
print(f'prompt length: {{len(PROMPT):,}} chars')
"""
    ))

    # ── 3 async run ─────────────────────────────────────────────────────────
    cells.append(md("## 3 — Run async (cached by SHA-1 of segment text)\n"))
    cells.append(code(
"""def text_hash(s: str) -> str:
    return hashlib.sha1(s.encode('utf-8')).hexdigest()[:16]

def load_cache():
    if CACHE.exists():
        c = pd.read_parquet(CACHE)
        return {r.text_hash: json.loads(r.result_json) for r in c.itertuples()}
    return {}

def save_cache(cache):
    rows = [{'text_hash': k, 'result_json': json.dumps(v, ensure_ascii=False)} for k, v in cache.items()]
    pd.DataFrame(rows).to_parquet(CACHE, index=False)

async def code_one(client, sem, text):
    async with sem:
        try:
            r = await client.chat.completions.create(
                model=MODEL,
                messages=[
                    {'role': 'system', 'content': PROMPT},
                    {'role': 'user', 'content': text[:4000]},
                ],
                temperature=0,
                max_tokens=1400,
                response_format={'type': 'json_object'},
            )
            return json.loads(r.choices[0].message.content)
        except Exception as e:
            return {'error': str(e)[:200]}

async def run_all():
    cache = load_cache()
    print(f'cache: {len(cache)} entries')
    todo = [(text_hash(t), t) for t in sample['_text'] if text_hash(t) not in cache]
    print(f'rows to code: {len(todo)} / {len(sample)}')
    if not todo:
        return cache
    client = AsyncOpenAI()
    sem = asyncio.Semaphore(CONCURRENCY)
    tasks = [code_one(client, sem, t) for _, t in todo]
    results = await atqdm.gather(*tasks)
    for (h, _), res in zip(todo, results):
        cache[h] = res
    save_cache(cache)
    print(f'cached: {len(cache)} entries')
    return cache

cache = await run_all()
"""
    ))

    # ── 4 validate ──────────────────────────────────────────────────────────
    cells.append(md("## 4 — Validation: enforce closed-vocab, repair gating violations\n"))
    cells.append(code(
"""with open(ROOT / 'temp' / 'human_content_headers.json') as f:
    HUMAN_HDRS = json.load(f)
HDR_BY_KEY = {}
for h in HUMAN_HDRS:
    first = h.split('\\n')[0].strip()
    if first.startswith('Q'):
        key = first.split('.')[0].strip().lower().replace(' ', '')
        HDR_BY_KEY[key] = h

def hdr(q): return HDR_BY_KEY[q.lower()]

def coerce_one(val, opts, default=''):
    if not isinstance(val, str): return default
    v = val.strip()
    if v in opts: return v
    for o in opts:
        if v.lower() == o.lower(): return o
    return default

def coerce_multi(val, opts):
    if not isinstance(val, list): return []
    seen = []
    for v in val:
        c = coerce_one(v, opts, default='')
        if c and c not in seen: seen.append(c)
    return seen

def sanitize_themes(p, s1, s2):
    pp  = coerce_one(p,  THEMES, 'Unclear')
    s1c = coerce_one(s1, THEMES, '')
    s2c = coerce_one(s2, THEMES, '')
    if pp == 'Unclear': return pp, '', ''
    if s1c == pp: s1c = ''
    if s2c == pp or s2c == s1c: s2c = ''
    return pp, s1c, s2c

YN = ['Yes', 'No']
fixlog = {
    'q4_consistency': 0, 'q6_q7_consistency': 0,
    'q1_q1a_consistency': 0, 'q18_q18a_consistency': 0,
    'other_gate_cleanups': 0,
}

rows = []
for _, r in sample.iterrows():
    h = text_hash(r['_text'])
    raw = cache.get(h, {})

    primary_theme, sec1, sec2 = sanitize_themes(
        raw.get('primary_theme'), raw.get('secondary_theme_1'), raw.get('secondary_theme_2'))

    masculinity_identity  = coerce_one(raw.get('masculinity_identity'),  YN, 'No')
    normative_orientation = coerce_one(raw.get('normative_orientation'), NORMATIVE_ORIENTATIONS, 'Unclear')
    target_of_claim       = coerce_one(raw.get('target_of_claim'),        TARGETS, 'Unclear')
    sentiment = coerce_one(raw.get('sentiment'), SENTIMENTS, 'Unclear')
    emotion   = coerce_one(raw.get('emotion'),   EMOTIONS,    'None of these')
    tone      = coerce_one(raw.get('tone'),      TONES,       'Detached')

    # === Q1 / Q1a / Q1b (RULE 5) ===
    q1  = coerce_one(raw.get('q1'),  Q1_OPTS, 'No')
    q1a = coerce_multi(raw.get('q1a'), Q1A_OPTS)
    q1b = (raw.get('q1b') or '').strip()
    if q1 == 'No' and q1a:
        q1 = 'Yes'; fixlog['q1_q1a_consistency'] += 1
    if q1 == 'No':
        q1a = []; q1b = ''
    if 'Other' not in q1a:
        q1b = ''
    q1a_str = '; '.join(q1a)

    # === Q2 / Q2a ===
    q2  = coerce_multi(raw.get('q2'), Q2_OPTS)
    q2a = (raw.get('q2a') or '').strip()
    if 'Other' not in q2: q2a = ''
    q2_str = '; '.join(q2)

    # === Q3 / Q3a ===
    q3  = coerce_one(raw.get('q3'), Q3_OPTS, 'Other')
    q3a = (raw.get('q3a') or '').strip()
    if q3 != 'Other': q3a = ''

    # === Q4 / Q5 / Q6 / Q7 (RULE 2 + 3) ===
    q4 = coerce_one(raw.get('q4'), Q4_OPTS, 'No')
    q5 = coerce_one(raw.get('q5'), Q5_OPTS, 'Does not address masculinity or gender norms')
    q6 = coerce_one(raw.get('q6'), Q6_OPTS, 'Not applicable')
    q7 = coerce_multi(raw.get('q7'), Q7_OPTS)
    q7a = (raw.get('q7a') or '').strip()

    q5_engaged = q5 != 'Does not address masculinity or gender norms'
    q6_engaged = q6 == 'Yes'
    q7_engaged = bool([x for x in q7 if x != 'Not applicable'])
    if q4 == 'No' and (q5_engaged or q6_engaged or q7_engaged):
        q4 = 'Yes, implicitly'; fixlog['q4_consistency'] += 1
    if q4 == 'No':
        q5 = 'Does not address masculinity or gender norms'
        q6 = 'Not applicable'
        q7 = ['Not applicable']
        q7a = ''
    if q6 != 'Yes':
        if q7_engaged:
            fixlog['q6_q7_consistency'] += 1
        q7 = ['Not applicable']
        q7a = ''
    if 'Other' not in q7: q7a = ''
    q7_str = '; '.join(q7) if q7 else 'Not applicable'

    # === Q8 / Q8a ===
    q8  = coerce_multi(raw.get('q8'),  Q8_OPTS)
    q8a = (raw.get('q8a') or '').strip()
    if 'Other' not in q8: q8a = ''
    q8_str = '; '.join(q8) if q8 else 'No clear problem is identified'

    # === Q9 / Q9a ===
    q9  = coerce_multi(raw.get('q9'),  Q9_OPTS)
    q9a = (raw.get('q9a') or '').strip()
    if 'Other' not in q9: q9a = ''
    q9_str = '; '.join(q9) if q9 else 'No clear solution'

    # === Q10 / Q10a ===
    q10  = coerce_multi(raw.get('q10'), Q10_OPTS)
    q10a = (raw.get('q10a') or '').strip()
    if 'Other' not in q10: q10a = ''
    q10_str = '; '.join(q10) if q10 else 'Other'

    # === Q11 ===
    q11 = coerce_multi(raw.get('q11'), Q11_OPTS)
    q11_str = '; '.join(q11) if q11 else 'None of these apply'

    # === Q12 / Q12a ===
    q12  = coerce_one(raw.get('q12'), Q12_OPTS, 'No support')
    q12a = (raw.get('q12a') or '').strip()
    if q12 != 'Other': q12a = ''

    # === Q13 / Q13a ===
    q13  = coerce_one(raw.get('q13'), Q13_OPTS, 'No justification')
    q13a = (raw.get('q13a') or '').strip()
    if q13 != 'Other': q13a = ''

    # === Q14 / Q15 / Q16 ===
    q14 = coerce_one(raw.get('q14'), Q14_OPTS, 'Not mentioned')
    q15 = coerce_one(raw.get('q15'), Q15_OPTS, 'Not mentioned')
    q16 = coerce_one(raw.get('q16'), Q16_OPTS, 'Not mentioned')

    # === Q17 ===
    q17 = coerce_one(raw.get('q17'), Q17_OPTS, 'No')

    # === Q18 / Q18a / Q18b (RULE 6) ===
    q18  = coerce_one(raw.get('q18'), Q18_OPTS, 'No')
    q18a = coerce_multi(raw.get('q18a'), Q18A_OPTS)
    q18b = (raw.get('q18b') or '').strip()
    if q18 == 'No' and q18a:
        q18 = 'Yes'; fixlog['q18_q18a_consistency'] += 1
    if q18 == 'No':
        q18a = []; q18b = ''
    if 'Other' not in q18a:
        q18b = ''
    q18a_str = '; '.join(q18a)

    rows.append({
        'Content ID':                 r['_id'],
        'Context':                    r['_context'],
        'Content Text / Description': r['_text'],
        hdr('q1'):  q1,  hdr('q1a'): q1a_str, hdr('q1b'): q1b,
        hdr('q2'):  q2_str, hdr('q2a'): q2a,
        hdr('q3'):  q3,  hdr('q3a'): q3a,
        hdr('q4'):  q4,  hdr('q5'):  q5,  hdr('q6'): q6,
        hdr('q7'):  q7_str, hdr('q7a'): q7a,
        hdr('q8'):  q8_str, hdr('q8a'): q8a,
        hdr('q9'):  q9_str, hdr('q9a'): q9a,
        hdr('q10'): q10_str, hdr('q10a'): q10a,
        hdr('q11'): q11_str,
        hdr('q12'): q12, hdr('q12a'): q12a,
        hdr('q13'): q13, hdr('q13a'): q13a,
        hdr('q14'): q14, hdr('q15'): q15, hdr('q16'): q16,
        hdr('q17'): q17,
        hdr('q18'): q18, hdr('q18a'): q18a_str, hdr('q18b'): q18b,
        'creator':  r['creator'],
    })

out = pd.DataFrame(rows)
print(f'coded rows: {len(out)}')
print(f'auto-fixes: {fixlog}')
print(f'\\nQ4 (addresses gender norms): {out[hdr("q4")].value_counts().to_dict()}')
print(f'Q14 (sentiment toward men):   {out[hdr("q14")].value_counts().to_dict()}')
print(f'Q15 (sentiment toward women): {out[hdr("q15")].value_counts().to_dict()}')
"""
    ))

    # ── 5 summary ───────────────────────────────────────────────────────────
    cells.append(md("## 5 — Summary stats by creator\n"))
    cells.append(code(
"""print('=== Q4 (addresses gender) × creator ===')
print(out.groupby(['creator', hdr('q4')]).size().unstack(fill_value=0).to_string())
print('\\n=== Q14 (sentiment toward men) × creator ===')
print(out.groupby(['creator', hdr('q14')]).size().unstack(fill_value=0).to_string())
print('\\n=== Q15 (sentiment toward women) × creator ===')
print(out.groupby(['creator', hdr('q15')]).size().unstack(fill_value=0).to_string())
print('\\n=== Q18 (calls to action) × creator ===')
print(out.groupby(['creator', hdr('q18')]).size().unstack(fill_value=0).to_string())
"""
    ))

    # ── 6 export ────────────────────────────────────────────────────────────
    cells.append(md(f"## 6 — Export to xlsx\n\nWrites sheet `{country} - LLM Coding` into `Codebooks/LLM Codebook/LLM Coding - Content Analysis.xlsx`, preserving the other country's sheet.\n"))
    cells.append(code(
f"""COUNTRY = '{country}'
SHEET_NAME = '{country} - LLM Coding'
SHEET_HEADER_COLOR = '{sheet_color}'

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

export_cols = [c for c in out.columns if c not in ('creator',)]
out_export = out[export_cols].copy()

if OUT_XLSX.exists():
    wb = openpyxl.load_workbook(OUT_XLSX)
else:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)
ws.append(list(out_export.columns))
for _, row in out_export.iterrows():
    ws.append([row[c] for c in out_export.columns])

header_fill = PatternFill('solid', fgColor=SHEET_HEADER_COLOR)
header_font = Font(bold=True, color='FFFFFF', size=10)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
ws.row_dimensions[1].height = 60
ws.freeze_panes = 'D2'   # freeze first 3 metadata cols

for col_idx, col_name in enumerate(out_export.columns, 1):
    letter = get_column_letter(col_idx)
    if col_name == 'Content Text / Description':
        ws.column_dimensions[letter].width = 70
    elif col_name == 'Context':
        ws.column_dimensions[letter].width = 45
    elif col_name.startswith('Q'):
        first_token = col_name.split('.')[0].strip()
        if first_token.lower().endswith(('a','b')) and first_token.lower() != 'qa':
            ws.column_dimensions[letter].width = 35
        else:
            ws.column_dimensions[letter].width = 22
    else:
        ws.column_dimensions[letter].width = 18

for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.font = Font(size=10)

# methodology sheet
if 'Methodology' in wb.sheetnames: del wb['Methodology']
mws = wb.create_sheet('Methodology')
mws.append(['country', 'metric', 'value'])
for row_data in [
    (COUNTRY, 'Total rows', len(out_export)),
    (COUNTRY, 'Per creator (cap)', N_PER_CREATOR),
    ('Both', 'Model', MODEL),
    ('Both', 'Seed', SEED),
    ('Both', 'Themes vocabulary', ', '.join(THEMES)),
    ('Both', 'Sentiment values', ', '.join(SENTIMENTS)),
    ('Both', 'Emotion values', ', '.join(EMOTIONS)),
    ('Both', 'Tone values', ', '.join(TONES)),
    ('Both', 'Normative Orientation values', ', '.join(NORMATIVE_ORIENTATIONS)),
    ('Both', 'Target of Claim values', ', '.join(TARGETS)),
]:
    mws.append([str(x) for x in row_data])
mws.column_dimensions['A'].width = 12
mws.column_dimensions['B'].width = 28
mws.column_dimensions['C'].width = 80
for cell in mws[1]: cell.font = Font(bold=True)

# preserve sheet order
order_pref = ['Nigeria - LLM Coding', 'Kenya - LLM Coding', 'Methodology']
ordered = [wb[n] for n in order_pref if n in wb.sheetnames]
wb._sheets = ordered + [s for s in wb._sheets if s not in ordered]

wb.save(OUT_XLSX)
print(f'wrote {{OUT_XLSX}} ({{OUT_XLSX.stat().st_size:,}} bytes)')
print(f'sheets: {{[s.title for s in wb.worksheets]}}')
"""
    ))

    # ── notes ───────────────────────────────────────────────────────────────
    cells.append(md(
f"""## Notes

- Cache is keyed by SHA-1 of the segment text and lives at `temp/{cache_filename}`. Re-running is free.
- To force a re-code, delete the cache file or change the prompt.
- Sample is up to 50 per creator (capped at availability) — smaller pools take everything.
- Same 13-theme + Sentiment + Emotion + Tone vocabularies as the audience workbook so the two are
  dimensionally comparable across content and audience.
- All answers are validated against closed vocabularies before export. Gating violations
  (Q1/Q1a, Q4/Q5/Q6/Q7, Q6/Q7, Q18/Q18a, "Other" open-text fields) are auto-corrected and
  counted in `fixlog`.
"""
    ))

    return {
        "cells": cells,
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "version": "3.11"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }


def main():
    out_nigeria = ROOT / "Nigeria" / "Notebooks" / "LLM Coding Notebook - Content Analysis.ipynb"
    out_kenya   = ROOT / "Kenya"   / "Notebooks" / "LLM Coding Notebook - Content Analysis.ipynb"
    out_kenya.parent.mkdir(parents=True, exist_ok=True)

    for country, out_path in [("Nigeria", out_nigeria), ("Kenya", out_kenya)]:
        nb = build_notebook(country)
        with open(out_path, "w") as f:
            json.dump(nb, f, indent=1, ensure_ascii=False)
        print(f"wrote {out_path} ({out_path.stat().st_size:,} bytes, {len(nb['cells'])} cells)")


if __name__ == "__main__":
    main()
