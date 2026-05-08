"""
Final pass on the Nigeria human-coder spreadsheets:

  1. Polish formatting/spacing on every coder file:
       - DIRECTIONS row 1  : wrap text, vertical center, height 50
       - Header row 2      : wrap text, center align, height 60
       - Missing col widths: audience C (URL) and content C (text)
       - Q column widths   : 28 (default ~10 was too cramped)
       - Blank separator rows 33-34 and 40-41: explicit height 18
       - Freeze panes      : A3
  2. Rename every coder file to the manager-facing convention:
       Human {A..F} - Audience Analysis Codebook.xlsx
       Human {A..F} - Content Analysis Codebook.xlsx
  3. Remove the _backup_pre_overlap folder (user already done with it).
  4. Remove the stale `*_unique_assignment_tracker.xlsx` trackers
     (those came from the original build script; they only contain the
     180 single-coded entries and are now superseded by the
     `*_full_assignment_tracker.xlsx` files which include overlap).
"""

from __future__ import annotations
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

ROOT = Path("/Users/sushildalavi/Desktop/NLC/Gates-Manfluencer-Project")
BASE = ROOT / "Codebooks" / "Human Codebooks"

CODERS = ["A", "B", "C", "D", "E", "F"]

TYPE_SPEC = {
    "audience": {
        "subdir": "audience",
        "data_cols": 4,         # ID, Commenter URL, OG URL, Comment Text
        "data_widths": {1: 14, 2: 60, 3: 60, 4: 80},
    },
    "content": {
        "subdir": "content",
        "data_cols": 3,         # ID, Context, Content Text/Description
        "data_widths": {1: 14, 2: 80, 3: 80},
    },
}

Q_COL_WIDTH = 28
DIRECTIONS_ROW_HEIGHT = 50
HEADER_ROW_HEIGHT = 60
BLANK_SEPARATOR_HEIGHT = 18

NEW_NAME = "Human {letter} - {Type} Analysis Codebook.xlsx"


def polish_file(path: Path, typ: str) -> None:
    spec = TYPE_SPEC[typ]
    wb = load_workbook(path)
    ws = wb.active
    n_cols = ws.max_column

    # 1. DIRECTIONS row (row 1)
    c = ws.cell(1, 1)
    c.alignment = Alignment(wrap_text=True, vertical="center",
                            horizontal="left", indent=1)
    ws.row_dimensions[1].height = DIRECTIONS_ROW_HEIGHT

    # 2. Header row (row 2) - wrap+center on EVERY header cell.
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(2, col_idx)
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal="center")
    ws.row_dimensions[2].height = HEADER_ROW_HEIGHT

    # 3. Column widths.
    # Data cols (per spec).
    for col_idx, width in spec["data_widths"].items():
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width
    # Q cols (everything after the data cols).
    for col_idx in range(spec["data_cols"] + 1, n_cols + 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = Q_COL_WIDTH

    # 4. Blank separator rows -- explicit small height (rows 33,34,40,41).
    for r in (33, 34, 40, 41):
        ws.row_dimensions[r].height = BLANK_SEPARATOR_HEIGHT

    # 5. Freeze panes.
    ws.freeze_panes = "A3"

    wb.save(path)


def rename_file(path: Path, typ: str, letter: str) -> Path:
    type_word = "Audience" if typ == "audience" else "Content"
    new_name = NEW_NAME.format(letter=letter, Type=type_word)
    new_path = path.parent / new_name
    if path != new_path:
        path.rename(new_path)
    return new_path


def main() -> None:
    print(f"Working in: {BASE.relative_to(ROOT)}\n")

    # ---- step 1+2: polish formatting and rename --------------------------
    for typ, spec in TYPE_SPEC.items():
        sub = BASE / spec["subdir"]
        print(f"[{typ}]  {sub.relative_to(ROOT)}")
        for letter in CODERS:
            old = sub / f"{typ}_nigeria_coder_{letter}_unique.xlsx"
            if not old.exists():
                print(f"  MISSING: {old.name}")
                continue
            polish_file(old, typ)
            new = rename_file(old, typ, letter)
            print(f"  polished + renamed -> {new.name}")
        print()

    # ---- step 3: remove backup folder ------------------------------------
    backup = BASE / "_backup_pre_overlap"
    if backup.exists():
        shutil.rmtree(backup)
        print(f"removed backup folder: {backup.relative_to(ROOT)}")
    else:
        print("backup folder already removed")

    # ---- step 4: remove stale unique trackers ----------------------------
    for stale in ("audience_nigeria_unique_assignment_tracker.xlsx",
                  "content_nigeria_unique_assignment_tracker.xlsx"):
        p = BASE / stale
        if p.exists():
            p.unlink()
            print(f"removed stale tracker: {stale}")

    # ---- final inventory --------------------------------------------------
    print("\nFinal inventory:")
    for sub in sorted(BASE.iterdir()):
        if sub.is_dir():
            for f in sorted(sub.iterdir()):
                if f.suffix == ".xlsx":
                    print(f"  {sub.name}/{f.name}")
        elif sub.suffix == ".xlsx":
            print(f"  {sub.name}")


if __name__ == "__main__":
    main()
