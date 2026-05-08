"""Final polish for the LLM Coding xlsx — ready for Google Sheets upload.

- Remove Run Info sheet
- Per-sheet header styling (bold, dark blue fill, white font, wrap text, 100px tall)
- Per-column widths tuned to content type
- Wrap text on all data cells, vertical=top
- Sensible row heights (default 45, taller for known-long-text columns)
- Freeze panes at E2 (Comment Text stays visible while scrolling)
- Center-align boolean/short-value columns; left-align text columns
- Methodology sheet also styled
- Idempotent: safe to re-run
"""
from __future__ import annotations
import hashlib
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Audience Analysis.xlsx"

# ─── widths per column-name pattern ─────────────────────────────────────────

def col_width(col_name: str) -> int:
    if not isinstance(col_name, str): return 18
    n = col_name.strip()
    n_lower = n.lower()

    # exact metadata
    if n == 'Comment ID':            return 14
    if n == 'Commenter Post URL':    return 36
    if n == "Influencer's OG Post URL": return 36
    if n == 'Comment Text':          return 60

    # front LLM summary
    if n in ('Primary Theme', 'Secondary Theme 1', 'Secondary Theme 2'):  return 24
    if n == 'Masculinity Identity':  return 14
    if n == 'Normative Orientation': return 16
    if n == 'Target of Claim':       return 22
    if n == 'Sentiment':             return 12
    if n == 'Emotion Detection':     return 16
    if n == 'Tone':                  return 14

    # Q-columns — match by question number prefix
    short_q = ('Q1.','Q4.','Q12.','Q13.','Q14.','Q15.','Q16.','Q17.','Q18.','Q19.','Q20.',
               'Q21.','Q21a.','Q21c.','Q21e.','Q21g.')   # mostly Yes/No
    long_select = ('Q3.','Q7.','Q10.')                    # multi-select with long options
    medium_select = ('Q2.','Q5.','Q6.','Q8.','Q21h.')      # single-pick longer options
    open_text = ('Q7A','Q9.','Q11.','Q14a','Q15a','Q16a','Q17a','Q18a','Q19a','Q20a','Q21b','Q21d','Q21f')

    for p in short_q:
        if n.startswith(p): return 16
    for p in long_select:
        if n.startswith(p): return 36
    for p in medium_select:
        if n.startswith(p): return 22
    for p in open_text:
        if n.startswith(p): return 38

    return 18

def is_short_value_col(col_name: str) -> bool:
    """columns that hold one-word values like Yes/No/Positive/Negative — center-align"""
    if not isinstance(col_name, str): return False
    if col_name in ('Sentiment','Emotion Detection','Tone','Masculinity Identity',
                    'Normative Orientation','Target of Claim'): return True
    short_prefixes = ('Q1.','Q4.','Q5.','Q6.','Q8.','Q12.','Q13.','Q14.','Q15.','Q16.','Q17.',
                      'Q18.','Q19.','Q20.','Q21.','Q21a.','Q21c.','Q21e.','Q21g.','Q21h.')
    return any(col_name.startswith(p) for p in short_prefixes)

# ─── styling primitives ──────────────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')      # deep blue
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
DATA_FONT   = Font(size=10, name='Calibri')
ZEBRA_FILL  = PatternFill('solid', fgColor='F2F2F2')      # very light gray
THIN_BORDER = Border(left=Side(style='thin', color='D9D9D9'),
                     right=Side(style='thin', color='D9D9D9'),
                     top=Side(style='thin', color='D9D9D9'),
                     bottom=Side(style='thin', color='D9D9D9'))

def style_data_sheet(ws):
    n_rows = ws.max_row
    n_cols = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, n_cols+1)]

    # column widths
    for ci, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_width(h)

    # header row
    ws.row_dimensions[1].height = 105
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical='center',
                                horizontal='left', indent=0)
        c.border = THIN_BORDER

    # data cells
    for ridx in range(2, n_rows+1):
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=ridx, column=ci)
            cell.font = DATA_FONT
            horiz = 'center' if is_short_value_col(h) else 'left'
            cell.alignment = Alignment(wrap_text=True, vertical='top',
                                       horizontal=horiz)
            cell.border = THIN_BORDER
            if ridx % 2 == 0:
                cell.fill = ZEBRA_FILL

    # row heights — generous default to accommodate wrapped text
    for r in range(2, n_rows+1):
        ws.row_dimensions[r].height = 60

    # freeze: row 1 + first 4 columns (Comment ID / URL / URL / Comment Text)
    ws.freeze_panes = 'E2'

def style_methodology_sheet(ws):
    n_rows = ws.max_row
    n_cols = ws.max_column

    # widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 90

    ws.row_dimensions[1].height = 28
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        c.border = THIN_BORDER

    for ridx in range(2, n_rows+1):
        for ci in range(1, n_cols+1):
            cell = ws.cell(row=ridx, column=ci)
            cell.font = DATA_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = THIN_BORDER
            if ridx % 2 == 0:
                cell.fill = ZEBRA_FILL
        ws.row_dimensions[ridx].height = 24

    ws.freeze_panes = 'A2'

# ─── apply ───────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX)
print(f"loaded; sheets: {wb.sheetnames}")

# 1. Remove Run Info if present
if 'Run Info' in wb.sheetnames:
    del wb['Run Info']
    print("  removed 'Run Info'")

# 2. Style data sheets
for sn in ['Nigeria - LLM Coding', 'Kenya - LLM Coding']:
    if sn in wb.sheetnames:
        style_data_sheet(wb[sn])
        print(f"  styled '{sn}'")

# 3. Style Methodology
if 'Methodology' in wb.sheetnames:
    style_methodology_sheet(wb['Methodology'])
    print(f"  styled 'Methodology'")

# 4. Order sheets: Nigeria, Kenya, Methodology
order_pref = ['Nigeria - LLM Coding', 'Kenya - LLM Coding', 'Methodology']
ordered = [wb[n] for n in order_pref if n in wb.sheetnames]
wb._sheets = ordered + [s for s in wb._sheets if s not in ordered]

wb.save(XLSX)
print(f"\nwrote {XLSX} ({XLSX.stat().st_size:,} bytes)")
print(f"final sheets: {[s.title for s in wb.worksheets]}")
print(f"sha-256: {hashlib.sha256(open(XLSX,'rb').read()).hexdigest()[:12]}")
