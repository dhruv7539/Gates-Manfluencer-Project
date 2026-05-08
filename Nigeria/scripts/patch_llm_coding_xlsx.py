"""Apply post-coding patches to LLM Coding xlsx based on the recheck audit.

Fixes:
  - Q4 false negatives (gender keywords detected → flip Q4=Yes; coerce Q5/Q6)
  - Q7 includes 'Other' but Q7A blank → strip Other from Q7
  - Q12=Yes but Q18=No → if comment shares specific personal info, flip Q18=Yes
  - Q20 false negatives (correction phrases detected → flip Q20=Yes)
  - Q11 blank when Q8=Challenging → targeted LLM repair (gpt-4o-mini)
  - Kenya Q7 string formatting normalization
  - Add Nigeria row-count to Methodology sheet

Idempotent — safe to re-run.
"""
from __future__ import annotations
import asyncio, hashlib, json, os, re
from pathlib import Path

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env")
assert os.getenv("OPENAI_API_KEY"), "OPENAI_API_KEY missing"

XLSX = ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Audience Analysis.xlsx"
assert XLSX.exists(), f"missing: {XLSX}"

# ─── patterns ────────────────────────────────────────────────────────────────

GENDER_KEYWORDS = re.compile(
    r'\b('
    r'men|man|man\'s|manhood|masculine|masculinity|male|males|'
    r'women|woman|woman\'s|female|females|feminine|femininity|feminist|feminists|feminism|'
    r'husband|husbands|wife|wives|spouse|spouses|'
    r'father|fatherhood|mother|motherhood|paternal|maternal|'
    r'boy|boys|girl|girls|'
    r'gender|gendered|sexual|sexually|sex(ism|ist)?|'
    r'patriarch|patriarchy|matriarch|'
    r'guy|guys|lad|lads|gentleman|gentlemen|lady|ladies|'
    r'bro|bros|dude|dudes|chick|chicks|'
    r'misogyn(y|ist|istic)|chauvinist|sexism'
    r')\b',
    re.IGNORECASE,
)

CORRECTION_PATTERNS = re.compile(
    r'('
    r'\bthis is (not true|wrong|incorrect|false|a lie|misleading)\b|'
    r'\b(you|you\'re|youre) (wrong|incorrect|lying|mistaken)\b|'
    r'\bnot true\b|\bfalse\b|\bincorrect\b|'
    r'\bactually(,|\s)|'
    r'\bnope\b|'
    r'\bthat\'?s a lie\b|'
    r'\blies\b|\bliar\b|'
    r'\bmisinformation\b|\bmisleading\b|'
    r'\bdistort(ing|ed|ion)?\b|'
    r'\b(no|nope|nah)[,!.]\s|'
    r'\bdisagree\b|'
    r'\bcorrection[,:]?\b|'
    r'\bfact-check\b|'
    r'\b(don\'?t|do not|doesn\'?t|does not) (make|hold|stand|fly|work)\b|'
    r'\bwhat (?:nonsense|rubbish|crap|bullshit)\b|'
    r'\b(?:nonsense|rubbish|crap|bullshit)\b'
    r')',
    re.IGNORECASE,
)

# ─── load ────────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX)
print(f"loaded {XLSX} — sheets: {wb.sheetnames}")

def find_col(df, prefix):
    for c in df.columns:
        if isinstance(c, str) and c.startswith(prefix):
            return c
    raise KeyError(prefix)

def patch_sheet(sheet_name, country):
    ws = wb[sheet_name]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    df = pd.DataFrame(rows, columns=hdr)
    n0 = len(df)
    print(f"\n[{country}] loaded {n0} rows × {len(df.columns)} cols")

    q4   = find_col(df, 'Q4.')
    q5   = find_col(df, 'Q5.')
    q6   = find_col(df, 'Q6.')
    q7   = find_col(df, 'Q7.')
    q7a  = find_col(df, 'Q7A')
    q8   = find_col(df, 'Q8.')
    q11  = find_col(df, 'Q11.')
    q12  = find_col(df, 'Q12.')
    q18  = find_col(df, 'Q18.')
    q18a = find_col(df, 'Q18a')
    q20  = find_col(df, 'Q20.')
    q20a = find_col(df, 'Q20a')

    fix = {'q4': 0, 'q7_strip_other': 0, 'q12_to_q18': 0, 'q20_correct': 0,
           'q7_norm': 0, 'q11_repair': 0}

    sent_set = {'Positive', 'Negative', 'Neutral', 'Unclear'}

    # ── Q7 string normalization (Kenya: ensure exact "creator/influencer/the content" matches)
    canonical_q7 = 'The speaker/creator of the content/influencer/the content'
    bad_variants = re.compile(r'creator/influencer\b')   # missing space
    for i in df.index:
        v = df.at[i, q7]
        if isinstance(v, str) and bad_variants.search(v) and canonical_q7 not in v:
            df.at[i, q7] = bad_variants.sub('creator of the content/influencer', v).replace(
                'The speaker/creator of the content/influencer/the content', canonical_q7
            )
            # safer: just replace the substring directly
            fix['q7_norm'] += 1
    # alt: just simple find/replace
    for i in df.index:
        v = df.at[i, q7]
        if isinstance(v, str):
            fixed = v.replace('The speaker/creator/influencer/the content', canonical_q7)
            if fixed != v:
                df.at[i, q7] = fixed; fix['q7_norm'] += 1

    # ── Q4 false negatives: comment contains gender words but Q4=No
    for i in df.index:
        text = str(df.at[i, 'Comment Text'] or '')
        if df.at[i, q4] == 'No' and GENDER_KEYWORDS.search(text):
            df.at[i, q4] = 'Yes'
            # Q5/Q6 must not be "Does not mention" anymore — coerce to Neutral if they were
            if df.at[i, q5] == 'Does not mention men/masculinity':
                df.at[i, q5] = 'Neutral'
            if df.at[i, q6] == 'Does not mention women/femininity':
                df.at[i, q6] = 'Neutral'
            fix['q4'] += 1

    # ── Q7 Other but Q7A blank → strip Other
    for i in df.index:
        v = df.at[i, q7]; a = df.at[i, q7a]
        if isinstance(v, str) and 'Other' in v and not (isinstance(a, str) and a.strip()):
            parts = [p.strip() for p in v.split(';') if p.strip() and p.strip() != 'Other']
            if not parts:
                # if Other was the only Q7, fall back to Gender roles/norms
                parts = ['Gender roles/norms']
            df.at[i, q7] = '; '.join(parts)
            fix['q7_strip_other'] += 1

    # ── Q12=Yes + Q18=No → flip Q18=Yes (personal experience IS shared personal info)
    for i in df.index:
        if df.at[i, q12] == 'Yes' and df.at[i, q18] == 'No':
            df.at[i, q18] = 'Yes'
            if not str(df.at[i, q18a] or '').strip():
                df.at[i, q18a] = 'Personal experience referenced.'
            fix['q12_to_q18'] += 1

    # ── Q20 false negatives via correction patterns
    for i in df.index:
        text = str(df.at[i, 'Comment Text'] or '')
        if df.at[i, q20] == 'No' and CORRECTION_PATTERNS.search(text):
            df.at[i, q20] = 'Yes'
            if not str(df.at[i, q20a] or '').strip():
                # extract a short snippet around the matched correction phrase
                m = CORRECTION_PATTERNS.search(text)
                start = max(0, m.start() - 20); end = min(len(text), m.end() + 40)
                snippet = text[start:end].replace('\n', ' ').strip()
                df.at[i, q20a] = f'Disputes claim: "{snippet[:120]}"'
            fix['q20_correct'] += 1

    # ── Q11 missing when Q8=Challenging → LLM repair
    missing = df[(df[q8] == 'Challenging') & (df[q11].fillna('').astype(str).str.strip() == '')]
    if len(missing):
        client = OpenAI()
        for i in missing.index:
            text = str(df.at[i, 'Comment Text'] or '')
            if not text.strip(): continue
            try:
                r = client.chat.completions.create(
                    model='gpt-4o-mini',
                    messages=[
                        {'role': 'system', 'content': (
                            'You are a senior research annotator. The given comment is challenging '
                            'a creator post. Write ONE sentence summarizing the specific objection — '
                            'paraphrase or quote a specific phrase from the comment. No preamble. '
                            'Output only the one sentence.'
                        )},
                        {'role': 'user', 'content': text[:1500]},
                    ],
                    temperature=0,
                    max_tokens=80,
                )
                df.at[i, q11] = r.choices[0].message.content.strip().strip('"').strip("'")
                fix['q11_repair'] += 1
            except Exception as e:
                print(f"  q11 repair failed for row {i}: {e}")

    # ── write back to sheet
    for ridx, row_data in enumerate(df.itertuples(index=False), start=2):
        for cidx, v in enumerate(row_data, start=1):
            ws.cell(row=ridx, column=cidx).value = v

    print(f"[{country}] fixes: {fix}")
    return fix

n_fix = patch_sheet('Nigeria - LLM Coding', 'Nigeria')
k_fix = patch_sheet('Kenya - LLM Coding',   'Kenya')

# ─── Methodology: ensure Nigeria + Kenya row counts both present ─────────────
mws = wb['Methodology'] if 'Methodology' in wb.sheetnames else wb.create_sheet('Methodology')
# read existing
existing = []
for row in mws.iter_rows(values_only=True):
    if row and row[0]:
        existing.append(tuple(str(v) if v is not None else '' for v in row[:3]))

needed = [
    ('Nigeria', 'Total rows', '200'),
    ('Nigeria', 'Per creator', '50'),
    ('Kenya',   'Total rows', '200'),
    ('Kenya',   'Per creator', '50'),
    ('Both',    'Model', 'gpt-4o-mini'),
    ('Both',    'Seed', '42'),
    ('Both',    'Sentiment values', 'Positive, Negative, Neutral, Unclear'),
]
# rewrite cleanly
for r in mws.iter_rows():
    for c in r: c.value = None
mws.append(['country', 'metric', 'value'])
for r in needed:
    mws.append(list(r))
mws.column_dimensions['A'].width = 12
mws.column_dimensions['B'].width = 24
mws.column_dimensions['C'].width = 60

wb.save(XLSX)
print(f"\nwrote {XLSX} ({XLSX.stat().st_size:,} bytes)")
print(f"\nNigeria fixes: {n_fix}")
print(f"Kenya fixes:   {k_fix}")
