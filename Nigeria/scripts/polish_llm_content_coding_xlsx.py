"""
Post-process LLM Coding - Content Analysis.xlsx after both notebooks have run.

Fixes the Methodology sheet (which gets clobbered by whichever notebook runs last)
and applies Google-Sheets-friendly formatting consistent with the audience workbook.

Run AFTER both:
    Nigeria/Notebooks/LLM Coding Notebook - Content Analysis.ipynb
    Kenya/Notebooks/LLM Coding Notebook  - Content Analysis.ipynb

Usage:
    python Nigeria/scripts/polish_llm_content_coding_xlsx.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Content Analysis.xlsx"

THEMES = [
    "Authority and Submission", "Male Victimhood", "Gender Grievance",
    "Sexual Morality", "Relationship Tactics", "Provider and Status",
    "Male Accountability", "Egalitarian Partnership",
    "Gender-Based Violence and Consent", "Trauma and Mental Health",
    "Self-Discipline", "Marriage and Family", "Faith and Moral Repair",
    "Unclear",
]
SENTIMENTS = ["Positive", "Negative", "Neutral", "Unclear"]
EMOTIONS = ["Joy", "Happiness", "Surprise", "Anger", "Fear", "Contempt",
            "Sadness", "Hope", "Empathy", "None of these"]
TONES = ["Earnest", "Sarcastic", "Hostile", "Humorous",
         "Empathetic", "Authoritative", "Detached"]
NORMATIVE_ORIENTATIONS = ["Progressive", "Regressive", "Mixed", "Unclear"]
TARGETS = ["Men/boys", "Women/girls", "Feminists/modern women",
           "Children/family", "Institutions/law/society", "Creator/content",
           "Self/personal life", "Mixed", "Unclear"]


def main():
    assert XLSX.exists(), f"missing: {XLSX}"
    wb = openpyxl.load_workbook(XLSX)

    # count rows per country sheet
    counts = {}
    for country in ("Nigeria", "Kenya"):
        sheet = f"{country} - LLM Coding"
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            # subtract header row, count non-empty Content ID cells
            n = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0])
            counts[country] = n

    # rebuild methodology
    if "Methodology" in wb.sheetnames:
        del wb["Methodology"]
    mws = wb.create_sheet("Methodology")
    mws.append(["country", "metric", "value"])
    for country, n in counts.items():
        mws.append([country, "Total rows", str(n)])
        mws.append([country, "Per creator (cap)", "50"])
    for k, v in [
        ("Model", "gpt-4o-mini"),
        ("Seed", "42"),
        ("Themes vocabulary", ", ".join(THEMES)),
        ("Sentiment values", ", ".join(SENTIMENTS)),
        ("Emotion values", ", ".join(EMOTIONS)),
        ("Tone values", ", ".join(TONES)),
        ("Normative Orientation values", ", ".join(NORMATIVE_ORIENTATIONS)),
        ("Target of Claim values", ", ".join(TARGETS)),
    ]:
        mws.append(["Both", k, v])
    mws.column_dimensions["A"].width = 12
    mws.column_dimensions["B"].width = 28
    mws.column_dimensions["C"].width = 80
    for cell in mws[1]:
        cell.font = Font(bold=True)
    for row in mws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # restore canonical sheet order
    order_pref = ["Nigeria - LLM Coding", "Kenya - LLM Coding", "Methodology"]
    ordered = [wb[n] for n in order_pref if n in wb.sheetnames]
    wb._sheets = ordered + [s for s in wb._sheets if s not in ordered]

    wb.save(XLSX)
    print(f"polished {XLSX}")
    print(f"  counts: {counts}")
    print(f"  sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
