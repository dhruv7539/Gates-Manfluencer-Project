"""
Triple-check verification of the Nigeria human-coder spreadsheets after
overlap rows were appended.  Runs every check that matters before
sending these files out.

Reads the actual on-disk files (NOT the script's in-memory state) and
compares against the master pools and against each other.
"""

from __future__ import annotations
import sys
from pathlib import Path
from collections import Counter

import pandas as pd
from openpyxl import load_workbook

ROOT = Path("/Users/sushildalavi/Desktop/NLC/Gates-Manfluencer-Project")
AUD_MASTER = ROOT / "Nigeria" / "Audience Analysis" / "Nigeria Audience Analysis Final.xlsx"
CON_MASTER = ROOT / "Nigeria" / "Content Analysis" / "Nigeria Content Analysis Final.xlsx"
OUT_BASE = ROOT / "Codebooks" / "Human Codebooks"

CODERS = ["A", "B", "C", "D", "E", "F"]

EXPECTED_LAYOUT = {
    "directions_row": 1,
    "header_row": 2,
    "unique_start": 3, "unique_end": 32,
    "blank_1_start": 33, "blank_1_end": 34,
    "pairwise_start": 35, "pairwise_end": 39,
    "blank_2_start": 40, "blank_2_end": 41,
    "allshared_start": 42, "allshared_end": 46,
    "total_rows": 46,
}

TYPE_SPEC = {
    "audience": {
        "id_col": "Comment ID",
        "text_col": "Comment Text",
        "data_cols": ["Comment ID", "Commenter Post URL",
                      "Influencer's OG Post URL", "Comment Text"],
        "expected_cols": 41,
        "master": AUD_MASTER,
        "master_text_col": "Comment",
        "master_id_col": "Comment ID",
        "creators": ["Banky Wellington", "Deyemi Okanlawon",
                     "Agba John Doe", "Shola"],
    },
    "content": {
        "id_col": "Content ID",
        "text_col": "Content Text / Description",
        "data_cols": ["Content ID", "Context", "Content Text / Description"],
        "expected_cols": 33,
        "master": CON_MASTER,
        "master_text_col_fn": lambda df: "Verbatim Text " if "Verbatim Text " in df.columns else "Tweet",
        "master_id_col": "Content ID",
        "creators": ["Banky Wellington", "Deyemi Okanlawon", "Ebuka Obi-Uchendu",
                     "Shola", "Wizarab", "Agba John Doe"],
    },
}

ERRORS: list[str] = []
PASSES: list[str] = []


def fail(msg: str) -> None:
    ERRORS.append(msg)
    print(f"  FAIL  {msg}")


def ok(msg: str) -> None:
    PASSES.append(msg)
    print(f"  ok    {msg}")


def coder_path(typ: str, letter: str) -> Path:
    type_word = "Audience" if typ == "audience" else "Content"
    return OUT_BASE / typ / f"Human {letter} - {type_word} Analysis Codebook.xlsx"


def load_coder_file(typ: str, letter: str) -> dict:
    """Pull everything we need from one file."""
    path = coder_path(typ, letter)
    wb = load_workbook(path)
    ws = wb.active
    spec = TYPE_SPEC[typ]
    headers = [c.value for c in ws[EXPECTED_LAYOUT["header_row"]]]

    def read_block(start: int, end: int) -> list[dict]:
        rows = []
        for r in range(start, end + 1):
            d = {}
            for col in spec["data_cols"]:
                if col in headers:
                    d[col] = ws.cell(r, headers.index(col) + 1).value
            rows.append(d)
        return rows

    info = {
        "path": path,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "directions": ws.cell(EXPECTED_LAYOUT["directions_row"], 1).value,
        "headers": headers,
        "freeze": ws.freeze_panes,
        "unique": read_block(EXPECTED_LAYOUT["unique_start"],
                             EXPECTED_LAYOUT["unique_end"]),
        "blank_1": read_block(EXPECTED_LAYOUT["blank_1_start"],
                              EXPECTED_LAYOUT["blank_1_end"]),
        "pairwise": read_block(EXPECTED_LAYOUT["pairwise_start"],
                               EXPECTED_LAYOUT["pairwise_end"]),
        "blank_2": read_block(EXPECTED_LAYOUT["blank_2_start"],
                              EXPECTED_LAYOUT["blank_2_end"]),
        "allshared": read_block(EXPECTED_LAYOUT["allshared_start"],
                                EXPECTED_LAYOUT["allshared_end"]),
        "header_fills": [c.fill.fgColor.rgb if c.fill.fgColor else None
                         for c in ws[EXPECTED_LAYOUT["header_row"]]],
        "header_bolds": [c.font.bold for c in ws[EXPECTED_LAYOUT["header_row"]]],
    }
    wb.close()
    return info


def check_one_type(typ: str) -> dict:
    print(f"\n========== {typ.upper()} ==========")
    spec = TYPE_SPEC[typ]
    files = {L: load_coder_file(typ, L) for L in CODERS}

    # 1. Files exist and open.
    for L in CODERS:
        if not coder_path(typ, L).exists():
            fail(f"{typ} coder {L}: file missing")
        else:
            ok(f"{typ} coder {L}: file exists")

    # 2. Each file has exactly 46 rows and the expected column count.
    for L, info in files.items():
        if info["max_row"] != EXPECTED_LAYOUT["total_rows"]:
            fail(f"{typ} coder {L}: max_row = {info['max_row']}, expected {EXPECTED_LAYOUT['total_rows']}")
        if info["max_col"] != spec["expected_cols"]:
            fail(f"{typ} coder {L}: max_col = {info['max_col']}, expected {spec['expected_cols']}")
    ok(f"{typ}: all 6 files have {EXPECTED_LAYOUT['total_rows']} rows x {spec['expected_cols']} cols")

    # 3. Headers match across all 6 files (identical templates).
    ref_headers = files["A"]["headers"]
    for L in CODERS:
        if files[L]["headers"] != ref_headers:
            fail(f"{typ} coder {L}: headers differ from A")
    ok(f"{typ}: all 6 files share identical {len(ref_headers)}-column header row")

    # 4. DIRECTIONS row present and identical across all 6 files.
    ref_dir = files["A"]["directions"]
    if not (isinstance(ref_dir, str) and ref_dir.startswith("DIRECTIONS")):
        fail(f"{typ}: A1 doesn't start with 'DIRECTIONS'")
    for L in CODERS:
        if files[L]["directions"] != ref_dir:
            fail(f"{typ} coder {L}: DIRECTIONS row differs from A")
    ok(f"{typ}: DIRECTIONS row present and identical across all 6 files")

    # 5. Header styling: yellow fill (FFC000) and bold on data column headers.
    for L, info in files.items():
        # Coder-facing data cols (first N) should be bold + filled.
        for i, col in enumerate(spec["data_cols"]):
            if not info["header_bolds"][i]:
                fail(f"{typ} coder {L}: header col {i+1} ({col}) not bold")
        # Fill should be present (any non-None) on every header cell.
        for i, fill in enumerate(info["header_fills"]):
            if not fill:
                fail(f"{typ} coder {L}: header col {i+1} has no fill")
    ok(f"{typ}: header row is bold + filled across all 6 files")

    # 6. Freeze panes set to A3.
    for L, info in files.items():
        if info["freeze"] != "A3":
            fail(f"{typ} coder {L}: freeze_panes = {info['freeze']}, expected A3")
    ok(f"{typ}: freeze panes = A3 across all 6 files")

    # 7. Blank separator rows (33-34 and 40-41) actually blank in data cols.
    for L, info in files.items():
        for blk_name in ("blank_1", "blank_2"):
            for ri, row in enumerate(info[blk_name]):
                for col, v in row.items():
                    if v not in (None, ""):
                        fail(f"{typ} coder {L}: {blk_name} row {ri} col {col} not blank: {v!r}")
    ok(f"{typ}: blank separator rows 33-34 and 40-41 are empty in all 6 files")

    # 8. Each coder has exactly 30 unique + 5 pairwise + 5 all-shared NON-NULL rows.
    id_col = spec["id_col"]
    for L, info in files.items():
        u = sum(1 for r in info["unique"] if r[id_col])
        p = sum(1 for r in info["pairwise"] if r[id_col])
        a = sum(1 for r in info["allshared"] if r[id_col])
        if (u, p, a) != (30, 5, 5):
            fail(f"{typ} coder {L}: counts (unique, pairwise, all-shared) = ({u},{p},{a}), expected (30,5,5)")
    ok(f"{typ}: every coder has exactly 30 + 5 + 5 populated rows")

    # 9. Within each coder file, no duplicate IDs across the three blocks.
    for L, info in files.items():
        ids = [r[id_col] for r in info["unique"] + info["pairwise"] + info["allshared"]
               if r[id_col]]
        if len(ids) != len(set(ids)):
            dups = [k for k, v in Counter(ids).items() if v > 1]
            fail(f"{typ} coder {L}: duplicate IDs within file: {dups}")
    ok(f"{typ}: no duplicate IDs within any coder file")

    # 10. All-shared rows: same 5 IDs across all 6 coders, in the SAME order.
    ref_shared_ids = [r[id_col] for r in files["A"]["allshared"]]
    for L in CODERS:
        ids_L = [r[id_col] for r in files[L]["allshared"]]
        if ids_L != ref_shared_ids:
            fail(f"{typ} coder {L}: all-shared IDs differ from A: {ids_L} vs {ref_shared_ids}")
    if len(set(ref_shared_ids)) != 5:
        fail(f"{typ}: all-shared has only {len(set(ref_shared_ids))} distinct IDs, expected 5")
    ok(f"{typ}: all-shared block has same 5 IDs in same order across all 6 coders: {ref_shared_ids}")

    # 11. All-shared TEXT also identical across coders (not just IDs).
    text_col = spec["text_col"]
    ref_texts = [r[text_col] for r in files["A"]["allshared"]]
    for L in CODERS:
        texts_L = [r[text_col] for r in files[L]["allshared"]]
        if texts_L != ref_texts:
            fail(f"{typ} coder {L}: all-shared TEXT differs from A")
    ok(f"{typ}: all-shared text content identical across all 6 coders")

    # 12. Single-coded set (180 IDs, union over coders) must NOT intersect with overlap IDs.
    single_ids = set()
    for L in CODERS:
        for r in files[L]["unique"]:
            if r[id_col]:
                single_ids.add(r[id_col])
    overlap_ids = set()
    for L in CODERS:
        for r in files[L]["pairwise"] + files[L]["allshared"]:
            if r[id_col]:
                overlap_ids.add(r[id_col])
    collisions = single_ids & overlap_ids
    if collisions:
        fail(f"{typ}: COLLISION between single-coded and overlap: {collisions}")
    else:
        ok(f"{typ}: zero collisions between single-coded ({len(single_ids)}) and overlap ({len(overlap_ids)}) IDs")
    if len(single_ids) != 180:
        fail(f"{typ}: single-coded distinct IDs = {len(single_ids)}, expected 180")
    if len(overlap_ids) != 20:
        fail(f"{typ}: overlap distinct IDs = {len(overlap_ids)}, expected 20 (15 pairwise + 5 shared)")

    # 13. Each single-coded ID appears in exactly 1 coder file.
    sc_count = Counter()
    for L in CODERS:
        for r in files[L]["unique"]:
            if r[id_col]:
                sc_count[r[id_col]] += 1
    bad_sc = {k: v for k, v in sc_count.items() if v != 1}
    if bad_sc:
        fail(f"{typ}: single-coded IDs not appearing exactly once: {bad_sc}")
    else:
        ok(f"{typ}: every single-coded ID appears in exactly 1 coder file (180/180)")

    # 14. Pairwise pattern: AB=3, BC=2, CD=3, DE=2, EF=3, FA=2.
    pw_ids_by_coder = {L: set(r[id_col] for r in files[L]["pairwise"] if r[id_col])
                       for L in CODERS}
    expected_pairs = [("A","B",3),("B","C",2),("C","D",3),("D","E",2),("E","F",3),("F","A",2)]
    actual_per_pair = {}
    for left, right, exp in expected_pairs:
        shared = pw_ids_by_coder[left] & pw_ids_by_coder[right]
        actual_per_pair[(left, right)] = len(shared)
        if len(shared) != exp:
            fail(f"{typ}: pair {left}+{right} share {len(shared)} rows, expected {exp}")
    ok(f"{typ}: pairwise distribution {actual_per_pair} matches expected pattern")

    # 15. Each pairwise ID is in exactly 2 coder files.
    pw_count = Counter()
    for L in CODERS:
        for id_ in pw_ids_by_coder[L]:
            pw_count[id_] += 1
    bad_pw = {k: v for k, v in pw_count.items() if v != 2}
    if bad_pw:
        fail(f"{typ}: pairwise IDs not in exactly 2 files: {bad_pw}")
    else:
        ok(f"{typ}: every pairwise ID appears in exactly 2 coder files (15/15)")

    # 16. Each all-shared ID is in exactly 6 coder files.
    sh_count = Counter()
    for L in CODERS:
        for r in files[L]["allshared"]:
            if r[id_col]:
                sh_count[r[id_col]] += 1
    bad_sh = {k: v for k, v in sh_count.items() if v != 6}
    if bad_sh:
        fail(f"{typ}: all-shared IDs not in exactly 6 files: {bad_sh}")
    else:
        ok(f"{typ}: every all-shared ID appears in exactly 6 coder files (5/5)")

    # 17. Master fidelity: every overlap ID actually exists in master AND its
    # text matches the master text (no rewriting/synthetic rows).
    master_index: dict[str, dict] = {}
    for creator in spec["creators"]:
        df = pd.read_excel(spec["master"], sheet_name=creator)
        if "master_text_col_fn" in spec:
            tcol = spec["master_text_col_fn"](df)
        else:
            tcol = spec["master_text_col"]
        for _, mrow in df.iterrows():
            master_index[mrow[spec["master_id_col"]]] = {
                "text": mrow[tcol], "creator": creator,
            }

    text_mismatches = []
    missing = []
    for L in CODERS:
        for r in files[L]["pairwise"] + files[L]["allshared"]:
            id_ = r[id_col]
            if not id_:
                continue
            if id_ not in master_index:
                missing.append((L, id_))
                continue
            file_text = r[text_col]
            master_text = master_index[id_]["text"]
            if str(file_text).strip() != str(master_text).strip():
                text_mismatches.append((L, id_))
    if missing:
        fail(f"{typ}: overlap IDs not in master: {missing[:10]}{'...' if len(missing)>10 else ''}")
    if text_mismatches:
        fail(f"{typ}: overlap rows with text NOT matching master: {text_mismatches[:10]}{'...' if len(text_mismatches)>10 else ''}")
    if not missing and not text_mismatches:
        ok(f"{typ}: every overlap row exists in master AND text matches master verbatim")

    # 18. Single-coded text fidelity.
    sc_mismatches = []
    sc_missing = []
    for L in CODERS:
        for r in files[L]["unique"]:
            id_ = r[id_col]
            if not id_:
                continue
            if id_ not in master_index:
                sc_missing.append((L, id_))
                continue
            if str(r[text_col]).strip() != str(master_index[id_]["text"]).strip():
                sc_mismatches.append((L, id_))
    if sc_missing:
        fail(f"{typ}: single-coded IDs not in master: {sc_missing[:10]}")
    if sc_mismatches:
        fail(f"{typ}: single-coded rows with text NOT matching master: {sc_mismatches[:10]}")
    if not sc_missing and not sc_mismatches:
        ok(f"{typ}: every single-coded row exists in master AND text matches master verbatim")

    # 19. Q columns are blank for the coder to fill (audience: Q1..Q21h, content: Q1..Q18b).
    # All cells beyond the data_cols block on every populated data row should be empty.
    n_data_cols = len(spec["data_cols"])
    for L, info in files.items():
        for blk in ("unique", "pairwise", "allshared"):
            for ri, row in enumerate(info[blk]):
                # we already loaded only data cols into row dict, so reload from file
                pass
        # actually re-open and check
        wb = load_workbook(coder_path(typ, L), read_only=True)
        ws = wb.active
        bad = 0
        for r in (list(range(EXPECTED_LAYOUT["unique_start"], EXPECTED_LAYOUT["unique_end"]+1))
                  + list(range(EXPECTED_LAYOUT["pairwise_start"], EXPECTED_LAYOUT["pairwise_end"]+1))
                  + list(range(EXPECTED_LAYOUT["allshared_start"], EXPECTED_LAYOUT["allshared_end"]+1))):
            for c in range(n_data_cols + 1, spec["expected_cols"] + 1):
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    bad += 1
        wb.close()
        if bad:
            fail(f"{typ} coder {L}: {bad} non-blank cells in Q-column region (should be empty for coders to fill)")
    ok(f"{typ}: Q-column region (cols {n_data_cols+1}..{spec['expected_cols']}) is fully blank in all 6 files")

    # 20. Orientation balance check on overlap pool (sanity).
    # ...

    return {
        "single_ids": single_ids, "overlap_ids": overlap_ids,
        "all_shared_ids": ref_shared_ids,
        "pairwise_pattern": actual_per_pair,
    }


def check_tracker(typ: str, single_ids: set, overlap_ids: set) -> None:
    print(f"\n--- TRACKER ({typ}) ---")
    path = OUT_BASE / f"{typ}_nigeria_full_assignment_tracker.xlsx"
    if not path.exists():
        fail(f"tracker missing: {path.name}")
        return
    df = pd.read_excel(path)
    if len(df) != 240:
        fail(f"tracker {path.name}: {len(df)} rows, expected 240")
    else:
        ok(f"tracker {path.name}: 240 assignment rows")
    # Counts per assignment_type
    by_type = df.groupby("assignment_type")["original_row_id"].nunique()
    if by_type.get("single_coded", 0) != 180:
        fail(f"tracker {path.name}: single distinct = {by_type.get('single_coded',0)}, expected 180")
    if by_type.get("pairwise_overlap", 0) != 15:
        fail(f"tracker {path.name}: pairwise distinct = {by_type.get('pairwise_overlap',0)}, expected 15")
    if by_type.get("all_coder_shared", 0) != 5:
        fail(f"tracker {path.name}: shared distinct = {by_type.get('all_coder_shared',0)}, expected 5")
    ok(f"tracker {path.name}: distinct counts 180 / 15 / 5 match")
    # Tracker IDs == coder file IDs
    tracker_ids = set(df["original_row_id"])
    file_ids = single_ids | overlap_ids
    if tracker_ids != file_ids:
        only_tracker = tracker_ids - file_ids
        only_files = file_ids - tracker_ids
        fail(f"tracker {path.name}: ID mismatch with coder files. only_in_tracker={only_tracker}, only_in_files={only_files}")
    else:
        ok(f"tracker {path.name}: 200 distinct IDs match coder files exactly")


def main() -> int:
    print("Reading on-disk files and running verification...\n")
    aud_state = check_one_type("audience")
    con_state = check_one_type("content")
    check_tracker("audience", aud_state["single_ids"], aud_state["overlap_ids"])
    check_tracker("content", con_state["single_ids"], con_state["overlap_ids"])

    # Combined tracker
    print(f"\n--- COMBINED TRACKER ---")
    combo = OUT_BASE / "nigeria_full_assignment_tracker_all.xlsx"
    if combo.exists():
        df = pd.read_excel(combo)
        if len(df) == 480:
            ok(f"combined tracker: 480 rows (240 audience + 240 content)")
        else:
            fail(f"combined tracker: {len(df)} rows, expected 480")
    else:
        fail("combined tracker missing")

    print(f"\n{'='*60}")
    print(f"VERIFICATION SUMMARY")
    print(f"{'='*60}")
    print(f"  passed: {len(PASSES)}")
    print(f"  failed: {len(ERRORS)}")
    if ERRORS:
        print(f"\nFAILURES:")
        for e in ERRORS:
            print(f"  - {e}")
        return 1
    print(f"\nALL CHECKS PASSED.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
