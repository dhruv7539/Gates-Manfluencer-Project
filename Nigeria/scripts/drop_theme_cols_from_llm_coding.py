"""
Drop the 9 LLM front-summary columns from BOTH:
    Codebooks/LLM Codebook/LLM Coding - Audience Analysis.xlsx
    Codebooks/LLM Codebook/LLM Coding - Content Analysis.xlsx

Columns dropped:
    Primary Theme, Secondary Theme 1, Secondary Theme 2,
    Masculinity Identity, Normative Orientation, Target of Claim,
    Sentiment, Emotion Detection, Tone

After drop:
    Audience: 4 metadata + 37 Q-cols = 41 cols
    Content : 3 metadata + 30 Q-cols = 33 cols

Methodology sheet is also stripped of theme/sentiment/etc. vocabulary rows.

Usage:
    python Nigeria/scripts/drop_theme_cols_from_llm_coding.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
TARGETS = [
    ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Audience Analysis.xlsx",
    ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Content Analysis.xlsx",
]

DROP_COLS = {
    "Primary Theme", "Secondary Theme 1", "Secondary Theme 2",
    "Masculinity Identity", "Normative Orientation", "Target of Claim",
    "Sentiment", "Emotion Detection", "Tone",
}

DROP_METHODOLOGY_METRICS = {
    "Themes vocabulary", "Sentiment values", "Emotion values", "Tone values",
    "Normative Orientation values", "Target of Claim values",
}


def drop_cols_from_sheet(ws, header_color="305496"):
    """Rebuild the worksheet without the DROP_COLS columns, preserving styling."""
    header = [c.value for c in ws[1]]
    keep_idx = [i for i, h in enumerate(header) if h not in DROP_COLS]
    new_header = [header[i] for i in keep_idx]

    # collect data rows
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row): continue
        data.append([row[i] for i in keep_idx])

    # wipe sheet, re-write
    ws.delete_rows(1, ws.max_row)
    for col_letter, _ in list(ws.column_dimensions.items()):
        del ws.column_dimensions[col_letter]

    ws.append(new_header)
    for r in data:
        ws.append(r)

    # restyle header
    fill = PatternFill("solid", fgColor=header_color)
    font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 60

    # body styling + col widths
    for col_idx, col_name in enumerate(new_header, 1):
        letter = get_column_letter(col_idx)
        if col_name in ("Comment Text", "Content Text / Description"):
            ws.column_dimensions[letter].width = 70 if col_name.startswith("Content") else 60
        elif col_name == "Context":
            ws.column_dimensions[letter].width = 45
        elif col_name in ("Commenter Post URL", "Influencer's OG Post URL"):
            ws.column_dimensions[letter].width = 35
        elif isinstance(col_name, str) and col_name.startswith("Q"):
            first_token = col_name.split(".")[0].strip()
            if first_token.lower().endswith(("a", "b")) and first_token.lower() != "qa":
                ws.column_dimensions[letter].width = 35
            else:
                ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = 18

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(size=10)

    # freeze first 3-4 metadata cols depending on whether 'Content ID' or 'Comment ID' is at A
    first_col = new_header[0] if new_header else ""
    if first_col == "Content ID":
        ws.freeze_panes = "D2"   # 3 metadata cols
    else:
        ws.freeze_panes = "E2"   # 4 metadata cols (audience)


def strip_methodology(ws):
    """Remove rows whose 'metric' is in DROP_METHODOLOGY_METRICS."""
    keep = []
    for row in ws.iter_rows(values_only=True):
        if not row: continue
        metric = row[1] if len(row) > 1 else None
        if metric in DROP_METHODOLOGY_METRICS:
            continue
        keep.append(row)
    ws.delete_rows(1, ws.max_row)
    for r in keep:
        ws.append(list(r))
    for cell in ws[1]:
        cell.font = Font(bold=True)


def main():
    for xlsx_path in TARGETS:
        if not xlsx_path.exists():
            print(f"skip {xlsx_path.name} — not found")
            continue
        wb = openpyxl.load_workbook(xlsx_path)
        for sheet_name in list(wb.sheetnames):
            ws = wb[sheet_name]
            if "LLM Coding" in sheet_name:
                # pick the existing header color (Nigeria=blue, Kenya=orange)
                header_color = "C65911" if sheet_name.startswith("Kenya") else "305496"
                before = ws.max_column
                drop_cols_from_sheet(ws, header_color=header_color)
                after = ws.max_column
                print(f"  {sheet_name}: {before} -> {after} cols")
            elif sheet_name == "Methodology":
                strip_methodology(ws)
                print(f"  {sheet_name}: stripped theme/sentiment vocab rows")

        # canonical sheet order
        order_pref = ["Nigeria - LLM Coding", "Kenya - LLM Coding", "Methodology"]
        ordered = [wb[n] for n in order_pref if n in wb.sheetnames]
        wb._sheets = ordered + [s for s in wb._sheets if s not in ordered]

        wb.save(xlsx_path)
        print(f"wrote {xlsx_path.name}")
        print()


if __name__ == "__main__":
    main()
