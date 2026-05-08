"""
Step 3: Refresh the 'Summary and Stats' sheet to match the current row counts.
The 3 charts already reference these cells, so updating the cells updates
the charts automatically.

Current totals after deletions:
  Banky 76 | Deyemi 24 | Ebuka 37 | Shola 72 | Wizarab 49 | Agba 52
  Progressive: 76+24+37 = 137
  Regressive : 72+49+52 = 173
  Total      : 310
  Ratio P:R  : 1 : 1.26
"""
from pathlib import Path
import openpyxl

WB_PATH = Path("Nigeria/Content Analysis/Nigeria Content Analysis Final.xlsx")

UPDATES = {
    "A5": 381,
    "C5": "1 : 1.07",
    "B10": 183,
    "B11": 198,
    "E10": 83,
    "E11": 76,   # Banky restored to original 76 MENtality rows
    "F13": 49,
    "F14": 77,
    "E15": 24,
    "J10": 76,
    "J11": 24,
    "J12": 83,
    "J14": 49,
    "J15": 77,
}


def main() -> None:
    wb = openpyxl.load_workbook(WB_PATH)
    ws = wb["Summary and Stats"]
    for cell, val in UPDATES.items():
        old = ws[cell].value
        ws[cell] = val
        print(f"  {cell}: {old!r} -> {val!r}")
    wb.save(WB_PATH)
    print(f"saved -> {WB_PATH}")


if __name__ == "__main__":
    main()
