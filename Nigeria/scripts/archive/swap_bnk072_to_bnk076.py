"""
One-off swap: replace BNK_072 (a thin reaction comment) with BNK_076
(a substantive 652-char comment) in the all-shared block of every
audience coder file.  Also update the audience tracker and the combined
tracker.

BNK_072 was at row 45 (the 4th of 5 all-shared rows) in every audience
coder file.  Same position, same coder set -- just better content.
"""

from __future__ import annotations
from pathlib import Path
import re
import zipfile
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

ROOT = Path("/Users/sushildalavi/Desktop/NLC/Gates-Manfluencer-Project")
AUD_MASTER = ROOT / "Nigeria" / "Audience Analysis" / "Nigeria Audience Analysis Final.xlsx"
OUT_BASE = ROOT / "Codebooks" / "human codebook"

OLD_ID = "BNK_072"
NEW_ID = "BNK_076"
ROW_INDEX = 45  # row 45 in every audience coder file (4th of 5 all-shared)

URL_RE = re.compile(r"https?://[^\s)]+")
HYPERLINK_FONT = Font(color="0563C1", underline="single")


def fetch_new_row() -> dict:
    df = pd.read_excel(AUD_MASTER, sheet_name="Banky Wellington")
    r = df[df["Comment ID"] == NEW_ID].iloc[0]
    src = r["Source URL"]
    matches = URL_RE.findall(str(src))
    influencer_url = matches[0] if matches else None
    return {
        "Comment ID": r["Comment ID"],
        "Commenter Post URL": None,  # YouTube row -> no commenter URL
        "Influencer's OG Post URL": influencer_url,
        "Comment Text": r["Comment"],
        "Source URL": src,
        "Influencer": r["Influencer"],
        "Platform": r["Platform"],
    }


def swap_in_coder_file(path: Path, new_row: dict) -> None:
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[2]]

    def col_for(name: str) -> int:
        return headers.index(name) + 1

    wrap = Alignment(wrap_text=True, vertical="top")

    # Sanity: confirm row 45 currently holds OLD_ID before touching it.
    cur = ws.cell(ROW_INDEX, col_for("Comment ID")).value
    if cur != OLD_ID:
        raise RuntimeError(f"{path.name}: row {ROW_INDEX} has id {cur!r}, expected {OLD_ID}")

    # Comment ID
    c = ws.cell(ROW_INDEX, col_for("Comment ID"), new_row["Comment ID"])
    c.alignment = wrap

    # Commenter Post URL -> blank for this YouTube row
    c = ws.cell(ROW_INDEX, col_for("Commenter Post URL"), None)
    c.hyperlink = None
    # reset to default font so leftover blue-underline styling doesn't persist
    c.font = Font()
    c.alignment = wrap

    # Influencer's OG Post URL -> the YouTube URL, hyperlinked
    url = new_row["Influencer's OG Post URL"]
    c = ws.cell(ROW_INDEX, col_for("Influencer's OG Post URL"), url)
    if url:
        c.hyperlink = url
        c.font = HYPERLINK_FONT
    c.alignment = wrap

    # Comment Text
    c = ws.cell(ROW_INDEX, col_for("Comment Text"), new_row["Comment Text"])
    c.font = Font()
    c.alignment = wrap

    wb.save(path)


def update_tracker(path: Path, new_row: dict) -> None:
    """Replace every row where original_row_id == OLD_ID with NEW_ID.
    Also update the surrounding metadata fields so the tracker stays
    in sync with the master."""
    df = pd.read_excel(path)
    mask = df["original_row_id"] == OLD_ID
    n = int(mask.sum())
    if n == 0:
        return
    df.loc[mask, "original_row_id"] = NEW_ID
    if "Comment Text" in df.columns:
        df.loc[mask, "Comment Text"] = new_row["Comment Text"]
    if "Source URL" in df.columns:
        df.loc[mask, "Source URL"] = new_row["Source URL"]
    if "Influencer" in df.columns:
        df.loc[mask, "Influencer"] = new_row["Influencer"]
    if "Platform" in df.columns:
        df.loc[mask, "Platform"] = new_row["Platform"]
    df.to_excel(path, index=False)
    # restore header styling
    wb = load_workbook(path)
    ws = wb.active
    from openpyxl.styles import PatternFill
    fill = PatternFill("solid", fgColor="FFD966")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  tracker {path.name}: replaced {n} rows ({OLD_ID} -> {NEW_ID})")


def main() -> None:
    new_row = fetch_new_row()
    print(f"Replacing {OLD_ID} with {NEW_ID} in audience all-shared block (row {ROW_INDEX})")
    print(f"New text: {new_row['Comment Text'][:120]}...")
    print()

    # Coder files
    for L in "ABCDEF":
        path = OUT_BASE / "audience" / f"audience_nigeria_coder_{L}_unique.xlsx"
        swap_in_coder_file(path, new_row)
        print(f"  swapped row {ROW_INDEX} in {path.name}")

    # Trackers
    print()
    update_tracker(OUT_BASE / "audience_nigeria_full_assignment_tracker.xlsx", new_row)
    update_tracker(OUT_BASE / "nigeria_full_assignment_tracker_all.xlsx", new_row)

    print("\nDONE")


if __name__ == "__main__":
    main()
