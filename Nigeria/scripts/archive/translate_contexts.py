"""
Step 2: For each row containing Pidgin or Yoruba in the verbatim text,
ensure the Context column includes inline English glosses for non-English
terms so a coder unfamiliar with Nigerian languages can interpret it.
Updates the workbook in place; preserves all formatting.
"""
from __future__ import annotations

import asyncio
import json
import os
import re
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openai import AsyncOpenAI
from tqdm.asyncio import tqdm as atqdm

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env")
assert os.getenv("OPENAI_API_KEY"), "OPENAI_API_KEY missing"
WB_PATH = Path("Nigeria/Content Analysis/Nigeria Content Analysis Final.xlsx")

PIDGIN_YORUBA = re.compile(
    r"\b(dey|wey|na|wahala|abeg|sabi|pikin|oga|biko|nau|jare|knack|sef|chai|kuro|gbam|ooo|abi|una|olo[sm]ho|àgbá|agba|olosho|asoebi|iwa|kwarupt|don dey|kwarup|naa|haba|kpomo|wetin|wia|fit|wotowoto|toto|gan|para|fvck|babes?)\b",
    re.IGNORECASE,
)
YORUBA_DIACRITICS = re.compile(r"[àáèéìíòóùúṣẹọḿǹẽĩõũ]")

LLM_MODEL = "gpt-4o"
CONCURRENCY = 5

PROMPT = """You are improving CONTEXT notes for human coders on a Nigerian masculinity content analysis (Norman Lear Center / Gates Foundation).

Each row has VERBATIM TEXT (a tweet or podcast snippet, may contain Nigerian Pidgin or Yoruba) and a CURRENT CONTEXT note. Your job:

- If the verbatim text contains Pidgin or Yoruba terms, slang, or idioms that a non-Nigerian coder would not understand, REWRITE the context so it includes inline English glosses for each such term, e.g. "knack (have sex)", "abeg (please)", "dey im lane (stay in one's lane)".
- Preserve the original analytical meaning of the context. Do not change the interpretation.
- Keep the context concise: 1 sentence, <=35 words.
- If the verbatim has NO Pidgin/Yoruba/idiomatic content needing gloss, return the existing context UNCHANGED.

Return STRICT JSON: {"context": "<the final context string>"}"""


async def improve_one(client: AsyncOpenAI, sem, sid: str, text: str, ctx: str) -> tuple[str, str]:
    async with sem:
        msg = (
            f"VERBATIM TEXT:\n{text}\n\n"
            f"CURRENT CONTEXT:\n{ctx}\n\n"
            "Return JSON only."
        )
        r = await client.chat.completions.create(
            model=LLM_MODEL,
            messages=[
                {"role": "system", "content": PROMPT},
                {"role": "user", "content": msg},
            ],
            response_format={"type": "json_object"},
            temperature=0.2,
        )
        data = json.loads(r.choices[0].message.content)
        return sid, data.get("context", ctx).strip()


async def main():
    wb = openpyxl.load_workbook(WB_PATH)
    candidates = []
    sid_to_cell = {}
    for sn in wb.sheetnames:
        if sn == "Summary and Stats":
            continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            sid = ws.cell(r, 1).value
            ctx = ws.cell(r, 6).value
            txt = ws.cell(r, 7).value
            if not sid or not txt:
                continue
            text = str(txt)
            if PIDGIN_YORUBA.search(text) or YORUBA_DIACRITICS.search(text):
                candidates.append((sid, text, str(ctx) if ctx else ""))
                sid_to_cell[sid] = (sn, r)
    print(f"Translating {len(candidates)} contexts...")

    client = AsyncOpenAI()
    sem = asyncio.Semaphore(CONCURRENCY)
    tasks = [improve_one(client, sem, sid, text, ctx) for sid, text, ctx in candidates]
    results = await atqdm.gather(*tasks)

    changed = 0
    for sid, new_ctx in results:
        sn, r = sid_to_cell[sid]
        ws = wb[sn]
        old = ws.cell(r, 6).value
        if new_ctx and str(old).strip() != new_ctx.strip():
            ws.cell(r, 6).value = new_ctx
            changed += 1
            print(f"  [{sn} {sid}] -> {new_ctx[:120]}")

    wb.save(WB_PATH)
    print(f"\nUpdated {changed}/{len(candidates)} context cells.")


if __name__ == "__main__":
    asyncio.run(main())
