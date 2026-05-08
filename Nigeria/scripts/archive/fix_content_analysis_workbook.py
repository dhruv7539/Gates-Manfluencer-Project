"""
Step 1: Renumber Content IDs (preserving yellow rows), rename column to
"Content ID", clean up any blank header cells. Updates the workbook IN PLACE
preserving all row formatting (banding, yellow highlights).
"""
from pathlib import Path
import openpyxl

WB_PATH = Path("Nigeria/Content Analysis/Nigeria Content Analysis Final.xlsx")

# Yellow row positions (sheet_row, preserved_id_number) determined by inspection.
YELLOW = {
    "Banky Wellington": (4, 3),       # BNK_003
    "Deyemi Okanlawon": (15, 21),     # DEY_021
    "Ebuka Obi-Uchendu": None,
    "Shola":            (46, 49),     # SHO_049
    "Wizarab":          (11, 14),     # WIZ_014
    "Agba John Doe":    (38, 54),     # AGB_054
}

PREFIX = {
    "Banky Wellington": "BNK",
    "Deyemi Okanlawon": "DEY",
    "Ebuka Obi-Uchendu": "EBK",
    "Shola": "SHO",
    "Wizarab": "WIZ",
    "Agba John Doe": "AGB",
}

INFLUENCER_NAME = {
    "Banky Wellington": "Banky Wellington",
    "Deyemi Okanlawon": "Deyemi Okanlawon",
    "Ebuka Obi-Uchendu": "Ebuka Obi-Uchendu",
    "Shola": "Shola",
    "Wizarab": "Wizarab",
    "Agba John Doe": "Agba John Doe",
}


def renumber_plan(yellow_data_idx: int | None, yellow_num: int | None, n: int) -> list[int]:
    if yellow_data_idx is None:
        return list(range(1, n + 1))
    if yellow_num <= n:
        others = [i for i in range(1, n + 1) if i != yellow_num]
    else:
        others = list(range(1, n))
    out = []
    it = iter(others)
    for data_idx in range(1, n + 1):
        if data_idx == yellow_data_idx:
            out.append(yellow_num)
        else:
            out.append(next(it))
    return out


def main() -> None:
    wb = openpyxl.load_workbook(WB_PATH)
    for sn, prefix in PREFIX.items():
        ws = wb[sn]
        # Find last data row (column A non-empty or column G non-empty)
        last = 1
        for r in range(2, ws.max_row + 1):
            v_a = ws.cell(r, 1).value
            v_g = ws.cell(r, 7).value
            if (v_a not in (None, "", "  ", "   ")) or (v_g not in (None, "")):
                last = r
        n = last - 1

        # Rename header A1 to "Content ID"
        ws.cell(1, 1).value = "Content ID"
        # Fix B1 if blank (Agba/Shola had blanks)
        if not ws.cell(1, 2).value or not str(ws.cell(1, 2).value).strip():
            ws.cell(1, 2).value = "Influencer"

        # Renumber
        yellow = YELLOW[sn]
        if yellow:
            sheet_row, ynum = yellow
            ydata_idx = sheet_row - 1
        else:
            ydata_idx, ynum = None, None
        ids = renumber_plan(ydata_idx, ynum, n)

        for data_idx, num in enumerate(ids, start=1):
            r = data_idx + 1
            ws.cell(r, 1).value = f"{prefix}_{num:03d}"
            # Also ensure Influencer column populated (some sheets had blanks)
            if not ws.cell(r, 2).value or not str(ws.cell(r, 2).value).strip():
                ws.cell(r, 2).value = INFLUENCER_NAME[sn]

        print(f"  {sn}: {n} rows renumbered, yellow={yellow}")

    wb.save(WB_PATH)
    print(f"saved -> {WB_PATH}")


if __name__ == "__main__":
    main()
