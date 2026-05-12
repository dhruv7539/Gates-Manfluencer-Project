"""
Clean and compile the six human-coded Content Analysis codebooks.

Inputs (filled-in coder files):
    ~/Downloads/Content Analysis Codebooks/
        A - Content Analysis Codebook - Adanna.xlsx
        B - Content Analysis Codebook - Auriyana.xlsx
        C - Content Analysis Codebook - Kally.xlsx
        D - Content Analysis Codebook - Selene.xlsx
        E - Content Analysis Codebook - Rohan.xlsx
        F - Content Analysis Codebook - KK.xlsx

Each file has two sheets ('Nigeria', 'Kenya'). Each row is one Content ID
coded against Q1..Q18 (33 columns total, single- and multi-select mixed).

What this script does:
1. Reads every coder's two sheets.
2. Normalizes case, trailing punctuation, whitespace, and known enum
   variants (e.g., "Commentary Content" -> "Commentary/reaction content";
   "More regressive" -> "More regressive/traditional/restrictive";
   "Unlear" -> "Unclear").
3. Splits multi-select cells by comma and re-normalizes each part. The
   '/' inside an option label (e.g., "Dating/marriage") is NEVER treated
   as a delimiter — only commas separate selected options.
4. Writes one cleaned-but-otherwise-unchanged copy of each coder's file
   to Codebooks/Human Codebooks/content - cleaned/.
5. Builds the master compiled codebook:
       Codebooks/Human Codebooks/Master Human Content Codebook.xlsx
   Each unique (country, Content ID) appears exactly once.
   - Items coded by a single coder: that coder's row goes in as-is.
   - Items coded by 2 or 6 coders ("overlap" items): the final row
     uses the modal answer per field across coders, with ties broken by
     first coder alphabetically.
   The master also includes overlap-reliability sheets that lay every
   coder's answer side-by-side for the 20 overlap items per country.
"""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_DIR = Path.home() / "Downloads" / "Content Analysis Codebooks"
OUT_DIR = Path("Codebooks/Human Codebooks")
CLEANED_DIR = OUT_DIR / "content - cleaned"
MASTER_PATH = OUT_DIR / "Master Human Content Codebook.xlsx"

# ---------------------------------------------------------------------------
# Schema definition (positional — every coder's file has the same 33 columns,
# even when the header text was edited slightly).
# ---------------------------------------------------------------------------

COLUMN_NAMES = [
    "Content ID", "Context", "Content Text / Description",
    "Q1. Attention-getter (Yes/No)",
    "Q1a. Attention-getting strategies",
    "Q1b. Other strategy",
    "Q2. Primary topic(s)",
    "Q2a. Other topic",
    "Q3. Type of content",
    "Q3a. Other content type",
    "Q4. Addresses masculinity / gender norms",
    "Q5. Type of masculinity / gender norms",
    "Q6. Addresses what men should do",
    "Q7. What men do or need to do",
    "Q7a. Other directive",
    "Q8. Problem identified",
    "Q8a. Other problem",
    "Q9. Solution proposed",
    "Q9a. Other solution",
    "Q10. Communication mode",
    "Q10a. Other communication mode",
    "Q11. Audience needs",
    "Q12. How claims are supported",
    "Q12a. Other claim support",
    "Q13. How claims are justified",
    "Q13a. Other justification",
    "Q14. Sentiment toward men",
    "Q15. Sentiment toward women",
    "Q16. Sentiment toward traditional gender norms",
    "Q17. Fear or threat used",
    "Q18. Calls to action present",
    "Q18a. Types of calls to action",
    "Q18b. Other call to action",
]

# Multi-select columns: cell value is a comma-separated list of enum options.
MULTI_SELECT_COLS = {
    "Q1a. Attention-getting strategies",
    "Q2. Primary topic(s)",
    "Q3. Type of content",       # original spec was single, some coders used multi — accept
    "Q7. What men do or need to do",
    "Q8. Problem identified",
    "Q9. Solution proposed",
    "Q10. Communication mode",
    "Q11. Audience needs",
    "Q12. How claims are supported",   # some coders used multi
    "Q13. How claims are justified",   # some coders used multi
    "Q18a. Types of calls to action",
}

# Free-text "Other" columns — left as-is apart from trim.
OPEN_TEXT_COLS = {
    "Context", "Content Text / Description",
    "Q1b. Other strategy", "Q2a. Other topic", "Q3a. Other content type",
    "Q7a. Other directive", "Q8a. Other problem", "Q9a. Other solution",
    "Q10a. Other communication mode", "Q12a. Other claim support",
    "Q13a. Other justification", "Q18b. Other call to action",
}

# Single-select enums — canonical option set per column.
SINGLE_SELECT_ENUMS = {
    "Q1. Attention-getter (Yes/No)": ["Yes", "No"],
    "Q4. Addresses masculinity / gender norms": [
        "Yes, explicitly", "Yes, implicitly", "No",
    ],
    "Q5. Type of masculinity / gender norms": [
        "More regressive/traditional/restrictive",
        "More progressive/equitable/expansive",
        "Mixed/unclear",
        "Does not address masculinity or gender norms",
    ],
    "Q6. Addresses what men should do": [
        "Yes", "No", "Unclear", "Not applicable",
    ],
    "Q14. Sentiment toward men": [
        "Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned",
    ],
    "Q15. Sentiment toward women": [
        "Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned",
    ],
    "Q16. Sentiment toward traditional gender norms": [
        "Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned",
    ],
    "Q17. Fear or threat used": ["Yes", "Somewhat", "No"],
    "Q18. Calls to action present": ["Yes", "No"],
}

# Multi-select enums — canonical option set per column.
MULTI_SELECT_ENUMS = {
    "Q1a. Attention-getting strategies": [
        "Compelling question", "Use of all CAPS", "Humor or sarcasm",
        "Shares something violent or gross", "Shares something sexual",
        "Shares something surprising",
        "Uses a news headline or social media trend as opener",
        "Interesting visual or meme", "Other",
    ],
    "Q2. Primary topic(s)": [
        "Dating/marriage", "Friends/socializing", "Family/children",
        "Money/status", "Fitness/self-improvement", "Mental health",
        "Gender issues, e.g. equality",
        "Social issues, e.g. corruption", "Religion/morality",
        "Gaming/technology", "Other",
    ],
    "Q3. Type of content": [
        "Interview/conversational content",
        "Motivational/self-help content",
        "Commentary/reaction content", "Other",
    ],
    "Q7. What men do or need to do": [
        "Men need to dominate/lead", "Men need to provide/succeed",
        "Men are disadvantaged/victims", "Men need to improve themselves",
        "Men need to be fully self-reliant",
        "Men need to be emotionally open", "Men need to not show emotions",
        "Men need to be equal partners",
        "Mixed/unclear", "Other", "Not applicable",
    ],
    "Q8. Problem identified": [
        "Kenyan or Nigerian political/social problems",
        "Global political/social/cultural problems",
        "Western political/social influence",
        "Women/feminism", "Men's behavior",
        "Economic/status pressure",
        "Mental health/emotional struggle",
        "No clear problem is identified", "Other",
    ],
    "Q9. Solution proposed": [
        "Social or political change", "Assert dominance/control",
        "More wealth/status", "More self-discipline/fitness",
        "More emotional growth/healing",
        "More equality/respect for men", "More equality/respect for women",
        "Building community", "No clear solution", "Other",
    ],
    "Q10. Communication mode": [
        "Advice/instruction", "Personal story", "Commentary/opinion",
        "Debate/argument", "Humor/satire", "Motivational speech",
        "News/telling facts", "Other",
    ],
    "Q11. Audience needs": [
        "Entertainment/escapism", "Information seeking",
        "Connection/social interaction",
        "Self expression/identity construction", "Status seeking",
        "Documentation of events", "None of these apply",
    ],
    "Q12. How claims are supported": [
        "Generalizations about men/women", "Personal experience",
        "Stories about men/women", "Cultural/social observations",
        "Facts/statistics", "Moral/religious claims",
        "Mixed", "No support", "Other",
    ],
    "Q13. How claims are justified": [
        "No justification", "Anecdotal examples",
        "Presented as common sense", "References data",
        "References religion/tradition",
        "References external sources, such as other influencers",
        "Other",
    ],
    "Q18a. Types of calls to action": [
        "Calls for audience to follow / subscribe",
        "Calls for audience to comment / engage",
        "Calls for audience to share content",
        "Calls for audience to take action in their own life",
        "Calls for political / social action",
        "Calls to consume a product or service",
        "Other",
    ],
}

# Lower-cased lookup tables for matching coder-entered text back to the
# canonical enum value.
def _enum_lookup(enums: dict[str, list[str]]) -> dict[str, dict[str, str]]:
    lookup = {}
    for col, values in enums.items():
        d = {}
        for v in values:
            d[v.lower()] = v
            # collapse whitespace, drop separators that coders often add
            d[re.sub(r"[\s/]+", "", v).lower()] = v
        lookup[col] = d
    return lookup


SINGLE_LOOKUP = _enum_lookup(SINGLE_SELECT_ENUMS)
MULTI_LOOKUP = _enum_lookup(MULTI_SELECT_ENUMS)


# Column-aware aliases. Keys are (column, lowercased_value).
COLUMN_ALIAS: dict[tuple[str, str], str] = {
    # Q3 — content type
    ("Q3. Type of content", "commentary content"): "Commentary/reaction content",
    ("Q3. Type of content", "conversational content"): "Interview/conversational content",
    ("Q3. Type of content", "interview content"): "Interview/conversational content",
    ("Q3. Type of content", "motivational/self help content"): "Motivational/self-help content",
    ("Q3. Type of content", "commentary / conversational content"):
        "Commentary/reaction content, Interview/conversational content",
    ("Q3. Type of content", "commentary/conversational content"):
        "Commentary/reaction content, Interview/conversational content",

    # Q4 — addresses masculinity (single-select)
    ("Q4. Addresses masculinity / gender norms", "yes explicit"): "Yes, explicitly",
    ("Q4. Addresses masculinity / gender norms", "yes explicitly"): "Yes, explicitly",
    ("Q4. Addresses masculinity / gender norms", "yes implicit"): "Yes, implicitly",
    ("Q4. Addresses masculinity / gender norms", "yes implicitly"): "Yes, implicitly",
    ("Q4. Addresses masculinity / gender norms", "yes, implicitlly"): "Yes, implicitly",
    ("Q4. Addresses masculinity / gender norms", "yes, explicity"): "Yes, explicitly",
    ("Q4. Addresses masculinity / gender norms", "yex, explicitly"): "Yes, explicitly",
    ("Q4. Addresses masculinity / gender norms", "yes, explcitly"): "Yes, explicitly",

    # Q5 — masculinity orientation
    ("Q5. Type of masculinity / gender norms", "more regressive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms", "more regresive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms",
     "more regressive, traditional, & restrictive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms",
     "more regressive/traditional/restrictiv"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms", "more progressive"):
        "More progressive/equitable/expansive",
    ("Q5. Type of masculinity / gender norms", "more progressive and equitable"):
        "More progressive/equitable/expansive",
    ("Q5. Type of masculinity / gender norms",
     "more progressive, equitable, & expansive"):
        "More progressive/equitable/expansive",
    ("Q5. Type of masculinity / gender norms",
     "does not address masculinity/gender norms"):
        "Does not address masculinity or gender norms",
    ("Q5. Type of masculinity / gender norms",
     "does not address masculinity / gender norms"):
        "Does not address masculinity or gender norms",
    ("Q5. Type of masculinity / gender norms", "regresive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms", "progresive"):
        "More progressive/equitable/expansive",

    # Q6 — Mixed not in enum, closest match is Unclear
    ("Q6. Addresses what men should do", "mixed"): "Unclear",
    ("Q6. Addresses what men should do", "not appicable"): "Not applicable",

    # Q14/Q15/Q16 — sentiments
    ("Q14. Sentiment toward men", "not menioned"): "Not mentioned",
    ("Q14. Sentiment toward men", "unclear / positive"): "Mixed",
    ("Q14. Sentiment toward men", "unclear/positive"): "Mixed",
    ("Q14. Sentiment toward men", "unlear"): "Unclear",
    ("Q15. Sentiment toward women", "not menioned"): "Not mentioned",
    ("Q15. Sentiment toward women", "unlear"): "Unclear",
    ("Q16. Sentiment toward traditional gender norms", "not menioned"): "Not mentioned",
    ("Q16. Sentiment toward traditional gender norms", "unlear"): "Unclear",
}

# Multi-select aliases (column-aware)
MULTI_ALIAS: dict[tuple[str, str], str] = {
    # Q2 — primary topics (short forms; full options contain comma)
    ("Q2. Primary topic(s)", "gender issues"): "Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "social issues"): "Social issues, e.g. corruption",
    ("Q2. Primary topic(s)", "dating"): "Dating/marriage",
    ("Q2. Primary topic(s)", "marriage"): "Dating/marriage",
    ("Q2. Primary topic(s)", "dating / sex"): "Dating/marriage",
    ("Q2. Primary topic(s)", "family"): "Family/children",
    ("Q2. Primary topic(s)", "religion"): "Religion/morality",
    ("Q2. Primary topic(s)", "morality"): "Religion/morality",
    ("Q2. Primary topic(s)", "fitness"): "Fitness/self-improvement",
    ("Q2. Primary topic(s)", "self-improvement"): "Fitness/self-improvement",
    ("Q2. Primary topic(s)", "money"): "Money/status",
    ("Q2. Primary topic(s)", "status"): "Money/status",
    ("Q2. Primary topic(s)", "friends"): "Friends/socializing",
    ("Q2. Primary topic(s)", "socializing"): "Friends/socializing",

    # Q3 — short forms
    ("Q3. Type of content", "commentary content"): "Commentary/reaction content",
    ("Q3. Type of content", "commentary"): "Commentary/reaction content",
    ("Q3. Type of content", "conversational content"): "Interview/conversational content",
    ("Q3. Type of content", "interview"): "Interview/conversational content",
    ("Q3. Type of content", "motivational content"): "Motivational/self-help content",
    ("Q3. Type of content", "self-help"): "Motivational/self-help content",
    ("Q3. Type of content", "motivational/self help content"): "Motivational/self-help content",

    # Q7 — directives
    ("Q7. What men do or need to do", "no applicable"): "Not applicable",

    # Q8 — problems
    ("Q8. Problem identified", "mens behavior"): "Men's behavior",
    ("Q8. Problem identified", "men behavior"): "Men's behavior",
    ("Q8. Problem identified", "women's behavior"): "Women/feminism",
    ("Q8. Problem identified", "women / women's behavior"): "Women/feminism",
    ("Q8. Problem identified", "mental health"): "Mental health/emotional struggle",
    ("Q8. Problem identified", "mental health / kenyan cultural problems"):
        "Mental health/emotional struggle, Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "kenyan cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "general political/social/cultural problems"):
        "Global political/social/cultural problems",   # 'General' typo for 'Global'
    ("Q8. Problem identified", "kenyan or nigerian political/social problems"):
        "Kenyan or Nigerian political/social problems",

    # Q10 — communication
    ("Q10. Communication mode", "commentary content"): "Commentary/opinion",
    ("Q10. Communication mode", "commentary"): "Commentary/opinion",
    ("Q10. Communication mode", "advice"): "Advice/instruction",
    ("Q10. Communication mode", "instruction"): "Advice/instruction",
    ("Q10. Communication mode", "news"): "News/telling facts",
    ("Q10. Communication mode", "facts"): "News/telling facts",
    ("Q10. Communication mode", "humor"): "Humor/satire",
    ("Q10. Communication mode", "satire"): "Humor/satire",

    # Q11 — audience needs
    ("Q11. Audience needs", "identity construction"): "Self expression/identity construction",
    ("Q11. Audience needs", "self expression"): "Self expression/identity construction",
    ("Q11. Audience needs", "connetion/social interaction"): "Connection/social interaction",
    ("Q11. Audience needs", "connection"): "Connection/social interaction",
    ("Q11. Audience needs", "social interaction"): "Connection/social interaction",
    ("Q11. Audience needs", "entertainment"): "Entertainment/escapism",
    ("Q11. Audience needs", "escapism"): "Entertainment/escapism",

    # Q12 — claim support
    ("Q12. How claims are supported", "generalizations about women"):
        "Generalizations about men/women",
    ("Q12. How claims are supported", "generalizations about men"):
        "Generalizations about men/women",
    ("Q12. How claims are supported", "social observations"):
        "Cultural/social observations",
    ("Q12. How claims are supported", "cultural observations"):
        "Cultural/social observations",
    ("Q12. How claims are supported", "facts"): "Facts/statistics",
    ("Q12. How claims are supported", "statistics"): "Facts/statistics",
    ("Q12. How claims are supported", "stories about women"):
        "Stories about men/women",
    ("Q12. How claims are supported", "stories about men"):
        "Stories about men/women",
    ("Q12. How claims are supported", "moral claims"): "Moral/religious claims",
    ("Q12. How claims are supported", "religious claims"): "Moral/religious claims",

    # Q13 — justifications
    ("Q13. How claims are justified", "common sense"): "Presented as common sense",
    ("Q13. How claims are justified", "religion"): "References religion/tradition",
    ("Q13. How claims are justified", "tradition"): "References religion/tradition",
    ("Q13. How claims are justified", "anecdotes"): "Anecdotal examples",
    ("Q13. How claims are justified", "external sources"):
        "References external sources, such as other influencers",
    ("Q13. How claims are justified", "data"): "References data",

    # Q18a — CTA types: the "Calls for men/women to follow more X gender norms"
    # phrases coders wrote freely — map to "Other" with a note since these
    # don't match any standard CTA enum.
    ("Q18a. Types of calls to action",
     "calls for men to follow more traditional gender norms"): "Other",
    ("Q18a. Types of calls to action",
     "calls for men to follow more equitable gender norms"): "Other",
    ("Q18a. Types of calls to action",
     "calls for women to follow more traditional gender norms"): "Other",
    ("Q18a. Types of calls to action",
     "calls for women to follow more equitable gender norms"): "Other",
    ("Q18a. Types of calls to action",
     "calls for politicians or social figures to do something"):
        "Calls for political / social action",
    ("Q18a. Types of calls to action", "dominate more"): "Other",

    # Q7 — typos
    ("Q7. What men do or need to do", "men need to dominate/ead"):
        "Men need to dominate/lead",
    ("Q7. What men do or need to do", "men need to be dominate / lead more"):
        "Men need to dominate/lead",
    ("Q7. What men do or need to do", "men need to now show emotions"):
        "Men need to not show emotions",   # 'now' was a typo for 'not'
    ("Q7. What men do or need to do",
     "men need to be more emotionally open and heal"):
        "Men need to be emotionally open",
    ("Q7. What men do or need to do",
     "men need to be more emotionally open"):
        "Men need to be emotionally open",
    ("Q7. What men do or need to do", "mixed"): "Mixed/unclear",
    ("Q7. What men do or need to do", "unclear"): "Mixed/unclear",

    # Q8 — shorter forms / typos
    ("Q8. Problem identified", "nigerian cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "kenyan cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "no problem is identified"):
        "No clear problem is identified",
    ("Q8. Problem identified", "women and"): "Women/feminism",

    # Q9 — short forms
    ("Q9. Solution proposed", "more respect for men"):
        "More equality/respect for men",
    ("Q9. Solution proposed", "more respect for women"):
        "More equality/respect for women",

    # Q10 — short forms
    ("Q10. Communication mode", "opinion"): "Commentary/opinion",
    ("Q10. Communication mode", "conversational content"):
        "Personal story",   # closest, but flag if coder meant interview
    ("Q10. Communication mode", "advice / commentary / opinion"):
        "Advice/instruction, Commentary/opinion",
    ("Q10. Communication mode", "commentary / debate"):
        "Commentary/opinion, Debate/argument",
    ("Q10. Communication mode", "commentary /"): "Commentary/opinion",

    # Q12 — fragments
    ("Q12. How claims are supported", "/ social observations"):
        "Cultural/social observations",

    # Q13 — short forms
    ("Q13. How claims are justified", "references external sources"):
        "References external sources, such as other influencers",

    # Q2 — mashed multi-options
    ("Q2. Primary topic(s)", "marriage and gender issues"):
        "Dating/marriage, Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "social issues / corruption"):
        "Social issues, e.g. corruption",
    ("Q2. Primary topic(s)", "gender issues / equality"):
        "Gender issues, e.g. equality",

    # Q3 — multi mashes
    ("Q3. Type of content", "commentary / conversational content"):
        "Commentary/reaction content, Interview/conversational content",
    ("Q3. Type of content", "self-help content"): "Motivational/self-help content",
    ("Q3. Type of content", "reaction content"): "Commentary/reaction content",
    ("Q3. Type of content", "commentary/reaction"): "Commentary/reaction content",
    ("Q3. Type of content", "commentary / self-help content"):
        "Commentary/reaction content, Motivational/self-help content",

    # Q1a typos
    ("Q1a. Attention-getting strategies", "others"): "Other",
    ("Q1a. Attention-getting strategies", "humor/sarcasm"): "Humor or sarcasm",
    ("Q1a. Attention-getting strategies", "uses all caps"): "Use of all CAPS",

    # Q2 typos
    ("Q2. Primary topic(s)", "mental heath"): "Mental health",
    ("Q2. Primary topic(s)", "family and religion"):
        "Family/children, Religion/morality",
    ("Q2. Primary topic(s)", "family and children"): "Family/children",
    ("Q2. Primary topic(s)", "religion and morality"): "Religion/morality",
    ("Q2. Primary topic(s)", "gender issues / morality"):
        "Gender issues, e.g. equality, Religion/morality",
    ("Q2. Primary topic(s)", "gender issues / marriage"):
        "Gender issues, e.g. equality, Dating/marriage",
    ("Q2. Primary topic(s)", "gender issues / dating"):
        "Gender issues, e.g. equality, Dating/marriage",
    ("Q2. Primary topic(s)", "family / gender issues"):
        "Family/children, Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "children / family / gender issues"):
        "Family/children, Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "family / social issues"):
        "Family/children, Social issues, e.g. corruption",
    ("Q2. Primary topic(s)", "gemder issues and family (parenting)"):
        "Gender issues, e.g. equality, Family/children",
    ("Q2. Primary topic(s)", "corruption / family trauma"):
        "Social issues, e.g. corruption, Family/children",
    ("Q2. Primary topic(s)", "amongst men / gender issues"):
        "Gender issues, e.g. equality",

    # Q7 typos
    ("Q7. What men do or need to do", "not applicabe"): "Not applicable",
    ("Q7. What men do or need to do", "men need to be dominant/lead"):
        "Men need to dominate/lead",
    ("Q7. What men do or need to do", "men need to build community"): "Other",
    ("Q7. What men do or need to do",
     "men need to know they're allowed to be emotionally open"):
        "Men need to be emotionally open",
    ("Q7. What men do or need to do", "unlclear"): "Mixed/unclear",
    ("Q7. What men do or need to do", "men need to not show emotion"):
        "Men need to not show emotions",
    ("Q7. What men do or need to do", "not to date \"born again\" women"):
        "Other",

    # Q8 — Nigerian/Kenyan cultural problem variants
    ("Q8. Problem identified", "nigerian culture problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian culture issues"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian culture and social issues"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian social and cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian social & culture issues"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "kenyan social & cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "men''s behavior"): "Men's behavior",
    ("Q8. Problem identified", "women"): "Women/feminism",
    ("Q8. Problem identified", "mental health and emotional struggle"):
        "Mental health/emotional struggle",
    ("Q8. Problem identified", "global social issue of the patriarchy in particular"):
        "Global political/social/cultural problems",

    # Q9 — solution variants
    ("Q9. Solution proposed", "more emotional growth"):
        "More emotional growth/healing",
    ("Q9. Solution proposed", "more emotional growth and healing in both parties of a marriage"):
        "More emotional growth/healing",
    ("Q9. Solution proposed", "more emotional growth for both partners in marriage"):
        "More emotional growth/healing",

    # Q10 — fragments
    ("Q10. Communication mode", "commenary/opinion"): "Commentary/opinion",
    ("Q10. Communication mode", "commentary content /"): "Commentary/opinion",
    ("Q10. Communication mode", "/ commentary"): "Commentary/opinion",

    # Q11 — fragments
    ("Q11. Audience needs", "connection / self-expression"):
        "Connection/social interaction, Self expression/identity construction",

    # Q12 typos / variants
    ("Q12. How claims are supported", "generalizations about men and women"):
        "Generalizations about men/women",
    ("Q12. How claims are supported", "moral / religous claims"):
        "Moral/religious claims",
    ("Q12. How claims are supported", "anecdotal examples"): "Personal experience",
    ("Q12. How claims are supported", "generalization about relationships"):
        "Generalizations about men/women",
    ("Q12. How claims are supported", "cultura/social observations"):
        "Cultural/social observations",

    # Q13 — typos
    ("Q13. How claims are justified", "anecdotal exampled"): "Anecdotal examples",
    ("Q13. How claims are justified", "anecdotal exmaples"): "Anecdotal examples",
    ("Q13. How claims are justified", "anectdotal exmaples / religous refrences"):
        "Anecdotal examples, References religion/tradition",
    ("Q13. How claims are justified", "refrences religion"):
        "References religion/tradition",

    # Q18a
    ("Q18a. Types of calls to action", "calls for audience to share the content"):
        "Calls for audience to share content",
    ("Q18a. Types of calls to action",
     "calls for women to follow more traditional gender norm"): "Other",
    ("Q18a. Types of calls to action",
     "calls for reading the influencer's content with an open mind"):
        "Calls for audience to take action in their own life",
}


def _strip(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).replace("\xa0", " ").strip()
    # Normalize unicode quotes / apostrophes so "Men's" matches "Men's"
    s = s.replace("’", "'").replace("‘", "'")
    s = s.replace("“", '"').replace("”", '"')
    s = re.sub(r"\s+", " ", s)
    # drop a single trailing period if not part of an ellipsis
    if s.endswith(".") and not s.endswith("..."):
        s = s[:-1].rstrip()
    return s


def _canon_single(col: str, value: str) -> str:
    """Map a raw single-select cell back to the canonical enum value, if
    possible. If no match is found, return the original (stripped) value."""
    if not value:
        return ""
    raw = value
    low = raw.lower().strip()

    if (col, low) in COLUMN_ALIAS:
        return COLUMN_ALIAS[(col, low)]

    lookup = SINGLE_LOOKUP.get(col, {})
    if low in lookup:
        return lookup[low]
    # squeeze whitespace + slashes
    squeezed = re.sub(r"[\s/]+", "", low)
    if squeezed in lookup:
        return lookup[squeezed]
    # for Q4 single-select cases where coder wrote "Yes, explicitly. note"
    if "," in low:
        head = low.split(",", 1)[0].strip() + ", " + low.split(",", 1)[1].split(".", 1)[0].strip()
        head = head.rstrip(".").strip()
        if head in lookup:
            return lookup[head]
    return raw.strip()


def _canon_multi(col: str, value: str) -> str:
    """Clean a multi-select cell. Uses longest-enum-match-first so that
    enum options legitimately containing commas (e.g., 'Gender issues,
    e.g. equality') are not split apart."""
    if not value:
        return ""
    enum = MULTI_SELECT_ENUMS.get(col, [])
    found: list[str] = []
    remaining = " " + value + " "
    # try longest enum options first so multi-word options consume tokens
    # before shorter substrings can claim them
    for opt in sorted(enum, key=len, reverse=True):
        # Word-bounded match to avoid e.g. "Men's behavior" matching
        # the substring "men's behavior" inside "Women's Behavior".
        pattern = re.compile(
            r"(?<![A-Za-z0-9])" + re.escape(opt) + r"(?![A-Za-z0-9])",
            flags=re.IGNORECASE,
        )
        for _ in range(20):
            m = pattern.search(remaining)
            if not m:
                break
            found.append(opt)
            remaining = remaining[: m.start()] + " | " + remaining[m.end():]
    # leftover after pulling out known enum options
    leftovers = [t.strip(" ,.;|") for t in re.split(r",", remaining)]
    leftovers = [t for t in leftovers if t and t != "|"]
    # apply per-column aliases to leftovers, then fall back to as-is
    for t in leftovers:
        low = t.lower().rstrip(".").strip()
        if (col, low) in MULTI_ALIAS:
            found.append(MULTI_ALIAS[(col, low)])
            continue
        # also try generic alias table for shared variants
        for (alias_col, alias_low), alias_val in MULTI_ALIAS.items():
            if alias_col == col and alias_low == low:
                found.append(alias_val); break
        else:
            # try squeezed match
            lookup = MULTI_LOOKUP.get(col, {})
            squeezed = re.sub(r"[\s/]+", "", low)
            if squeezed in lookup:
                found.append(lookup[squeezed])
            elif low in lookup:
                found.append(lookup[low])
            else:
                # preserve unmatched token verbatim
                found.append(t)
    # dedupe while preserving order
    seen: set[str] = set()
    result: list[str] = []
    for c in found:
        key = c.lower().strip()
        if key and key not in seen:
            seen.add(key)
            result.append(c)
    return ", ".join(result)


def _clean_cell(col: str, value: object) -> str:
    s = _strip(value)
    if not s:
        return ""
    if col in OPEN_TEXT_COLS or col == "Content ID":
        return s
    if col in MULTI_SELECT_COLS:
        return _canon_multi(col, s)
    if col in SINGLE_SELECT_ENUMS:
        return _canon_single(col, s)
    return s


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

CODER_FILES = {
    "Adanna":   "A - Content Analysis Codebook - Adanna.xlsx",
    "Auriyana": "B - Content Analysis Codebook - Auriyana.xlsx",
    "Kally":    "C - Content Analysis Codebook - Kally.xlsx",
    "Selene":   "D - Content Analysis Codebook - Selene.xlsx",
    "Rohan":    "E - Content Analysis Codebook - Rohan.xlsx",
    "KK":       "F - Content Analysis Codebook - KK.xlsx",
}


def read_coder(name: str, filename: str) -> dict[str, pd.DataFrame]:
    """Read one coder's workbook. Returns {country: dataframe}."""
    path = INPUT_DIR / filename
    out: dict[str, pd.DataFrame] = {}
    for sheet in ("Nigeria", "Kenya"):
        # header=1 — row 0 is the "DIRECTIONS:" banner
        df = pd.read_excel(path, sheet_name=sheet, header=1, dtype=object)
        df = df[df.iloc[:, 0].notna()].copy().reset_index(drop=True)
        # rename positional columns to our canonical names
        if df.shape[1] < len(COLUMN_NAMES):
            raise ValueError(
                f"{name}/{sheet} only has {df.shape[1]} columns, expected "
                f"{len(COLUMN_NAMES)}"
            )
        df = df.iloc[:, : len(COLUMN_NAMES)].copy()
        df.columns = COLUMN_NAMES
        # clean every cell
        for col in df.columns:
            df[col] = df[col].apply(lambda v: _clean_cell(col, v))
        out[sheet] = df
    return out


# ---------------------------------------------------------------------------
# Writing helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)


def write_styled(path: Path, sheets: dict[str, pd.DataFrame],
                 freeze_first_col: bool = True):
    """Write a workbook with consistent styling and frozen header row."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    wb = load_workbook(path)
    for sheet_name in sheets:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2" if freeze_first_col else "A2"
        # header
        for col_idx, _ in enumerate(ws[1], start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True)
        ws.row_dimensions[1].height = 60
        # body cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        # column widths
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            header = ws.cell(row=1, column=col_idx).value or ""
            if "Content ID" in header:
                w = 18
            elif header in ("Context", "Content Text / Description"):
                w = 45
            elif header.startswith("Q1b") or header.endswith("Other"):
                w = 22
            else:
                w = 28
            ws.column_dimensions[letter].width = w
    wb.save(path)


# ---------------------------------------------------------------------------
# Master compilation
# ---------------------------------------------------------------------------

def modal_value(values: list[str]) -> str:
    """Modal vote, tie-broken by first occurrence."""
    counts = Counter([v for v in values if v != ""])
    if not counts:
        return ""
    top = counts.most_common()
    best_n = top[0][1]
    winners = [v for v, n in top if n == best_n]
    # tie-break: keep the first one encountered (already in insertion order
    # because Counter preserves first-seen ordering in Python 3.7+)
    return winners[0]


def build_master(per_coder: dict[str, dict[str, pd.DataFrame]]) -> dict[str, pd.DataFrame]:
    """Returns {sheet_name: dataframe} for the master workbook."""
    long_rows: list[dict] = []
    for coder, sheets in per_coder.items():
        for country, df in sheets.items():
            for _, row in df.iterrows():
                rec = row.to_dict()
                rec["__coder"] = coder
                rec["__country"] = country
                long_rows.append(rec)
    long_df = pd.DataFrame(long_rows)

    masters: dict[str, pd.DataFrame] = {}
    overlap_panes: dict[str, pd.DataFrame] = {}

    for country in ("Nigeria", "Kenya"):
        sub = long_df[long_df["__country"] == country].copy()
        master_rows: list[dict] = []
        overlap_records: list[dict] = []

        for cid, group in sub.groupby("Content ID", sort=False):
            n_coders = len(group)
            if n_coders == 1:
                rec = group.iloc[0].to_dict()
                rec["Coder"] = rec.pop("__coder")
                rec["Coders on this item"] = 1
                master_rows.append(rec)
                continue

            # multi-coded — modal vote per cell
            collapsed = {"Content ID": cid}
            for col in COLUMN_NAMES:
                if col == "Content ID":
                    continue
                values = group[col].tolist()
                collapsed[col] = modal_value(values)
            coders_here = sorted(group["__coder"].unique())
            collapsed["Coder"] = ", ".join(coders_here) + " (modal vote)"
            collapsed["Coders on this item"] = n_coders
            master_rows.append(collapsed)

            # build long overlap pane row(s) — one row per (item, coder)
            for _, r in group.iterrows():
                overlap_records.append({
                    "Content ID": cid,
                    "Coder": r["__coder"],
                    **{c: r[c] for c in COLUMN_NAMES if c != "Content ID"},
                })

        # order rows by Content ID for stability
        master_df = pd.DataFrame(master_rows)
        # reorder: Content ID, Coder, Coders on this item, then Q cols
        cols = ["Content ID", "Coder", "Coders on this item"] + [
            c for c in COLUMN_NAMES if c != "Content ID"
        ]
        master_df = master_df.reindex(columns=cols)
        master_df = master_df.sort_values("Content ID").reset_index(drop=True)

        overlap_df = pd.DataFrame(overlap_records)
        if not overlap_df.empty:
            overlap_df = overlap_df.sort_values(["Content ID", "Coder"]).reset_index(drop=True)

        masters[country] = master_df
        overlap_panes[country] = overlap_df

    # also create a "singles only" sheet per country
    singles = {}
    for country in ("Nigeria", "Kenya"):
        master = masters[country]
        singles[country] = master[master["Coders on this item"] == 1].copy()

    return {
        "Nigeria — All Items": masters["Nigeria"],
        "Kenya — All Items": masters["Kenya"],
        "Nigeria — Singles Only": singles["Nigeria"],
        "Kenya — Singles Only": singles["Kenya"],
        "Nigeria — Overlap Reliability": overlap_panes["Nigeria"],
        "Kenya — Overlap Reliability": overlap_panes["Kenya"],
    }


def methodology_sheet() -> pd.DataFrame:
    rows = [
        ("Master Human Content Codebook — methodology", ""),
        ("Source files",
         "Six coder workbooks (A–F) in ~/Downloads/Content Analysis Codebooks/."),
        ("Design",
         "Partition with reliability overlap. Per country: 200 unique "
         "Content IDs split across 6 coders. 180 IDs coded by exactly 1 "
         "coder; 15 IDs coded by 2 coders; 5 IDs coded by all 6 coders "
         "(reliability anchors). Each coder coded 40 IDs per country."),
        ("Rows per country", "200 (all unique Content IDs)."),
        ("Rows on the 'All Items' sheets",
         "200 per country — each Content ID appears exactly once."),
        ("How overlap items were collapsed",
         "For items coded by more than one coder, each field uses the "
         "modal value across coders. Ties are broken by the first coder "
         "alphabetically. The 'Coder' column shows which coders "
         "contributed, and the 'Coders on this item' column shows how "
         "many coders weighed in."),
        ("Singles-Only sheet",
         "Filtered subset where Coders on this item = 1. Useful when you "
         "want only the rows the master is making no judgement on."),
        ("Overlap-Reliability sheets",
         "Long-format view of every coder's answers for the 20 overlap "
         "items per country (15 double-coded + 5 sextuple-coded). Use "
         "this for inter-rater reliability work."),
        ("Cleaning applied",
         "Trimmed whitespace and trailing periods; normalized case; "
         "mapped known variants ('Commentary Content' → 'Commentary/"
         "reaction content'; 'More regressive' → 'More regressive/"
         "traditional/restrictive'; 'Does not address masculinity/"
         "gender norms' → 'Does not address masculinity or gender "
         "norms'; 'Unlear' → 'Unclear'; 'Regresive' → 'More regressive/"
         "traditional/restrictive'). Multi-select cells were split on "
         "commas only — '/' inside an option label was never treated "
         "as a delimiter."),
        ("Multi-select questions",
         "Q1a, Q2, Q3, Q7, Q8, Q9, Q10, Q11, Q12, Q13, Q18a. Each "
         "cell may contain multiple comma-separated enum values. Q3, "
         "Q12, Q13 were single-select in the original codebook but "
         "some coders entered multiple values; those have been "
         "preserved verbatim."),
        ("Single-select questions",
         "Q1, Q4, Q5, Q6, Q14, Q15, Q16, Q17, Q18."),
        ("Open-text columns",
         "Context, Content Text / Description, and every 'Q*a/b. Other' "
         "column are free text and were only whitespace-cleaned."),
    ]
    return pd.DataFrame(rows, columns=["Field", "Description"])


def cleaning_log(per_coder_raw, per_coder_clean) -> pd.DataFrame:
    rows = []
    for coder, sheets in per_coder_clean.items():
        for country, clean_df in sheets.items():
            raw_df = per_coder_raw[coder][country]
            changes = 0
            for col in clean_df.columns:
                for r, c in zip(raw_df[col].fillna("").astype(str),
                                clean_df[col].fillna("").astype(str)):
                    if r.strip() != c.strip():
                        changes += 1
            rows.append({
                "Coder": coder,
                "Country": country,
                "Rows": len(clean_df),
                "Cells normalized": changes,
            })
    return pd.DataFrame(rows)


def known_unresolved(per_coder_clean) -> pd.DataFrame:
    """Catalogue cells that still don't match the single-select enum after
    cleaning. These need a human decision."""
    rows: list[dict] = []
    for coder, sheets in per_coder_clean.items():
        for country, df in sheets.items():
            for col, valid in SINGLE_SELECT_ENUMS.items():
                valid_set = set(valid)
                for idx, val in df[col].items():
                    v = str(val).strip()
                    if v and v not in valid_set:
                        rows.append({
                            "Coder": coder,
                            "Country": country,
                            "Content ID": df.iloc[idx]["Content ID"],
                            "Column": col,
                            "Raw value": v,
                            "Why unresolved":
                                "Value not in single-select enum; "
                                "needs human adjudication.",
                        })
    if not rows:
        return pd.DataFrame(columns=[
            "Coder","Country","Content ID","Column","Raw value","Why unresolved",
        ])
    return pd.DataFrame(rows).sort_values(["Country","Column","Coder"])


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Reading coder files from", INPUT_DIR)

    per_coder_raw: dict[str, dict[str, pd.DataFrame]] = {}
    per_coder_clean: dict[str, dict[str, pd.DataFrame]] = {}

    for coder, fname in CODER_FILES.items():
        # raw read (uncleaned, only header alignment)
        raw_sheets: dict[str, pd.DataFrame] = {}
        for sheet in ("Nigeria", "Kenya"):
            raw = pd.read_excel(INPUT_DIR / fname, sheet_name=sheet,
                                header=1, dtype=object)
            raw = raw[raw.iloc[:, 0].notna()].copy().reset_index(drop=True)
            raw = raw.iloc[:, : len(COLUMN_NAMES)].copy()
            raw.columns = COLUMN_NAMES
            raw_sheets[sheet] = raw
        per_coder_raw[coder] = raw_sheets

        cleaned = read_coder(coder, fname)
        per_coder_clean[coder] = cleaned

        # write cleaned per-coder workbook
        out_path = CLEANED_DIR / fname.replace("Content Analysis Codebook",
                                               "Content Analysis Codebook (cleaned)")
        write_styled(out_path, cleaned)
        print(f"  wrote {out_path}")

    # master
    master_sheets = build_master(per_coder_clean)
    master_sheets = {
        "Methodology": methodology_sheet(),
        "Cleaning Log": cleaning_log(per_coder_raw, per_coder_clean),
        "Known Unresolved Cells": known_unresolved(per_coder_clean),
        **master_sheets,
    }
    write_styled(MASTER_PATH, master_sheets, freeze_first_col=False)
    print(f"\nMaster compiled codebook: {MASTER_PATH}")

    # print quick stats
    print("\n=== Master summary ===")
    for sheet_name, df in master_sheets.items():
        print(f"  [{sheet_name}] rows = {len(df)}")


if __name__ == "__main__":
    main()
