"""
Build the Nigeria-only single-coded / unique portion of the human-coder
spreadsheets for the Gates / NLC masculinity project.

Per analysis type (audience, content):
  - 180 high-quality, orientation-balanced Nigeria rows
  - split into 6 non-overlapping coder sets of 30
  - 15 progressive + 15 regressive per coder

Output:
  Codebooks/human codebook/audience/audience_nigeria_coder_{A..F}_unique.xlsx
  Codebooks/human codebook/content/content_nigeria_coder_{A..F}_unique.xlsx
  Codebooks/human codebook/audience_nigeria_unique_assignment_tracker.xlsx
  Codebooks/human codebook/content_nigeria_unique_assignment_tracker.xlsx
"""

from __future__ import annotations
import os
import random
import re
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font

# URL regex that excludes the wrapping ')' from the master's
# "(reply on Creator's post URL)" suffix.
URL_RE = re.compile(r"https?://[^\s)]+")
HYPERLINK_FONT = Font(color="0563C1", underline="single")
LINK_INLINE_FONT = InlineFont(color="0563C1", u="single")

ROOT = Path("/Users/sushildalavi/Desktop/NLC/Gates-Manfluencer-Project")
AUD_MASTER = ROOT / "Nigeria" / "Audience Analysis" / "Nigeria Audience Analysis Final.xlsx"
CON_MASTER = ROOT / "Nigeria" / "Content Analysis" / "Nigeria Content Analysis Final.xlsx"

OUT_BASE = ROOT / "Codebooks" / "human codebook"
OUT_AUD = OUT_BASE / "audience"
OUT_CON = OUT_BASE / "content"

SEED = 42
CODERS = ["Coder_A", "Coder_B", "Coder_C", "Coder_D", "Coder_E", "Coder_F"]
ROWS_PER_CODER = 30
TOTAL_ROWS = len(CODERS) * ROWS_PER_CODER  # 180
HALF = TOTAL_ROWS // 2  # 90 per orientation

ORIENTATION_AUD = {
    "Banky Wellington": "Progressive",
    "Deyemi Okanlawon": "Progressive",
    "Agba John Doe":    "Regressive",
    "Shola":            "Regressive",
}
ORIENTATION_CON = {
    "Banky Wellington":   "Progressive",
    "Deyemi Okanlawon":   "Progressive",
    "Ebuka Obi-Uchendu":  "Progressive",
    "Agba John Doe":      "Regressive",
    "Shola":              "Regressive",
    "Wizarab":            "Regressive",
}

# How many rows to take from each creator within its orientation.
# Sums to 90 per orientation, 180 overall, balanced.
QUOTA_AUD = {
    "Banky Wellington": 45,
    "Deyemi Okanlawon": 45,
    "Agba John Doe":    45,
    "Shola":            45,
}
QUOTA_CON = {
    "Banky Wellington":   33,
    "Deyemi Okanlawon":   24,   # max available
    "Ebuka Obi-Uchendu":  33,
    "Agba John Doe":      30,
    "Shola":              30,
    "Wizarab":            30,
}


def _split_two_urls(combined: object) -> tuple[str | None, str | None]:
    """Return (commenter_url, influencer_post_url) from the master's combined
    'Source URL' string. YouTube rows only have one URL → goes to the
    influencer slot, commenter slot stays None."""
    if not isinstance(combined, str) or not combined.strip():
        return (None, None)
    matches = URL_RE.findall(combined)
    if not matches:
        return (None, None)
    if len(matches) == 1:
        return (None, matches[0])
    return (matches[0], matches[-1])


def load_audience_pool() -> pd.DataFrame:
    """Each URL is split into its own column so each cell carries a single
    hyperlink — that's the only way the links survive a copy-paste into
    Google Sheets / another workbook (Excel's clipboard exposes one
    hyperlink per cell)."""
    frames = []
    for creator, quota in QUOTA_AUD.items():
        df = pd.read_excel(AUD_MASTER, sheet_name=creator)
        df = df.head(quota).copy()
        if len(df) < quota:
            raise RuntimeError(f"Audience: {creator} has only {len(df)} rows, need {quota}")
        urls = df["Source URL"].apply(_split_two_urls)
        df["Commenter Post URL"] = urls.apply(lambda t: t[0])
        df["Influencer's OG Post URL"] = urls.apply(lambda t: t[1])
        df["original_row_id"] = df["Comment ID"]
        df["Comment Text"] = df["Comment"]
        df["orientation"] = ORIENTATION_AUD[creator]
        frames.append(df[[
            "original_row_id", "Comment ID",
            "Commenter Post URL", "Influencer's OG Post URL",
            "Comment Text",
            "Influencer", "Platform", "Source URL", "orientation",
        ]])
    return pd.concat(frames, ignore_index=True)


def load_content_pool() -> pd.DataFrame:
    """Same shape as audience: minimal coder-facing columns
    (Content ID / Context / Content Text / Description) plus extra metadata
    kept on the DataFrame for the tracker."""
    frames = []
    for creator, quota in QUOTA_CON.items():
        df = pd.read_excel(CON_MASTER, sheet_name=creator)
        df = df.head(quota).copy()
        if len(df) < quota:
            raise RuntimeError(f"Content: {creator} has only {len(df)} rows, need {quota}")
        text_col = "Verbatim Text " if "Verbatim Text " in df.columns else "Tweet"
        df["original_row_id"] = df["Content ID"]
        df["Context"] = df["Context "]                       # strip trailing space
        df["Content Text / Description"] = df[text_col]
        df["orientation"] = ORIENTATION_CON[creator]
        frames.append(df[[
            "original_row_id", "Content ID", "Context",
            "Content Text / Description",
            "Influencer", "Platform", "Content Type", "Source URL",
            "orientation",
        ]])
    return pd.concat(frames, ignore_index=True)


def assign_to_coders(pool: pd.DataFrame, analysis_type: str) -> pd.DataFrame:
    """Within each orientation, shuffle (fixed seed) and round-robin assign to
    6 coders so each coder gets 15 progressive + 15 regressive."""
    rng = random.Random(SEED)
    assigned = []
    for orientation in ("Progressive", "Regressive"):
        sub = pool[pool["orientation"] == orientation].copy()
        idx = list(sub.index)
        rng.shuffle(idx)
        for slot, row_idx in enumerate(idx):
            coder = CODERS[slot % len(CODERS)]
            row = sub.loc[row_idx].to_dict()
            row["coder_id"] = coder
            assigned.append(row)
    out = pd.DataFrame(assigned)
    out["assignment_type"] = "single_coded"
    out["country"] = "Nigeria"
    out["analysis_type"] = analysis_type
    return out


URL_COLUMNS = ("Source URL", "Commenter Post URL", "Influencer's OG Post URL")
TEXT_COLUMNS = ("Comment", "Comment Text", "Verbatim Text ", "Tweet",
                "Content Text / Description", "Context")

AUD_CODER_COLS = ["Comment ID", "Commenter Post URL",
                  "Influencer's OG Post URL", "Comment Text"]
CON_CODER_COLS = ["Content ID", "Context", "Content Text / Description"]


def _xml_escape(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;")
             .replace(">", "&gt;").replace('"', "&quot;"))


def _inject_secondary_hyperlinks(path: Path,
                                 secondary: dict[str, list[str]]) -> None:
    """openpyxl supports only one hyperlink per cell. For cells with more URLs
    we open the saved xlsx as a zip and append extra <hyperlink> + matching
    <Relationship> entries directly. Excel uses each entry's `display`
    attribute to decide which substring of the cell text is clickable."""
    if not secondary:
        return
    extra = [(ref, url) for ref, urls in secondary.items() for url in urls]
    if not extra:
        return

    sheet_name = "xl/worksheets/sheet1.xml"
    rels_name = "xl/worksheets/_rels/sheet1.xml.rels"

    with zipfile.ZipFile(path, "r") as z:
        files = {n: z.read(n) for n in z.namelist()}

    sheet_xml = files[sheet_name].decode()
    rels_xml = files.get(
        rels_name,
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
    ).decode()

    max_rid = max(
        (int(m.group(1)) for m in re.finditer(r'Id="rId(\d+)"', rels_xml)),
        default=0,
    )

    new_hl = []
    new_rel = []
    rid = max_rid + 1
    for ref, url in extra:
        u = _xml_escape(url)
        new_hl.append(
            '<hyperlink xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            f'ref="{ref}" r:id="rId{rid}" display="{u}"/>'
        )
        new_rel.append(
            '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
            f'Target="{u}" TargetMode="External" Id="rId{rid}"/>'
        )
        rid += 1
    new_hl_xml = "".join(new_hl)
    new_rel_xml = "".join(new_rel)

    if "</hyperlinks>" in sheet_xml:
        sheet_xml = sheet_xml.replace("</hyperlinks>",
                                      new_hl_xml + "</hyperlinks>")
    else:
        block = f"<hyperlinks>{new_hl_xml}</hyperlinks>"
        if "<pageMargins" in sheet_xml:
            sheet_xml = sheet_xml.replace("<pageMargins", block + "<pageMargins")
        else:
            sheet_xml = sheet_xml.replace("</worksheet>", block + "</worksheet>")

    if "</Relationships>" in rels_xml:
        rels_xml = rels_xml.replace("</Relationships>",
                                    new_rel_xml + "</Relationships>")
    else:
        # self-closing <Relationships .../>
        rels_xml = rels_xml.replace(
            "/>", f">{new_rel_xml}</Relationships>", 1)

    files[sheet_name] = sheet_xml.encode()
    files[rels_name] = rels_xml.encode()

    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, path)


def _style_url_cell(cell, full_text: str,
                    pending: dict[str, list[str]]) -> None:
    """If the cell contains 1 URL: set cell.hyperlink, blue+underline font.
    If it contains 2+ URLs: build CellRichText with each URL styled as a link,
    set cell.hyperlink to the first, and queue the rest for XML injection."""
    matches = list(URL_RE.finditer(full_text))
    if not matches:
        cell.value = full_text
        return
    if len(matches) == 1:
        cell.value = full_text
        cell.hyperlink = matches[0].group(0)
        cell.font = HYPERLINK_FONT
        return

    parts = []
    last = 0
    urls = []
    for m in matches:
        if m.start() > last:
            parts.append(full_text[last:m.start()])
        url = m.group(0)
        parts.append(TextBlock(font=LINK_INLINE_FONT, text=url))
        urls.append(url)
        last = m.end()
    if last < len(full_text):
        parts.append(full_text[last:])

    cell.value = CellRichText(parts)
    cell.hyperlink = urls[0]
    pending.setdefault(cell.coordinate, []).extend(urls[1:])


HEADER_FILL = "FFD966"   # warm yellow, matches user's reference image
HEADER_FONT = Font(bold=True, color="000000", size=12)


def _style_header(ws) -> None:
    from openpyxl.styles import PatternFill, Border, Side
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    border_side = Side(style="thin", color="000000")
    border = Border(left=border_side, right=border_side,
                    top=border_side, bottom=border_side)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 56


def apply_hyperlinks(path: Path) -> None:
    """Style URL cells (multi-URL → rich text + secondary hyperlink injection),
    style the header row, set sensible widths + wrap-text on body cells."""
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    src_idxs = [i + 1 for i, h in enumerate(header) if h in URL_COLUMNS]

    wrap = Alignment(wrap_text=True, vertical="top")
    pending: dict[str, list[str]] = {}

    for row in ws.iter_rows(min_row=2):
        for col_idx in src_idxs:
            cell = row[col_idx - 1]
            v = cell.value
            if isinstance(v, str) and v.strip():
                _style_url_cell(cell, v, pending)
                cell.alignment = wrap
        for i, h in enumerate(header, start=1):
            if h in TEXT_COLUMNS:
                row[i - 1].alignment = wrap

    for i, h in enumerate(header, start=1):
        col = ws.cell(row=1, column=i).column_letter
        if h in URL_COLUMNS:
            ws.column_dimensions[col].width = 60
        elif h in TEXT_COLUMNS:
            ws.column_dimensions[col].width = 80
        elif h in ("Influencer", "Platform", "Content Type"):
            ws.column_dimensions[col].width = 20
        elif h in ("Comment ID", "Content ID", "original_row_id"):
            ws.column_dimensions[col].width = 14

    _style_header(ws)
    ws.freeze_panes = "A2"
    wb.save(path)
    _inject_secondary_hyperlinks(path, pending)


def write_coder_files(assigned: pd.DataFrame, out_dir: Path, prefix: str) -> None:
    """Write coder files with ONLY the 3 user-facing columns.
    Admin metadata (coder_id, country, etc.) lives in the tracker, not here —
    the filename already encodes the coder."""
    out_dir.mkdir(parents=True, exist_ok=True)
    if "Comment Text" in assigned.columns:
        cols = AUD_CODER_COLS
    else:
        cols = CON_CODER_COLS
    for coder in CODERS:
        sub = assigned[assigned["coder_id"] == coder][cols].copy()
        letter = coder.split("_")[1]
        out_path = out_dir / f"{prefix}_nigeria_coder_{letter}_unique.xlsx"
        sub.to_excel(out_path, index=False)
        apply_hyperlinks(out_path)
        print(f"  wrote {out_path.relative_to(ROOT)}  ({len(sub)} rows)")


def write_tracker(assigned: pd.DataFrame, path: Path) -> None:
    """Tracker keeps full QA metadata that's hidden from the coder files."""
    if "Comment Text" in assigned.columns:
        text_col = "Comment Text"
        ref_cols = ["Reference Text/Context"]
    else:
        text_col = "Content Text / Description"
        ref_cols = ["Context", "Content Type"]
    cols = [
        "original_row_id", "coder_id", "assignment_type", "country",
        "analysis_type", "orientation", "Influencer", "Platform",
        "Source URL", text_col, *ref_cols,
    ]
    cols = [c for c in cols if c in assigned.columns]
    tracker = assigned[cols].sort_values(["coder_id", "original_row_id"])
    path.parent.mkdir(parents=True, exist_ok=True)
    tracker.to_excel(path, index=False)
    apply_hyperlinks(path)
    print(f"  wrote {path.relative_to(ROOT)}  ({len(tracker)} rows)")


def qa(assigned: pd.DataFrame, label: str) -> None:
    print(f"\n--- QA: {label} ---")
    # 1. row count per coder
    counts = assigned.groupby("coder_id").size()
    assert (counts == ROWS_PER_CODER).all(), f"per-coder count mismatch: {counts.to_dict()}"
    print(f"  ✓ each coder file has exactly {ROWS_PER_CODER} rows")
    # 2. total rows
    assert len(assigned) == TOTAL_ROWS, f"total rows = {len(assigned)}"
    print(f"  ✓ total assigned = {TOTAL_ROWS}")
    # 3. no duplicate original_row_id across coders
    dup = assigned["original_row_id"].duplicated().sum()
    assert dup == 0, f"{dup} duplicate original_row_id values"
    print(f"  ✓ no duplicate original_row_id across coder unique sets")
    # 4. all single_coded
    assert (assigned["assignment_type"] == "single_coded").all()
    print(f"  ✓ all assignment_type values are 'single_coded'")
    # 5. orientation balance overall + per coder
    o = assigned["orientation"].value_counts().to_dict()
    print(f"  ✓ overall orientation: {o}")
    per = assigned.groupby(["coder_id", "orientation"]).size().unstack(fill_value=0)
    print("  ✓ per-coder orientation:")
    print(per.to_string().replace("\n", "\n      "))


def summary_table(aud: pd.DataFrame, con: pd.DataFrame) -> None:
    combined = pd.concat([aud, con], ignore_index=True)
    summary = (combined.groupby(
        ["analysis_type", "country", "coder_id", "assignment_type"]
    ).size().rename("rows").reset_index())
    print("\n=== FINAL SUMMARY TABLE ===")
    print(summary.to_string(index=False))
    print(f"\nGrand total: {len(combined)} rows  "
          f"(audience {len(aud)}, content {len(con)})")


def main() -> None:
    print(f"Random seed: {SEED}")

    print("\nLoading audience pool …")
    aud_pool = load_audience_pool()
    print(f"  audience pool size: {len(aud_pool)}")
    aud_assigned = assign_to_coders(aud_pool, "audience_analysis")

    print("\nLoading content pool …")
    con_pool = load_content_pool()
    print(f"  content pool size: {len(con_pool)}")
    con_assigned = assign_to_coders(con_pool, "content_analysis")

    print("\nWriting audience coder files …")
    write_coder_files(aud_assigned, OUT_AUD, "audience")
    print("\nWriting content coder files …")
    write_coder_files(con_assigned, OUT_CON, "content")

    print("\nWriting trackers …")
    write_tracker(aud_assigned, OUT_BASE / "audience_nigeria_unique_assignment_tracker.xlsx")
    write_tracker(con_assigned, OUT_BASE / "content_nigeria_unique_assignment_tracker.xlsx")

    qa(aud_assigned, "audience")
    qa(con_assigned, "content")
    summary_table(aud_assigned, con_assigned)


if __name__ == "__main__":
    main()
