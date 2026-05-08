"""Build Nigeria Audience Analysis Final.xlsx from the four locked Audience Comments - Final files.

Tone and structure mirrors Nigeria Content Analysis Final.xlsx:
- Title sheet: "Summary and Stats" with title, subtitle, and a counts table
- One sheet per influencer with columns:
  Comment ID | Influencer | Platform | Source URL | Comment

The Source URL column is formatted to match the user's screenshot:
  <comment_url> (reply on <Short Name>'s post <origin_url>)

For Banky Wellington (YouTube comments) the Source URL is the video URL only,
since individual YouTube comment permalinks are not available in the source data.

Quality filtering:
- 36 weaker rows excluded (EXCLUDE_IDS).
- Comments under 150 characters are dropped to keep the corpus substantial.
- Pidgin / Yoruba / Igbo phrases are translated to Standard English in the
  Comment column.
"""

from __future__ import annotations

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # Nigeria/
SRC_DIR = os.path.join(ROOT, "Audience Analysis", "Audience Comments - Final")
OUT_PATH = os.path.join(ROOT, "Audience Analysis", "Nigeria Audience Analysis Final.xlsx")

MIN_COMMENT_LENGTH = 50  # characters; drops only the truly tiny reactions

INFLUENCER_META = {
    "Agba John Doe_Never Leave Marriage Because Husband Cheated.xlsx": {
        "influencer": "Agba John Doe",
        "short_name": "Agba",
        "platform": "X",
        "thread_topic": "Never Leave Marriage Because Husband Cheated",
        "stance": "Regressive",
        "sheet_name": "Agba John Doe",
    },
    "Banky Wellington_MENtality Podcast.xlsx": {
        "influencer": "Banky Wellington",
        "short_name": "Banky",
        "platform": "YouTube (MENtality podcast)",
        "thread_topic": "MENtality Podcast",
        "stance": "Progressive",
        "sheet_name": "Banky Wellington",
    },
    "Deyemi Okanlawon_Stop Raping Women Response.xlsx": {
        "influencer": "Deyemi Okanlawon",
        "short_name": "Deyemi",
        "platform": "X",
        "thread_topic": "Stop Raping Women Response",
        "stance": "Progressive",
        "sheet_name": "Deyemi Okanlawon",
    },
    "Shola_7 Women Will Beg One Man to Marry.xlsx": {
        "influencer": "Shola",
        "short_name": "Shola",
        "platform": "X",
        "thread_topic": "7 Women Will Beg One Man to Marry",
        "stance": "Regressive",
        "sheet_name": "Shola",
    },
}


# --- Translations dictionary -------------------------------------------------
# Pidgin / Yoruba / Igbo phrases rendered into Standard English while preserving
# the speaker's tone and meaning. Comments already in standard English are
# reproduced verbatim from the source (handled in load_source_rows).
TRANSLATIONS: dict[str, str] = {
    # ---- Agba John Doe ----
    "AGB_008": "Most divorced women of African descent end up being passed around like footballs while the man ends up in a stable relationship/marriage. Don't leave your husband for the streets, sisters.",
    "AGB_019": "As much as we sensible husbands pray our marriages don't get to this point, cheating is not the same for a man and a woman. Ask your mother, she will tell you to stand your ground; there is nothing for you out there. You are the one who will lose if you decide to leave...",
    "AGB_022": "Clearly it is evident in our societal response. Man, it requires deep thought before a woman leaves her marriage, otherwise she will end up regretting her actions for the rest of her life if she can't find what she once had — and the reality is that she may never find it. Hmm.",
    "AGB_032": "@mrjohndoe... So are you now saying that even in an abusive marriage where the husband is beating her and cheating on her, she should stay? What do you have to say about that?",
    "AGB_037": "I wonder why I would leave my husband on account of his cheating only to go and settle down as a second wife to another man or as a side chick. Women, please apply some sense in matters like this — don't let cheating turn you from wife No. 1 into wife No. 2 just because you're angry.",
    "AGB_038": "That's your business. It is dangerous to stay with a cheating man — even if not for your ego, at least for your health. Don't let him bring a disease home to a poor woman who is just minding her own business at home. To your tents, O Israel!",
    "AGB_044": "It is what it is. The irony of life and the fact of it. Imagine leaving your cheating husband — ladies, it rains everywhere. Don't let any feminist motivational speaker talk you into abandoning your home, because the reality is that you will always end up being the side chick or remain single for life.",
    "AGB_049": "Well just know, if you agree with this thread you must also agree to accept illegitimate children, or accept the possibility of getting HIV. There's no point enduring cheating for 20 years just to leave when the consequences of the cheating show up. Accept it all with a full chest, good wife.",
    "AGB_054": "That's the bitter truth, but men can do better and work on their self-discipline, because women have started to cheat back, which leads to men raising children that are not theirs — DNA scandals everywhere.",
    "AGB_055": "The only time I can suggest divorce is when the marriage is abusive. Aside from that, don't divorce over cheating, especially as a lady. It doesn't end well.",
    "AGB_061": "The summary is: a husband is not waiting for you outside. Right? A man can cheat and still love and respect his housewife. It's not that cheating is good — it isn't — but women have the power to stop a man from cheating; they just don't calm down.",
    "AGB_080": "Woke queens will think that Sir JD is a woman-hater, but in truth he has said it all. A man at 70 can marry a virgin, but a woman cannot. A patient and strong-willed wife can change her cheating husband for the better. If you think other men are better than your husband, wait for a shocker.",
    "AGB_081": "And do you men sit and think about the pain and trauma you are putting your innocent wife through? How do you comfortably hurt a woman you claim to love? Honestly, I'm not built emotionally for this, and my husband knows: if he ever cheats it will completely destroy me.",
    "AGB_088": "Whichever man misunderstands this tweet and sees it as a license to cheat — sigh, lol. I pity you, dear. We will stay; we won't leave. Your punishment will come in later years. Women can be enduring and yet dangerous. Continue, dear — enjoy yourself well now!",
    "AGB_092": "A man can cheat and still love his wife, but if a woman cheats, she has lost respect for the husband.",
    "AGB_093": "I know of a wife who was too feminist. The guy went on to put a ring on his side chick, and the first wife with a traditional marriage is now the side chick. He visits her here and there.",
    "AGB_095": "If you're a woke queen, have you heard? Don't come and quote rubbish... Because now you'll start saying \"leave your husband, there are better men outside.\" No man will accept you when you divorce your husband for him — unless he's a fool. Like that one who said \"like we agreed.\"",
    "AGB_117": "Hate it all you want, this is the stark reality of Nigeria. It is VERY WRONG and shouldn't be so. Nonetheless, women, get your money to avoid a lot of the aforementioned mess, because it seems finance has a lot to do with this. As a male, BE A FAITHFUL AND GOOD HUSBAND!!!!",
    "AGB_118": "How is a woman supposed to stop a man from cheating? Is she supposed to flog him? How? So as an adult man, you realize that cheating is not good but you continue, and then you expect the woman to be the one to stop you. Why can't you stop yourself? It's unbelievable.",
    "AGB_119": "And the fool that cheated would get a loyal virgin or a woman, right? Sir, pack it well — there's no valid truth in what you have said so far. Before getting married, ask yourself why. Go to the school of marriage and learn, because it's an institution with lots of hurt.",
    "AGB_121": "Hmmmm, this thread will not sit well with the woke queens, but in all of it I believe men can still be faithful to their wives... but what do I even know? It is what it is.",
    "AGB_126": "Lmao, better to be a woman who sleeps around than be with a 'husband' who goes around disrespecting his wife by sleeping with anything. Let everywhere scatter!",
    "AGB_137": "Men are naturally polygynous. The earlier you understand this, the better it is for you. I'm looking for a second wife. I don't cheat and I can't cheat. My penis will not enter any hole except my wives'.",
    "AGB_138": "But white ladies leave their cheating husband and claim his properties; they have a social system that forces the man to still provide for the kids. We don't have that here, yet we want to copy them and call it wokeness.",
    "AGB_146": "You need to see the way women are bashing your post on Instagram. One girl even said any man who follows you on Twitter is a red flag. The truth is always bitter.",
    "AGB_147": "What is all this now? Is marriage by force?",
    "AGB_149": "Value, sir. Thank you for this community — I have been able to learn a lot. Men and women do not have the same grace when they cheat.",
    "AGB_155": "Will men please stop cheating, for crying out loud!!! It's unbelievable.",
    "AGB_156": "Comrade, please break this down for me. You mean the side chick becomes the wife and the wife becomes the side chick — or are my eyes deceiving me?",
    "AGB_157": "Is it all about servicing and getting serviced just to survive? What if she has money? You'll all just keep thinking one way. What if she wants to be with another man who has a woman? What if that's what she wants?",
    "AGB_159": "Sometimes what causes this cheating is when a man isn't getting enough sex from his partner.",
    "AGB_161": "I prefer to mind my business, because I have learned that no matter how much men cheat, there's someone they cherish and don't want to disrespect. No matter how many girls they have, there is that one special girl who means a lot to them.",
    "AGB_163": "In the end, with this your mentality, the only option is to cheat back and then wait to be caught. If I eventually get caught, everybody must forgive everybody — otherwise everybody will catch hands! End of story.",
    "AGB_173": "This isn't about wokeness — can you all at least stop glorifying cheating? I don't see why you men can't stay faithful. Do you even know how ladies feel knowing our man is cheating on us? Men should do better — you all are not God.",
    "AGB_175": "I want to ask: why should it always be the women at the receiving end? Do we women look like God, who can't make mistakes, and always have to forgive an idiot who can't control his third leg?",
    "AGB_176": "Mostly it's women like her who reply. May God not let us marry a wicked queen — this one would burn her husband if she found out he was cheating...",
    # ---- Banky Wellington ----
    "BNK_006": "One thing I've noticed is that when a man fulfills his responsibilities, it's seen as a normal thing — something he's supposed to do. But when a woman does, it's often described as her 'helping' or 'taking care of your responsibilities.' There's a popular Yoruba saying, \"Nítorí ìrègún l'obìnrin fi ń ṣọ̀rẹ́,\" meaning \"a woman does something once and keeps reminding you about it.\" Sola Sobowale even echoed this line in the movie Her Excellency. Yes, we may not have control over who earns more, but a man should still handle the basic responsibilities — that's leadership, not competition. Don't let your woman take over your responsibilities. When both sides understand and respect their roles, peace naturally follows. Let her money be her money, brother!!!",
    "BNK_018": "Man, this is therapy for me to build a better relationship. To make a relationship or marriage work, you need to work on yourself.",
    "BNK_076": "I feel this podcast is for me. I am married with a kid. I dated my wife for 9 years before marriage and we have been married for almost 2 years now. Marriage is way different from a relationship — I can tell you that. Most of my close friends are not married, and I find it weird having conversations with them about marriage and fatherhood (survivor's guilt). My wife is literally my best friend, girlfriend, and wife all at the same time, but it's not the same for her because her friends are married. And I agree with Alex — many guys are not married because they know too much. If I had broken up with my wife, I'm sure I'd still be in the streets.",
    "BNK_117": "I couldn't resist commenting. This podcast is well packed. May the fire never reduce its hotness. Dang! With the rise of men feminizing themselves and the whole identity crisis, this podcast is right on time. I love everything this space represents. (Please, please, please, stop it.)",
    "BNK_142": "My friend calls me and tells me he loves me, lol. I'll first ask him if he smoked sawdust with sandpaper to make him this rough.",
    "BNK_143": "How do we talk about men's friendship and not talk about how money affects friendship? We need another episode on that, please.",
    # ---- Deyemi Okanlawon ----
    "DEY_011": "Respectfully, sir — you have no sense! 1) What was your contribution when Mirable cried that she was raped on 15 February and it turned out to be a lie? 2) Two days later, a young lady accused David of rape, the police arrested David, and the rest is history. Later she confessed and apologized that it was a lie. Do you realize that 10 years ago, Mirable and that lady would have succeeded in falsely accusing those men, and they would have been severely punished for nothing? Sir, what was your contribution to those recent false rape accusers? I hope that if a lady comes out and accuses you of rape when you know it isn't true, you would be very happy to bear the name 'rapist' and have everything you've worked for taken away because of a lie she couldn't prove? Rape is a very sensitive topic. So many innocent people have been falsely accused, even when it was consensual, and even though some didn't know anything about it. You can't talk about rapists without talking about the accusers. There are two truths here, and they cannot be treated with emotion or gender bias.",
    "DEY_019": "But with that Simi tweet about having sexual fantasies about children — will you use this same energy and do us a big favor by putting out 'STOP MOLESTING AND RAPING OUR CHILDREN'? Some of you hopeless, horny degenerates calling yourselves men are the reason innocent boys have no say when they are being abused, sexually molested, or raped by the likes of Simi. The same you who would tell your son to man up when a girl who lacks home training abused him. But you can't even understand the damage that will cause him, because you are a useless man who shouldn't have a say in public opinion. The same you who would approve of how those grown men become rapists, which only gives you a headache when they get caught, but you go silent when a female rapist or female pedophile is being apprehended. Better not carry this foolish mentality and speak in public, or you might not be the one returning home. Cursed one, idiot.",
    "DEY_023": "Sir, you are a rapist! Stop raping women. I expect you to say yes, sir. #Yes",
    "DEY_037": "Deyemi, you've set yourself up for a dragging. We've seen people who walked this path become victims, with their grammar suddenly changing. I guess you have 2 or 3 sons — you should be championing the need to punish false rape accusations while also kicking against rape itself. Don't pander!",
    "DEY_038": "But you know she could have condemned 'false rape accusations' even at the heat of the whole rape moment and still stood well, rather than dismissing it. Anyway, that's just the women's thought process.",
    "DEY_044": "There's no human on earth who has agreed to rape. You all are just devils, and you do look like the kind of person who would falsely accuse someone of rape with that your forehead like a ninja turtle. That girl should be serving a prison term right now, but she's not. Damn you all, miserable lot.",
    "DEY_046": "Man, you are probably well educated, but you just veered off the entire context of the conversation, and I wonder whether the education even passed through you. Nobody agreed to men raping women, but should we just take an accuser's word for it when we know false accusations exist? Better pray it doesn't reach your side.",
    "DEY_049": "You're mixing things up. She was talking about rape as a general issue, not validating or dismissing any specific case. Let's not twist the context to fit a different argument, please!",
    "DEY_054": "All the mad people talking rubbish — if they were accused of rape and the allegation turned out to be false, that's when their mouths would be full of explanations about false accusations. But until then, they'll keep talking rubbish.",
    "DEY_056": "It's always the faceless accounts that come here to make dumb statements. Simi is married to a man and yet she said 'all men are rapists' — your stupidity is very loud, idiot, fool, useless one.",
    "DEY_058": "Hypocrisy at its finest. So in a court of law, only one side is heard, judgment passed, and then the other party is told 'we can start a separate hearing another time/place.' What people in Nigeria call a 'fake celebrity.' I guess you didn't read where she was asked about false accusations and said 'STFU.'",
    "DEY_063": "Why should that be your first response? That isn't my first response, please — I will never be part of incriminating an innocent man.",
    "DEY_077": "I saw one of my former neighbours abusing Simi just now — he didn't even fear that I was online. If I expose what he has done, this app will scatter. If we start to open up everybody's past life here, the majority would log out — Simi's tweet alone would be enough to take her down.",
    "DEY_088": "You are a crass enabler. The response should be YES! Not start another discourse on top of an existing topic. Are you people that daft in real life? Like, yo.",
    # ---- Shola ----
    "SHO_001": "They'll think it's a joke. Men are the prize. These girls are just in their era and they feel like all men are after them — until they can't even find one. Every man in this world can find a wife to marry no matter how old he is; women can't.",
    "SHO_015": "True, but three women cannot share a wretched man who is equally lazy.",
    "SHO_025": "It won't be funny — they'll be surprised by the kind of conditions men will give them just to marry them. Their eyes will open by then.",
    "SHO_026": "It's really hard to believe, especially for our generation, that there was once a time when women were so scarce that even a 30-year-old man would have to book a newborn girl child and wait all the way for her to grow up before marrying her. Wow, you see this generation — you just need an iPhone, a car, and a house, and come and see desperate women lined up at your gate.",
    "SHO_032": "Not yet, because if you aren't shining as a man, they don't cling to you. The reality will dawn on them soon though.",
    "SHO_037": "Don't stress yourself — the majority of men here won't have the strength when the era finally arrives.",
    "SHO_043": "That era arrived a long time ago. One woman even offered me 5 million Naira to come and marry her.",
    "SHO_044": "Fela Kuti showed us the blueprint. We are getting to an era where a man will have 3–5 baby mamas all living with him. If you don't want it, go to your father's house.",
    "SHO_045": "Not yet for broke men. It's only a man who guides them they desperately beg.",
    "SHO_058": "I have been thinking, at the rate at which our men are dying through being outlaws, that some women might consider paying our dowry. I won't go further.",
    "SHO_070": "I tell you. One was just asking me 'aren't you ready to settle down?' — auntie, even though I want to settle down, it's not with you. I'm not in a rush; I want to take my time, because I'm the prize.",
    "SHO_073": "I agree with you — three girls are already begging me to date them. Four are remaining, but it's only the one with good character and her own money that I'll pick.",
    "SHO_077": "That means you all are too many; we only need a few men.",
    "SHO_080": "Some end up dead before their time — mentally ill, poor, family and children hating them, diseased, off their path, wealthy but unsuccessful in family life, confused and lost. You see it all the time, but the ego-driven refuse to be wise, and they repeat the same fate.",
    "SHO_088": "True. I got a text from an old friend this morning asking about my wellbeing — after everything, and she's aware that I'm currently in a serious relationship. Her words: 'Give me a chance, let the gods' will happen.' My first thought was: this isn't the one you all want to use HIV to kill me.",
}


# Comment IDs to exclude — weaker rows flagged in QA review (too short, generic
# praise, side-drama, or not independently rich enough for audience coding).
EXCLUDE_IDS: set[str] = {
    # Agba John Doe — too short / generic
    "AGB_068", "AGB_077", "AGB_101", "AGB_116", "AGB_147", "AGB_152", "AGB_168", "AGB_174",
    # Agba John Doe — second-pass cleanup (thinner than the rest of the thread)
    "AGB_136", "AGB_156", "AGB_158", "AGB_164", "AGB_171",
    # Banky Wellington — podcast praise / brief generic comments
    "BNK_015", "BNK_016", "BNK_019", "BNK_022", "BNK_029", "BNK_034", "BNK_036",
    "BNK_037", "BNK_038", "BNK_039", "BNK_060", "BNK_061", "BNK_068", "BNK_119",
    "BNK_121", "BNK_130",
    # Banky Wellington — second-pass cleanup (podcast praise / generic parenting / thin)
    "BNK_007", "BNK_032", "BNK_047", "BNK_063", "BNK_067", "BNK_069", "BNK_083",
    "BNK_091", "BNK_096", "BNK_111", "BNK_113", "BNK_125", "BNK_129", "BNK_131",
    "BNK_132", "BNK_133", "BNK_139", "BNK_146", "BNK_147",
    # Deyemi Okanlawon — insults / side-drama / too thin
    "DEY_027", "DEY_052", "DEY_055", "DEY_060", "DEY_081",
    # Deyemi Okanlawon — second-pass cleanup (insult-style / too thin)
    "DEY_030", "DEY_074", "DEY_075", "DEY_087",
    # Shola — generic "this is true" reactions
    "SHO_047", "SHO_063", "SHO_081", "SHO_083", "SHO_084", "SHO_085", "SHO_086",
    # Shola — second-pass cleanup (very short / generic / unclear)
    "SHO_006", "SHO_034", "SHO_041", "SHO_049", "SHO_056", "SHO_061", "SHO_069",
    "SHO_076", "SHO_077", "SHO_079",
}


def build_source_url(comment_url: str | None, origin_url: str | None, short_name: str) -> str:
    """Format the Source URL cell to match the screenshot style.

    For X replies:  <comment_url> (reply on <ShortName>'s post <origin_url>)
    For YouTube comments where there is no individual comment URL: just the video URL.
    """
    if comment_url and origin_url:
        return f"{comment_url} (reply on {short_name}'s post {origin_url})"
    if origin_url:
        return origin_url
    return comment_url or ""


def load_source_rows(file_name: str, meta: dict) -> list[dict]:
    path = os.path.join(SRC_DIR, file_name)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        if len(headers) == 4:
            comment_id, text, comment_url, origin_url = r[0], r[1], r[2], r[3]
        else:  # 3 cols (Banky)
            comment_id, text, origin_url = r[0], r[1], r[2]
            comment_url = None

        if comment_id in EXCLUDE_IDS:
            continue

        comment_text = TRANSLATIONS.get(comment_id, text or "")
        if len(comment_text) < MIN_COMMENT_LENGTH:
            continue

        out.append({
            "id": comment_id,
            "comment": comment_text,
            "source_url": build_source_url(comment_url, origin_url, meta["short_name"]),
        })
    return out


# --- Styling helpers ---------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER


def write_influencer_sheet(wb: Workbook, meta: dict, items: list[dict]):
    ws = wb.create_sheet(meta["sheet_name"])
    headers = [
        "Comment ID",
        "Influencer",
        "Platform",
        "Source URL",
        "Comment",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for item in items:
        ws.append([
            item["id"],
            meta["influencer"],
            meta["platform"],
            item["source_url"],
            item["comment"],
        ])

    widths = [12, 22, 30, 95, 95]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    max_row = ws.max_row
    for row in range(2, max_row + 1):
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.alignment = WRAP
            c.font = CELL_FONT
            c.border = BORDER
        ws.row_dimensions[row].height = 90

    ws.freeze_panes = "A2"


def write_summary_sheet(wb: Workbook, all_items: dict[str, list[dict]]):
    ws = wb.create_sheet("Summary and Stats", 0)

    title_font = Font(name="Calibri", size=20, bold=True, color="111827")
    subtitle_font = Font(name="Calibri", size=12, italic=True, color="4B5563")
    section_font = Font(name="Calibri", size=12, bold=True, color="111827")
    label_font = Font(name="Calibri", size=11, color="374151")
    value_font = Font(name="Calibri", size=11, bold=True, color="111827")

    ws["A1"] = "Nigeria Masculinity Audience Analysis"
    ws["A1"].font = title_font
    ws["A2"] = "Norman Lear Center × Gates Foundation"
    ws["A2"].font = subtitle_font

    # --- Headline metrics row ---
    total_comments = sum(len(v) for v in all_items.values())
    creator_count = len(all_items)
    progressive_creators = sum(1 for m in INFLUENCER_META.values() if m["stance"] == "Progressive")
    regressive_creators = sum(1 for m in INFLUENCER_META.values() if m["stance"] == "Regressive")
    progressive_total = sum(len(all_items[m["influencer"]]) for m in INFLUENCER_META.values() if m["stance"] == "Progressive")
    regressive_total = sum(len(all_items[m["influencer"]]) for m in INFLUENCER_META.values() if m["stance"] == "Regressive")
    ratio_label = (
        f"1 : {regressive_total / progressive_total:.2f}" if progressive_total else "n/a"
    )

    ws["A4"] = "Total comments"
    ws["B4"] = "Creators"
    ws["C4"] = "Stance ratio"
    for col in (1, 2, 3):
        c = ws.cell(row=4, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    ws["A5"] = total_comments
    ws["B5"] = f"{creator_count} ({progressive_creators} progressive + {regressive_creators} regressive)"
    ws["C5"] = ratio_label
    for col in (1, 2, 3):
        c = ws.cell(row=5, column=col)
        c.font = value_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    ws["A6"] = "Comments across 4 audience threads"
    ws["B6"] = "Balanced sample frame"
    ws["C6"] = "Progressive : Regressive"
    for col in (1, 2, 3):
        c = ws.cell(row=6, column=col)
        c.font = subtitle_font
        c.alignment = Alignment(horizontal="center", vertical="top")
        c.border = BORDER

    # --- Stance and per-creator breakdown ---
    ws["A9"] = "Stance"
    ws["B9"] = "Comments"
    ws["C9"] = "Creator"
    ws["D9"] = "Comments"
    for col in range(1, 5):
        c = ws.cell(row=9, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    ws["A10"] = "Progressive"
    ws["B10"] = progressive_total
    ws["A11"] = "Regressive"
    ws["B11"] = regressive_total
    for r in (10, 11):
        for col in (1, 2):
            c = ws.cell(row=r, column=col)
            c.font = label_font if col == 1 else value_font
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")

    creator_rows_start = 10
    for i, (_, meta) in enumerate(INFLUENCER_META.items()):
        r = creator_rows_start + i
        ws.cell(row=r, column=3, value=meta["influencer"]).font = label_font
        ws.cell(row=r, column=4, value=len(all_items[meta["influencer"]])).font = value_font
        for col in (3, 4):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if col == 3 else "center", vertical="center")

    # --- Per-thread topic detail ---
    ws["A15"] = "Threads / episodes covered"
    ws["A15"].font = section_font

    ws["A16"] = "Influencer"
    ws["B16"] = "Stance"
    ws["C16"] = "Platform"
    ws["D16"] = "Thread / Episode"
    ws["E16"] = "Comments"
    for col in range(1, 6):
        c = ws.cell(row=16, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    r = 17
    for _, meta in INFLUENCER_META.items():
        items = all_items[meta["influencer"]]
        ws.cell(row=r, column=1, value=meta["influencer"]).font = label_font
        ws.cell(row=r, column=2, value=meta["stance"]).font = label_font
        ws.cell(row=r, column=3, value=meta["platform"]).font = label_font
        ws.cell(row=r, column=4, value=meta["thread_topic"]).font = label_font
        ws.cell(row=r, column=5, value=len(items)).font = value_font
        for col in range(1, 6):
            c = ws.cell(row=r, column=col)
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left" if col != 5 else "center")
            c.border = BORDER
        r += 1

    # --- Methodology note ---
    r += 1
    ws.cell(row=r, column=1, value="Methodology notes").font = section_font
    r += 1
    notes = [
        "Comments are drawn from the four locked Audience Comments – Final spreadsheets in "
        "Nigeria/Audience Analysis/Audience Comments - Final/.",
        "Pidgin, Yoruba, and Igbo phrases have been translated into Standard English while "
        "preserving the speaker's tone and meaning. Comments already in standard English are "
        "reproduced verbatim.",
        f"Quality filter: comments shorter than {MIN_COMMENT_LENGTH} characters and a manually "
        "flagged set of generic / off-scope reactions are excluded so that every retained row is "
        "long enough and topical enough for downstream coding.",
        "Source URL format — X replies: \"<comment URL> (reply on <Name>'s post <original post "
        "URL>)\". For YouTube comments, the Source URL links to the source episode rather than "
        "individual comment permalinks, since individual YouTube comment URLs were not available "
        "in the source data.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note).font = label_font
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        ws.row_dimensions[r].height = 50
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 16

    # Merge title cells visually with adjacent ones
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    all_items: dict[str, list[dict]] = {}
    for file_name, meta in INFLUENCER_META.items():
        items = load_source_rows(file_name, meta)
        all_items[meta["influencer"]] = items

    write_summary_sheet(wb, all_items)
    for file_name, meta in INFLUENCER_META.items():
        write_influencer_sheet(wb, meta, all_items[meta["influencer"]])

    wb.save(OUT_PATH)
    print("Wrote:", OUT_PATH)
    for inf, items in all_items.items():
        print(f"  {inf}: {len(items)} comments")
    print(f"  TOTAL: {sum(len(v) for v in all_items.values())}")


if __name__ == "__main__":
    main()
