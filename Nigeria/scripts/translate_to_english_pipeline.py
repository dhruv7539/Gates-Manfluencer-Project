"""
Translate audience Complete + content Final datasets to English for downstream
exploratory analysis. Adds three columns to every row:

    text_original       — verbatim text exactly as scraped/transcribed
    text_english        — fully English version (verbatim if already English)
    language_detected   — short label: "English" / "Nigerian Pidgin" /
                          "Yoruba" / "Igbo" / "Mixed (Pidgin + English)" / etc.

Outputs:
    Nigeria/Audience Analysis/Translated/audience_complete_translated.parquet
    Nigeria/Audience Analysis/Translated/audience_complete_translated.xlsx   (one sheet per source file)
    Nigeria/Audience Analysis/Translated/audience_final_translated.parquet
    Nigeria/Audience Analysis/Translated/audience_final_translated.xlsx
    Nigeria/Content Analysis/Translated/content_final_translated.parquet
    Nigeria/Content Analysis/Translated/content_final_translated.xlsx        (one sheet per creator)

Caching: temp/translation_cache.parquet keyed by sha1(text). Re-running is free.
LLM calls only happen for rows that are NOT trivially pure English.
"""
from __future__ import annotations

import asyncio
import hashlib
import json
import os
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openai import AsyncOpenAI
from tqdm.asyncio import tqdm as atqdm

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env")
assert os.getenv("OPENAI_API_KEY"), "OPENAI_API_KEY missing"

AUDIENCE_DIR  = ROOT / "Nigeria" / "Audience Analysis" / "Audience Comments - Complete"
AUDIENCE_FINAL_WB = ROOT / "Nigeria" / "Audience Analysis" / "Nigeria Audience Analysis Final.xlsx"
CONTENT_WB    = ROOT / "Nigeria" / "Content Analysis" / "Nigeria Content Analysis Final.xlsx"

OUT_AUDIENCE_PARQUET       = ROOT / "Nigeria" / "Audience Analysis" / "Translated" / "audience_complete_translated.parquet"
OUT_AUDIENCE_XLSX          = ROOT / "Nigeria" / "Audience Analysis" / "Translated" / "audience_complete_translated.xlsx"
OUT_AUDIENCE_FINAL_PARQUET = ROOT / "Nigeria" / "Audience Analysis" / "Translated" / "audience_final_translated.parquet"
OUT_AUDIENCE_FINAL_XLSX    = ROOT / "Nigeria" / "Audience Analysis" / "Translated" / "audience_final_translated.xlsx"
OUT_CONTENT_PARQUET        = ROOT / "Nigeria" / "Content Analysis" / "Translated" / "content_final_translated.parquet"
OUT_CONTENT_XLSX           = ROOT / "Nigeria" / "Content Analysis" / "Translated" / "content_final_translated.xlsx"

for _p in (OUT_AUDIENCE_PARQUET, OUT_CONTENT_PARQUET):
    _p.parent.mkdir(parents=True, exist_ok=True)

CACHE_PATH = ROOT / "temp" / "translation_cache.parquet"
CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)

LLM_MODEL   = "gpt-4o-mini"
CONCURRENCY = 12

# ─── language fast-path ──────────────────────────────────────────────────────
# Pure-English heuristic: if the text is short, in basic Latin-1 range, has no
# known Pidgin/Yoruba markers, and contains common English function words,
# we tag it as English without an LLM call.

NON_ENGLISH_MARKERS = re.compile(
    r"\b(dey|wey|na|wahala|abeg|sabi|pikin|oga|biko|nau|jare|knack|sef|chai|"
    r"kuro|gbam|ooo|abi|una|olo[sm]ho|àgbá|agba|olosho|asoebi|iwa|kwarupt|"
    r"don dey|kwarup|naa|haba|kpomo|wetin|wia|fit|wotowoto|toto|gan|para|"
    r"omo|aboki|alaye|babes?|innit|fvck|wahala|jor|sha|shey|shey?|ojoro|"
    r"mumu|tatafo|gbege|gbegborun|aje|wahala|para|mago|maga|bizniz|gbasa|"
    r"akata|igbati|kpe|kparapo|odogwu|werey|orisirisi|gbe|igba|olohun|"
    r"abikoji|alagbala|bawo|bawo ni|báwo|ekaaro|ekaale|ekuro|igba|kilode|"
    r"nko|odiabo|opetebete|orogun|otun|oti|owo|oya|paaa|para|sebi|seun|"
    r"shebi|shey|sho|shogo|shuga|sista|so|sokoto|tata|tete|tinz|tope|"
    r"tutu|wadi|wahalla|warri|wayo|woli|wuruwuru|yawa|yoruba|"
    r"chineke|nkechi|adanma|nna|nne|biko|ndo|ndi|ozugbo|onye|igbo)\b",
    re.IGNORECASE,
)
YORUBA_DIACRITICS = re.compile(r"[àáèéìíòóùúṣẹọḿǹẽĩõũÀÁÈÉÌÍÒÓÙÚṢẸỌḾǸẼĨÕŨ]")
# only specific Hausa markers; common short particles overlap with English so omitted
HAUSA_MARKERS = re.compile(r"\b(inshallah|wallahi|sannu|kuma|amma|wani|wata|abinda)\b", re.IGNORECASE)

ENGLISH_FUNCTION_WORDS = re.compile(r"\b(the|and|is|of|to|in|that|it|for|on|with|you|i|this|but|are|was|be|have|has|not|so|do|will|would|could|should|just|like|all|what|when|why|how|who)\b", re.IGNORECASE)


def is_obviously_english(text: str) -> bool:
    if not text or len(text.strip()) < 3:
        return True  # trivial — skip LLM
    if YORUBA_DIACRITICS.search(text):
        return False
    if NON_ENGLISH_MARKERS.search(text):
        return False
    if HAUSA_MARKERS.search(text) and not ENGLISH_FUNCTION_WORDS.search(text):
        return False
    # mostly ASCII + common punctuation/emoji + has English function words
    ascii_ratio = sum(1 for c in text if ord(c) < 128) / len(text)
    if ascii_ratio > 0.85 and ENGLISH_FUNCTION_WORDS.search(text):
        return True
    return False


# ─── caching ─────────────────────────────────────────────────────────────────

def text_hash(t: str) -> str:
    return hashlib.sha1(t.strip().encode("utf-8")).hexdigest()


def load_cache() -> dict[str, dict]:
    if not CACHE_PATH.exists():
        return {}
    df = pd.read_parquet(CACHE_PATH)
    return {r["hash"]: {"text_english": r["text_english"], "language_detected": r["language_detected"]} for _, r in df.iterrows()}


def save_cache(cache: dict[str, dict]):
    rows = [{"hash": h, **v} for h, v in cache.items()]
    pd.DataFrame(rows).to_parquet(CACHE_PATH, index=False)


# ─── LLM translation ─────────────────────────────────────────────────────────

PROMPT = """You translate Nigerian social-media text into clear English for academic analysis.

For each numbered item, return:
  - language: a SHORT label such as "English", "Nigerian Pidgin",
    "Yoruba", "Igbo", "Hausa", "Mixed (Pidgin + English)",
    "Mixed (Yoruba + English)", etc.
  - text_english: the same content rewritten in clear, natural English.
    Preserve the speaker's tone (sarcastic, angry, supportive, etc.) and meaning.
    Keep proper nouns, hashtags, @-mentions, and emojis as-is.
    If the item is already entirely in English, return it UNCHANGED.

Return STRICT JSON: {"results": [{"id": <int>, "language": "...", "text_english": "..."}]}
"""


async def translate_batch(client: AsyncOpenAI, sem, batch: list[tuple[int, str]]) -> list[dict]:
    async with sem:
        items = "\n".join(f"{i}. {t}" for i, t in batch)
        try:
            r = await client.chat.completions.create(
                model=LLM_MODEL,
                messages=[
                    {"role": "system", "content": PROMPT},
                    {"role": "user", "content": items},
                ],
                response_format={"type": "json_object"},
                temperature=0.0,
            )
            data = json.loads(r.choices[0].message.content)
            results = data.get("results", [])
        except Exception as e:
            print(f"  batch failed: {e}", file=sys.stderr)
            results = []
        # index back to original
        out = []
        by_id = {x.get("id"): x for x in results if isinstance(x, dict)}
        for idx, _t in batch:
            x = by_id.get(idx) or by_id.get(str(idx)) or {}
            out.append({
                "language": x.get("language", "Unknown"),
                "text_english": x.get("text_english"),
            })
        return out


async def translate_all(texts: list[str]) -> list[dict]:
    """Translate texts; returns list of {language_detected, text_english} aligned to input."""
    cache = load_cache()
    results: list[dict | None] = [None] * len(texts)

    # Fast path: pure English without LLM
    pending: list[tuple[int, str]] = []
    for i, t in enumerate(texts):
        if not t or not str(t).strip():
            results[i] = {"language_detected": "Unknown", "text_english": ""}
            continue
        t = str(t)
        h = text_hash(t)
        if h in cache:
            results[i] = {"language_detected": cache[h]["language_detected"], "text_english": cache[h]["text_english"]}
            continue
        if is_obviously_english(t):
            results[i] = {"language_detected": "English", "text_english": t}
            cache[h] = {"language_detected": "English", "text_english": t}
            continue
        pending.append((i, t))

    print(f"  Pure-English fast path: {sum(1 for r in results if r is not None and r.get('language_detected') == 'English')} / {len(texts)}")
    print(f"  Pending LLM translation: {len(pending)}")

    if pending:
        client = AsyncOpenAI()
        sem = asyncio.Semaphore(CONCURRENCY)
        BATCH = 15
        batches = [pending[i:i+BATCH] for i in range(0, len(pending), BATCH)]
        # batch indices use local 0..len(batch)-1 ids
        async def go(batch):
            local = [(j, t) for j, (_orig_i, t) in enumerate(batch)]
            r = await translate_batch(client, sem, local)
            return batch, r
        all_results = await atqdm.gather(*[go(b) for b in batches])
        for batch, r in all_results:
            for (orig_i, t), entry in zip(batch, r):
                txt_en = entry["text_english"] or t  # fallback to original on failure
                lang = entry["language"] or "Unknown"
                results[orig_i] = {"language_detected": lang, "text_english": txt_en}
                cache[text_hash(t)] = {"language_detected": lang, "text_english": txt_en}

    save_cache(cache)
    return [r if r is not None else {"language_detected": "Unknown", "text_english": ""} for r in results]


# ─── data loaders ────────────────────────────────────────────────────────────

def load_audience() -> pd.DataFrame:
    rows = []
    for creator_dir in sorted(AUDIENCE_DIR.iterdir()):
        if not creator_dir.is_dir():
            continue
        creator = creator_dir.name
        # Banky has YouTube/Instagram/MENtality subdirs; others are flat
        if any(p.is_dir() for p in creator_dir.iterdir()):
            for platform_dir in sorted(creator_dir.iterdir()):
                if not platform_dir.is_dir():
                    continue
                for f in sorted(platform_dir.glob("*.xlsx")):
                    if f.name.startswith("~"):
                        continue
                    rows.extend(_load_xlsx(f, creator, platform=platform_dir.name))
        else:
            for f in sorted(creator_dir.glob("*.xlsx")):
                if f.name.startswith("~"):
                    continue
                rows.extend(_load_xlsx(f, creator, platform=None))
    df = pd.DataFrame(rows)
    df["text_original"] = df["text_original"].astype(str).str.strip()
    df = df[df["text_original"].str.len() > 0].reset_index(drop=True)
    return df


def load_audience_final() -> pd.DataFrame:
    """Load the consolidated 417-row audience workbook (Nigeria Audience Analysis Final.xlsx)."""
    wb = openpyxl.load_workbook(AUDIENCE_FINAL_WB, read_only=True)
    rows = []
    for sn in wb.sheetnames:
        if sn == "Summary and Stats":
            continue
        ws = wb[sn]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = [str(h).strip() if h else "" for h in data[0]]
        def find(name):
            for i, h in enumerate(headers):
                if h.lower().startswith(name.lower()):
                    return i
            return None
        i_id  = find("Comment ID")
        i_inf = find("Influencer")
        i_pl  = find("Platform")
        i_url = find("Source URL")
        # exact match for the comment text column (avoid prefix collision with "Comment ID")
        i_txt = next((i for i, h in enumerate(headers) if h.strip().lower() == "comment"), None)
        if i_txt is None:
            print(f"  WARN: no Comment col in {sn}, skipping")
            continue
        for r in data[1:]:
            if not r or len(r) <= i_txt or not r[i_txt]:
                continue
            rows.append({
                "comment_id":    r[i_id]  if i_id  is not None else None,
                "creator":       r[i_inf] if i_inf is not None else sn,
                "platform":      r[i_pl]  if i_pl  is not None else None,
                "source_url":    r[i_url] if i_url is not None else None,
                "text_original": r[i_txt],
            })
    wb.close()
    df = pd.DataFrame(rows)
    df["text_original"] = df["text_original"].astype(str).str.strip()
    df = df[df["text_original"].str.len() > 0].reset_index(drop=True)
    return df


def _load_xlsx(path: Path, creator: str, platform: str | None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    out = []
    for r in ws.iter_rows(values_only=True, min_row=2):
        if r and r[0]:
            out.append({
                "creator": creator,
                "platform": platform,
                "source_file": path.stem,
                "text_original": r[0],
            })
    wb.close()
    return out


def load_content() -> pd.DataFrame:
    wb = openpyxl.load_workbook(CONTENT_WB, read_only=True)
    rows = []
    for sn in wb.sheetnames:
        if sn == "Summary and Stats":
            continue
        ws = wb[sn]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = [str(h).strip() if h else "" for h in data[0]]
        # Find columns
        def find(col_name):
            for i, h in enumerate(headers):
                if h.lower().startswith(col_name.lower()):
                    return i
            return None
        idx_id = find("Content ID")
        idx_inf = find("Influencer")
        idx_plat = find("Platform")
        idx_type = find("Content Type")
        idx_url = find("Source URL")
        idx_ctx = find("Context")
        # text col is "Verbatim Text" or "Tweet"
        idx_text = find("Verbatim")
        if idx_text is None:
            idx_text = find("Tweet")
        if idx_text is None:
            print(f"  WARN: no text column in {sn}, skipping")
            continue
        for r in data[1:]:
            if not r or len(r) <= idx_text or not r[idx_text]:
                continue
            rows.append({
                "content_id": r[idx_id] if idx_id is not None else None,
                "creator": r[idx_inf] if idx_inf is not None else sn,
                "platform": r[idx_plat] if idx_plat is not None else None,
                "content_type": r[idx_type] if idx_type is not None else None,
                "source_url": r[idx_url] if idx_url is not None else None,
                "context": r[idx_ctx] if idx_ctx is not None else None,
                "text_original": r[idx_text],
            })
    wb.close()
    df = pd.DataFrame(rows)
    df["text_original"] = df["text_original"].astype(str).str.strip()
    df = df[df["text_original"].str.len() > 0].reset_index(drop=True)
    return df


# ─── writers ─────────────────────────────────────────────────────────────────

def write_audience(df: pd.DataFrame):
    df.to_parquet(OUT_AUDIENCE_PARQUET, index=False)
    # one sheet per source_file
    with pd.ExcelWriter(OUT_AUDIENCE_XLSX, engine="openpyxl") as xw:
        # summary
        summary = df.groupby(["creator", "platform", "source_file"]).size().reset_index(name="comments")
        summary.to_excel(xw, sheet_name="Summary", index=False)
        for sf, sub in df.groupby("source_file"):
            name = sf[:31]
            sub[["creator", "platform", "source_file", "text_original", "text_english", "language_detected"]].to_excel(xw, sheet_name=name, index=False)
    print(f"  wrote {OUT_AUDIENCE_PARQUET}")
    print(f"  wrote {OUT_AUDIENCE_XLSX}")


def write_audience_final(df: pd.DataFrame):
    df.to_parquet(OUT_AUDIENCE_FINAL_PARQUET, index=False)
    with pd.ExcelWriter(OUT_AUDIENCE_FINAL_XLSX, engine="openpyxl") as xw:
        summary = df.groupby("creator").size().reset_index(name="comments")
        summary.to_excel(xw, sheet_name="Summary", index=False)
        cols = ["comment_id", "creator", "platform", "source_url", "text_original", "text_english", "language_detected"]
        df[cols].to_excel(xw, sheet_name="audience", index=False)
    print(f"  wrote {OUT_AUDIENCE_FINAL_PARQUET}")
    print(f"  wrote {OUT_AUDIENCE_FINAL_XLSX}")


def write_content(df: pd.DataFrame):
    df.to_parquet(OUT_CONTENT_PARQUET, index=False)
    with pd.ExcelWriter(OUT_CONTENT_XLSX, engine="openpyxl") as xw:
        summary = df.groupby("creator").size().reset_index(name="segments")
        summary.to_excel(xw, sheet_name="Summary", index=False)
        for cr, sub in df.groupby("creator"):
            name = str(cr)[:31]
            cols = ["content_id", "creator", "platform", "content_type", "source_url", "context", "text_original", "text_english", "language_detected"]
            sub[cols].to_excel(xw, sheet_name=name, index=False)
    print(f"  wrote {OUT_CONTENT_PARQUET}")
    print(f"  wrote {OUT_CONTENT_XLSX}")


# ─── main ────────────────────────────────────────────────────────────────────

async def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else "all"

    if mode in ("all", "audience"):
        print("=" * 60)
        print("AUDIENCE COMMENTS — COMPLETE TIER")
        print("=" * 60)
        df_a = load_audience()
        print(f"  loaded {len(df_a)} rows from {df_a['source_file'].nunique()} files")
        translations = await translate_all(df_a["text_original"].tolist())
        df_a["text_english"]      = [t["text_english"] for t in translations]
        df_a["language_detected"] = [t["language_detected"] for t in translations]
        write_audience(df_a)
        print("  language breakdown:")
        print(df_a["language_detected"].value_counts().head(15).to_string())

    if mode in ("all", "audience_final"):
        print("=" * 60)
        print("AUDIENCE — FINAL (consolidated 417-row workbook)")
        print("=" * 60)
        df_a = load_audience_final()
        print(f"  loaded {len(df_a)} rows across {df_a['creator'].nunique()} creators")
        translations = await translate_all(df_a["text_original"].tolist())
        df_a["text_english"]      = [t["text_english"] for t in translations]
        df_a["language_detected"] = [t["language_detected"] for t in translations]
        write_audience_final(df_a)
        print("  language breakdown:")
        print(df_a["language_detected"].value_counts().head(15).to_string())

    if mode in ("all", "content"):
        print("=" * 60)
        print("CONTENT — FINAL")
        print("=" * 60)
        df_c = load_content()
        print(f"  loaded {len(df_c)} rows across {df_c['creator'].nunique()} creators")
        translations = await translate_all(df_c["text_original"].tolist())
        df_c["text_english"]      = [t["text_english"] for t in translations]
        df_c["language_detected"] = [t["language_detected"] for t in translations]
        write_content(df_c)
        print("  language breakdown:")
        print(df_c["language_detected"].value_counts().head(15).to_string())

    print("\nDone.")


if __name__ == "__main__":
    asyncio.run(main())
