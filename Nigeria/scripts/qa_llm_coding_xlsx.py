"""
15-check QA suite for the LLM Coding workbooks.

Runs against:
    Codebooks/LLM Codebook/LLM Coding - Audience Analysis.xlsx
    Codebooks/LLM Codebook/LLM Coding - Content Analysis.xlsx

Reports PASS / FAIL per check with row-level detail when something fails.

Usage:
    python Nigeria/scripts/qa_llm_coding_xlsx.py
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Iterable

import openpyxl

ROOT = Path(__file__).resolve().parents[2]

AUD_XLSX = ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Audience Analysis.xlsx"
CON_XLSX = ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Content Analysis.xlsx"

HUM_AUD = ROOT / "Codebooks" / "Human Codebooks" / "audience" / "Human A - Audience Analysis Codebook.xlsx"
HUM_CON = ROOT / "Codebooks" / "Human Codebooks" / "content"  / "Human A - Content Analysis Codebook.xlsx"


# ──────────────────────────────────────────────────────────────────────────────
# small helpers
# ──────────────────────────────────────────────────────────────────────────────

class Result:
    def __init__(self):
        self.checks: list[tuple[int, str, bool, str]] = []

    def record(self, num: int, name: str, ok: bool, detail: str = ""):
        self.checks.append((num, name, ok, detail))

    def report(self):
        passed = sum(1 for _, _, ok, _ in self.checks if ok)
        total = len(self.checks)
        print(f"\n{'=' * 78}")
        print(f"QA RESULT: {passed}/{total} checks passed")
        print('=' * 78)
        for num, name, ok, detail in self.checks:
            mark = '✓ PASS' if ok else '✗ FAIL'
            print(f"  {num:2d}. {mark}  {name}")
            if detail:
                for line in detail.splitlines():
                    print(f"          {line}")


def headers_of(path: Path, sheet: str | None = None, header_row: int = 1) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    return [c.value for c in ws[header_row]]


def rows_of(path: Path, sheet: str) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or all(v is None for v in r):
            continue
        rows.append(dict(zip(header, r)))
    return header, rows


def pick(d: dict, prefix: str) -> tuple[str, object]:
    """find the first column whose first token == prefix (e.g., 'Q4', 'Q21a')."""
    for k, v in d.items():
        if not isinstance(k, str): continue
        first = k.split('.')[0].strip()
        if first.lower() == prefix.lower():
            return k, v
    raise KeyError(f"no column starting with {prefix!r}")


# ──────────────────────────────────────────────────────────────────────────────
# vocab definitions
# ──────────────────────────────────────────────────────────────────────────────

YN = {'Yes', 'No', '', None}

AUD_Q1_OPTS = {'Positive', 'Negative', 'Neutral', 'Unclear'}
AUD_Q5_OPTS = {'Positive', 'Negative', 'Neutral', 'Unclear', 'Does not mention men/masculinity'}
AUD_Q6_OPTS = {'Positive', 'Negative', 'Neutral', 'Unclear', 'Does not mention women/femininity'}
AUD_Q8_OPTS = {'Supporting', 'Challenging', 'Neutral', 'Unclear'}
AUD_Q21H_OPTS = {'Female', 'Male', 'Non-binary', 'Other', 'Unclear', '', None}

CON_SENT_OPTS = {'Negative', 'Positive', 'Mixed', 'Neutral', 'Unclear', 'Not mentioned'}
CON_Q4_OPTS = {'Yes, explicitly', 'Yes, implicitly', 'No'}
CON_Q5_OPTS = {
    'More regressive/traditional/restrictive',
    'More progressive/equitable/expansive',
    'Mixed/unclear',
    'Does not address masculinity or gender norms',
}
CON_Q6_OPTS = {'Yes', 'No', 'Unclear', 'Not applicable'}
CON_Q17_OPTS = {'Yes', 'Somewhat', 'No'}
CON_Q18_OPTS = {'Yes', 'No'}


# ──────────────────────────────────────────────────────────────────────────────
# checks
# ──────────────────────────────────────────────────────────────────────────────

def check_01_aud_headers_match_human(res: Result):
    hum = headers_of(HUM_AUD, header_row=2)
    llm_ng = headers_of(AUD_XLSX, sheet='Nigeria - LLM Coding')
    llm_ke = headers_of(AUD_XLSX, sheet='Kenya - LLM Coding')
    diffs = []
    if llm_ng != hum:
        diffs.append(f'Nigeria sheet ≠ human ({len(llm_ng)} vs {len(hum)} cols)')
    if llm_ke != hum:
        diffs.append(f'Kenya sheet ≠ human ({len(llm_ke)} vs {len(hum)} cols)')
    res.record(1, 'Audience headers match human codebook (41 cols, identical text)',
               not diffs, '\n'.join(diffs))


def check_02_con_headers_match_human(res: Result):
    hum = headers_of(HUM_CON, header_row=2)
    llm_ng = headers_of(CON_XLSX, sheet='Nigeria - LLM Coding')
    llm_ke = headers_of(CON_XLSX, sheet='Kenya - LLM Coding')
    diffs = []
    if llm_ng != hum:
        diffs.append(f'Nigeria sheet ≠ human ({len(llm_ng)} vs {len(hum)} cols)')
    if llm_ke != hum:
        diffs.append(f'Kenya sheet ≠ human ({len(llm_ke)} vs {len(hum)} cols)')
    res.record(2, 'Content headers match human codebook (33 cols, identical text)',
               not diffs, '\n'.join(diffs))


def check_03_per_workbook_sheet_consistency(res: Result):
    """Both country sheets in each workbook have identical schemas."""
    bad = []
    for name, path in [('Audience', AUD_XLSX), ('Content', CON_XLSX)]:
        ng = headers_of(path, sheet='Nigeria - LLM Coding')
        ke = headers_of(path, sheet='Kenya - LLM Coding')
        if ng != ke:
            bad.append(f'{name}: Nigeria/Kenya sheets differ ({len(ng)} vs {len(ke)} cols)')
    res.record(3, 'Each workbook: Nigeria + Kenya sheets share identical headers',
               not bad, '\n'.join(bad))


def check_04_aud_q1_vocab(res: Result):
    """Audience Q1 (sentiment) ∈ {Positive, Negative, Neutral, Unclear}."""
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(AUD_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            _, v = pick(r, 'Q1')
            if v not in AUD_Q1_OPTS:
                bad.append(f'{sheet} row {i}: Q1={v!r}')
    res.record(4, 'Audience Q1 (overall sentiment) — values within closed vocab',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_05_con_q14_q15_q16_vocab(res: Result):
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(CON_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            for q in ('Q14', 'Q15', 'Q16'):
                _, v = pick(r, q)
                if v not in CON_SENT_OPTS:
                    bad.append(f'{sheet} row {i}: {q}={v!r}')
    res.record(5, 'Content Q14/Q15/Q16 (sentiment) — values within closed 6-value vocab',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_06_yes_no_fields(res: Result):
    """All bare Yes/No fields contain only 'Yes' or 'No'."""
    bad = []
    # audience yes/no fields
    aud_yn = ['Q4', 'Q12', 'Q13', 'Q14', 'Q15', 'Q16', 'Q17', 'Q18', 'Q19', 'Q20',
              'Q21', 'Q21a', 'Q21c', 'Q21e', 'Q21g']
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(AUD_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            for q in aud_yn:
                _, v = pick(r, q)
                if v not in YN:
                    bad.append(f'AUD/{sheet} row {i}: {q}={v!r}')
    # content yes/no fields
    con_yn = ['Q1', 'Q18']
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(CON_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            for q in con_yn:
                _, v = pick(r, q)
                if v not in YN:
                    bad.append(f'CON/{sheet} row {i}: {q}={v!r}')
    res.record(6, 'All Yes/No fields contain only Yes / No (audience + content)',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_07_aud_q4_consistency(res: Result):
    """Audience: if Q4=No → Q5='Does not mention men/masculinity', Q6='Does not mention women/femininity'.
       Conversely if Q5/Q6 carry a sentiment label, Q4 must = Yes."""
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(AUD_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            _, q4 = pick(r, 'Q4')
            _, q5 = pick(r, 'Q5')
            _, q6 = pick(r, 'Q6')
            if q4 == 'No':
                if q5 != 'Does not mention men/masculinity':
                    bad.append(f'{sheet} r{i}: Q4=No but Q5={q5!r}')
                if q6 != 'Does not mention women/femininity':
                    bad.append(f'{sheet} r{i}: Q4=No but Q6={q6!r}')
            else:
                # Q4 = Yes → Q5/Q6 should not be the "Does not mention" sentinel
                if q5 == 'Does not mention men/masculinity' and q6 == 'Does not mention women/femininity':
                    bad.append(f'{sheet} r{i}: Q4={q4!r} but both Q5/Q6 = "Does not mention"')
    res.record(7, 'Audience Q4 ↔ Q5/Q6 internal consistency',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_08_aud_q11_required_when_challenging(res: Result):
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(AUD_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            _, q8 = pick(r, 'Q8')
            _, q11 = pick(r, 'Q11')
            if q8 == 'Challenging' and (q11 is None or str(q11).strip() == ''):
                bad.append(f'{sheet} r{i}: Q8=Challenging but Q11 empty')
    res.record(8, 'Audience Q11 required (non-empty) when Q8 = Challenging',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_09_aud_q21_consistency(res: Result):
    """Q21 must = Yes whenever any of Q21a/c/e/g = Yes."""
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(AUD_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            _, q21 = pick(r, 'Q21')
            sub = []
            for q in ('Q21a', 'Q21c', 'Q21e', 'Q21g'):
                _, v = pick(r, q)
                sub.append(v)
            if 'Yes' in sub and q21 != 'Yes':
                bad.append(f'{sheet} r{i}: Q21={q21!r} but sub={sub}')
            if q21 == 'No' and 'Yes' in sub:
                bad.append(f'{sheet} r{i}: Q21=No but sub has Yes: {sub}')
    res.record(9, 'Audience Q21 ↔ Q21a/c/e/g internal consistency',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_10_aud_q21_text_gates(res: Result):
    """Q21b/d/f/h text fields must be empty when their gates (Q21a/c/e/g) = No."""
    bad = []
    pairs = [('Q21a', 'Q21b'), ('Q21c', 'Q21d'), ('Q21e', 'Q21f'), ('Q21g', 'Q21h')]
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(AUD_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            for gate, txt in pairs:
                _, gv = pick(r, gate)
                _, tv = pick(r, txt)
                if gv != 'Yes' and tv not in ('', None):
                    bad.append(f'{sheet} r{i}: {gate}={gv!r} but {txt}={tv!r}')
    res.record(10, 'Audience Q21b/d/f/h text fields empty when their gate = No',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_11_con_q4_q5_q6_q7_consistency(res: Result):
    """Content: Q4=No → Q5='Does not address...', Q6='Not applicable', Q7 contains only 'Not applicable'."""
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(CON_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            _, q4 = pick(r, 'Q4')
            _, q5 = pick(r, 'Q5')
            _, q6 = pick(r, 'Q6')
            _, q7 = pick(r, 'Q7')
            if q4 == 'No':
                if q5 != 'Does not address masculinity or gender norms':
                    bad.append(f'{sheet} r{i}: Q4=No but Q5={q5!r}')
                if q6 != 'Not applicable':
                    bad.append(f'{sheet} r{i}: Q4=No but Q6={q6!r}')
                # Q7 is multi-select stored as semi-colon string; should be exactly 'Not applicable'
                q7_items = [x.strip() for x in str(q7 or '').split(';') if x.strip()]
                if q7_items and q7_items != ['Not applicable']:
                    bad.append(f'{sheet} r{i}: Q4=No but Q7={q7_items}')
    res.record(11, 'Content Q4=No → Q5/Q6/Q7 set to their "not applicable" defaults',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_12_con_other_gates(res: Result):
    """Content: open-text 'a' / 'b' columns must be empty unless the gate parent contains 'Other'."""
    pairs = [
        ('Q1a', 'Q1b'),    # Q1b empty unless Q1a contains Other
        ('Q2', 'Q2a'),
        ('Q3', 'Q3a'),
        ('Q7', 'Q7a'),
        ('Q8', 'Q8a'),
        ('Q9', 'Q9a'),
        ('Q10', 'Q10a'),
        ('Q12', 'Q12a'),
        ('Q13', 'Q13a'),
        ('Q18a', 'Q18b'),  # Q18b empty unless Q18a contains Other
    ]
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(CON_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            for gate, txt in pairs:
                _, gv = pick(r, gate)
                _, tv = pick(r, txt)
                tv_stripped = str(tv or '').strip()
                gate_str = str(gv or '')
                if 'Other' not in gate_str and tv_stripped:
                    bad.append(f'{sheet} r{i}: {gate}={gv!r} but {txt}={tv!r}')
    res.record(12, 'Content "Other" open-text gates: a/b column empty unless parent contains Other',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_13_con_q18_gates(res: Result):
    """Content: Q18=No → Q18a empty list AND Q18b empty."""
    bad = []
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(CON_XLSX, sheet)
        for i, r in enumerate(rows, start=2):
            _, q18 = pick(r, 'Q18')
            _, q18a = pick(r, 'Q18a')
            _, q18b = pick(r, 'Q18b')
            if q18 == 'No':
                if str(q18a or '').strip():
                    bad.append(f'{sheet} r{i}: Q18=No but Q18a={q18a!r}')
                if str(q18b or '').strip():
                    bad.append(f'{sheet} r{i}: Q18=No but Q18b={q18b!r}')
    res.record(13, 'Content Q18=No → Q18a + Q18b empty',
               not bad, f'{len(bad)} violations: {bad[:3]}' if bad else '')


def check_14_sample_coverage(res: Result):
    """Audience: 50/creator (200/sheet). Content: ≤50/creator and Σ matches."""
    detail = []
    ok = True

    # audience: source-row count is the LLM-sample comment_id; group by URL→creator is fragile,
    # so just assert exactly 200 rows per sheet.
    for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
        _, rows = rows_of(AUD_XLSX, sheet)
        if len(rows) != 200:
            ok = False
            detail.append(f'AUD/{sheet}: {len(rows)} rows (expected 200)')

    # content: per-creator counts ≤ 50 and totals match what the notebook reported
    expected_total = {'Nigeria - LLM Coding': 260, 'Kenya - LLM Coding': 241}
    for sheet, expect in expected_total.items():
        _, rows = rows_of(CON_XLSX, sheet)
        # use Content ID prefix (AGB_, BNK_, etc. for Nigeria; KIBE-, RIX-, etc. for Kenya)
        # plus a numeric tweet ID for Eric/Amerix
        per = Counter()
        for r in rows:
            cid = str(r.get('Content ID') or '')
            # strip after first non-letter so AGB_001 → AGB, KIBE-005 → KIBE
            prefix = ''
            for ch in cid:
                if ch.isalpha(): prefix += ch
                else: break
            per[prefix or 'NUMERIC'] += 1
        if len(rows) != expect:
            ok = False
            detail.append(f'CON/{sheet}: {len(rows)} rows (expected {expect}); per-prefix={dict(per)}')
        elif any(v > 50 for v in per.values()):
            ok = False
            detail.append(f'CON/{sheet}: per-creator over 50 cap: {dict(per)}')
    res.record(14, 'Sample-size coverage: audience 200/sheet, content per-creator ≤ 50',
               ok, '\n'.join(detail))


def check_15_no_blank_text_no_dup_ids(res: Result):
    """Every row has non-empty primary text + ID. No duplicate IDs within a sheet."""
    bad = []
    cases = [
        (AUD_XLSX, 'Comment ID', 'Comment Text'),
        (CON_XLSX, 'Content ID', 'Content Text / Description'),
    ]
    for path, id_col, text_col in cases:
        for sheet in ('Nigeria - LLM Coding', 'Kenya - LLM Coding'):
            _, rows = rows_of(path, sheet)
            ids = []
            for i, r in enumerate(rows, start=2):
                cid = r.get(id_col)
                txt = r.get(text_col)
                if cid is None or str(cid).strip() == '':
                    bad.append(f'{path.name}/{sheet} r{i}: blank {id_col}')
                if txt is None or str(txt).strip() == '':
                    bad.append(f'{path.name}/{sheet} r{i}: blank {text_col}')
                ids.append(str(cid))
            dups = [k for k, v in Counter(ids).items() if v > 1]
            if dups:
                bad.append(f'{path.name}/{sheet}: duplicate {id_col}s: {dups[:5]}')
    res.record(15, 'No blank Comment/Content text or ID; no duplicate IDs within a sheet',
               not bad, '\n'.join(bad[:5]) if bad else '')


# ──────────────────────────────────────────────────────────────────────────────
# main
# ──────────────────────────────────────────────────────────────────────────────

def main():
    res = Result()
    for fn in [
        check_01_aud_headers_match_human,
        check_02_con_headers_match_human,
        check_03_per_workbook_sheet_consistency,
        check_04_aud_q1_vocab,
        check_05_con_q14_q15_q16_vocab,
        check_06_yes_no_fields,
        check_07_aud_q4_consistency,
        check_08_aud_q11_required_when_challenging,
        check_09_aud_q21_consistency,
        check_10_aud_q21_text_gates,
        check_11_con_q4_q5_q6_q7_consistency,
        check_12_con_other_gates,
        check_13_con_q18_gates,
        check_14_sample_coverage,
        check_15_no_blank_text_no_dup_ids,
    ]:
        try:
            fn(res)
        except Exception as e:
            res.record(len(res.checks) + 1, fn.__name__, False, f'EXCEPTION: {e}')
    res.report()


if __name__ == "__main__":
    main()
