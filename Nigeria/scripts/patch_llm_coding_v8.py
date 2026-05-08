"""v8 patches based on the latest recheck audit:
1. Q20 broader correction-pattern detection (insults + negations + explicit rejections)
2. Q21=Yes with no sub-field=Yes → set Q21=No (Nigeria AGB_112, AGB_081)
"""
from __future__ import annotations
import re, hashlib
from pathlib import Path
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Codebooks" / "LLM Codebook" / "LLM Coding - Audience Analysis.xlsx"

# higher-confidence Q20 correction patterns
# only these flips will be auto-applied; everything else stays as-is
Q20_HIGH_CONF = re.compile(
    r'('
    # insult addressed at someone (signals correction of their claim)
    r'\byou\s+(?:dumb|stupid|idiot|foolish|fool|silly|toad|moron|imbecile|delusional|brainwashed|crazy|mad|sick|ridiculous|pathetic)\b|'
    r'\b(?:dumb|stupid|silly|foolish)\s+(?:toad|fool|idiot|woman|man|guy|girl|comment|statement|argument|take|opinion|reasoning|thread|post|tweet)\b|'
    # explicit rejection / common-sense pushback phrases
    r"\bthat'?s\s+(?:not|absolutely\s+not|simply\s+not|hardly|absurd|ridiculous|untrue|false|nonsense|rubbish|crap|bullshit|trash|stupid)\b|"
    r"\bwhat\s+(?:nonsense|rubbish|crap|bullshit|trash|absurdity|garbage)\b|"
    r"\bmakes?\s+no\s+sense\b|"
    r"\bdoesn'?t\s+make\s+sense\b|"
    r"\bnot\s+how\s+it\s+works\b|"
    r"\bthis\s+is\s+(?:wrong|incorrect|false|misleading|absurd)\b|"
    r"\b(?:you|u)'?re\s+(?:wrong|incorrect|mistaken|lying|delusional)\b|"
    # direct "X are not Y" rejection of a claim/analogy
    r"\b(?:women|men|wives|husbands|girls|boys|people)\s+are\s+not\s+\w+\b|"
    # disagree
    r"\bi\s+strongly\s+disagree\b|"
    r"\bi\s+disagree\b|"
    r"\b(?:nope|nah)[,!.\s]"
    r')',
    re.IGNORECASE,
)

wb = openpyxl.load_workbook(XLSX)
fixes_total = {'q20_flip': 0, 'q21_to_no': 0}

for sheet_name in ['Nigeria - LLM Coding', 'Kenya - LLM Coding']:
    ws = wb[sheet_name]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    df = pd.DataFrame(rows, columns=hdr)

    q20  = [c for c in df.columns if c.startswith('Q20.')][0]
    q20a = [c for c in df.columns if c.startswith('Q20a')][0]
    q21  = [c for c in df.columns if c.startswith('Q21.')][0]
    q21a = [c for c in df.columns if c.startswith('Q21a.')][0]
    q21c = [c for c in df.columns if c.startswith('Q21c.')][0]
    q21e = [c for c in df.columns if c.startswith('Q21e.')][0]
    q21g = [c for c in df.columns if c.startswith('Q21g.')][0]

    fix = {'q20_flip': 0, 'q21_to_no': 0, 'examples_q20': []}

    # 1. Q20 broader correction detection
    for i in df.index:
        if df.at[i, q20] != 'No':
            continue
        text = str(df.at[i, 'Comment Text'] or '')
        m = Q20_HIGH_CONF.search(text)
        if m:
            df.at[i, q20] = 'Yes'
            if not str(df.at[i, q20a] or '').strip():
                # quote a brief snippet around the matched phrase
                start = max(0, m.start() - 15); end = min(len(text), m.end() + 50)
                snip = text[start:end].replace('\n', ' ').strip()
                df.at[i, q20a] = f'Disputes/rejects: "{snip[:120]}"'
            fix['q20_flip'] += 1
            if len(fix['examples_q20']) < 3:
                fix['examples_q20'].append(f"  [{df.at[i,'Comment ID']}] {text[:120]}")

    # 2. Q21=Yes but all subs=No → flip Q21=No
    for i in df.index:
        if df.at[i, q21] == 'Yes' and all(df.at[i, c] != 'Yes' for c in (q21a, q21c, q21e, q21g)):
            df.at[i, q21] = 'No'
            fix['q21_to_no'] += 1

    # write back
    for ridx, row_data in enumerate(df.itertuples(index=False), start=2):
        for cidx, v in enumerate(row_data, start=1):
            ws.cell(row=ridx, column=cidx).value = v

    print(f"[{sheet_name}] q20 flips: {fix['q20_flip']}; q21 -> No: {fix['q21_to_no']}")
    for ex in fix['examples_q20']:
        print(ex)
    fixes_total['q20_flip'] += fix['q20_flip']
    fixes_total['q21_to_no'] += fix['q21_to_no']

# update Run Info — bump version + new sha
ri = wb['Run Info']
for row in ri.iter_rows():
    if row[0].value == 'VERSION':
        row[1].value = 'v8 — final (post-recheck-audit-2 patches: q20 broader + q21 reconciled)'
    elif row[0].value == 'Audit fixes applied (v7)':
        row[1].value = 'v7 — Q4 keyword repair, Q7-Other-Q7A gating, Q12->Q18 cross-check, Q20 correction-phrase regex, Q11 LLM repair, Methodology completed'

# add v8 row if not present
new_v8_label = 'Audit fixes applied (v8)'
new_v8_value = 'v8 — Q20 broader pattern detection (insults + negations + explicit rejections); Q21=Yes but all sub-fields=No reconciled to Q21=No (AGB_112, AGB_081)'
found = False
for row in ri.iter_rows():
    if row[0].value == new_v8_label:
        row[1].value = new_v8_value; found = True
if not found:
    ri.append([new_v8_label, new_v8_value])

wb.save(XLSX)

# update SHA in Run Info too
new_sha = hashlib.sha256(open(XLSX, 'rb').read()).hexdigest()[:12]
wb2 = openpyxl.load_workbook(XLSX)
ri2 = wb2['Run Info']
for row in ri2.iter_rows():
    if row[0].value == 'SHA-256 (first 12)':
        row[1].value = new_sha
wb2.save(XLSX)
print(f"\ntotal: {fixes_total}")
print(f"new SHA-256: {hashlib.sha256(open(XLSX,'rb').read()).hexdigest()[:12]}")
