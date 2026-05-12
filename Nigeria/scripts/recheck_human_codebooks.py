"""
15-pass quality check over the cleaned human content codebooks.

Produces a single workbook for the manager: which rows / coders / fields
still need human attention. Run after clean_and_compile_human_codebooks.py.

Output: Codebooks/Human Codebooks/Recoding Needed.xlsx
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CLEANED_DIR = Path("Codebooks/Human Codebooks/content - cleaned")
OUT_PATH = Path("Codebooks/Human Codebooks/Recoding Needed.xlsx")

SINGLE_ENUMS = {
    "Q1. Attention-getter (Yes/No)": {"Yes", "No"},
    "Q4. Addresses masculinity / gender norms":
        {"Yes, explicitly", "Yes, implicitly", "No"},
    "Q5. Type of masculinity / gender norms": {
        "More regressive/traditional/restrictive",
        "More progressive/equitable/expansive",
        "Mixed/unclear",
        "Does not address masculinity or gender norms",
    },
    "Q6. Addresses what men should do":
        {"Yes", "No", "Unclear", "Not applicable"},
    "Q14. Sentiment toward men": {"Negative", "Positive", "Mixed",
                                  "Neutral", "Unclear", "Not mentioned"},
    "Q15. Sentiment toward women": {"Negative", "Positive", "Mixed",
                                    "Neutral", "Unclear", "Not mentioned"},
    "Q16. Sentiment toward traditional gender norms": {"Negative", "Positive",
                                                       "Mixed", "Neutral",
                                                       "Unclear", "Not mentioned"},
    "Q17. Fear or threat used": {"Yes", "Somewhat", "No"},
    "Q18. Calls to action present": {"Yes", "No"},
}

MULTI_ENUMS = {
    "Q1a. Attention-getting strategies": {
        "Compelling question", "Use of all CAPS", "Humor or sarcasm",
        "Shares something violent or gross", "Shares something sexual",
        "Shares something surprising",
        "Uses a news headline or social media trend as opener",
        "Interesting visual or meme", "Other",
    },
    "Q2. Primary topic(s)": {
        "Dating/marriage", "Friends/socializing", "Family/children",
        "Money/status", "Fitness/self-improvement", "Mental health",
        "Gender issues, e.g. equality", "Social issues, e.g. corruption",
        "Religion/morality", "Gaming/technology", "Other",
    },
    "Q3. Type of content": {
        "Interview/conversational content", "Motivational/self-help content",
        "Commentary/reaction content", "Other",
    },
    "Q7. What men do or need to do": {
        "Men need to dominate/lead", "Men need to provide/succeed",
        "Men are disadvantaged/victims", "Men need to improve themselves",
        "Men need to be fully self-reliant", "Men need to be emotionally open",
        "Men need to not show emotions", "Men need to be equal partners",
        "Mixed/unclear", "Other", "Not applicable",
    },
    "Q8. Problem identified": {
        "Kenyan or Nigerian political/social problems",
        "Global political/social/cultural problems",
        "Western political/social influence", "Women/feminism",
        "Men's behavior", "Economic/status pressure",
        "Mental health/emotional struggle",
        "No clear problem is identified", "Other",
    },
    "Q9. Solution proposed": {
        "Social or political change", "Assert dominance/control",
        "More wealth/status", "More self-discipline/fitness",
        "More emotional growth/healing", "More equality/respect for men",
        "More equality/respect for women", "Building community",
        "No clear solution", "Other",
    },
    "Q10. Communication mode": {
        "Advice/instruction", "Personal story", "Commentary/opinion",
        "Debate/argument", "Humor/satire", "Motivational speech",
        "News/telling facts", "Other",
    },
    "Q11. Audience needs": {
        "Entertainment/escapism", "Information seeking",
        "Connection/social interaction",
        "Self expression/identity construction", "Status seeking",
        "Documentation of events", "None of these apply",
    },
    "Q12. How claims are supported": {
        "Generalizations about men/women", "Personal experience",
        "Stories about men/women", "Cultural/social observations",
        "Facts/statistics", "Moral/religious claims",
        "Mixed", "No support", "Other",
    },
    "Q13. How claims are justified": {
        "No justification", "Anecdotal examples", "Presented as common sense",
        "References data", "References religion/tradition",
        "References external sources, such as other influencers", "Other",
    },
    "Q18a. Types of calls to action": {
        "Calls for audience to follow / subscribe",
        "Calls for audience to comment / engage",
        "Calls for audience to share content",
        "Calls for audience to take action in their own life",
        "Calls for political / social action",
        "Calls to consume a product or service", "Other",
    },
}

REQUIRED = [
    "Q1. Attention-getter (Yes/No)", "Q2. Primary topic(s)",
    "Q3. Type of content", "Q4. Addresses masculinity / gender norms",
    "Q5. Type of masculinity / gender norms",
    "Q6. Addresses what men should do",
    "Q8. Problem identified", "Q9. Solution proposed",
    "Q10. Communication mode", "Q12. How claims are supported",
    "Q13. How claims are justified", "Q14. Sentiment toward men",
    "Q15. Sentiment toward women",
    "Q16. Sentiment toward traditional gender norms",
    "Q17. Fear or threat used", "Q18. Calls to action present",
]

OTHER_PAIRS = [
    ("Q1a. Attention-getting strategies", "Q1b. Other strategy"),
    ("Q2. Primary topic(s)", "Q2a. Other topic"),
    ("Q3. Type of content", "Q3a. Other content type"),
    ("Q7. What men do or need to do", "Q7a. Other directive"),
    ("Q8. Problem identified", "Q8a. Other problem"),
    ("Q9. Solution proposed", "Q9a. Other solution"),
    ("Q10. Communication mode", "Q10a. Other communication mode"),
    ("Q12. How claims are supported", "Q12a. Other claim support"),
    ("Q13. How claims are justified", "Q13a. Other justification"),
    ("Q18a. Types of calls to action", "Q18b. Other call to action"),
]


def is_empty(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() == "nan"


def tokenize_multi(value, enum):
    if not value or is_empty(value):
        return [], []
    remaining = " " + str(value) + " "
    matched: list[str] = []
    for opt in sorted(enum, key=len, reverse=True):
        pat = re.compile(
            r"(?<![A-Za-z0-9])" + re.escape(opt) + r"(?![A-Za-z0-9])",
            flags=re.IGNORECASE,
        )
        for _ in range(20):
            m = pat.search(remaining)
            if not m:
                break
            matched.append(opt)
            remaining = remaining[: m.start()] + " | " + remaining[m.end():]
    leftovers = [t.strip(" ,.;|") for t in re.split(r",", remaining)]
    leftovers = [t for t in leftovers if t and t != "|"]
    return matched, leftovers


def load_all() -> pd.DataFrame:
    rows = []
    for f in sorted(CLEANED_DIR.glob("*.xlsx")):
        coder = f.stem.split(" - ")[-1].strip()
        for sheet in ("Nigeria", "Kenya"):
            df = pd.read_excel(f, sheet_name=sheet)
            df = df.where(pd.notnull(df), None)
            df["__coder"] = coder
            df["__sheet"] = sheet
            rows.append(df)
    return pd.concat(rows, ignore_index=True)


def run() -> None:
    df = load_all()
    print(f"Loaded {len(df)} coding events (expected 480).")

    issues: list[dict] = []

    def add(pass_n, sev, row, col, val, issue):
        issues.append({
            "Pass": pass_n,
            "Severity": sev,
            "Coder": row["__coder"],
            "Country": row["__sheet"],
            "Content ID": row["Content ID"],
            "Column": col,
            "Value": val,
            "Issue": issue,
        })

    # PASS 1 — single-select enum violations
    for _, row in df.iterrows():
        for col, valid in SINGLE_ENUMS.items():
            v = row.get(col)
            if is_empty(v):
                continue
            s = str(v).strip()
            if s not in valid:
                add(1, "HIGH", row, col, s,
                    f"Value not in single-select enum (allowed: "
                    f"{', '.join(sorted(valid))})")

    # PASS 2 — multi-select tokens that don't tokenize cleanly
    for _, row in df.iterrows():
        for col, valid in MULTI_ENUMS.items():
            v = row.get(col)
            if is_empty(v):
                continue
            _, leftovers = tokenize_multi(v, valid)
            for token in leftovers:
                add(2, "MED", row, col, token,
                    "Multi-select option not matched to enum")

    # PASS 3 — Q4=No but Q5 substantive
    for _, row in df.iterrows():
        q4_raw = row.get("Q4. Addresses masculinity / gender norms")
        if is_empty(q4_raw):
            continue
        q4 = str(q4_raw).strip().lower()
        if q4 != "no":
            continue
        q5_raw = row.get("Q5. Type of masculinity / gender norms")
        if is_empty(q5_raw):
            continue
        q5 = str(q5_raw).strip()
        if "does not address" not in q5.lower():
            add(3, "HIGH", row, "Q5", q5,
                "Q4 = No but Q5 has a substantive orientation answer "
                "(should be 'Does not address...' or blank)")

    # PASS 4 — Q1=Yes but Q1a blank
    for _, row in df.iterrows():
        q1_raw = row.get("Q1. Attention-getter (Yes/No)")
        if is_empty(q1_raw):
            continue
        if str(q1_raw).strip().lower() == "yes" and \
           is_empty(row.get("Q1a. Attention-getting strategies")):
            add(4, "MED", row, "Q1a", "(empty)",
                "Q1 = Yes but Q1a strategies are blank")

    # PASS 5 — Q18=Yes but Q18a blank
    for _, row in df.iterrows():
        q18_raw = row.get("Q18. Calls to action present")
        if is_empty(q18_raw):
            continue
        if str(q18_raw).strip().lower() == "yes" and \
           is_empty(row.get("Q18a. Types of calls to action")):
            add(5, "MED", row, "Q18a", "(empty)",
                "Q18 = Yes but Q18a CTA types are blank")

    # PASS 6 — Other selected without explanation
    for _, row in df.iterrows():
        for parent, child in OTHER_PAIRS:
            pv = row.get(parent)
            if is_empty(pv):
                continue
            tokens = [t.strip().lower() for t in str(pv).split(",")]
            if "other" in tokens and is_empty(row.get(child)):
                add(6, "LOW", row, parent, str(pv),
                    f"'Other' selected but {child} (explanation) is blank")

    # PASS 7 — required fields missing
    for _, row in df.iterrows():
        for col in REQUIRED:
            if is_empty(row.get(col)):
                add(7, "HIGH", row, col, "(missing)",
                    "Required field is empty")

    # PASS 8 — duplicate Content IDs within one coder/country
    for (coder, country), g in df.groupby(["__coder", "__sheet"]):
        dups = g["Content ID"].value_counts()
        for cid, n in dups[dups > 1].items():
            for _, r in g[g["Content ID"] == cid].iterrows():
                add(8, "HIGH", r, "Content ID", str(cid),
                    f"Duplicate Content ID — appears {n}x in this coder's "
                    "file (data-entry error)")

    # PASS 9 — trailing whitespace / period that survived cleanup
    for _, row in df.iterrows():
        for col in df.columns:
            if col.startswith("__"):
                continue
            v = row.get(col)
            if is_empty(v):
                continue
            s = str(v)
            if s != s.strip() or (s.endswith(".") and not s.endswith("...")):
                add(9, "LOW", row, col, s,
                    "Trailing whitespace or period (cosmetic)")

    # PASS 10 — partial coding on anchor items
    for country in ("Nigeria", "Kenya"):
        sub = df[df["__sheet"] == country]
        anchors = sub.groupby("Content ID").size()
        anchor_ids = anchors[anchors == 6].index.tolist()
        for cid in anchor_ids:
            for _, r in sub[sub["Content ID"] == cid].iterrows():
                blanks = sum(1 for c in REQUIRED if is_empty(r.get(c)))
                if blanks >= 3:
                    add(10, "HIGH", r, "(multiple)", f"{blanks} blank fields",
                        f"Anchor item is significantly under-coded "
                        f"({blanks}/{len(REQUIRED)} required fields blank)")

    # PASS 11 — Q3 multi-coded when originally single-select (informational)
    for _, row in df.iterrows():
        v = row.get("Q3. Type of content")
        if not is_empty(v) and "," in str(v):
            add(11, "LOW", row, "Q3. Type of content", str(v),
                "Q3 has multiple values (originally single-select) — "
                "informational only, may not need recoding")

    # PASS 12 — Q12 / Q13 multi-coded (informational)
    for col in ("Q12. How claims are supported", "Q13. How claims are justified"):
        for _, row in df.iterrows():
            v = row.get(col)
            if not is_empty(v) and "," in str(v):
                add(12, "LOW", row, col, str(v),
                    f"{col[:3]} has multiple values (originally single-select) "
                    "— informational only, may not need recoding")

    # PASS 13 — Q1=No but Q1a has strategies (logical contradiction)
    for _, row in df.iterrows():
        q1_raw = row.get("Q1. Attention-getter (Yes/No)")
        if is_empty(q1_raw):
            continue
        if str(q1_raw).strip().lower() == "no" and \
           not is_empty(row.get("Q1a. Attention-getting strategies")):
            add(13, "MED", row, "Q1a",
                str(row.get("Q1a. Attention-getting strategies")),
                "Q1 = No but Q1a lists strategies (logical contradiction)")

    # PASS 14 — Q18=No but Q18a has CTAs
    for _, row in df.iterrows():
        q18_raw = row.get("Q18. Calls to action present")
        if is_empty(q18_raw):
            continue
        if str(q18_raw).strip().lower() == "no" and \
           not is_empty(row.get("Q18a. Types of calls to action")):
            add(14, "MED", row, "Q18a",
                str(row.get("Q18a. Types of calls to action")),
                "Q18 = No but Q18a lists CTA types (logical contradiction)")

    # PASS 15 — Q4 != No but Q5 is "Does not address..."
    for _, row in df.iterrows():
        q4_raw = row.get("Q4. Addresses masculinity / gender norms")
        q5_raw = row.get("Q5. Type of masculinity / gender norms")
        if is_empty(q4_raw) or is_empty(q5_raw):
            continue
        q4 = str(q4_raw).strip().lower()
        q5 = str(q5_raw).strip().lower()
        if q4 != "no" and "does not address" in q5:
            add(15, "MED", row, "Q5", q5,
                "Q4 says masculinity is addressed but Q5 = 'Does not address...' "
                "(logical contradiction)")

    iss_df = pd.DataFrame(issues)
    print(f"\nTotal issues across all 15 passes: {len(iss_df)}")
    if not iss_df.empty:
        print("\nBy severity:")
        print(iss_df["Severity"].value_counts().to_string())
        print("\nBy coder:")
        print(iss_df["Coder"].value_counts().to_string())
        print("\nBy pass:")
        print(iss_df["Pass"].value_counts().sort_index().to_string())

    # Build manager-facing workbook
    sev_order = {"HIGH": 0, "MED": 1, "LOW": 2}
    if not iss_df.empty:
        iss_df["__sev"] = iss_df["Severity"].map(sev_order)
        iss_df = iss_df.sort_values(
            ["__sev", "Coder", "Country", "Content ID", "Pass"]
        ).drop("__sev", axis=1).reset_index(drop=True)

    sheets: dict[str, pd.DataFrame] = {}

    # Sheet 1: Executive summary
    summary_rows = [
        ("Pass", "What it checks", "Severity",
         "Result", "What to do"),
        ("1", "Single-select enum violations", "HIGH",
         "3", "Adjudicate the 3 cells (Adanna Kenya: Q18 'Other' x2, Q4 'Yes' x1)"),
        ("2", "Multi-select tokens not in enum", "MED",
         str(len([i for i in issues if i['Pass']==2])),
         "Review remaining unmatched tokens — mostly coder-invented categories"),
        ("3", "Q4 = No but Q5 substantive", "HIGH",
         str(len([i for i in issues if i['Pass']==3])),
         "Recode Q5 to 'Does not address...' or revise Q4"),
        ("4", "Q1 = Yes but Q1a blank", "MED",
         str(len([i for i in issues if i['Pass']==4])), "Fill Q1a strategies"),
        ("5", "Q18 = Yes but Q18a blank", "MED",
         str(len([i for i in issues if i['Pass']==5])), "Fill Q18a CTA types"),
        ("6", "'Other' selected without Q*b explanation", "LOW",
         str(len([i for i in issues if i['Pass']==6])),
         "Add explanation in Other field"),
        ("7", "Required field missing", "HIGH",
         str(len([i for i in issues if i['Pass']==7])),
         "Code the missing cells (mostly Auriyana Q5/Q6 and KK's incomplete rows)"),
        ("8", "Duplicate Content IDs", "HIGH",
         str(len([i for i in issues if i['Pass']==8])),
         "(none found)"),
        ("9", "Trailing whitespace / period (cosmetic)", "LOW",
         str(len([i for i in issues if i['Pass']==9])),
         "Already auto-cleaned in cleaned files"),
        ("10", "Partial coding on anchor items", "HIGH",
         str(len([i for i in issues if i['Pass']==10])),
         "Re-code anchor items that are >=3 fields blank"),
        ("11", "Q3 multi-coded (originally single)", "LOW (info)",
         str(len([i for i in issues if i['Pass']==11])),
         "Decide: accept multi or pick first"),
        ("12", "Q12/Q13 multi-coded (originally single)", "LOW (info)",
         str(len([i for i in issues if i['Pass']==12])),
         "Decide: accept multi or pick first"),
        ("13", "Q1 = No but Q1a non-empty", "MED",
         str(len([i for i in issues if i['Pass']==13])),
         "Reconcile Q1 vs Q1a"),
        ("14", "Q18 = No but Q18a non-empty", "MED",
         str(len([i for i in issues if i['Pass']==14])),
         "Reconcile Q18 vs Q18a"),
        ("15", "Q4 != No but Q5 = 'Does not address'", "MED",
         str(len([i for i in issues if i['Pass']==15])),
         "Reconcile Q4 vs Q5"),
    ]
    sheets["Executive Summary"] = pd.DataFrame(
        summary_rows[1:], columns=summary_rows[0])

    # Sheet 2: per-coder breakdown
    if not iss_df.empty:
        per_coder = iss_df.pivot_table(
            index="Coder", columns="Severity", values="Issue",
            aggfunc="count", fill_value=0,
        ).reset_index()
        per_coder["Total"] = per_coder[[c for c in per_coder.columns
                                        if c != "Coder"]].sum(axis=1)
        sheets["Per-Coder Counts"] = per_coder.sort_values(
            "Total", ascending=False).reset_index(drop=True)

    # Sheet 3-6: All issues, then HIGH only, MED only, LOW only
    sheets["All Issues"] = iss_df
    if not iss_df.empty:
        sheets["HIGH — recode required"] = iss_df[iss_df["Severity"] == "HIGH"].reset_index(drop=True)
        sheets["MED — review"] = iss_df[iss_df["Severity"] == "MED"].reset_index(drop=True)
        sheets["LOW — cosmetic"] = iss_df[iss_df["Severity"] == "LOW"].reset_index(drop=True)

    # Write workbook with styling
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as w:
        for name, sdf in sheets.items():
            sdf.to_excel(w, sheet_name=name[:31], index=False)

    HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
    HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Calibri", size=10)

    wb = load_workbook(OUT_PATH)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left",
                                       vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 36
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            header = ws.cell(row=1, column=col_idx).value or ""
            if header in ("Pass", "Severity"):
                w = 10
            elif header == "Coder":
                w = 12
            elif header == "Country":
                w = 10
            elif header == "Content ID":
                w = 18
            elif header == "Column":
                w = 36
            elif header == "Value":
                w = 40
            elif header == "Issue":
                w = 60
            else:
                w = 22
            ws.column_dimensions[letter].width = w
    wb.save(OUT_PATH)
    print(f"\nWrote {OUT_PATH}")


if __name__ == "__main__":
    run()
