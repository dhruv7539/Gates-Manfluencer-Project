"""
Append overlap rows (5 pairwise + 5 all-coder shared) to each Nigeria
human-coder spreadsheet, using the new template layout introduced in
coder A (DIRECTIONS banner row + Q-column coding form headers).

Layout per coder file (Option B - two blank-row gaps):
  Row 1            DIRECTIONS banner
  Row 2            Headers (data cols + Q columns)
  Rows 3..32       30 single-coded rows  (preserved from existing files)
  Rows 33..34      2 blank separator rows
  Rows 35..39      5 pairwise_overlap rows
  Rows 40..41      2 blank separator rows
  Rows 42..46      5 all_coder_shared rows  (same 5 across all 6 coders)

Final shape:  46 rows, full Q column form, 40 codable rows per coder.

Pairwise pattern (15 distinct rows per analysis type):
  AB x3, BC x2, CD x3, DE x2, EF x3, FA x2

Each coder ends up with 5 pairwise rows:
  A: 3 with B + 2 with F
  B: 3 with A + 2 with C
  C: 2 with B + 3 with D
  D: 3 with C + 2 with E
  E: 2 with D + 3 with F
  F: 3 with E + 2 with A

Overlap rows are picked from the same master pools the unique rows came
from but from rows BEYOND the head(quota) cutoff so they cannot collide
with the 180 single-coded rows already assigned.  A hard QA check
asserts no overlap original_row_id appears in any coder's unique set.
"""

from __future__ import annotations
import os
import random
import re
import zipfile
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ----------------------------- paths ---------------------------------------

ROOT = Path("/Users/sushildalavi/Desktop/NLC/Gates-Manfluencer-Project")
AUD_MASTER = ROOT / "Nigeria" / "Audience Analysis" / "Nigeria Audience Analysis Final.xlsx"
CON_MASTER = ROOT / "Nigeria" / "Content Analysis" / "Nigeria Content Analysis Final.xlsx"
OUT_BASE = ROOT / "Codebooks" / "human codebook"
OUT_AUD = OUT_BASE / "audience"
OUT_CON = OUT_BASE / "content"


# ----------------------------- constants -----------------------------------

SEED = 42
CODERS = ["A", "B", "C", "D", "E", "F"]

# Same quotas the original build script used (rows 0..quota-1 per creator
# went into the 180 single-coded set).  Overlap rows must come from rows
# quota..N-1.
QUOTA_AUD = {
    "Banky Wellington": 45,
    "Deyemi Okanlawon": 45,
    "Agba John Doe":    45,
    "Shola":            45,
}
QUOTA_CON = {
    "Banky Wellington":   33,
    "Deyemi Okanlawon":   24,
    "Ebuka Obi-Uchendu":  33,
    "Agba John Doe":      30,
    "Shola":              30,
    "Wizarab":            30,
}

ORIENTATION_AUD = {
    "Banky Wellington": "Progressive",
    "Deyemi Okanlawon": "Progressive",
    "Agba John Doe":    "Regressive",
    "Shola":            "Regressive",
}
ORIENTATION_CON = {
    "Banky Wellington":   "Progressive",
    "Deyemi Okanlawon":   "Progressive",
    "Ebuka Obi-Uchendu":  "Progressive",
    "Agba John Doe":      "Regressive",
    "Shola":              "Regressive",
    "Wizarab":            "Regressive",
}

# How many overlap rows (above the head(quota) cutoff) to pull per creator
# per analysis type.  20 total per type = 15 pairwise + 5 all-coder shared.
OVERLAP_PICK_AUD = {
    "Banky Wellington": 5,
    "Deyemi Okanlawon": 5,
    "Agba John Doe":    5,
    "Shola":            5,
}
OVERLAP_PICK_CON = {
    "Banky Wellington":   6,
    "Ebuka Obi-Uchendu":  4,
    "Agba John Doe":      4,
    "Shola":              3,
    "Wizarab":            3,
    # Deyemi has 0 rows beyond its quota - skipped.
}

# Pairwise pattern (15 rows total, sums to 5 per coder).
PAIRWISE_PATTERN = [
    ("A", "B", 3),
    ("B", "C", 2),
    ("C", "D", 3),
    ("D", "E", 2),
    ("E", "F", 3),
    ("F", "A", 2),
]

# Layout row indices.
ROW_DIRECTIONS = 1
ROW_HEADER = 2
ROW_UNIQUE_START = 3
ROW_UNIQUE_END = 32
ROW_PAIRWISE_START = 35
ROW_PAIRWISE_END = 39
ROW_ALLCODER_START = 42
ROW_ALLCODER_END = 46
TOTAL_ROWS = ROW_ALLCODER_END  # 46

# Columns that should get hyperlink styling when their value is a URL.
URL_COLUMNS = ("Source URL", "Commenter Post URL", "Influencer's OG Post URL")
URL_RE = re.compile(r"https?://[^\s)]+")
HYPERLINK_FONT = Font(color="0563C1", underline="single")
LINK_INLINE_FONT = InlineFont(color="0563C1", u="single")


# ----------------------------- master pool loading -------------------------

def _split_two_urls(combined: object) -> tuple[str | None, str | None]:
    if not isinstance(combined, str) or not combined.strip():
        return (None, None)
    matches = URL_RE.findall(combined)
    if not matches:
        return (None, None)
    if len(matches) == 1:
        return (None, matches[0])
    return (matches[0], matches[-1])


def load_audience_overlap_pool() -> pd.DataFrame:
    """Audience rows BEYOND the head(45) cutoff per creator -- the
    highest-quality rows that were not pulled into the single-coded set."""
    frames = []
    for creator, pick in OVERLAP_PICK_AUD.items():
        df = pd.read_excel(AUD_MASTER, sheet_name=creator)
        cutoff = QUOTA_AUD[creator]
        avail = df.iloc[cutoff:cutoff + pick].copy()
        if len(avail) < pick:
            raise RuntimeError(
                f"Audience: {creator} only has {len(avail)} rows beyond "
                f"the head({cutoff}) cutoff, need {pick}")
        urls = avail["Source URL"].apply(_split_two_urls)
        avail["Commenter Post URL"] = urls.apply(lambda t: t[0])
        avail["Influencer's OG Post URL"] = urls.apply(lambda t: t[1])
        avail["original_row_id"] = avail["Comment ID"]
        avail["Comment Text"] = avail["Comment"]
        avail["orientation"] = ORIENTATION_AUD[creator]
        frames.append(avail[[
            "original_row_id", "Comment ID",
            "Commenter Post URL", "Influencer's OG Post URL",
            "Comment Text",
            "Influencer", "Platform", "Source URL", "orientation",
        ]])
    return pd.concat(frames, ignore_index=True)


def load_content_overlap_pool() -> pd.DataFrame:
    """Content rows BEYOND the head(quota) cutoff per creator."""
    frames = []
    for creator, pick in OVERLAP_PICK_CON.items():
        df = pd.read_excel(CON_MASTER, sheet_name=creator)
        cutoff = QUOTA_CON[creator]
        avail = df.iloc[cutoff:cutoff + pick].copy()
        if len(avail) < pick:
            raise RuntimeError(
                f"Content: {creator} only has {len(avail)} rows beyond "
                f"the head({cutoff}) cutoff, need {pick}")
        text_col = "Verbatim Text " if "Verbatim Text " in avail.columns else "Tweet"
        avail["original_row_id"] = avail["Content ID"]
        avail["Context"] = avail["Context "]
        avail["Content Text / Description"] = avail[text_col]
        avail["orientation"] = ORIENTATION_CON[creator]
        frames.append(avail[[
            "original_row_id", "Content ID", "Context",
            "Content Text / Description",
            "Influencer", "Platform", "Content Type", "Source URL",
            "orientation",
        ]])
    return pd.concat(frames, ignore_index=True)


# ----------------------------- existing-coder reading ----------------------

def read_existing_coder_data(typ: str) -> dict[str, list[dict]]:
    """For each coder file, return the existing 30 single-coded rows as
    list-of-dicts keyed by header.  Source files use either the OLD layout
    (header row 1, data rows 2..31) or the NEW layout (DIRECTIONS row 1,
    header row 2, data rows 3..32).  Auto-detect by checking row 1."""
    out = {}
    for letter in CODERS:
        path = OUT_AUD / f"{typ}_nigeria_coder_{letter}_unique.xlsx" \
            if typ == "audience" \
            else OUT_CON / f"{typ}_nigeria_coder_{letter}_unique.xlsx"
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        first = ws.cell(1, 1).value
        if isinstance(first, str) and first.startswith("DIRECTIONS"):
            header_row, data_start = 2, 3
        else:
            header_row, data_start = 1, 2

        headers = [c.value for c in ws[header_row]]
        # Only keep the data columns the coder actually fills (the leading
        # ID/URL/text cols).  Q columns are blank in the existing files.
        if typ == "audience":
            data_cols = ["Comment ID", "Commenter Post URL",
                         "Influencer's OG Post URL", "Comment Text"]
        else:
            data_cols = ["Content ID", "Context", "Content Text / Description"]

        col_index = {h: i for i, h in enumerate(headers)}
        rows = []
        for ri in range(data_start, data_start + 30):
            row = ws[ri]
            d = {}
            for col in data_cols:
                if col in col_index:
                    d[col] = row[col_index[col]].value
                else:
                    d[col] = None
            rows.append(d)
        out[letter] = rows
        wb.close()
    return out


# ----------------------------- template extraction -------------------------

def extract_template(typ: str) -> dict:
    """Pull DIRECTIONS text + every header cell + styling from coder A."""
    path = OUT_AUD / f"{typ}_nigeria_coder_A_unique.xlsx" \
        if typ == "audience" \
        else OUT_CON / f"{typ}_nigeria_coder_A_unique.xlsx"
    wb = load_workbook(path)
    ws = wb.active
    directions_text = ws.cell(1, 1).value
    directions_font = copy(ws.cell(1, 1).font)
    directions_fill = copy(ws.cell(1, 1).fill)
    directions_align = copy(ws.cell(1, 1).alignment)

    headers = []
    header_styles = []
    for c in ws[2]:
        headers.append(c.value)
        header_styles.append({
            "font": copy(c.font), "fill": copy(c.fill),
            "align": copy(c.alignment), "border": copy(c.border),
        })
    col_widths = {k: v.width for k, v in ws.column_dimensions.items()}
    row2_height = ws.row_dimensions[2].height
    wb.close()
    return {
        "directions_text": directions_text,
        "directions_font": directions_font,
        "directions_fill": directions_fill,
        "directions_align": directions_align,
        "headers": headers,
        "header_styles": header_styles,
        "col_widths": col_widths,
        "row2_height": row2_height,
    }


# ----------------------------- overlap allocation --------------------------

def allocate_overlap(pool: pd.DataFrame, label: str
                     ) -> tuple[list[dict], dict[str, list[dict]],
                                list[tuple[str, str]]]:
    """Shuffle deterministically, take first 5 as all-coder shared, next 15
    as pairwise (assigned by PAIRWISE_PATTERN order).  Returns:
      - all_shared_rows: list of 5 dicts
      - pairwise_by_pair: {'A_B': [..3 rows..], 'B_C': [..2..], ...}
      - pairwise_seq: ordered list of (pair_label, group_id) per row
    """
    rng = random.Random(SEED if label == "audience" else SEED + 1)
    idx = list(pool.index)
    rng.shuffle(idx)
    if len(idx) < 20:
        raise RuntimeError(f"{label}: overlap pool only has {len(idx)} rows, need 20")

    shared_idx = idx[:5]
    pair_idx = idx[5:20]

    all_shared = pool.loc[shared_idx].to_dict("records")
    pairwise_by_pair: dict[str, list[dict]] = {}
    cursor = 0
    for left, right, n in PAIRWISE_PATTERN:
        key = f"{left}_{right}"
        rows = pool.loc[pair_idx[cursor:cursor + n]].to_dict("records")
        pairwise_by_pair[key] = rows
        cursor += n
    return all_shared, pairwise_by_pair


def coder_pairwise_assignments(pairwise_by_pair: dict[str, list[dict]]
                               ) -> dict[str, list[tuple[dict, str]]]:
    """For each coder, return list of (row_dict, group_id) -- 5 rows per coder.
    group_id is e.g. AB_01, AB_02, AB_03 / BC_01, BC_02 / etc."""
    out: dict[str, list[tuple[dict, str]]] = {c: [] for c in CODERS}
    for left, right, _ in PAIRWISE_PATTERN:
        key = f"{left}_{right}"
        for i, row in enumerate(pairwise_by_pair[key], start=1):
            gid = f"{left}{right}_{i:02d}"
            out[left].append((row, gid))
            out[right].append((row, gid))
    return out


# ----------------------------- writing -------------------------------------

def _xml_escape(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;")
             .replace(">", "&gt;").replace('"', "&quot;"))


def _inject_secondary_hyperlinks(path: Path,
                                 secondary: dict[str, list[str]]) -> None:
    """openpyxl supports only one hyperlink per cell.  Append extra hyperlink
    XML entries directly for cells with 2+ URLs."""
    if not secondary:
        return
    extra = [(ref, url) for ref, urls in secondary.items() for url in urls]
    if not extra:
        return
    sheet_name = "xl/worksheets/sheet1.xml"
    rels_name = "xl/worksheets/_rels/sheet1.xml.rels"
    with zipfile.ZipFile(path, "r") as z:
        files = {n: z.read(n) for n in z.namelist()}
    sheet_xml = files[sheet_name].decode()
    rels_xml = files.get(
        rels_name,
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
    ).decode()
    max_rid = max(
        (int(m.group(1)) for m in re.finditer(r'Id="rId(\d+)"', rels_xml)),
        default=0,
    )
    new_hl, new_rel = [], []
    rid = max_rid + 1
    for ref, url in extra:
        u = _xml_escape(url)
        new_hl.append(
            '<hyperlink xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            f'ref="{ref}" r:id="rId{rid}" display="{u}"/>'
        )
        new_rel.append(
            '<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
            f'Target="{u}" TargetMode="External" Id="rId{rid}"/>'
        )
        rid += 1
    new_hl_xml = "".join(new_hl)
    new_rel_xml = "".join(new_rel)
    if "</hyperlinks>" in sheet_xml:
        sheet_xml = sheet_xml.replace("</hyperlinks>",
                                      new_hl_xml + "</hyperlinks>")
    else:
        block = f"<hyperlinks>{new_hl_xml}</hyperlinks>"
        if "<pageMargins" in sheet_xml:
            sheet_xml = sheet_xml.replace("<pageMargins", block + "<pageMargins")
        else:
            sheet_xml = sheet_xml.replace("</worksheet>", block + "</worksheet>")
    if "</Relationships>" in rels_xml:
        rels_xml = rels_xml.replace("</Relationships>",
                                    new_rel_xml + "</Relationships>")
    else:
        rels_xml = rels_xml.replace(
            "/>", f">{new_rel_xml}</Relationships>", 1)
    files[sheet_name] = sheet_xml.encode()
    files[rels_name] = rels_xml.encode()
    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, path)


def _style_url_cell(cell, full_text: str,
                    pending: dict[str, list[str]]) -> None:
    matches = list(URL_RE.finditer(full_text))
    if not matches:
        cell.value = full_text
        return
    if len(matches) == 1:
        cell.value = full_text
        cell.hyperlink = matches[0].group(0)
        cell.font = HYPERLINK_FONT
        return
    parts = []
    last = 0
    urls = []
    for m in matches:
        if m.start() > last:
            parts.append(full_text[last:m.start()])
        url = m.group(0)
        parts.append(TextBlock(font=LINK_INLINE_FONT, text=url))
        urls.append(url)
        last = m.end()
    if last < len(full_text):
        parts.append(full_text[last:])
    cell.value = CellRichText(parts)
    cell.hyperlink = urls[0]
    pending.setdefault(cell.coordinate, []).extend(urls[1:])


def write_coder_file(path: Path, template: dict, typ: str,
                     unique_rows: list[dict],
                     pairwise_rows: list[dict],
                     allshared_rows: list[dict]) -> None:
    """Build the file fresh: directions row, headers, 30 unique rows,
    blank gap, 5 pairwise, blank gap, 5 all-shared.  Apply A's styling."""
    if typ == "audience":
        data_cols = ["Comment ID", "Commenter Post URL",
                     "Influencer's OG Post URL", "Comment Text"]
    else:
        data_cols = ["Content ID", "Context", "Content Text / Description"]

    wb = Workbook()
    ws = wb.active
    headers = template["headers"]

    # Row 1 -- DIRECTIONS banner.
    c = ws.cell(ROW_DIRECTIONS, 1, template["directions_text"])
    c.font = template["directions_font"]
    c.fill = template["directions_fill"]
    c.alignment = template["directions_align"]
    # Extend the orange fill across the full header width so the banner
    # visually spans all columns even though only A1 holds text.
    fill_color = template["directions_fill"].fgColor.rgb
    if fill_color:
        banner_fill = PatternFill("solid", fgColor=fill_color)
        for col_idx in range(2, len(headers) + 1):
            ws.cell(ROW_DIRECTIONS, col_idx).fill = banner_fill

    # Row 2 -- header row.
    for col_idx, (h, style) in enumerate(zip(headers, template["header_styles"]), start=1):
        cell = ws.cell(ROW_HEADER, col_idx, h)
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["align"]
        cell.border = style["border"]
    if template["row2_height"]:
        ws.row_dimensions[ROW_HEADER].height = template["row2_height"]

    # Column widths.
    for col_letter, width in template["col_widths"].items():
        if width:
            ws.column_dimensions[col_letter].width = width

    # Helper -- map data column name -> 1-based column index in this sheet.
    def col_idx_for(name: str) -> int:
        return headers.index(name) + 1

    pending: dict[str, list[str]] = {}
    wrap = Alignment(wrap_text=True, vertical="top")

    def write_data_row(ri: int, row: dict) -> None:
        for col in data_cols:
            ci = col_idx_for(col)
            val = row.get(col)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            cell = ws.cell(ri, ci, val)
            if col in URL_COLUMNS and isinstance(val, str) and val.strip():
                _style_url_cell(cell, val, pending)
                cell.alignment = wrap
            else:
                cell.alignment = wrap

    # Rows 3..32 -- 30 single-coded.
    for offset, row in enumerate(unique_rows):
        write_data_row(ROW_UNIQUE_START + offset, row)
    # Rows 35..39 -- 5 pairwise.
    for offset, row in enumerate(pairwise_rows):
        write_data_row(ROW_PAIRWISE_START + offset, row)
    # Rows 42..46 -- 5 all-coder shared.
    for offset, row in enumerate(allshared_rows):
        write_data_row(ROW_ALLCODER_START + offset, row)

    ws.freeze_panes = "A3"
    wb.save(path)
    _inject_secondary_hyperlinks(path, pending)


# ----------------------------- tracker -------------------------------------

def write_tracker(typ: str, assignments: list[dict], path: Path) -> None:
    """One row per (coder, original_row_id) assignment."""
    df = pd.DataFrame(assignments)
    cols = ["analysis_type", "country", "coder_id", "assignment_type",
            "overlap_group_id", "assigned_pair", "original_row_id",
            "Influencer", "Platform", "orientation", "Source URL"]
    if typ == "audience":
        cols += ["Comment Text"]
    else:
        cols += ["Context", "Content Text / Description", "Content Type"]
    cols = [c for c in cols if c in df.columns]
    df = df[cols].sort_values(["coder_id", "assignment_type", "original_row_id"])
    df.to_excel(path, index=False)
    # Light styling on the tracker header.
    wb = load_workbook(path)
    ws = wb.active
    fill = PatternFill("solid", fgColor="FFD966")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
    ws.freeze_panes = "A2"
    wb.save(path)


# ----------------------------- QA ------------------------------------------

def qa(typ: str, assignments: list[dict],
       single_ids: set, overlap_ids: set) -> None:
    print(f"\n--- QA: {typ} ---")
    df = pd.DataFrame(assignments)

    # Hard guard: no overlap row id collides with single_coded set.
    collisions = overlap_ids & single_ids
    assert not collisions, f"overlap collides with single-coded: {collisions}"
    print(f"  ok no overlap_id collides with single-coded set "
          f"(|single|={len(single_ids)}, |overlap|={len(overlap_ids)})")

    # 200 distinct source rows total.
    distinct = df["original_row_id"].nunique()
    assert distinct == 200, f"distinct row count = {distinct}, expected 200"
    print(f"  ok 200 distinct original_row_id values")

    # Per assignment_type counts.
    by_type = df.groupby("assignment_type")["original_row_id"].nunique()
    assert by_type.get("single_coded", 0) == 180, by_type
    assert by_type.get("pairwise_overlap", 0) == 15, by_type
    assert by_type.get("all_coder_shared", 0) == 5, by_type
    print(f"  ok 180 single + 15 pairwise + 5 all-shared distinct rows")

    # Total assignments = 240.
    assert len(df) == 240, f"total assignments = {len(df)}"
    print(f"  ok 240 total coder assignments")

    # Per coder: 40 rows = 30 single + 5 pairwise + 5 all-shared.
    per_coder = df.groupby(["coder_id", "assignment_type"]).size().unstack(fill_value=0)
    for coder in CODERS:
        full = f"Coder_{coder}"
        assert per_coder.loc[full, "single_coded"] == 30, per_coder
        assert per_coder.loc[full, "pairwise_overlap"] == 5, per_coder
        assert per_coder.loc[full, "all_coder_shared"] == 5, per_coder
    print(f"  ok every coder has 30 + 5 + 5")

    # Pairwise distribution AB=3, BC=2, CD=3, DE=2, EF=3, FA=2.
    pw = df[df["assignment_type"] == "pairwise_overlap"]
    pair_counts = pw.groupby("assigned_pair")["original_row_id"].nunique()
    expected = {"A_B": 3, "B_C": 2, "C_D": 3, "D_E": 2, "E_F": 3, "F_A": 2}
    for k, v in expected.items():
        got = int(pair_counts.get(k, 0))
        assert got == v, f"{k}: expected {v}, got {got}"
    print(f"  ok pairwise distribution {expected}")

    # Each pairwise row appears in exactly 2 coder files.
    pw_rep = pw.groupby("original_row_id")["coder_id"].nunique()
    assert (pw_rep == 2).all(), pw_rep[pw_rep != 2]
    print(f"  ok every pairwise row assigned to exactly 2 coders")

    # Each all-shared row appears in exactly 6 coder files.
    sh = df[df["assignment_type"] == "all_coder_shared"]
    sh_rep = sh.groupby("original_row_id")["coder_id"].nunique()
    assert (sh_rep == 6).all(), sh_rep[sh_rep != 6]
    print(f"  ok every all-shared row assigned to exactly 6 coders")

    # Each single_coded row appears in exactly 1 coder file.
    sc = df[df["assignment_type"] == "single_coded"]
    sc_rep = sc.groupby("original_row_id")["coder_id"].nunique()
    assert (sc_rep == 1).all(), sc_rep[sc_rep != 1]
    print(f"  ok every single-coded row assigned to exactly 1 coder")


# ----------------------------- main ----------------------------------------

def process_one(typ: str) -> tuple[list[dict], set, set]:
    print(f"\n========== {typ.upper()} ==========")
    template = extract_template(typ)
    print(f"  template headers: {len(template['headers'])} cols")

    existing = read_existing_coder_data(typ)
    single_ids = set()
    for letter, rows in existing.items():
        for r in rows:
            id_col = "Comment ID" if typ == "audience" else "Content ID"
            single_ids.add(r[id_col])
    print(f"  loaded {sum(len(v) for v in existing.values())} existing single-coded rows "
          f"({len(single_ids)} distinct ids)")

    if typ == "audience":
        pool = load_audience_overlap_pool()
    else:
        pool = load_content_overlap_pool()
    print(f"  overlap pool: {len(pool)} rows")
    # Guard before allocation: pool ids must not overlap with single ids.
    pool_ids = set(pool["original_row_id"])
    collisions = pool_ids & single_ids
    assert not collisions, f"pool collides with single-coded: {collisions}"

    all_shared, pairwise_by_pair = allocate_overlap(pool, typ)
    coder_pairwise = coder_pairwise_assignments(pairwise_by_pair)

    out_dir = OUT_AUD if typ == "audience" else OUT_CON
    assignments = []
    overlap_ids = set()

    id_col = "Comment ID" if typ == "audience" else "Content ID"

    for letter in CODERS:
        full = f"Coder_{letter}"
        unique_rows = existing[letter]
        # build coder's pairwise rows (5) -- preserve order from PAIRWISE_PATTERN
        pw_pairs = coder_pairwise[letter]
        pw_rows = [r for r, _ in pw_pairs]
        sh_rows = all_shared

        # write file
        path = out_dir / f"{typ}_nigeria_coder_{letter}_unique.xlsx"
        write_coder_file(path, template, typ, unique_rows, pw_rows, sh_rows)
        print(f"  wrote {path.relative_to(ROOT)}")

        # tracker assignments for this coder
        for r in unique_rows:
            assignments.append({
                "analysis_type": f"{typ}_analysis",
                "country": "Nigeria",
                "coder_id": full,
                "assignment_type": "single_coded",
                "overlap_group_id": "",
                "assigned_pair": "",
                "original_row_id": r[id_col],
            })
        for (row, gid) in pw_pairs:
            left, right = gid[0], gid[1]
            overlap_ids.add(row["original_row_id"])
            assignments.append({
                "analysis_type": f"{typ}_analysis",
                "country": "Nigeria",
                "coder_id": full,
                "assignment_type": "pairwise_overlap",
                "overlap_group_id": gid,
                "assigned_pair": f"{left}_{right}",
                "original_row_id": row["original_row_id"],
                "Influencer": row.get("Influencer"),
                "Platform": row.get("Platform"),
                "orientation": row.get("orientation"),
                "Source URL": row.get("Source URL"),
                "Comment Text": row.get("Comment Text"),
                "Context": row.get("Context"),
                "Content Text / Description": row.get("Content Text / Description"),
                "Content Type": row.get("Content Type"),
            })
        for i, row in enumerate(sh_rows, start=1):
            gid = f"ALL_{i:02d}"
            overlap_ids.add(row["original_row_id"])
            assignments.append({
                "analysis_type": f"{typ}_analysis",
                "country": "Nigeria",
                "coder_id": full,
                "assignment_type": "all_coder_shared",
                "overlap_group_id": gid,
                "assigned_pair": "ALL_CODERS",
                "original_row_id": row["original_row_id"],
                "Influencer": row.get("Influencer"),
                "Platform": row.get("Platform"),
                "orientation": row.get("orientation"),
                "Source URL": row.get("Source URL"),
                "Comment Text": row.get("Comment Text"),
                "Context": row.get("Context"),
                "Content Text / Description": row.get("Content Text / Description"),
                "Content Type": row.get("Content Type"),
            })

    qa(typ, assignments, single_ids, overlap_ids)

    tracker_path = OUT_BASE / f"{typ}_nigeria_full_assignment_tracker.xlsx"
    write_tracker(typ, assignments, tracker_path)
    print(f"  wrote tracker: {tracker_path.relative_to(ROOT)}")
    return assignments, single_ids, overlap_ids


def main() -> None:
    print(f"Random seed: {SEED}  (audience uses {SEED}, content uses {SEED+1})")
    aud_assign, _, _ = process_one("audience")
    con_assign, _, _ = process_one("content")

    # Combined tracker.
    combined = pd.DataFrame(aud_assign + con_assign)
    cols = ["analysis_type", "country", "coder_id", "assignment_type",
            "overlap_group_id", "assigned_pair", "original_row_id",
            "Influencer", "Platform", "orientation"]
    cols = [c for c in cols if c in combined.columns]
    out = OUT_BASE / "nigeria_full_assignment_tracker_all.xlsx"
    combined[cols].sort_values(
        ["analysis_type", "coder_id", "assignment_type", "original_row_id"]
    ).to_excel(out, index=False)
    wb = load_workbook(out)
    ws = wb.active
    fill = PatternFill("solid", fgColor="FFD966")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
    ws.freeze_panes = "A2"
    wb.save(out)
    print(f"\nwrote combined tracker: {out.relative_to(ROOT)}")

    # Summary tables.
    print("\n=== SUMMARY: rows by analysis_type x coder_id x assignment_type ===")
    summary = combined.groupby(
        ["analysis_type", "coder_id", "assignment_type"]
    ).size().unstack(fill_value=0)
    print(summary.to_string())

    print("\n=== SUMMARY: pairwise distribution by analysis_type x assigned_pair ===")
    pw = combined[combined["assignment_type"] == "pairwise_overlap"]
    pw_summary = pw.groupby(["analysis_type", "assigned_pair"])["original_row_id"].nunique()
    print(pw_summary.to_string())

    print("\nDONE")


if __name__ == "__main__":
    main()
