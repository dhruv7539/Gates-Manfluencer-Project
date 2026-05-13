"""
Process the human-coded codebooks (Content + Audience) end to end.

Inputs (filled-in coder files from the team):
    ~/Downloads/Content Analysis Codebooks/        (6 files, 33-col schema)
    ~/Downloads/Audience Analysis Codebooks/       (6 files, 41-col schema)

For each set, the script:
1. Reads every coder's two sheets ("Nigeria" and "Kenya").
2. Cleans every cell — trim whitespace, drop trailing periods, normalize
   unicode quotes, map known short-forms and typos back to the canonical
   enum values, and tokenize multi-select cells using longest-enum-match
   so that options legitimately containing commas (e.g.
   "Gender issues, e.g. equality") are not split apart.
3. Writes the cleaned per-coder files into a "<set> - cleaned/" subfolder.
4. Builds a master compiled workbook:
       Codebooks/Human Codebooks/Master Human <Set> Codebook.xlsx
   Each unique (country, Item ID) appears exactly once. For items coded by
   multiple coders ("overlap" items: 15 doubles + 5 sextuples per country),
   the final row uses the modal answer per field across coders.
5. Builds a manager-facing recoding report:
       Codebooks/Human Codebooks/Recoding Needed - <Set>.xlsx
   listing every cell that needs human attention, with severity, the
   originating coder, country, item ID, and a one-line description.

Coder identity is taken from the filename (after the last " - "), not the
A/B/C/D/E/F prefix, because the prefix-to-name mapping differs between
Content (D=Selene, E=Rohan) and Audience (D=Rohan, E=Selene).
"""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CONTENT_IN = Path.home() / "Downloads" / "Content Analysis Codebooks"
AUDIENCE_IN = Path.home() / "Downloads" / "Audience Analysis Codebooks"
OUT_DIR = Path("Codebooks/Human Codebooks")


# =============================================================================
# CONTENT schema (33 cols)
# =============================================================================

CONTENT_COLS = [
    "Content ID", "Context", "Content Text / Description",
    "Q1. Attention-getter (Yes/No)",
    "Q1a. Attention-getting strategies",
    "Q1b. Other strategy",
    "Q2. Primary topic(s)",
    "Q2a. Other topic",
    "Q3. Type of content",
    "Q3a. Other content type",
    "Q4. Addresses masculinity / gender norms",
    "Q5. Type of masculinity / gender norms",
    "Q6. Addresses what men should do",
    "Q7. What men do or need to do",
    "Q7a. Other directive",
    "Q8. Problem identified",
    "Q8a. Other problem",
    "Q9. Solution proposed",
    "Q9a. Other solution",
    "Q10. Communication mode",
    "Q10a. Other communication mode",
    "Q11. Audience needs",
    "Q12. How claims are supported",
    "Q12a. Other claim support",
    "Q13. How claims are justified",
    "Q13a. Other justification",
    "Q14. Sentiment toward men",
    "Q15. Sentiment toward women",
    "Q16. Sentiment toward traditional gender norms",
    "Q17. Fear or threat used",
    "Q18. Calls to action present",
    "Q18a. Types of calls to action",
    "Q18b. Other call to action",
]

CONTENT_ID_COL = "Content ID"
CONTENT_TEXT_COL = "Content Text / Description"

CONTENT_SINGLE_ENUMS = {
    "Q1. Attention-getter (Yes/No)": ["Yes", "No"],
    "Q4. Addresses masculinity / gender norms":
        ["Yes, explicitly", "Yes, implicitly", "No"],
    "Q5. Type of masculinity / gender norms": [
        "More regressive/traditional/restrictive",
        "More progressive/equitable/expansive",
        "Mixed/unclear",
        "Does not address masculinity or gender norms",
    ],
    "Q6. Addresses what men should do":
        ["Yes", "No", "Unclear", "Not applicable"],
    "Q14. Sentiment toward men":
        ["Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned"],
    "Q15. Sentiment toward women":
        ["Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned"],
    "Q16. Sentiment toward traditional gender norms":
        ["Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned"],
    "Q17. Fear or threat used": ["Yes", "Somewhat", "No"],
    "Q18. Calls to action present": ["Yes", "No"],
}

CONTENT_MULTI_ENUMS = {
    "Q1a. Attention-getting strategies": [
        "Compelling question", "Use of all CAPS", "Humor or sarcasm",
        "Shares something violent or gross", "Shares something sexual",
        "Shares something surprising",
        "Uses a news headline or social media trend as opener",
        "Interesting visual or meme", "Other",
    ],
    "Q2. Primary topic(s)": [
        "Dating/marriage", "Friends/socializing", "Family/children",
        "Money/status", "Fitness/self-improvement", "Mental health",
        "Gender issues, e.g. equality", "Social issues, e.g. corruption",
        "Religion/morality", "Gaming/technology", "Other",
    ],
    "Q3. Type of content": [
        "Interview/conversational content", "Motivational/self-help content",
        "Commentary/reaction content", "Other",
    ],
    "Q7. What men do or need to do": [
        "Men need to dominate/lead", "Men need to provide/succeed",
        "Men are disadvantaged/victims", "Men need to improve themselves",
        "Men need to be fully self-reliant", "Men need to be emotionally open",
        "Men need to not show emotions", "Men need to be equal partners",
        "Mixed/unclear", "Other", "Not applicable",
    ],
    "Q8. Problem identified": [
        "Kenyan or Nigerian political/social problems",
        "Global political/social/cultural problems",
        "Western political/social influence", "Women/feminism",
        "Men's behavior", "Economic/status pressure",
        "Mental health/emotional struggle",
        "No clear problem is identified", "Other",
    ],
    "Q9. Solution proposed": [
        "Social or political change", "Assert dominance/control",
        "More wealth/status", "More self-discipline/fitness",
        "More emotional growth/healing", "More equality/respect for men",
        "More equality/respect for women", "Building community",
        "No clear solution", "Other",
    ],
    "Q10. Communication mode": [
        "Advice/instruction", "Personal story", "Commentary/opinion",
        "Debate/argument", "Humor/satire", "Motivational speech",
        "News/telling facts", "Other",
    ],
    "Q11. Audience needs": [
        "Entertainment/escapism", "Information seeking",
        "Connection/social interaction",
        "Self expression/identity construction", "Status seeking",
        "Documentation of events", "None of these apply",
    ],
    "Q12. How claims are supported": [
        "Generalizations about men/women", "Personal experience",
        "Stories about men/women", "Cultural/social observations",
        "Facts/statistics", "Moral/religious claims",
        "Mixed", "No support", "Other",
    ],
    "Q13. How claims are justified": [
        "No justification", "Anecdotal examples", "Presented as common sense",
        "References data", "References religion/tradition",
        "References external sources, such as other influencers", "Other",
    ],
    "Q18a. Types of calls to action": [
        "Calls for audience to follow / subscribe",
        "Calls for audience to comment / engage",
        "Calls for audience to share content",
        "Calls for audience to take action in their own life",
        "Calls for political / social action",
        "Calls to consume a product or service", "Other",
    ],
}

CONTENT_OPEN_TEXT = {
    "Context", "Content Text / Description",
    "Q1b. Other strategy", "Q2a. Other topic", "Q3a. Other content type",
    "Q7a. Other directive", "Q8a. Other problem", "Q9a. Other solution",
    "Q10a. Other communication mode", "Q12a. Other claim support",
    "Q13a. Other justification", "Q18b. Other call to action",
}

CONTENT_REQUIRED = [
    "Q1. Attention-getter (Yes/No)", "Q2. Primary topic(s)",
    "Q3. Type of content", "Q4. Addresses masculinity / gender norms",
    "Q5. Type of masculinity / gender norms",
    "Q6. Addresses what men should do",
    "Q8. Problem identified", "Q9. Solution proposed",
    "Q10. Communication mode", "Q12. How claims are supported",
    "Q13. How claims are justified", "Q14. Sentiment toward men",
    "Q15. Sentiment toward women",
    "Q16. Sentiment toward traditional gender norms",
    "Q17. Fear or threat used", "Q18. Calls to action present",
]

CONTENT_OTHER_PAIRS = [
    ("Q1a. Attention-getting strategies", "Q1b. Other strategy"),
    ("Q2. Primary topic(s)", "Q2a. Other topic"),
    ("Q3. Type of content", "Q3a. Other content type"),
    ("Q7. What men do or need to do", "Q7a. Other directive"),
    ("Q8. Problem identified", "Q8a. Other problem"),
    ("Q9. Solution proposed", "Q9a. Other solution"),
    ("Q10. Communication mode", "Q10a. Other communication mode"),
    ("Q12. How claims are supported", "Q12a. Other claim support"),
    ("Q13. How claims are justified", "Q13a. Other justification"),
    ("Q18a. Types of calls to action", "Q18b. Other call to action"),
]


# =============================================================================
# AUDIENCE schema (41 cols)
# =============================================================================

AUDIENCE_COLS = [
    "Comment ID", "Commenter Post URL", "Influencer's OG Post URL",
    "Comment Text",
    "Q1. Overall sentiment of comment",
    "Q2. Primary emotional tone",
    "Q3. Commenter's emotional response to content",
    "Q4. Mentions men, women, or gender norms",
    "Q5. Sentiment toward men / masculinity",
    "Q6. Sentiment toward women / femininity",
    "Q7. Main topic of comment",
    "Q7a. Other topic",
    "Q8. Commenter's stance toward the content",
    "Q9. If supporting, explain why",
    "Q10. Need the content is serving",
    "Q11. If challenging, explain why",
    "Q12. References personal experience",
    "Q13. Includes sexist / derogatory language",
    "Q14. Acquired new knowledge",
    "Q14a. If yes, what did they learn",
    "Q15. Changed attitudes",
    "Q15a. If yes, how did attitude change",
    "Q16. Opinion reinforced by content",
    "Q16a. If yes, what opinion was reinforced",
    "Q17. Calls to action present",
    "Q17a. If yes, what action is urged",
    "Q18. Shares information (fact, link, etc.)",
    "Q18a. If yes, what information",
    "Q19. Advocates for something",
    "Q19a. If yes, what do they advocate for",
    "Q20. Corrects content or other comments",
    "Q20a. If yes, what is incorrect / correction",
    "Q21. Commenter self-identifies",
    "Q21a. Profession mentioned",
    "Q21b. If yes, what profession",
    "Q21c. Location mentioned",
    "Q21d. If yes, what location",
    "Q21e. Race / ethnicity mentioned",
    "Q21f. If yes, what race / ethnicity",
    "Q21g. Gender mentioned",
    "Q21h. If yes, what gender",
]

AUDIENCE_ID_COL = "Comment ID"
AUDIENCE_TEXT_COL = "Comment Text"

AUDIENCE_SINGLE_ENUMS = {
    "Q1. Overall sentiment of comment":
        ["Positive", "Negative", "Neutral", "Unclear"],
    "Q2. Primary emotional tone": [
        "Joy", "Happiness", "Surprise", "Anger", "Fear", "Contempt",
        "Sadness", "Hope", "Empathy", "None of these",
    ],
    "Q4. Mentions men, women, or gender norms": ["Yes", "No"],
    "Q5. Sentiment toward men / masculinity": [
        "Positive", "Negative", "Neutral", "Unclear",
        "Does not mention men/masculinity",
    ],
    "Q6. Sentiment toward women / femininity": [
        "Positive", "Negative", "Neutral", "Unclear",
        "Does not mention women/femininity",
    ],
    "Q8. Commenter's stance toward the content":
        ["Supporting", "Challenging", "Neutral", "Unclear"],
    "Q12. References personal experience": ["Yes", "No"],
    "Q13. Includes sexist / derogatory language": ["Yes", "No"],
    "Q14. Acquired new knowledge": ["Yes", "No"],
    "Q15. Changed attitudes": ["Yes", "No"],
    "Q16. Opinion reinforced by content": ["Yes", "No"],
    "Q17. Calls to action present": ["Yes", "No"],
    "Q18. Shares information (fact, link, etc.)": ["Yes", "No"],
    "Q19. Advocates for something": ["Yes", "No"],
    "Q20. Corrects content or other comments": ["Yes", "No"],
    "Q21. Commenter self-identifies": ["Yes", "No"],
    "Q21a. Profession mentioned": ["Yes", "No"],
    "Q21c. Location mentioned": ["Yes", "No"],
    "Q21e. Race / ethnicity mentioned": ["Yes", "No"],
    "Q21g. Gender mentioned": ["Yes", "No"],
    "Q21h. If yes, what gender":
        ["Female", "Male", "Non-binary", "Other", "Unclear"],
}

AUDIENCE_MULTI_ENUMS = {
    "Q3. Commenter's emotional response to content": [
        "Feeling seen/understood", "Feeling unseen/misunderstood",
        "Feeling attacked", "Feeling objectified", "None of these",
    ],
    "Q7. Main topic of comment": [
        "The speaker/creator of the content/influencer/the content",
        "Politics / social issues", "Dating/relationships/marriage",
        "Money/status", "Fitness", "Media/video games",
        "Mental health/emotions", "Gender roles/norms", "Other",
    ],
    "Q10. Need the content is serving": [
        "Entertainment/escapism", "Information seeking",
        "Connection/social interaction",
        "Self expression/identity construction", "Status seeking",
        "Documentation of events", "None of these apply",
    ],
}

AUDIENCE_OPEN_TEXT = {
    "Commenter Post URL", "Influencer's OG Post URL", "Comment Text",
    "Q7a. Other topic", "Q9. If supporting, explain why",
    "Q11. If challenging, explain why",
    "Q14a. If yes, what did they learn",
    "Q15a. If yes, how did attitude change",
    "Q16a. If yes, what opinion was reinforced",
    "Q17a. If yes, what action is urged",
    "Q18a. If yes, what information",
    "Q19a. If yes, what do they advocate for",
    "Q20a. If yes, what is incorrect / correction",
    "Q21b. If yes, what profession",
    "Q21d. If yes, what location",
    "Q21f. If yes, what race / ethnicity",
}

AUDIENCE_REQUIRED = [
    "Q1. Overall sentiment of comment",
    "Q2. Primary emotional tone",
    "Q3. Commenter's emotional response to content",
    "Q4. Mentions men, women, or gender norms",
    "Q7. Main topic of comment",
    "Q8. Commenter's stance toward the content",
    "Q12. References personal experience",
    "Q13. Includes sexist / derogatory language",
    "Q14. Acquired new knowledge",
    "Q15. Changed attitudes",
    "Q16. Opinion reinforced by content",
    "Q17. Calls to action present",
    "Q18. Shares information (fact, link, etc.)",
    "Q19. Advocates for something",
    "Q20. Corrects content or other comments",
    "Q21. Commenter self-identifies",
]

AUDIENCE_OTHER_PAIRS = [
    ("Q7. Main topic of comment", "Q7a. Other topic"),
]


# =============================================================================
# Column-aware alias table
# =============================================================================

ALIASES: dict[tuple[str, str], str] = {
    # ---------------- CONTENT side ----------------
    # Q3 content type
    ("Q3. Type of content", "commentary content"): "Commentary/reaction content",
    ("Q3. Type of content", "commentary"): "Commentary/reaction content",
    ("Q3. Type of content", "conversational content"): "Interview/conversational content",
    ("Q3. Type of content", "interview"): "Interview/conversational content",
    ("Q3. Type of content", "motivational content"): "Motivational/self-help content",
    ("Q3. Type of content", "self-help"): "Motivational/self-help content",
    ("Q3. Type of content", "self-help content"): "Motivational/self-help content",
    ("Q3. Type of content", "reaction content"): "Commentary/reaction content",
    ("Q3. Type of content", "commentary/reaction"): "Commentary/reaction content",
    ("Q3. Type of content", "motivational/self help content"): "Motivational/self-help content",
    ("Q3. Type of content", "commentary / conversational content"):
        "Commentary/reaction content, Interview/conversational content",
    ("Q3. Type of content", "commentary / self-help content"):
        "Commentary/reaction content, Motivational/self-help content",

    # Q4 — addresses masculinity (typos)
    ("Q4. Addresses masculinity / gender norms", "yes, implicitlly"): "Yes, implicitly",
    ("Q4. Addresses masculinity / gender norms", "yes, explicity"): "Yes, explicitly",
    ("Q4. Addresses masculinity / gender norms", "yex, explicitly"): "Yes, explicitly",
    ("Q4. Addresses masculinity / gender norms", "yes, explcitly"): "Yes, explicitly",

    # Q5 — orientation short forms
    ("Q5. Type of masculinity / gender norms", "more regressive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms", "more regresive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms",
     "more regressive, traditional, & restrictive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms",
     "more regressive/traditional/restrictiv"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms", "more progressive"):
        "More progressive/equitable/expansive",
    ("Q5. Type of masculinity / gender norms", "more progressive and equitable"):
        "More progressive/equitable/expansive",
    ("Q5. Type of masculinity / gender norms",
     "more progressive, equitable, & expansive"):
        "More progressive/equitable/expansive",
    ("Q5. Type of masculinity / gender norms",
     "does not address masculinity/gender norms"):
        "Does not address masculinity or gender norms",
    ("Q5. Type of masculinity / gender norms",
     "does not address masculinity / gender norms"):
        "Does not address masculinity or gender norms",
    ("Q5. Type of masculinity / gender norms", "regresive"):
        "More regressive/traditional/restrictive",
    ("Q5. Type of masculinity / gender norms", "progresive"):
        "More progressive/equitable/expansive",

    # Q6 — short forms / typos
    ("Q6. Addresses what men should do", "mixed"): "Unclear",
    ("Q6. Addresses what men should do", "not appicable"): "Not applicable",

    # Sentiment columns — typos
    ("Q14. Sentiment toward men", "not menioned"): "Not mentioned",
    ("Q14. Sentiment toward men", "unlear"): "Unclear",
    ("Q14. Sentiment toward men", "unclear / positive"): "Mixed",
    ("Q15. Sentiment toward women", "not menioned"): "Not mentioned",
    ("Q15. Sentiment toward women", "unlear"): "Unclear",
    ("Q16. Sentiment toward traditional gender norms", "not menioned"): "Not mentioned",
    ("Q16. Sentiment toward traditional gender norms", "unlear"): "Unclear",

    # Q2 — primary topic short forms
    ("Q2. Primary topic(s)", "gender issues"): "Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "social issues"): "Social issues, e.g. corruption",
    ("Q2. Primary topic(s)", "dating"): "Dating/marriage",
    ("Q2. Primary topic(s)", "marriage"): "Dating/marriage",
    ("Q2. Primary topic(s)", "dating / sex"): "Dating/marriage",
    ("Q2. Primary topic(s)", "family"): "Family/children",
    ("Q2. Primary topic(s)", "religion"): "Religion/morality",
    ("Q2. Primary topic(s)", "morality"): "Religion/morality",
    ("Q2. Primary topic(s)", "fitness"): "Fitness/self-improvement",
    ("Q2. Primary topic(s)", "self-improvement"): "Fitness/self-improvement",
    ("Q2. Primary topic(s)", "money"): "Money/status",
    ("Q2. Primary topic(s)", "status"): "Money/status",
    ("Q2. Primary topic(s)", "friends"): "Friends/socializing",
    ("Q2. Primary topic(s)", "socializing"): "Friends/socializing",
    ("Q2. Primary topic(s)", "mental heath"): "Mental health",
    ("Q2. Primary topic(s)", "family and religion"):
        "Family/children, Religion/morality",
    ("Q2. Primary topic(s)", "family and children"): "Family/children",
    ("Q2. Primary topic(s)", "religion and morality"): "Religion/morality",
    ("Q2. Primary topic(s)", "gender issues / morality"):
        "Gender issues, e.g. equality, Religion/morality",
    ("Q2. Primary topic(s)", "gender issues / marriage"):
        "Gender issues, e.g. equality, Dating/marriage",
    ("Q2. Primary topic(s)", "gender issues / dating"):
        "Gender issues, e.g. equality, Dating/marriage",
    ("Q2. Primary topic(s)", "family / gender issues"):
        "Family/children, Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "children / family / gender issues"):
        "Family/children, Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "family / social issues"):
        "Family/children, Social issues, e.g. corruption",
    ("Q2. Primary topic(s)", "gemder issues and family (parenting)"):
        "Gender issues, e.g. equality, Family/children",
    ("Q2. Primary topic(s)", "corruption / family trauma"):
        "Social issues, e.g. corruption, Family/children",
    ("Q2. Primary topic(s)", "amongst men / gender issues"):
        "Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "marriage and gender issues"):
        "Dating/marriage, Gender issues, e.g. equality",
    ("Q2. Primary topic(s)", "social issues / corruption"):
        "Social issues, e.g. corruption",
    ("Q2. Primary topic(s)", "gender issues / equality"):
        "Gender issues, e.g. equality",

    # Q7 — what men do
    ("Q7. What men do or need to do", "no applicable"): "Not applicable",
    ("Q7. What men do or need to do", "not applicabe"): "Not applicable",
    ("Q7. What men do or need to do", "men need to be dominant/lead"):
        "Men need to dominate/lead",
    ("Q7. What men do or need to do", "men need to dominate/ead"):
        "Men need to dominate/lead",
    ("Q7. What men do or need to do", "men need to be dominate / lead more"):
        "Men need to dominate/lead",
    ("Q7. What men do or need to do", "men need to now show emotions"):
        "Men need to not show emotions",
    ("Q7. What men do or need to do", "men need to not show emotion"):
        "Men need to not show emotions",
    ("Q7. What men do or need to do",
     "men need to be more emotionally open and heal"):
        "Men need to be emotionally open",
    ("Q7. What men do or need to do", "men need to be more emotionally open"):
        "Men need to be emotionally open",
    ("Q7. What men do or need to do",
     "men need to know they're allowed to be emotionally open"):
        "Men need to be emotionally open",
    ("Q7. What men do or need to do", "mixed"): "Mixed/unclear",
    ("Q7. What men do or need to do", "unclear"): "Mixed/unclear",
    ("Q7. What men do or need to do", "unlclear"): "Mixed/unclear",
    ("Q7. What men do or need to do", "men need to build community"): "Other",
    ("Q7. What men do or need to do", "not to date \"born again\" women"): "Other",

    # Q8 — problem
    ("Q8. Problem identified", "mens behavior"): "Men's behavior",
    ("Q8. Problem identified", "men behavior"): "Men's behavior",
    ("Q8. Problem identified", "men''s behavior"): "Men's behavior",
    ("Q8. Problem identified", "women"): "Women/feminism",
    ("Q8. Problem identified", "women's behavior"): "Women/feminism",
    ("Q8. Problem identified", "women / women's behavior"): "Women/feminism",
    ("Q8. Problem identified", "women and"): "Women/feminism",
    ("Q8. Problem identified", "mental health"): "Mental health/emotional struggle",
    ("Q8. Problem identified", "mental health and emotional struggle"):
        "Mental health/emotional struggle",
    ("Q8. Problem identified", "no problem is identified"):
        "No clear problem is identified",
    ("Q8. Problem identified", "kenyan or nigerian political/social problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "kenyan cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian culture problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian culture issues"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian culture and social issues"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian social and cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "nigerian social & culture issues"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "kenyan social & cultural problems"):
        "Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified", "mental health / kenyan cultural problems"):
        "Mental health/emotional struggle, Kenyan or Nigerian political/social problems",
    ("Q8. Problem identified",
     "global social issue of the patriarchy in particular"):
        "Global political/social/cultural problems",
    ("Q8. Problem identified", "general political/social/cultural problems"):
        "Global political/social/cultural problems",

    # Q9 — solutions
    ("Q9. Solution proposed", "more respect for men"):
        "More equality/respect for men",
    ("Q9. Solution proposed", "more respect for women"):
        "More equality/respect for women",
    ("Q9. Solution proposed", "more emotional growth"):
        "More emotional growth/healing",
    ("Q9. Solution proposed",
     "more emotional growth and healing in both parties of a marriage"):
        "More emotional growth/healing",
    ("Q9. Solution proposed",
     "more emotional growth for both partners in marriage"):
        "More emotional growth/healing",

    # Q10 communication mode
    ("Q10. Communication mode", "opinion"): "Commentary/opinion",
    ("Q10. Communication mode", "commentary content"): "Commentary/opinion",
    ("Q10. Communication mode", "commentary"): "Commentary/opinion",
    ("Q10. Communication mode", "commenary/opinion"): "Commentary/opinion",
    ("Q10. Communication mode", "advice"): "Advice/instruction",
    ("Q10. Communication mode", "instruction"): "Advice/instruction",
    ("Q10. Communication mode", "news"): "News/telling facts",
    ("Q10. Communication mode", "facts"): "News/telling facts",
    ("Q10. Communication mode", "humor"): "Humor/satire",
    ("Q10. Communication mode", "satire"): "Humor/satire",
    ("Q10. Communication mode", "advice / commentary / opinion"):
        "Advice/instruction, Commentary/opinion",
    ("Q10. Communication mode", "commentary / debate"):
        "Commentary/opinion, Debate/argument",
    ("Q10. Communication mode", "commentary /"): "Commentary/opinion",
    ("Q10. Communication mode", "commentary content /"): "Commentary/opinion",
    ("Q10. Communication mode", "/ commentary"): "Commentary/opinion",
    ("Q10. Communication mode", "conversational content"): "Personal story",

    # Q11 audience needs
    ("Q11. Audience needs", "identity construction"):
        "Self expression/identity construction",
    ("Q11. Audience needs", "self expression"):
        "Self expression/identity construction",
    ("Q11. Audience needs", "connetion/social interaction"):
        "Connection/social interaction",
    ("Q11. Audience needs", "connection"): "Connection/social interaction",
    ("Q11. Audience needs", "social interaction"): "Connection/social interaction",
    ("Q11. Audience needs", "entertainment"): "Entertainment/escapism",
    ("Q11. Audience needs", "escapism"): "Entertainment/escapism",
    ("Q11. Audience needs", "connection / self-expression"):
        "Connection/social interaction, Self expression/identity construction",

    # Q12 claim support
    ("Q12. How claims are supported", "generalizations about women"):
        "Generalizations about men/women",
    ("Q12. How claims are supported", "generalizations about men"):
        "Generalizations about men/women",
    ("Q12. How claims are supported", "generalizations about men and women"):
        "Generalizations about men/women",
    ("Q12. How claims are supported", "social observations"):
        "Cultural/social observations",
    ("Q12. How claims are supported", "cultural observations"):
        "Cultural/social observations",
    ("Q12. How claims are supported", "cultura/social observations"):
        "Cultural/social observations",
    ("Q12. How claims are supported", "facts"): "Facts/statistics",
    ("Q12. How claims are supported", "statistics"): "Facts/statistics",
    ("Q12. How claims are supported", "stories about women"):
        "Stories about men/women",
    ("Q12. How claims are supported", "stories about men"):
        "Stories about men/women",
    ("Q12. How claims are supported", "moral claims"): "Moral/religious claims",
    ("Q12. How claims are supported", "religious claims"): "Moral/religious claims",
    ("Q12. How claims are supported", "moral / religous claims"):
        "Moral/religious claims",
    ("Q12. How claims are supported", "anecdotal examples"): "Personal experience",
    ("Q12. How claims are supported", "generalization about relationships"):
        "Generalizations about men/women",

    # Q13 justifications
    ("Q13. How claims are justified", "common sense"): "Presented as common sense",
    ("Q13. How claims are justified", "religion"): "References religion/tradition",
    ("Q13. How claims are justified", "tradition"): "References religion/tradition",
    ("Q13. How claims are justified", "anecdotes"): "Anecdotal examples",
    ("Q13. How claims are justified", "anecdotal exampled"): "Anecdotal examples",
    ("Q13. How claims are justified", "anecdotal exmaples"): "Anecdotal examples",
    ("Q13. How claims are justified", "anectdotal exmaples / religous refrences"):
        "Anecdotal examples, References religion/tradition",
    ("Q13. How claims are justified", "refrences religion"):
        "References religion/tradition",
    ("Q13. How claims are justified", "external sources"):
        "References external sources, such as other influencers",
    ("Q13. How claims are justified", "references external sources"):
        "References external sources, such as other influencers",
    ("Q13. How claims are justified", "data"): "References data",

    # Q18a CTA types
    ("Q18a. Types of calls to action",
     "calls for politicians or social figures to do something"):
        "Calls for political / social action",
    ("Q18a. Types of calls to action", "calls for audience to share the content"):
        "Calls for audience to share content",

    # Q1a attention strategies
    ("Q1a. Attention-getting strategies", "others"): "Other",
    ("Q1a. Attention-getting strategies", "humor/sarcasm"): "Humor or sarcasm",
    ("Q1a. Attention-getting strategies", "uses all caps"): "Use of all CAPS",

    # ---------------- AUDIENCE side ----------------
    # Q1, Q5, Q6 — sentiment short forms (parallel to Q14/15/16 on content)
    ("Q1. Overall sentiment of comment", "unlear"): "Unclear",
    ("Q5. Sentiment toward men / masculinity", "unlear"): "Unclear",
    ("Q5. Sentiment toward men / masculinity",
     "does not mention men / masculinity"):
        "Does not mention men/masculinity",
    ("Q6. Sentiment toward women / femininity", "unlear"): "Unclear",
    ("Q6. Sentiment toward women / femininity",
     "does not mention women / femininity"):
        "Does not mention women/femininity",
    # Q7 audience topics (multi-select)
    ("Q7. Main topic of comment", "the speaker/creator of the content/ influencer/the content"):
        "The speaker/creator of the content/influencer/the content",
    ("Q7. Main topic of comment",
     "the speaker/creator of the content/influencer/the content"):
        "The speaker/creator of the content/influencer/the content",
    ("Q7. Main topic of comment", "speaker/creator"):
        "The speaker/creator of the content/influencer/the content",
    ("Q7. Main topic of comment", "the influencer"):
        "The speaker/creator of the content/influencer/the content",
    ("Q7. Main topic of comment", "politics"): "Politics / social issues",
    ("Q7. Main topic of comment", "social issues"): "Politics / social issues",
    ("Q7. Main topic of comment", "dating"): "Dating/relationships/marriage",
    ("Q7. Main topic of comment", "relationships"): "Dating/relationships/marriage",
    ("Q7. Main topic of comment", "marriage"): "Dating/relationships/marriage",
    ("Q7. Main topic of comment", "mental health"): "Mental health/emotions",
    ("Q7. Main topic of comment", "emotions"): "Mental health/emotions",
    ("Q7. Main topic of comment", "gender norms"): "Gender roles/norms",
    ("Q7. Main topic of comment", "gender roles"): "Gender roles/norms",
    # Q21h
    ("Q21h. If yes, what gender", "non binary"): "Non-binary",
    ("Q21h. If yes, what gender", "nonbinary"): "Non-binary",
}


# =============================================================================
# Cleaning utilities
# =============================================================================

def _strip(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).replace("\xa0", " ").strip()
    s = s.replace("’", "'").replace("‘", "'")
    s = s.replace("“", '"').replace("”", '"')
    s = re.sub(r"\s+", " ", s)
    if s.endswith(".") and not s.endswith("..."):
        s = s[:-1].rstrip()
    return s


def _enum_lookup(enums: dict[str, list[str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for col, values in enums.items():
        d: dict[str, str] = {}
        for v in values:
            d[v.lower()] = v
            d[re.sub(r"[\s/]+", "", v).lower()] = v
        out[col] = d
    return out


def _canon_single(col: str, value: str, lookup: dict[str, dict[str, str]]) -> str:
    if not value:
        return ""
    low = value.lower().strip()
    if (col, low) in ALIASES:
        return ALIASES[(col, low)]
    col_lookup = lookup.get(col, {})
    if low in col_lookup:
        return col_lookup[low]
    squeezed = re.sub(r"[\s/]+", "", low)
    if squeezed in col_lookup:
        return col_lookup[squeezed]
    if "," in low:
        head_split = low.split(",", 1)
        head = head_split[0].strip() + ", " + head_split[1].split(".", 1)[0].strip()
        head = head.rstrip(".").strip()
        if head in col_lookup:
            return col_lookup[head]
    return value


def _canon_multi(col: str, value: str, enum: list[str],
                 lookup: dict[str, dict[str, str]]) -> str:
    if not value:
        return ""
    remaining = " " + value + " "
    found: list[str] = []
    for opt in sorted(enum, key=len, reverse=True):
        pat = re.compile(
            r"(?<![A-Za-z0-9])" + re.escape(opt) + r"(?![A-Za-z0-9])",
            flags=re.IGNORECASE,
        )
        for _ in range(20):
            m = pat.search(remaining)
            if not m:
                break
            found.append(opt)
            remaining = remaining[: m.start()] + " | " + remaining[m.end():]
    leftovers = [t.strip(" ,.;|") for t in re.split(r",", remaining)]
    leftovers = [t for t in leftovers if t and t != "|"]
    for t in leftovers:
        low = t.lower().rstrip(".").strip()
        if (col, low) in ALIASES:
            found.append(ALIASES[(col, low)])
            continue
        col_lookup = lookup.get(col, {})
        if low in col_lookup:
            found.append(col_lookup[low])
            continue
        squeezed = re.sub(r"[\s/]+", "", low)
        if squeezed in col_lookup:
            found.append(col_lookup[squeezed])
            continue
        found.append(t)
    seen: set[str] = set()
    result: list[str] = []
    for c in found:
        key = c.lower().strip()
        if key and key not in seen:
            seen.add(key)
            result.append(c)
    return ", ".join(result)


def clean_cell(col: str, value: object, *, single_enums: dict[str, list[str]],
               multi_enums: dict[str, list[str]],
               open_text: set[str], id_col: str) -> str:
    s = _strip(value)
    if not s:
        return ""
    if col in open_text or col == id_col:
        return s
    if col in multi_enums:
        return _canon_multi(col, s, multi_enums[col], _enum_lookup(multi_enums))
    if col in single_enums:
        return _canon_single(col, s, _enum_lookup(single_enums))
    return s


def coder_from_filename(path: Path) -> str:
    """Return the coder name (the part after the last ' - ')."""
    stem = path.stem
    if " - " in stem:
        return stem.split(" - ")[-1].strip()
    return stem


def read_coder_file(path: Path, *, columns: list[str], id_col: str,
                    single_enums: dict[str, list[str]],
                    multi_enums: dict[str, list[str]],
                    open_text: set[str]) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    for sheet in ("Nigeria", "Kenya"):
        raw = pd.read_excel(path, sheet_name=sheet, header=1, dtype=object)
        raw = raw[raw.iloc[:, 0].notna()].copy().reset_index(drop=True)
        if raw.shape[1] < len(columns):
            raise ValueError(
                f"{path.name}/{sheet}: only {raw.shape[1]} columns, "
                f"expected at least {len(columns)}"
            )
        raw = raw.iloc[:, : len(columns)].copy()
        raw.columns = columns
        cleaned = raw.copy()
        for col in cleaned.columns:
            cleaned[col] = cleaned[col].apply(
                lambda v: clean_cell(col, v, single_enums=single_enums,
                                     multi_enums=multi_enums,
                                     open_text=open_text, id_col=id_col)
            )
        out[sheet] = cleaned
    return out


# =============================================================================
# Excel-writing helpers
# =============================================================================

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 48
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            header = ws.cell(row=1, column=col_idx).value or ""
            if header == "Issue":
                w = 60
            elif header in ("Content ID", "Comment ID", "Item ID"):
                w = 22
            elif "URL" in str(header):
                w = 40
            elif header in ("Context", "Content Text / Description",
                            "Comment Text"):
                w = 50
            elif "Open text" in str(header) or header.endswith("Other"):
                w = 25
            elif header in ("Pass", "Severity"):
                w = 10
            elif header == "Coder":
                w = 12
            elif header == "Country":
                w = 10
            elif header == "Column":
                w = 32
            elif header == "Value":
                w = 38
            else:
                w = 26
            ws.column_dimensions[letter].width = w
    wb.save(path)


# =============================================================================
# Master compilation
# =============================================================================

def modal(values: list[str]) -> str:
    counts = Counter([v for v in values if v != ""])
    if not counts:
        return ""
    top_n = counts.most_common(1)[0][1]
    winners = [v for v, n in counts.most_common() if n == top_n]
    return winners[0]


def build_master(per_coder: dict[str, dict[str, pd.DataFrame]], *,
                 columns: list[str], id_col: str) -> dict[str, pd.DataFrame]:
    long_rows: list[dict] = []
    for coder, sheets in per_coder.items():
        for country, df in sheets.items():
            for _, row in df.iterrows():
                rec = row.to_dict()
                rec["__coder"] = coder
                rec["__country"] = country
                long_rows.append(rec)
    long_df = pd.DataFrame(long_rows)

    masters: dict[str, pd.DataFrame] = {}
    overlap_panes: dict[str, pd.DataFrame] = {}

    for country in ("Nigeria", "Kenya"):
        sub = long_df[long_df["__country"] == country].copy()
        master_rows: list[dict] = []
        overlap_records: list[dict] = []

        for cid, group in sub.groupby(id_col, sort=False):
            n_coders = len(group)
            if n_coders == 1:
                rec = group.iloc[0].to_dict()
                rec["Coder"] = rec.pop("__coder")
                rec.pop("__country")
                rec["Coders on this item"] = 1
                master_rows.append(rec)
                continue
            collapsed: dict[str, object] = {id_col: cid}
            for col in columns:
                if col == id_col:
                    continue
                collapsed[col] = modal(group[col].tolist())
            coders_here = sorted(group["__coder"].unique())
            collapsed["Coder"] = ", ".join(coders_here) + " (modal vote)"
            collapsed["Coders on this item"] = n_coders
            master_rows.append(collapsed)

            for _, r in group.iterrows():
                overlap_records.append({
                    id_col: cid,
                    "Coder": r["__coder"],
                    **{c: r[c] for c in columns if c != id_col},
                })

        master_df = pd.DataFrame(master_rows)
        cols = [id_col, "Coder", "Coders on this item"] + [
            c for c in columns if c != id_col
        ]
        master_df = master_df.reindex(columns=cols)
        master_df = master_df.sort_values(id_col).reset_index(drop=True)

        overlap_df = pd.DataFrame(overlap_records)
        if not overlap_df.empty:
            overlap_df = overlap_df.sort_values([id_col, "Coder"]).reset_index(drop=True)
        masters[country] = master_df
        overlap_panes[country] = overlap_df

    singles = {
        country: masters[country][masters[country]["Coders on this item"] == 1]
        .copy() for country in ("Nigeria", "Kenya")
    }

    return {
        "Nigeria — All Items": masters["Nigeria"],
        "Kenya — All Items": masters["Kenya"],
        "Nigeria — Singles Only": singles["Nigeria"],
        "Kenya — Singles Only": singles["Kenya"],
        "Nigeria — Overlap Reliability": overlap_panes["Nigeria"],
        "Kenya — Overlap Reliability": overlap_panes["Kenya"],
    }


def methodology_sheet(set_name: str, id_col: str) -> pd.DataFrame:
    rows = [
        (f"Master Human {set_name} Codebook — methodology", ""),
        ("Source files",
         f"Six coder workbooks in ~/Downloads/{set_name} Analysis Codebooks/."),
        ("Design",
         "Partition with reliability overlap. Per country: 200 unique "
         "items split across 6 coders. 180 items coded by exactly 1 "
         "coder; 15 items coded by 2 coders; 5 items coded by all 6 "
         "coders (reliability anchors). Each coder coded 40 items per "
         "country."),
        ("Rows on the 'All Items' sheets",
         "200 per country — each item ID appears exactly once."),
        ("How overlap items were collapsed",
         "For items coded by more than one coder, each field uses the "
         "modal value across coders. Ties are broken by the first "
         "coder alphabetically. The 'Coder' column shows which coders "
         "contributed; 'Coders on this item' shows how many."),
        ("Singles-Only sheet",
         "The subset where Coders on this item = 1 — useful when you "
         "want only rows where the master makes no judgement."),
        ("Overlap-Reliability sheets",
         "Long-format view of every coder's answers for the 20 overlap "
         "items per country (15 double-coded + 5 sextuple-coded). Use "
         "this for inter-rater reliability work."),
        ("Cleaning applied",
         "Trimmed whitespace and trailing periods; normalized case; "
         "mapped known variants and typos to the canonical enum "
         "(e.g. 'Commentary Content' -> 'Commentary/reaction content'; "
         "'More regressive' -> 'More regressive/traditional/restrictive'; "
         "'Unlear' -> 'Unclear'). Multi-select cells were tokenized "
         "using longest-enum-match against the canonical option list, "
         "so options that legitimately contain commas (e.g. 'Gender "
         "issues, e.g. equality') are not split apart."),
        ("ID column", id_col),
    ]
    return pd.DataFrame(rows, columns=["Field", "Description"])


def cleaning_log(per_coder_raw, per_coder_clean) -> pd.DataFrame:
    rows = []
    for coder, sheets in per_coder_clean.items():
        for country, clean_df in sheets.items():
            raw_df = per_coder_raw[coder][country]
            changes = 0
            for col in clean_df.columns:
                for r, c in zip(raw_df[col].fillna("").astype(str),
                                clean_df[col].fillna("").astype(str)):
                    if r.strip() != c.strip():
                        changes += 1
            rows.append({
                "Coder": coder,
                "Country": country,
                "Rows": len(clean_df),
                "Cells normalized": changes,
            })
    return pd.DataFrame(rows)


# =============================================================================
# Recoding-needed report
# =============================================================================

def is_empty(v) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return s == "" or s.lower() == "nan"


def tokenize_multi(value, enum: list[str]) -> tuple[list[str], list[str]]:
    if is_empty(value):
        return [], []
    remaining = " " + str(value) + " "
    matched: list[str] = []
    for opt in sorted(enum, key=len, reverse=True):
        pat = re.compile(
            r"(?<![A-Za-z0-9])" + re.escape(opt) + r"(?![A-Za-z0-9])",
            flags=re.IGNORECASE,
        )
        for _ in range(20):
            m = pat.search(remaining)
            if not m:
                break
            matched.append(opt)
            remaining = remaining[: m.start()] + " | " + remaining[m.end():]
    leftovers = [t.strip(" ,.;|") for t in re.split(r",", remaining)]
    leftovers = [t for t in leftovers if t and t != "|"]
    return matched, leftovers


def build_recoding_report(per_coder_clean, *, columns, id_col,
                          single_enums, multi_enums,
                          required, other_pairs,
                          set_name: str) -> dict[str, pd.DataFrame]:
    """Run all consistency checks and produce a workbook of issues."""
    rows = []
    for coder, sheets in per_coder_clean.items():
        for country, df in sheets.items():
            for _, row in df.iterrows():
                rec = row.to_dict()
                rec["__coder"] = coder
                rec["__country"] = country
                rows.append(rec)
    df_all = pd.DataFrame(rows)

    issues: list[dict] = []

    def add(passno, sev, row, col, val, issue):
        issues.append({
            "Pass": passno, "Severity": sev,
            "Coder": row["__coder"], "Country": row["__country"],
            "Item ID": row[id_col],
            "Column": col, "Value": val, "Issue": issue,
        })

    # 1 — single-select enum violations
    for _, row in df_all.iterrows():
        for col, valid in single_enums.items():
            v = row.get(col)
            if is_empty(v):
                continue
            s = str(v).strip()
            if s not in set(valid):
                add(1, "HIGH", row, col, s,
                    "Value not in single-select enum (allowed: "
                    + ", ".join(valid) + ")")

    # 2 — multi-select tokens that don't match enum
    for _, row in df_all.iterrows():
        for col, valid in multi_enums.items():
            v = row.get(col)
            if is_empty(v):
                continue
            _, leftover = tokenize_multi(v, valid)
            for t in leftover:
                add(2, "MED", row, col, t,
                    "Multi-select option not matched to enum")

    # 3 — required fields missing
    for _, row in df_all.iterrows():
        for col in required:
            if is_empty(row.get(col)):
                add(3, "HIGH", row, col, "(missing)",
                    "Required field is empty")

    # 4 — Other selected but Q*b/follow-up blank
    for _, row in df_all.iterrows():
        for parent, child in other_pairs:
            pv = row.get(parent)
            if is_empty(pv):
                continue
            tokens = [t.strip().lower() for t in str(pv).split(",")]
            if "other" in tokens and is_empty(row.get(child)):
                add(4, "LOW", row, parent, str(pv),
                    f"'Other' selected but '{child}' is blank")

    # 5 — duplicate IDs within a coder
    for (coder, country), g in df_all.groupby(["__coder", "__country"]):
        dups = g[id_col].value_counts()
        for cid, n in dups[dups > 1].items():
            for _, r in g[g[id_col] == cid].iterrows():
                add(5, "HIGH", r, id_col, str(cid),
                    f"Duplicate ID — appears {n}x in this coder's file")

    # 6 — anchor items significantly under-coded (>=3 required blanks)
    for country in ("Nigeria", "Kenya"):
        sub = df_all[df_all["__country"] == country]
        anchors = sub.groupby(id_col).size()
        anchor_ids = anchors[anchors == 6].index.tolist()
        for cid in anchor_ids:
            for _, r in sub[sub[id_col] == cid].iterrows():
                blanks = sum(1 for c in required if is_empty(r.get(c)))
                if blanks >= 3:
                    add(6, "HIGH", r, "(multiple)", f"{blanks} blank",
                        f"Anchor item is significantly under-coded "
                        f"({blanks}/{len(required)} required fields blank)")

    # 7 — Audience Q5 substantive even though Q4 = No (and parallel: content)
    # Use convention: any single-select column whose name contains "Q5"
    # with a "Does not address" / "Does not mention" sentinel — for both
    # content and audience.
    sentinel_pairs = [
        ("Q4. Addresses masculinity / gender norms",
         "Q5. Type of masculinity / gender norms", "does not address"),
        ("Q4. Mentions men, women, or gender norms",
         "Q5. Sentiment toward men / masculinity", "does not mention"),
        ("Q4. Mentions men, women, or gender norms",
         "Q6. Sentiment toward women / femininity", "does not mention"),
    ]
    for q4_col, q5_col, sentinel in sentinel_pairs:
        if q4_col not in df_all.columns or q5_col not in df_all.columns:
            continue
        for _, row in df_all.iterrows():
            q4 = row.get(q4_col)
            q5 = row.get(q5_col)
            if is_empty(q4) or is_empty(q5):
                continue
            if str(q4).strip().lower() == "no":
                if sentinel not in str(q5).lower():
                    add(7, "HIGH", row, q5_col, str(q5),
                        f"{q4_col.split('.')[0]}=No but {q5_col.split('.')[0]} "
                        f"has substantive answer (should reference '{sentinel}' "
                        "or be blank)")

    # 8 — Q21 (audience) cascade contradictions
    # Q21=No but Q21a/Q21c/Q21e/Q21g = Yes
    if "Q21. Commenter self-identifies" in df_all.columns:
        sub_cols = [
            "Q21a. Profession mentioned",
            "Q21c. Location mentioned",
            "Q21e. Race / ethnicity mentioned",
            "Q21g. Gender mentioned",
        ]
        for _, row in df_all.iterrows():
            q21 = row.get("Q21. Commenter self-identifies")
            if is_empty(q21):
                continue
            if str(q21).strip().lower() == "no":
                for sc in sub_cols:
                    sv = row.get(sc)
                    if not is_empty(sv) and str(sv).strip().lower() == "yes":
                        add(8, "MED", row, sc, str(sv),
                            "Q21=No but " + sc[:6] + " says Yes — contradiction")

    # 9 — conditional follow-up missing
    follow_up = [
        ("Q14. Acquired new knowledge", "Q14a. If yes, what did they learn"),
        ("Q15. Changed attitudes", "Q15a. If yes, how did attitude change"),
        ("Q16. Opinion reinforced by content", "Q16a. If yes, what opinion was reinforced"),
        ("Q17. Calls to action present", "Q17a. If yes, what action is urged"),
        ("Q18. Shares information (fact, link, etc.)", "Q18a. If yes, what information"),
        ("Q19. Advocates for something", "Q19a. If yes, what do they advocate for"),
        ("Q20. Corrects content or other comments", "Q20a. If yes, what is incorrect / correction"),
    ]
    for parent, follow in follow_up:
        if parent not in df_all.columns or follow not in df_all.columns:
            continue
        for _, row in df_all.iterrows():
            pv = row.get(parent)
            if is_empty(pv):
                continue
            if str(pv).strip().lower() == "yes" and is_empty(row.get(follow)):
                add(9, "MED", row, follow, "(empty)",
                    f"{parent.split('.')[0]}=Yes but '{follow}' is blank")

    # 10 — trailing whitespace / period that survived
    for _, row in df_all.iterrows():
        for col in columns:
            v = row.get(col)
            if is_empty(v):
                continue
            s = str(v)
            if s != s.strip() or (s.endswith(".") and not s.endswith("...")):
                add(10, "LOW", row, col, s,
                    "Trailing whitespace or period (cosmetic)")

    # 11 — Q1/Q1a contradictions (content only)
    if ("Q1. Attention-getter (Yes/No)" in df_all.columns
            and "Q1a. Attention-getting strategies" in df_all.columns):
        for _, row in df_all.iterrows():
            q1 = row.get("Q1. Attention-getter (Yes/No)")
            q1a = row.get("Q1a. Attention-getting strategies")
            if is_empty(q1):
                continue
            yn = str(q1).strip().lower()
            if yn == "yes" and is_empty(q1a):
                add(11, "MED", row, "Q1a. Attention-getting strategies", "(empty)",
                    "Q1=Yes but Q1a strategies are blank")
            if yn == "no" and not is_empty(q1a):
                add(11, "MED", row, "Q1a. Attention-getting strategies", str(q1a),
                    "Q1=No but Q1a lists strategies (contradiction)")

    # 12 — Q18/Q18a contradictions (content only)
    if ("Q18. Calls to action present" in df_all.columns
            and "Q18a. Types of calls to action" in df_all.columns):
        for _, row in df_all.iterrows():
            q18 = row.get("Q18. Calls to action present")
            q18a = row.get("Q18a. Types of calls to action")
            if is_empty(q18):
                continue
            yn = str(q18).strip().lower()
            if yn == "yes" and is_empty(q18a):
                add(12, "MED", row, "Q18a. Types of calls to action", "(empty)",
                    "Q18=Yes but Q18a CTA types are blank")
            if yn == "no" and not is_empty(q18a):
                add(12, "MED", row, "Q18a. Types of calls to action", str(q18a),
                    "Q18=No but Q18a lists CTA types (contradiction)")

    df_iss = pd.DataFrame(issues)
    sev_order = {"HIGH": 0, "MED": 1, "LOW": 2}
    if not df_iss.empty:
        df_iss["__sev"] = df_iss["Severity"].map(sev_order)
        df_iss = df_iss.sort_values(
            ["__sev", "Coder", "Country", "Item ID", "Pass"]
        ).drop("__sev", axis=1).reset_index(drop=True)

    # Build sheets
    sheets: dict[str, pd.DataFrame] = {}

    # Pass descriptions for executive summary
    pass_descr = {
        1: "Single-select enum violation",
        2: "Multi-select option not in enum",
        3: "Required field missing",
        4: "'Other' selected without explanation",
        5: "Duplicate Item ID within a coder",
        6: "Anchor item significantly under-coded",
        7: "Q4=No but Q5/Q6 substantive answer",
        8: "Q21 sub-question contradiction (audience only)",
        9: "Yes/No follow-up blank (audience only)",
        10: "Trailing whitespace/period (cosmetic)",
        11: "Q1/Q1a contradiction (content only)",
        12: "Q18/Q18a contradiction (content only)",
    }

    sev_by_pass = {}
    if not df_iss.empty:
        for p in df_iss["Pass"].unique():
            sub = df_iss[df_iss["Pass"] == p]
            sev_by_pass[p] = sub["Severity"].value_counts().to_dict()

    summary_rows = []
    for p in sorted(pass_descr):
        counts = sev_by_pass.get(p, {})
        summary_rows.append({
            "Pass": p,
            "What it checks": pass_descr[p],
            "HIGH": counts.get("HIGH", 0),
            "MED":  counts.get("MED", 0),
            "LOW":  counts.get("LOW", 0),
            "Total": sum(counts.values()),
        })
    sheets["Executive Summary"] = pd.DataFrame(summary_rows)

    if not df_iss.empty:
        per_coder = df_iss.pivot_table(
            index="Coder", columns="Severity", values="Issue",
            aggfunc="count", fill_value=0,
        ).reset_index()
        for c in ("HIGH", "MED", "LOW"):
            if c not in per_coder.columns:
                per_coder[c] = 0
        per_coder["Total"] = per_coder[["HIGH", "MED", "LOW"]].sum(axis=1)
        per_coder = per_coder[["Coder", "HIGH", "MED", "LOW", "Total"]]
        sheets["Per-Coder Counts"] = per_coder.sort_values(
            "Total", ascending=False).reset_index(drop=True)

    sheets["All Issues"] = df_iss
    if not df_iss.empty:
        sheets["HIGH — recode required"] = df_iss[df_iss["Severity"] == "HIGH"].reset_index(drop=True)
        sheets["MED — review"] = df_iss[df_iss["Severity"] == "MED"].reset_index(drop=True)
        sheets["LOW — cosmetic"] = df_iss[df_iss["Severity"] == "LOW"].reset_index(drop=True)

    return sheets, df_iss


# =============================================================================
# Orchestrator
# =============================================================================

def process_set(set_name: str, input_dir: Path, *, columns,
                id_col, text_col, single_enums, multi_enums,
                open_text, required, other_pairs) -> tuple[int, int]:
    cleaned_dir = OUT_DIR / f"{set_name.lower()} - cleaned"
    cleaned_dir.mkdir(parents=True, exist_ok=True)

    per_coder_raw: dict[str, dict[str, pd.DataFrame]] = {}
    per_coder_clean: dict[str, dict[str, pd.DataFrame]] = {}

    coder_files = sorted(input_dir.glob("*.xlsx"))
    print(f"\n{'='*70}\nProcessing {set_name} — {len(coder_files)} coder files")
    print("="*70)

    for path in coder_files:
        coder = coder_from_filename(path)

        # Raw read (uncleaned, only header alignment)
        raw_sheets: dict[str, pd.DataFrame] = {}
        for sheet in ("Nigeria", "Kenya"):
            raw = pd.read_excel(path, sheet_name=sheet, header=1, dtype=object)
            raw = raw[raw.iloc[:, 0].notna()].copy().reset_index(drop=True)
            raw = raw.iloc[:, : len(columns)].copy()
            raw.columns = columns
            raw_sheets[sheet] = raw
        per_coder_raw[coder] = raw_sheets

        # Clean
        cleaned = read_coder_file(path, columns=columns, id_col=id_col,
                                  single_enums=single_enums,
                                  multi_enums=multi_enums,
                                  open_text=open_text)
        per_coder_clean[coder] = cleaned

        # Write cleaned per-coder file
        new_name = path.name.replace(
            f"{set_name} Analysis Codebook",
            f"{set_name} Analysis Codebook (cleaned)",
        )
        out_path = cleaned_dir / new_name
        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            for s, df in cleaned.items():
                df.to_excel(w, sheet_name=s, index=False)
        style_workbook(out_path)
        print(f"  cleaned -> {out_path}")

    # Master
    master_sheets = build_master(per_coder_clean, columns=columns, id_col=id_col)
    master_sheets = {
        "Methodology": methodology_sheet(set_name, id_col),
        "Cleaning Log": cleaning_log(per_coder_raw, per_coder_clean),
        **master_sheets,
    }
    master_path = OUT_DIR / f"Master Human {set_name} Codebook.xlsx"
    with pd.ExcelWriter(master_path, engine="openpyxl") as w:
        for s, df in master_sheets.items():
            df.to_excel(w, sheet_name=s, index=False)
    style_workbook(master_path)
    print(f"\nMaster: {master_path}")
    for s, df in master_sheets.items():
        print(f"  [{s}] {len(df)} rows")

    # Recoding report
    report_sheets, df_iss = build_recoding_report(
        per_coder_clean,
        columns=columns, id_col=id_col,
        single_enums=single_enums, multi_enums=multi_enums,
        required=required, other_pairs=other_pairs,
        set_name=set_name,
    )
    report_path = OUT_DIR / f"Recoding Needed - {set_name}.xlsx"
    with pd.ExcelWriter(report_path, engine="openpyxl") as w:
        for s, df in report_sheets.items():
            df.to_excel(w, sheet_name=s[:31], index=False)
    style_workbook(report_path)
    print(f"\nRecoding report: {report_path}")
    print(f"  Total issues: {len(df_iss)}")
    if not df_iss.empty:
        print("  By severity:", df_iss["Severity"].value_counts().to_dict())
        print("  By coder:   ", df_iss["Coder"].value_counts().to_dict())
        print("  By pass:    ", df_iss["Pass"].value_counts().sort_index().to_dict())

    return len(df_iss), int((df_iss["Severity"] == "HIGH").sum()) if not df_iss.empty else 0


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    process_set(
        "Content", CONTENT_IN,
        columns=CONTENT_COLS, id_col=CONTENT_ID_COL,
        text_col=CONTENT_TEXT_COL,
        single_enums=CONTENT_SINGLE_ENUMS, multi_enums=CONTENT_MULTI_ENUMS,
        open_text=CONTENT_OPEN_TEXT, required=CONTENT_REQUIRED,
        other_pairs=CONTENT_OTHER_PAIRS,
    )

    process_set(
        "Audience", AUDIENCE_IN,
        columns=AUDIENCE_COLS, id_col=AUDIENCE_ID_COL,
        text_col=AUDIENCE_TEXT_COL,
        single_enums=AUDIENCE_SINGLE_ENUMS, multi_enums=AUDIENCE_MULTI_ENUMS,
        open_text=AUDIENCE_OPEN_TEXT, required=AUDIENCE_REQUIRED,
        other_pairs=AUDIENCE_OTHER_PAIRS,
    )


if __name__ == "__main__":
    main()
