from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(
    "/Users/dhruvbhanderi/Documents/USC/New Research Engineer/"
    "Audience analysis kenya/Final/Kenya Audience Analysis Comments.xlsx"
)
OUT_JSON = Path(
    "/Users/dhruvbhanderi/Documents/USC/New Research Engineer/Codex/"
    "outputs/kenya_audience_cleaned/cleaned_data.json"
)

HEADERS = ["Comment ID", "Influencer", "Platform", "Source URL", "Comment"]

SHEET_META = {
    "Andrew I wonder how men": {
        "influencer": "Andrew",
        "audience_type": "Regressive",
        "platform": "X",
        "id_col": "id",
        "url_col": "url",
        "comment_col": "comment_text",
        "generated_ids": False,
    },
    "Rixpoet My voice was beaten": {
        "influencer": "Rixpoet",
        "audience_type": "Progressive",
        "platform": "YouTube",
        "id_col": None,
        "url_col": "pageUrl",
        "comment_col": "comment_text",
        "generated_ids": True,
        "id_prefix": "RIXPOET",
    },
    "Eddy Men are evolving": {
        "influencer": "Eddy",
        "audience_type": "Progressive",
        "platform": "TikTok",
        "id_col": "cid",
        "url_col": "videoWebUrl",
        "comment_col": "comment_text",
        "generated_ids": False,
    },
    "EricA woman cant love": {
        "influencer": "EricA",
        "audience_type": "Regressive",
        "platform": "X",
        "id_col": "id",
        "url_col": "url",
        "comment_col": "comment_text",
        "generated_ids": False,
    },
}

TRANSLATIONS = {
    ("Andrew I wonder how men", 13): (
        "That mindset of yours is expensive, man. Do not ever lose your focus because of "
        "some few inches of body part, I mean we have other great body parts too like the "
        "brain, the nose and Many others"
    ),
    ("Andrew I wonder how men", 34): (
        "Health wise broh, it is not about sleeping around with everyone; apart from HIV "
        "there is herpes. It is not about anything else, but the market is dirty."
    ),
    ("Andrew I wonder how men", 42): (
        "Women are not cars so your line of reasoning is already questionable. But Anto "
        "Daher wants three."
    ),
    ("Andrew I wonder how men", 50): (
        "Man, the only thing you usually think about is sleeping with women 😂 🫴 there is "
        "nothing else,"
    ),
    ("Andrew I wonder how men", 59): (
        "I said this and the man raising a child who is not his insulted me badly"
    ),
    ("Andrew I wonder how men", 70): (
        "I like those men who know what true love is; they are talking with dignity"
    ),
    ("Andrew I wonder how men", 71): (
        "When my brokenness is cured, this sickness will disappear instantly. I pile them "
        "on a tray like eggs. Gaddamit."
    ),
    ("Rixpoet My voice was beaten", 64): (
        "And yea, Mine was also a Church Guy who behind closed doors was like an absolute "
        "LUNATIC!🙆🙆 Honestly I was so confused! Luckily God revealed himself to me "
        "personally and I was able to separate between who God really was and what He was "
        "not. That kept me from Commiting suicide all the times I contemplated and even "
        "attempted. That helped me make it through coz Man! I could never have lived to "
        "tell the story."
    ),
    ("Rixpoet My voice was beaten", 69): (
        "I am very sorry, my younger sibling. I feel your pain. So sad but yet so "
        "inspirational. Thank you for sharing your story, #You have inspired many, me as "
        "well.\n\nThen someone just commented that your delusional..?? Unbelievable! I mean, "
        "first of all, it has made me so angry that I have flagged and reported her comment "
        "as Negative in this community, youtu.be should block him/her...\nNegative comments "
        "towards an individual's personal story not acceptable, period!!!!"
    ),
    ("Rixpoet My voice was beaten", 71): (
        "Sorry brother, may God favor u As a lady not a minute I will allow my husband to "
        "overdiscipline my child Normal discipline is enough"
    ),
    ("Rixpoet My voice was beaten", 92): (
        "Nothing like that... Some men are self destracting. They do not listen to anything "
        "and apparently they don't like being advised, they like being encouraged....\nThat "
        "means,\nHonestly choose to do something good and people clap clap clap.\nBefore then "
        "you shut up.\nMy thoughts, encouraging and advising should go hand in hand."
    ),
    ("Eddy Men are evolving", 5): (
        "Most men used to go to clubs to grow and bond with fellow men away from the other "
        "gender until women started following them to clubs again😅"
    ),
    ("Eddy Men are evolving", 7): (
        "I think we have done it too much partied too much,,clubs,events,roadtrips..."
        "Everything....Nowadays when it reaches 9PM you have already started dozing..And no, "
        "it is not lack of money"
    ),
    ("Eddy Men are evolving", 9): (
        "I can almost guarantee you, it is not evolving, probably it is being broke. You "
        "guard the little you have so you can eat 😂😂. This is not evolving, is called "
        "adaptation 😂😂"
    ),
    ("Eddy Men are evolving", 31): (
        "We are taking course back on our purpose. taking back our masculinity we won't let "
        "peace be taken away. For any woman who feels offended, let her feel bad at her own "
        "place"
    ),
    ("Eddy Men are evolving", 40): (
        "i know most of my fellow men go into the house by 6PM if there is no business "
        "outside😂Personal peace means a lot"
    ),
    ("Eddy Men are evolving", 45): (
        "look... it is not clubbing, i stopped even going to church you see those small "
        "meetings after church tumeeet 5minutes heee... when you enter there it is pledges "
        "and contributions and church rent plzz men wake up save that idol coin in a sacco "
        "somewhere drink your tea spent on your family"
    ),
    ("Eddy Men are evolving", 47): (
        "me ileft my wife and children n let them go for good...for my mental calmeness..."
        "inearly went crazy...mad..I used to go on the road talking to myself...iwas so "
        "tired of proving to be a man enough...3yrs down the line..ilive in peace..."
    ),
    ("Eddy Men are evolving", 50): (
        "Men, let us just tell the truth, we have 2 months waiting for us, Dec and January "
        "plus wives 😅"
    ),
    ("Eddy Men are evolving", 56): (
        "ya some are saying w n don't have money ya we do not have it, that is why we have "
        "settled down and are trying to find ourselves. Real men start family and starting "
        "family need tactics 🤣🤣"
    ),
    ("Eddy Men are evolving", 61): (
        "now the women who said they can do without mehn now are complaining that men don't "
        "attend clubs anymore? daaaaamn... anyway am not going back to clubs. I order take "
        "away and go camping with the boys; when we leave there, we get into codm"
    ),
    ("Eddy Men are evolving", 64): (
        "So-called \"Evolving\"😭😂😂😂they don't believe most women are awake and have known "
        "how powerful they are and women can't be controlled anymore like the ancient days "
        "and women have money nowadays 😂😂😂😂😂🤭, that story you are saying is yours and "
        "your people. Mscheeeeeew"
    ),
    ("Eddy Men are evolving", 70): (
        "😂😂sometimes me as a man i hear some men talk, and i wonder what kind of twilight "
        "zone or delusional world do some live. So-called men evolved? when was this 😂😂If "
        "that was to happen, this counry would have less gbv, safer, less corruption, "
        "development increase etc Eish here you have really lied, just say some men are "
        "broke and cant afford events. 😫😩"
    ),
    ("Eddy Men are evolving", 94): (
        "We are okay.....women will not understand our issues....we always come to clubs to "
        "relax and have fun with other men not women....so you have moved into the club....."
        "we are leaving it to you.....we are doing things....."
    ),
    ("Eddy Men are evolving", 96): (
        "😎You think Kikuyu Men will spend 15k in a day for one event while the know that "
        "15k for just 3 beers 🍻 @250, roast meat and ugali 350bob atleast 2 - 3days per "
        "week, then the other days watch football, gym and busy working on themselves, it's "
        "enough to last for a whole month 😂 while on the other hand these strange creatures "
        "eat ugali and sukuma for three months to save 25k for only one event because that "
        "day there is a new outfit, hair done and nails 💅 Kikuyu Men are futuristic 😄"
    ),
    ("Eddy Men are evolving", 99): (
        "Those who say men are broke, the road are full of v8, buildings rising every "
        "morning, financial institutions count profits, men corners are full with bottles on "
        "the table .. this gender never wakes up 😂😂"
    ),
    ("Eddy Men are evolving", 101): (
        "Men over 28 years are done clubbing and galavanting. They are now starting to build "
        "wealth and make important decisions about their life. That man ladies are waiting to "
        "engage is already engaged in money, peace, family and calling! We left the club "
        "quietly and went home. We have a beer over sports,football ! We are not there. We "
        "are busy!"
    ),
    ("Eddy Men are evolving", 111): (
        "Since I left stories about women and partying alone I can say I have seen tremendous "
        "self improvement financially mentally and spirituality all my projects are going on "
        "well and am just okay with being indoors"
    ),
    ("EricA woman cant love", 14): (
        "@amerix Let me choose someone who loves me so that I can give her respect"
    ),
    ("EricA woman cant love", 32): (
        "@amerix On the man's roles, you forgot the most important...he is to be the PRIEST "
        "of the home"
    ),
    ("EricA woman cant love", 129): (
        "@amerix You set up CCTV to catch your CHEATING WIFE in the act but you have mixed "
        "feelings after watching the video because of the way you are being cheated on; her "
        "screams and moans leave you wondering whether to GET ANGRY or GET AROUSED 🤣....... "
        "life is hard!\n\n#RutoMustGoNow"
    ),
    ("EricA woman cant love", 133): (
        "@amerix This reverse psychology confuses me, man 😂😂"
    ),
}


def clean_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "", name)[:31]


def col_map(ws):
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1)}


def main() -> None:
    wb = load_workbook(SOURCE, data_only=False)
    sheets = []
    summary_by_influencer = []
    by_platform = Counter()
    by_type = Counter()
    total_comments = 0
    total_generated_ids = 0
    translated_counts = Counter()
    generated_counts = Counter()

    for ws in wb.worksheets:
        meta = SHEET_META[ws.title]
        headers = col_map(ws)
        keep_col = headers.get("keep")
        comment_col = headers[meta["comment_col"]]
        url_col = headers[meta["url_col"]]
        id_col = headers.get(meta["id_col"]) if meta["id_col"] else None
        rows = []
        generated_index = 1
        translated = 0

        for row_idx in range(2, ws.max_row + 1):
            keep = ws.cell(row_idx, keep_col).value if keep_col else True
            comment = ws.cell(row_idx, comment_col).value
            if keep is not True or comment is None or not str(comment).strip():
                continue

            if id_col:
                comment_id = str(ws.cell(row_idx, id_col).value).strip()
            else:
                comment_id = f"{meta['id_prefix']}-{generated_index:04d}"
                generated_index += 1

            text = str(comment)
            if (ws.title, row_idx) in TRANSLATIONS:
                text = TRANSLATIONS[(ws.title, row_idx)]
                translated += 1

            rows.append(
                [
                    comment_id,
                    meta["influencer"],
                    meta["platform"],
                    str(ws.cell(row_idx, url_col).value or "").strip(),
                    text,
                ]
            )

        comment_count = len(rows)
        total_comments += comment_count
        by_platform[meta["platform"]] += comment_count
        by_type[meta["audience_type"]] += comment_count
        translated_counts[meta["influencer"]] = translated
        if meta["generated_ids"]:
            generated_counts[meta["influencer"]] = comment_count
            total_generated_ids += comment_count

        summary_by_influencer.append(
            {
                "Influencer": meta["influencer"],
                "Audience Type": meta["audience_type"],
                "Platform": meta["platform"],
                "Comments": comment_count,
                "Generated IDs": comment_count if meta["generated_ids"] else 0,
                "Translated Comments": translated,
            }
        )
        sheets.append(
            {
                "name": clean_sheet_name(ws.title),
                "headers": HEADERS,
                "rows": rows,
            }
        )

    summary = {
        "overview": [
            ["Metric", "Value"],
            ["Total comments", total_comments],
            ["Progressive comments", by_type["Progressive"]],
            ["Regressive comments", by_type["Regressive"]],
            ["Generated comment IDs", total_generated_ids],
            ["Translated/mixed-language comments", sum(translated_counts.values())],
        ],
        "by_influencer": [
            ["Influencer", "Audience Type", "Platform", "Comments", "Generated IDs", "Translated Comments"],
            *[
                [
                    row["Influencer"],
                    row["Audience Type"],
                    row["Platform"],
                    row["Comments"],
                    row["Generated IDs"],
                    row["Translated Comments"],
                ]
                for row in summary_by_influencer
            ],
        ],
        "by_platform": [["Platform", "Comments"], *[[k, v] for k, v in by_platform.items()]],
        "by_type": [["Audience Type", "Comments"], *[[k, v] for k, v in by_type.items()]],
    }

    payload = {
        "source": str(SOURCE),
        "sheets": sheets,
        "summary": summary,
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
