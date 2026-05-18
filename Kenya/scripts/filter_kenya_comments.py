 #!/usr/bin/env python3
"""Filter comment workbooks using keywords from a selected keyword sheet."""

from __future__ import annotations

import argparse
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook


HERE = Path(__file__).resolve().parent
FILTERING_ROOT = HERE.parent
DEFAULT_INPUT_DIR = FILTERING_ROOT / "inputs" / "Kenya"
DEFAULT_OUTPUT_ROOT = FILTERING_ROOT / "outputs" / "Kenya"
DEFAULT_KEYWORD_WORKBOOK = FILTERING_ROOT / "keywords" / "NLC Proposed keywords.xlsx"

KEYWORD_WORKBOOK_NAME = "NLC Proposed keywords.xlsx"
DEFAULT_KEYWORD_SHEET_NAME = "Kenya"
OUTPUT_DIR_NAME = "filtered_output"
EXPANDED_OUTPUT_DIR_NAME = "filtered_output_expanded_variants"
RUN_SUMMARY_NAME = "_run_summary.xlsx"
EXPECTED_KEYWORD_COUNT = 295
SUMMARY_TOP_KEYWORD_LIMIT = 20
ACTIVE_RELEVANCE = {"Highly relevant", "Moderately relevant"}
VALID_RELEVANCE = ACTIVE_RELEVANCE | {"Not relevant"}
MATCH_MODES = {"conservative", "expanded"}
ADDED_COLUMNS = [
    "matched_keywords",
    "matched_keyword_count",
    "source_file",
    "source_sheet",
    "content_column",
]

APOSTROPHE_TRANSLATION = str.maketrans(
    {
        "\u2018": "'",
        "\u2019": "'",
        "\u02bc": "'",
        "\u0060": "'",
        "\u00b4": "'",
    }
)

PUNCTUATION_TO_SPACE = re.compile(r"[^\w\s#']+")
WHITESPACE_RE = re.compile(r"\s+")


@dataclass(frozen=True)
class KeywordMatcher:
    keyword: str
    hashtag_patterns: tuple[re.Pattern[str], ...] = ()
    token_patterns: tuple[re.Pattern[str], ...] = ()
    phrase_variants: tuple[str, ...] = ()
    compact_phrase_variants: tuple[str, ...] = ()

    def matches(self, raw_text: str, normalized_text: str, compact_text: str) -> bool:
        if any(pattern.search(raw_text) for pattern in self.hashtag_patterns):
            return True
        if any(pattern.search(normalized_text) for pattern in self.token_patterns):
            return True
        if any(variant in normalized_text for variant in self.phrase_variants):
            return True
        if any(variant in compact_text for variant in self.compact_phrase_variants):
            return True
        return False


@dataclass(frozen=True)
class KeywordEntry:
    keyword: str
    relevance: str | None
    nlc_comment: str | None


def normalize_for_matching(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    text = text.translate(APOSTROPHE_TRANSLATION).lower()
    text = PUNCTUATION_TO_SPACE.sub(" ", text)
    text = WHITESPACE_RE.sub(" ", text).strip()
    return text


def normalize_for_hashtag_search(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    text = text.translate(APOSTROPHE_TRANSLATION).lower()
    return WHITESPACE_RE.sub(" ", text)


def normalize_for_compact_matching(value: str) -> str:
    text = unicodedata.normalize("NFKC", value)
    text = text.translate(APOSTROPHE_TRANSLATION).lower()
    return "".join(char for char in text if char.isalnum())


def normalize_relevance(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = WHITESPACE_RE.sub(" ", str(value)).strip().lower()
    if normalized == "highly relevant":
        return "Highly relevant"
    if normalized == "moderately relevant":
        return "Moderately relevant"
    if normalized == "not relevant":
        return "Not relevant"
    return None


def infer_relevance_from_nlc_comment(value: str | None) -> str | None:
    if not value:
        return None
    normalized = str(value).strip().lower()
    if "not relevant" in normalized:
        return "Not relevant"
    if "highly relevant" in normalized:
        return "Highly relevant"
    if "moderately relevant" in normalized:
        return "Moderately relevant"
    return None


def find_keyword_sheet_columns(worksheet) -> tuple[int, int | None, int | None]:
    headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    normalized_headers = [
        "" if value is None else WHITESPACE_RE.sub(" ", str(value)).strip().lower()
        for value in headers
    ]

    keyword_idx = next(
        (idx for idx, header in enumerate(normalized_headers) if header == "keyword"),
        None,
    )
    relevance_idx = next(
        (idx for idx, header in enumerate(normalized_headers) if "relevance" in header),
        None,
    )
    nlc_comment_idx = next(
        (idx for idx, header in enumerate(normalized_headers) if "nlc" in header and "comment" in header),
        None,
    )

    if keyword_idx is None:
        raise ValueError("Keyword sheet is missing a 'Keyword' header.")
    return keyword_idx, relevance_idx, nlc_comment_idx


def load_keywords(keyword_workbook: Path, keyword_sheet_name: str) -> list[KeywordEntry]:
    workbook = load_workbook(keyword_workbook, read_only=True, data_only=True)
    worksheet = workbook[keyword_sheet_name]
    keyword_idx, relevance_idx, nlc_comment_idx = find_keyword_sheet_columns(worksheet)
    keywords: list[KeywordEntry] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        values = list(row)
        keyword_value = values[keyword_idx] if keyword_idx < len(values) else None
        relevance_value = values[relevance_idx] if relevance_idx is not None and relevance_idx < len(values) else None
        nlc_comment_value = values[nlc_comment_idx] if nlc_comment_idx is not None and nlc_comment_idx < len(values) else None
        if keyword_value is None:
            continue
        keyword = str(keyword_value).strip()
        if keyword:
            keywords.append(
                KeywordEntry(
                    keyword=keyword,
                    relevance=normalize_relevance(relevance_value),
                    nlc_comment=None if nlc_comment_value is None else str(nlc_comment_value).strip(),
                )
            )
    return keywords


def load_excluded_keywords(exclude_keywords_file: Path | None) -> set[str]:
    if exclude_keywords_file is None:
        return set()
    excluded: set[str] = set()
    for line in exclude_keywords_file.read_text(encoding="utf-8").splitlines():
        value = line.strip()
        if not value or value.startswith("#"):
            continue
        excluded.add(value)
    return excluded


def resolve_relevance(entry: KeywordEntry) -> str | None:
    if entry.relevance in VALID_RELEVANCE:
        return entry.relevance
    return infer_relevance_from_nlc_comment(entry.nlc_comment)


def generate_morphological_variants(word: str) -> set[str]:
    variants: set[str] = set()
    if len(word) <= 2 or not word.isalpha():
        return variants

    if "'" in word:
        variants.add(word.replace("'", ""))

    irregulars = {
        "man": "men",
        "men": "man",
        "woman": "women",
        "women": "woman",
    }
    if word in irregulars:
        variants.add(irregulars[word])

    if word.endswith("y") and len(word) > 3:
        variants.add(word[:-1] + "ies")
    elif word.endswith("ies") and len(word) > 4:
        variants.add(word[:-3] + "y")

    if word.endswith("es") and len(word) > 4:
        variants.add(word[:-2])
    elif word.endswith("s") and not word.endswith("ss") and len(word) > 3:
        variants.add(word[:-1])
    else:
        variants.add(word + "s")

    if word.endswith(("s", "x", "z", "sh", "ch")):
        variants.add(word + "es")

    variants.discard(word)
    return variants


def generate_single_word_variants(word: str, active_keyword_norms: set[str]) -> tuple[str, ...]:
    variants = [word]
    for candidate in sorted(generate_morphological_variants(word)):
        if candidate not in active_keyword_norms:
            variants.append(candidate)
    return tuple(dict.fromkeys(variants))


def generate_phrase_variants(
    normalized_phrase: str, active_keyword_norms: set[str]
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    phrase_variants = [normalized_phrase]
    compact_variants = [normalize_for_compact_matching(normalized_phrase)]

    tokens = normalized_phrase.split()
    if tokens:
        last_token = tokens[-1]
        for candidate in sorted(generate_morphological_variants(last_token)):
            alt_phrase = " ".join(tokens[:-1] + [candidate])
            if alt_phrase not in active_keyword_norms:
                phrase_variants.append(alt_phrase)
                compact_variants.append(normalize_for_compact_matching(alt_phrase))

    return (
        tuple(dict.fromkeys(phrase_variants)),
        tuple(dict.fromkeys(value for value in compact_variants if value)),
    )


def build_matcher(keyword: str, match_mode: str, active_keyword_norms: set[str]) -> KeywordMatcher:
    normalized_keyword = normalize_for_matching(keyword)

    if keyword.startswith("#"):
        if match_mode == "expanded":
            hashtag_body = normalize_for_matching(keyword[1:])
            hashtag_tokens = hashtag_body.split()
            if hashtag_tokens:
                separator = r"[\s_\-]*"
                hashtag_pattern = separator.join(re.escape(token) for token in hashtag_tokens)
                patterns = (
                    re.compile(rf"(?<![\w#])#{hashtag_pattern}(?!\w)", re.IGNORECASE),
                )
            else:
                patterns = (
                    re.compile(rf"(?<![\w#]){re.escape(keyword.lower())}(?!\w)", re.IGNORECASE),
                )
        else:
            patterns = (
                re.compile(rf"(?<![\w#]){re.escape(keyword.lower())}(?!\w)", re.IGNORECASE),
            )
        return KeywordMatcher(keyword=keyword, hashtag_patterns=patterns)

    if " " in normalized_keyword:
        phrase_variants = (normalized_keyword,)
        compact_variants: tuple[str, ...] = ()
        if match_mode == "expanded":
            phrase_variants, compact_variants = generate_phrase_variants(
                normalized_keyword, active_keyword_norms
            )
        return KeywordMatcher(
            keyword=keyword,
            phrase_variants=phrase_variants,
            compact_phrase_variants=compact_variants,
        )

    variants = (normalized_keyword,)
    if match_mode == "expanded":
        variants = generate_single_word_variants(normalized_keyword, active_keyword_norms)
    token_patterns = tuple(
        re.compile(rf"(?<![a-z0-9]){re.escape(variant)}(?![a-z0-9])", re.IGNORECASE)
        for variant in variants
    )
    return KeywordMatcher(keyword=keyword, token_patterns=token_patterns)


def find_content_column(headers: Sequence[str]) -> str:
    lowered = {header.lower(): header for header in headers}
    if "comment" in lowered:
        return lowered["comment"]
    if "text" in lowered:
        return lowered["text"]
    raise ValueError("Could not find a 'comment' or 'text' column.")


def match_keywords(matchers: Sequence[KeywordMatcher], content: str | None) -> list[str]:
    if content is None:
        return []
    text = str(content)
    if not text.strip():
        return []

    raw_text = normalize_for_hashtag_search(text)
    normalized_text = normalize_for_matching(text)
    compact_text = normalize_for_compact_matching(text)
    matches: list[str] = []
    for matcher in matchers:
        if matcher.matches(
            raw_text=raw_text,
            normalized_text=normalized_text,
            compact_text=compact_text,
        ):
            matches.append(matcher.keyword)
    return matches


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        letter = column[0].column_letter
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 60)


def write_filtered_workbook(
    output_path: Path,
    match_mode: str,
    source_file: str,
    source_sheet: str,
    content_column: str,
    headers: Sequence[str],
    matched_rows: Sequence[Sequence[object]],
    total_rows: int,
    keyword_counter: Counter[str],
) -> None:
    workbook = Workbook()

    filtered_sheet = workbook.active
    filtered_sheet.title = "Filtered"
    filtered_headers = list(headers) + ADDED_COLUMNS
    filtered_sheet.append(filtered_headers)
    for row in matched_rows:
        filtered_sheet.append(list(row))

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["match_mode", match_mode])
    summary_sheet.append(["source_file", source_file])
    summary_sheet.append(["source_sheet", source_sheet])
    summary_sheet.append(["content_column", content_column])
    summary_sheet.append(["total_rows", total_rows])
    summary_sheet.append(["matched_rows", len(matched_rows)])
    match_rate = (len(matched_rows) / total_rows) if total_rows else 0
    summary_sheet.append(["match_rate", round(match_rate, 6)])
    summary_sheet.append(["unique_matched_keywords", len(keyword_counter)])
    summary_sheet.append([])
    summary_sheet.append(["Top keyword", "Match count"])
    if keyword_counter:
        for keyword, count in keyword_counter.most_common(SUMMARY_TOP_KEYWORD_LIMIT):
            summary_sheet.append([keyword, count])
    else:
        summary_sheet.append(["No matches", 0])

    autosize_columns(filtered_sheet)
    autosize_columns(summary_sheet)
    workbook.save(output_path)


def write_run_summary(output_dir: Path, file_summaries: Sequence[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Run Summary"
    headers = [
        "match_mode",
        "source_file",
        "source_sheet",
        "content_column",
        "total_rows",
        "matched_rows",
        "match_rate",
        "unique_matched_keywords",
        "output_file",
    ]
    sheet.append(headers)

    for summary in file_summaries:
        total_rows = int(summary["total_rows"])
        matched_rows = int(summary["matched_rows"])
        match_rate = matched_rows / total_rows if total_rows else 0
        sheet.append(
            [
                summary["match_mode"],
                summary["source_file"],
                summary["source_sheet"],
                summary["content_column"],
                total_rows,
                matched_rows,
                round(match_rate, 6),
                summary["unique_matched_keywords"],
                summary["output_file"],
            ]
        )

    autosize_columns(sheet)
    workbook.save(output_dir / RUN_SUMMARY_NAME)


def iter_source_workbooks(input_dir: Path, keyword_workbook: Path) -> Iterable[Path]:
    for path in sorted(input_dir.glob("*.xlsx")):
        if path.resolve() == keyword_workbook.resolve():
            continue
        yield path


def run_validations(
    keyword_entries: Sequence[KeywordEntry],
    matchers: Sequence[KeywordMatcher],
    match_mode: str,
    keyword_sheet_name: str,
) -> tuple[int, int, int]:
    keywords = [entry.keyword for entry in keyword_entries]
    unique_keywords = len(set(keywords))
    if keyword_sheet_name == DEFAULT_KEYWORD_SHEET_NAME and (
        len(keywords) != EXPECTED_KEYWORD_COUNT or unique_keywords != EXPECTED_KEYWORD_COUNT
    ):
        raise ValueError(
            "Expected exactly "
            f"{EXPECTED_KEYWORD_COUNT} unique Kenya keywords, "
            f"found {len(keywords)} rows and {unique_keywords} unique values."
        )
    if not keywords:
        raise ValueError(f"No keywords found in sheet '{keyword_sheet_name}'.")

    active_keywords = sum(1 for entry in keyword_entries if resolve_relevance(entry) in ACTIVE_RELEVANCE)
    not_relevant_keywords = sum(1 for entry in keyword_entries if resolve_relevance(entry) == "Not relevant")
    unresolved_keywords = sum(1 for entry in keyword_entries if resolve_relevance(entry) is None)

    keyword_names = {matcher.keyword for matcher in matchers}

    if "a woman can't love a man" in keyword_names:
        phrase_matches = match_keywords(matchers, "A woman can’t love a man, but the phrase should match.")
        if "a woman can't love a man" not in phrase_matches:
            raise ValueError("Phrase normalization validation failed.")

    if "#MasculinitySaturday" in keyword_names:
        hashtag_matches = match_keywords(matchers, "This is #masculinitysaturday in action.")
        if "#MasculinitySaturday" not in hashtag_matches:
            raise ValueError("Hashtag validation failed.")

    if "boy" in keyword_names:
        token_matches = match_keywords(matchers, "My boyfriend is here.")
        if "boy" in token_matches:
            raise ValueError("Whole-word token validation failed for 'boy'.")

    if match_mode == "expanded" and "body count" in keyword_names:
        compact_phrase_matches = match_keywords(matchers, "People keep arguing about bodycount online.")
        if "body count" not in compact_phrase_matches:
            raise ValueError("Expanded compact phrase validation failed for 'body count'.")

    return active_keywords, not_relevant_keywords, unresolved_keywords


def process_workbook(
    workbook_path: Path,
    output_dir: Path,
    matchers: Sequence[KeywordMatcher],
    match_mode: str,
) -> dict[str, object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    source_sheet = workbook.sheetnames[0]
    worksheet = workbook[source_sheet]

    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = ["" if value is None else str(value).strip() for value in header_values]
    content_column = find_content_column(headers)
    content_index = headers.index(content_column)

    matched_rows: list[list[object]] = []
    keyword_counter: Counter[str] = Counter()
    total_rows = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        values = list(row)
        if len(values) < len(headers):
            values.extend([None] * (len(headers) - len(values)))

        content = values[content_index] if content_index < len(values) else None
        matched_keywords = match_keywords(matchers, content)
        if not matched_keywords:
            continue

        keyword_counter.update(matched_keywords)
        matched_rows.append(
            values
            + [
                " | ".join(matched_keywords),
                len(matched_keywords),
                workbook_path.name,
                source_sheet,
                content_column,
            ]
        )

    output_name = f"{workbook_path.stem}__filtered.xlsx"
    output_path = output_dir / output_name
    write_filtered_workbook(
        output_path=output_path,
        match_mode=match_mode,
        source_file=workbook_path.name,
        source_sheet=source_sheet,
        content_column=content_column,
        headers=headers,
        matched_rows=matched_rows,
        total_rows=total_rows,
        keyword_counter=keyword_counter,
    )

    return {
        "match_mode": match_mode,
        "source_file": workbook_path.name,
        "source_sheet": source_sheet,
        "content_column": content_column,
        "total_rows": total_rows,
        "matched_rows": len(matched_rows),
        "unique_matched_keywords": len(keyword_counter),
        "output_file": output_name,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Filter Kenya dataset workbooks using keywords from the Kenya sheet."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="Directory containing the source workbooks.",
    )
    parser.add_argument(
        "--keyword-workbook",
        type=Path,
        default=DEFAULT_KEYWORD_WORKBOOK,
        help=f"Path to the keyword workbook. Defaults to keywords/{KEYWORD_WORKBOOK_NAME}.",
    )
    parser.add_argument(
        "--keyword-sheet",
        default=DEFAULT_KEYWORD_SHEET_NAME,
        help="Worksheet name inside the keyword workbook to use for filtering.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help=f"Output directory. Defaults to outputs/Kenya/{OUTPUT_DIR_NAME} "
        f"(or {EXPANDED_OUTPUT_DIR_NAME} for --match-mode expanded).",
    )
    parser.add_argument(
        "--exclude-keywords-file",
        type=Path,
        default=None,
        help="Optional newline-delimited keyword exclusion file.",
    )
    parser.add_argument(
        "--match-mode",
        choices=sorted(MATCH_MODES),
        default="conservative",
        help="Matching mode to use for comment filtering.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_dir = args.input_dir.resolve()
    keyword_workbook = args.keyword_workbook.resolve()
    output_dir = (
        args.output_dir.resolve()
        if args.output_dir is not None
        else DEFAULT_OUTPUT_ROOT
        / (EXPANDED_OUTPUT_DIR_NAME if args.match_mode == "expanded" else OUTPUT_DIR_NAME)
    )
    exclude_keywords_file = (
        args.exclude_keywords_file.resolve()
        if args.exclude_keywords_file is not None
        else None
    )

    if not keyword_workbook.exists():
        raise FileNotFoundError(f"Keyword workbook not found: {keyword_workbook}")
    if exclude_keywords_file is not None and not exclude_keywords_file.exists():
        raise FileNotFoundError(f"Exclude-keywords file not found: {exclude_keywords_file}")

    output_dir.mkdir(parents=True, exist_ok=True)

    keyword_entries = load_keywords(keyword_workbook, args.keyword_sheet)
    excluded_keywords = load_excluded_keywords(exclude_keywords_file)
    active_keywords = [
        entry.keyword
        for entry in keyword_entries
        if resolve_relevance(entry) in ACTIVE_RELEVANCE and entry.keyword not in excluded_keywords
    ]
    active_keyword_norms = {normalize_for_matching(keyword) for keyword in active_keywords}
    matchers = [
        build_matcher(
            keyword=keyword,
            match_mode=args.match_mode,
            active_keyword_norms=active_keyword_norms,
        )
        for keyword in active_keywords
    ]
    active_count, excluded_not_relevant_count, unresolved_count = run_validations(
        keyword_entries, matchers, args.match_mode, args.keyword_sheet
    )

    file_summaries: list[dict[str, object]] = []
    for workbook_path in iter_source_workbooks(input_dir, keyword_workbook):
        file_summaries.append(
            process_workbook(
                workbook_path=workbook_path,
                output_dir=output_dir,
                matchers=matchers,
                match_mode=args.match_mode,
            )
        )

    write_run_summary(output_dir, file_summaries)

    total_rows = sum(int(summary["total_rows"]) for summary in file_summaries)
    total_matches = sum(int(summary["matched_rows"]) for summary in file_summaries)
    print(f"Match mode: {args.match_mode}")
    print(f"Keyword sheet: {args.keyword_sheet}")
    print(f"Keywords loaded: {len(keyword_entries)}")
    print(f"Keywords excluded by file: {len(excluded_keywords)}")
    print(f"Relevant keywords before exclusions: {active_count}")
    print(f"Active keywords used: {len(active_keywords)}")
    print(f"Excluded as not relevant: {excluded_not_relevant_count}")
    print(f"Excluded as unresolved: {unresolved_count}")
    print(f"Source workbooks processed: {len(file_summaries)}")
    print(f"Rows scanned: {total_rows}")
    print(f"Matched rows exported: {total_matches}")
    print(f"Output directory: {output_dir}")


if __name__ == "__main__":
    main()
