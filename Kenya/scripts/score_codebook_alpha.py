#!/usr/bin/env python3
"""
Compare human and LLM codebook coding by Content ID.

Outputs:
  - A cleaned merged file with Human_* and LLM_* columns.
  - Question-level scores.
  - Overall summary scores.

The scorer follows the requested split:
  - Single-answer columns: percent agreement and nominal Krippendorff's alpha.
  - Multi-select columns: exact agreement and average Jaccard similarity.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Iterable, Sequence


IDENTITY_COLUMNS = {
    "content id",
    "content_id",
    "id",
    "context",
    "content text / description",
    "content text",
    "description",
}

DEFAULT_ID_COLUMN = "Content ID"


@dataclass(frozen=True)
class QuestionScore:
    column: str
    question_id: str
    question: str
    column_type: str
    n_valid: int
    percent_or_exact: float | None
    alpha: float | None
    jaccard: float | None

    @property
    def main_metric(self) -> str:
        if self.column_type == "Single-answer":
            return "Krippendorff's alpha"
        return "Jaccard"

    @property
    def main_score(self) -> float | None:
        if self.column_type == "Single-answer":
            return self.alpha
        return self.jaccard


def normalize_spaces(value: object) -> str:
    if value is None:
        return ""
    value = str(value).replace("\u00a0", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return value


def normalize_key(value: object) -> str:
    return normalize_spaces(value).casefold()


def canonical_header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize_key(value)).strip()


def clean_single(value: object) -> str:
    return normalize_key(value)


def display_single(value: object, canonical_labels: dict[str, str]) -> str:
    key = clean_single(value)
    return canonical_labels.get(key, normalize_spaces(value))


def parse_options(header: str) -> list[str]:
    if "Options:" not in header:
        return []

    option_text = header.split("Options:", 1)[1]
    option_text = re.sub(r"\bOpen text\b", "", option_text, flags=re.IGNORECASE)
    parts = re.split(r"\n| / ", option_text)

    options: list[str] = []
    seen: set[str] = set()
    for part in parts:
        label = normalize_spaces(part).strip(" /")
        if not label:
            continue
        key = normalize_key(label)
        if key not in seen:
            seen.add(key)
            options.append(label)
    return options


def split_multi_value(value: object, known_labels: Sequence[str] | None = None) -> list[str]:
    value = normalize_spaces(value)
    if not value:
        return []

    if known_labels:
        found: list[tuple[int, str]] = []
        value_key = normalize_key(value)
        for label in known_labels:
            label_key = normalize_key(label)
            if not label_key:
                continue
            index = value_key.find(label_key)
            if index >= 0:
                found.append((index, label))
        if found:
            found.sort()
            return [label for _, label in found]

    # The codebook convention is semicolon-separated multi-select answers.
    # Also accept common pasted separators. Commas are a fallback only when
    # the header did not provide option labels, because some labels contain commas.
    parts = re.split(r"\s*(?:;|,|\||\n|\r|\t)\s*", value)
    return [normalize_spaces(part) for part in parts if normalize_spaces(part)]


def clean_multi_set(value: object, known_labels: Sequence[str] | None = None) -> set[str]:
    return {normalize_key(label) for label in split_multi_value(value, known_labels) if normalize_key(label)}


def display_multi(
    value: object,
    canonical_labels: dict[str, str],
    known_labels: Sequence[str] | None = None,
) -> str:
    labels = clean_multi_set(value, known_labels)
    return "; ".join(canonical_labels.get(label, label) for label in sorted(labels))


def question_id(header: str) -> str:
    match = re.match(r"\s*(Q\d+[a-z]?)\.", header, flags=re.IGNORECASE)
    return match.group(1).upper() if match else normalize_spaces(header)


def question_text(header: str) -> str:
    return normalize_spaces(header.split("Options:", 1)[0])


def is_open_text(header: str) -> bool:
    key = normalize_key(header)
    return "open text" in key or "if other, specify" in key


def looks_multi_select(header: str, values: Iterable[str]) -> bool:
    key = normalize_key(header)
    if (
        "choose all that apply" in key
        or "select all that apply" in key
        or "multi-select" in key
        or "multiselect" in key
        or "multiple" in key
    ):
        return True

    # A data-based fallback for compact column names like "topics" or
    # "themes__themes" where labels are stored as "A; B".
    sample = [normalize_spaces(value) for value in values if normalize_spaces(value)]
    return any(";" in value for value in sample)


def parse_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        return [row for row in csv.reader(file)]


def parse_xlsx(path: Path, sheet_name: str | None = None) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading .xlsx files requires openpyxl. Install requirements.txt first.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows: list[list[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append([normalize_spaces(cell) for cell in row])
    return rows


def read_rows(path: Path, sheet_name: str | None = None) -> list[list[str]]:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return parse_csv(path)
    if suffix == ".tsv":
        with path.open(newline="", encoding="utf-8-sig") as file:
            return [row for row in csv.reader(file, delimiter="\t")]
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(path, sheet_name)
    raise ValueError(f"Unsupported file type for {path}. Use CSV, TSV, XLSX, or XLSM.")


def find_header_row(rows: Sequence[Sequence[str]], id_column: str) -> int:
    id_key = canonical_header_key(id_column)
    for index, row in enumerate(rows[:10]):
        keys = {canonical_header_key(cell) for cell in row}
        if id_key in keys:
            return index
    raise ValueError(f"Could not find an '{id_column}' column in the first 10 rows")


def dedupe_headers(headers: Sequence[str]) -> list[str]:
    seen: Counter[str] = Counter()
    output: list[str] = []
    for index, header in enumerate(headers):
        header = normalize_spaces(header) or f"Unnamed column {index + 1}"
        seen[header] += 1
        if seen[header] == 1:
            output.append(header)
        else:
            output.append(f"{header} ({seen[header]})")
    return output


def read_coding_file(
    path: Path,
    id_column: str,
    sheet_name: str | None = None,
) -> tuple[list[str], dict[str, dict[str, str]], str]:
    rows = read_rows(path, sheet_name)
    if not rows:
        raise ValueError(f"{path} is empty")

    header_index = find_header_row(rows, id_column)
    headers = dedupe_headers(rows[header_index])
    id_lookup = {canonical_header_key(header): header for header in headers}
    actual_id_column = id_lookup[canonical_header_key(id_column)]
    id_index = headers.index(actual_id_column)

    data: dict[str, dict[str, str]] = {}
    for row in rows[header_index + 1 :]:
        if not row or all(not normalize_spaces(cell) for cell in row):
            continue
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        content_id = normalize_spaces(padded[id_index])
        if not content_id:
            continue
        if content_id in data:
            raise ValueError(f"Duplicate Content ID '{content_id}' found in {path}")
        data[content_id] = {
            header: normalize_spaces(padded[index]) if index < len(padded) else ""
            for index, header in enumerate(headers)
        }

    return headers, data, actual_id_column


def shared_codebook_columns(
    human_headers: Sequence[str],
    llm_headers: Sequence[str],
    id_column: str,
    include_open_text: bool,
) -> list[tuple[str, str]]:
    llm_lookup = {canonical_header_key(header): header for header in llm_headers}
    pairs: list[tuple[str, str]] = []
    for human_header in human_headers:
        key = canonical_header_key(human_header)
        if key == canonical_header_key(id_column) or normalize_key(human_header) in IDENTITY_COLUMNS:
            continue
        if is_open_text(human_header) and not include_open_text:
            continue
        if key in llm_lookup:
            pairs.append((human_header, llm_lookup[key]))
    return pairs


def parse_column_list(raw: str | None) -> set[str]:
    if not raw:
        return set()
    return {canonical_header_key(part) for part in re.split(r"\s*,\s*", raw) if normalize_spaces(part)}


def build_label_map(
    ids: Sequence[str],
    human_rows: dict[str, dict[str, str]],
    llm_rows: dict[str, dict[str, str]],
    human_header: str,
    llm_header: str,
    multi: bool,
    known_labels: Sequence[str] | None = None,
) -> dict[str, str]:
    labels: dict[str, str] = {}
    for label in known_labels or []:
        key = normalize_key(label)
        if key:
            labels[key] = normalize_spaces(label)
    for content_id in ids:
        values = [
            human_rows[content_id].get(human_header, ""),
            llm_rows[content_id].get(llm_header, ""),
        ]
        for value in values:
            raw_labels = split_multi_value(value, known_labels) if multi else [normalize_spaces(value)]
            for raw_label in raw_labels:
                key = normalize_key(raw_label)
                if key and key not in labels:
                    labels[key] = normalize_spaces(raw_label)
    return labels


def nominal_krippendorff_alpha(pairs: Iterable[tuple[str, str]]) -> tuple[float | None, int]:
    clean_pairs = [(a, b) for a, b in pairs if a and b]
    n_items = len(clean_pairs)
    if n_items == 0:
        return None, 0

    observed_disagreement = sum(1 for a, b in clean_pairs if a != b) / n_items

    counts: Counter[str] = Counter()
    for a, b in clean_pairs:
        counts[a] += 1
        counts[b] += 1

    total_values = sum(counts.values())
    if total_values < 2:
        return None, n_items

    expected_agreement = sum(count * (count - 1) for count in counts.values()) / (
        total_values * (total_values - 1)
    )
    expected_disagreement = 1 - expected_agreement
    if math.isclose(expected_disagreement, 0.0):
        return None, n_items

    return 1 - (observed_disagreement / expected_disagreement), n_items


def interpretation(score: float | None) -> str:
    if score is None:
        return "Not available"
    if score >= 0.80:
        return "Strong agreement"
    if score >= 0.67:
        return "Acceptable agreement"
    if score >= 0.50:
        return "Moderate / needs caution"
    return "Weak agreement"


def fmt_score(value: float | None) -> str:
    return "" if value is None else f"{value:.6f}"


def write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def score_single_column(
    ids: Sequence[str],
    human_rows: dict[str, dict[str, str]],
    llm_rows: dict[str, dict[str, str]],
    human_header: str,
    llm_header: str,
) -> tuple[QuestionScore, list[bool | None]]:
    pairs: list[tuple[str, str]] = []
    exact_by_row: list[bool | None] = []

    for content_id in ids:
        human_value = clean_single(human_rows[content_id].get(human_header, ""))
        llm_value = clean_single(llm_rows[content_id].get(llm_header, ""))
        if not human_value or not llm_value:
            exact_by_row.append(None)
            continue
        pairs.append((human_value, llm_value))
        exact_by_row.append(human_value == llm_value)

    alpha, n_valid = nominal_krippendorff_alpha(pairs)
    percent_agreement = mean([1.0 if matched else 0.0 for matched in exact_by_row if matched is not None]) if n_valid else None

    return (
        QuestionScore(
            column=human_header,
            question_id=question_id(human_header),
            question=question_text(human_header),
            column_type="Single-answer",
            n_valid=n_valid,
            percent_or_exact=percent_agreement,
            alpha=alpha,
            jaccard=None,
        ),
        exact_by_row,
    )


def score_multi_column(
    ids: Sequence[str],
    human_rows: dict[str, dict[str, str]],
    llm_rows: dict[str, dict[str, str]],
    human_header: str,
    llm_header: str,
    known_labels: Sequence[str] | None = None,
) -> tuple[QuestionScore, list[bool | None], list[float | None]]:
    exact_by_row: list[bool | None] = []
    jaccard_by_row: list[float | None] = []

    for content_id in ids:
        human_set = clean_multi_set(human_rows[content_id].get(human_header, ""), known_labels)
        llm_set = clean_multi_set(llm_rows[content_id].get(llm_header, ""), known_labels)
        union = human_set | llm_set
        if not union:
            exact_by_row.append(None)
            jaccard_by_row.append(None)
            continue
        exact_by_row.append(human_set == llm_set)
        jaccard_by_row.append(len(human_set & llm_set) / len(union))

    valid_exact = [1.0 if matched else 0.0 for matched in exact_by_row if matched is not None]
    valid_jaccard = [score for score in jaccard_by_row if score is not None]
    exact_agreement = mean(valid_exact) if valid_exact else None
    avg_jaccard = mean(valid_jaccard) if valid_jaccard else None

    return (
        QuestionScore(
            column=human_header,
            question_id=question_id(human_header),
            question=question_text(human_header),
            column_type="Multi-select",
            n_valid=len(valid_jaccard),
            percent_or_exact=exact_agreement,
            alpha=None,
            jaccard=avg_jaccard,
        ),
        exact_by_row,
        jaccard_by_row,
    )


def output_paths(output: Path) -> tuple[Path, Path, Path]:
    suffix = output.suffix or ".csv"
    if suffix.casefold() != ".csv":
        raise ValueError("--output must be a .csv path")
    question_path = output
    merged_path = output.with_name(f"{output.stem}_merged_cleaned.csv")
    summary_path = output.with_name(f"{output.stem}_summary.csv")
    return question_path, merged_path, summary_path


def calculate_scores(
    human_path: Path,
    llm_path: Path,
    output_path: Path,
    include_open_text: bool,
    id_column: str,
    human_sheet: str | None,
    llm_sheet: str | None,
    force_single: set[str],
    force_multi: set[str],
) -> None:
    human_headers, human_rows, human_id_column = read_coding_file(human_path, id_column, human_sheet)
    llm_headers, llm_rows, llm_id_column = read_coding_file(llm_path, id_column, llm_sheet)

    common_ids = sorted(set(human_rows) & set(llm_rows))
    if not common_ids:
        raise ValueError("No overlapping Content ID values found between the two files")

    missing_from_human = sorted(set(llm_rows) - set(human_rows))
    missing_from_llm = sorted(set(human_rows) - set(llm_rows))
    column_pairs = shared_codebook_columns(human_headers, llm_headers, human_id_column, include_open_text)
    if not column_pairs:
        raise ValueError("No shared codebook columns found after excluding identity/open-text columns")

    question_path, merged_path, summary_path = output_paths(output_path)
    question_scores: list[QuestionScore] = []
    all_column_exact_scores: list[float] = []
    final_component_scores: list[float] = []

    merged_rows: list[dict[str, object]] = []
    per_row_extra: dict[str, dict[str, object]] = {content_id: {DEFAULT_ID_COLUMN: content_id} for content_id in common_ids}

    for human_header, llm_header in column_pairs:
        human_key = canonical_header_key(human_header)
        values = [
            human_rows[content_id].get(human_header, "")
            for content_id in common_ids
        ] + [
            llm_rows[content_id].get(llm_header, "")
            for content_id in common_ids
        ]

        if human_key in force_multi:
            is_multi = True
        elif human_key in force_single:
            is_multi = False
        else:
            is_multi = looks_multi_select(human_header, values)

        known_labels = parse_options(human_header) if is_multi else []
        label_map = build_label_map(
            common_ids,
            human_rows,
            llm_rows,
            human_header,
            llm_header,
            is_multi,
            known_labels,
        )
        clean_human_col = f"Human_{question_id(human_header)}"
        clean_llm_col = f"LLM_{question_id(human_header)}"

        for content_id in common_ids:
            if is_multi:
                per_row_extra[content_id][clean_human_col] = display_multi(
                    human_rows[content_id].get(human_header, ""),
                    label_map,
                    known_labels,
                )
                per_row_extra[content_id][clean_llm_col] = display_multi(
                    llm_rows[content_id].get(llm_header, ""),
                    label_map,
                    known_labels,
                )
            else:
                per_row_extra[content_id][clean_human_col] = display_single(
                    human_rows[content_id].get(human_header, ""),
                    label_map,
                )
                per_row_extra[content_id][clean_llm_col] = display_single(
                    llm_rows[content_id].get(llm_header, ""),
                    label_map,
                )

        if is_multi:
            score, exact_by_row, jaccard_by_row = score_multi_column(
                common_ids,
                human_rows,
                llm_rows,
                human_header,
                llm_header,
                known_labels,
            )
            for content_id, exact, jaccard in zip(common_ids, exact_by_row, jaccard_by_row, strict=True):
                per_row_extra[content_id][f"Exact_{question_id(human_header)}"] = "" if exact is None else int(exact)
                per_row_extra[content_id][f"Jaccard_{question_id(human_header)}"] = fmt_score(jaccard)
            if score.percent_or_exact is not None:
                all_column_exact_scores.append(score.percent_or_exact)
            if score.jaccard is not None:
                final_component_scores.append(score.jaccard)
        else:
            score, exact_by_row = score_single_column(
                common_ids,
                human_rows,
                llm_rows,
                human_header,
                llm_header,
            )
            for content_id, exact in zip(common_ids, exact_by_row, strict=True):
                per_row_extra[content_id][f"Exact_{question_id(human_header)}"] = "" if exact is None else int(exact)
            if score.percent_or_exact is not None:
                all_column_exact_scores.append(score.percent_or_exact)
                final_component_scores.append(score.percent_or_exact)

        question_scores.append(score)

    merged_rows = [per_row_extra[content_id] for content_id in common_ids]
    merged_fieldnames: list[str] = []
    for row in merged_rows:
        for key in row:
            if key not in merged_fieldnames:
                merged_fieldnames.append(key)

    question_rows = []
    for score in question_scores:
        question_rows.append(
            {
                "Codebook question": score.question,
                "Question ID": score.question_id,
                "Type": score.column_type,
                "Agreement metric": score.main_metric,
                "Score": fmt_score(score.main_score),
                "Interpretation": interpretation(score.main_score),
                "Percent / exact agreement": fmt_score(score.percent_or_exact),
                "Krippendorff's alpha": fmt_score(score.alpha),
                "Jaccard": fmt_score(score.jaccard),
                "Valid rows": score.n_valid,
                "Column": score.column,
            }
        )

    single_agreements = [
        score.percent_or_exact
        for score in question_scores
        if score.column_type == "Single-answer" and score.percent_or_exact is not None
    ]
    single_alphas = [
        score.alpha
        for score in question_scores
        if score.column_type == "Single-answer" and score.alpha is not None
    ]
    multi_jaccards = [
        score.jaccard
        for score in question_scores
        if score.column_type == "Multi-select" and score.jaccard is not None
    ]

    summary_rows = [
        {
            "Score": "Overall exact agreement",
            "How calculated": "Average exact match across all scored columns",
            "Value": fmt_score(mean(all_column_exact_scores) if all_column_exact_scores else None),
            "Interpretation": interpretation(mean(all_column_exact_scores) if all_column_exact_scores else None),
        },
        {
            "Score": "Average Krippendorff's alpha",
            "How calculated": "Average alpha across single-answer columns",
            "Value": fmt_score(mean(single_alphas) if single_alphas else None),
            "Interpretation": interpretation(mean(single_alphas) if single_alphas else None),
        },
        {
            "Score": "Average Jaccard",
            "How calculated": "Average Jaccard across multi-select columns",
            "Value": fmt_score(mean(multi_jaccards) if multi_jaccards else None),
            "Interpretation": interpretation(mean(multi_jaccards) if multi_jaccards else None),
        },
        {
            "Score": "Final combined agreement",
            "How calculated": "Average of single-answer percent agreements and multi-select Jaccard scores",
            "Value": fmt_score(mean(final_component_scores) if final_component_scores else None),
            "Interpretation": interpretation(mean(final_component_scores) if final_component_scores else None),
        },
        {
            "Score": "Overlapping Content IDs",
            "How calculated": "Rows present in both files by Content ID",
            "Value": str(len(common_ids)),
            "Interpretation": "",
        },
        {
            "Score": "Rows only in human file",
            "How calculated": f"Using ID column '{human_id_column}'",
            "Value": str(len(missing_from_llm)),
            "Interpretation": "",
        },
        {
            "Score": "Rows only in LLM file",
            "How calculated": f"Using ID column '{llm_id_column}'",
            "Value": str(len(missing_from_human)),
            "Interpretation": "",
        },
    ]

    write_csv(merged_path, merged_fieldnames, merged_rows)
    write_csv(
        question_path,
        [
            "Codebook question",
            "Question ID",
            "Type",
            "Agreement metric",
            "Score",
            "Interpretation",
            "Percent / exact agreement",
            "Krippendorff's alpha",
            "Jaccard",
            "Valid rows",
            "Column",
        ],
        question_rows,
    )
    write_csv(summary_path, ["Score", "How calculated", "Value", "Interpretation"], summary_rows)

    print(f"Wrote question-level scores to {question_path}")
    print(f"Wrote overall summary to {summary_path}")
    print(f"Wrote cleaned merged comparison file to {merged_path}")
    print(f"Compared {len(common_ids)} overlapping Content ID values")
    if missing_from_llm:
        print(f"Warning: {len(missing_from_llm)} Content IDs were only in the human file")
    if missing_from_human:
        print(f"Warning: {len(missing_from_human)} Content IDs were only in the LLM file")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compare human and LLM codebook coding by Content ID."
    )
    parser.add_argument("--human", required=True, type=Path, help="Path to the human-coded file")
    parser.add_argument("--llm", required=True, type=Path, help="Path to the LLM-coded file")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("agreement_question_scores.csv"),
        help="Path for the question-level score CSV. Summary and merged CSVs use the same stem.",
    )
    parser.add_argument(
        "--id-column",
        default=DEFAULT_ID_COLUMN,
        help="Shared ID column name. Defaults to 'Content ID'.",
    )
    parser.add_argument("--human-sheet", help="Sheet name for the human workbook, if using XLSX")
    parser.add_argument("--llm-sheet", help="Sheet name for the LLM workbook, if using XLSX")
    parser.add_argument(
        "--single-cols",
        help="Comma-separated column names to force as single-answer columns.",
    )
    parser.add_argument(
        "--multi-cols",
        help="Comma-separated column names to force as multi-select columns.",
    )
    parser.add_argument(
        "--include-open-text",
        action="store_true",
        help="Also score open-text fields. Usually leave this off.",
    )
    args = parser.parse_args()

    calculate_scores(
        human_path=args.human,
        llm_path=args.llm,
        output_path=args.output,
        include_open_text=args.include_open_text,
        id_column=args.id_column,
        human_sheet=args.human_sheet,
        llm_sheet=args.llm_sheet,
        force_single=parse_column_list(args.single_cols),
        force_multi=parse_column_list(args.multi_cols),
    )


if __name__ == "__main__":
    main()
