from __future__ import annotations

import shutil
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE = Path(
    "/Users/dhruvbhanderi/Documents/USC/New Research Engineer/"
    "Audience analysis kenya/Final/Kenya Content Analysis Snippets.xlsx"
)
OUTPUT = Path(
    "/Users/dhruvbhanderi/Documents/USC/New Research Engineer/"
    "Audience analysis kenya/Final/Kenya Content Analysis Snippets - With IDs.xlsx"
)

PREFIXES = {
    "Eddy Kimani": "EDDY",
    "Onyango Otieno (Rixpoet)": "RIX",
    "Eric (Amerix)": "AMX",
    "Philip Karanja": "PHIL",
    "Andrew Kibe": "KIBE",
}

AUDIENCE_TYPES = {
    "Eddy Kimani": "Progressive",
    "Onyango Otieno (Rixpoet)": "Progressive",
    "Eric (Amerix)": "Regressive",
    "Philip Karanja": "Progressive",
    "Andrew Kibe": "Regressive",
}

TRANSLATIONS = {
    ("Eddy Kimani", 4): (
        "Because at the end of it all, when all is said and done, we must recognize that "
        "being perfect is a fairy tale. Failure, scarcity, and shortcomings are all part "
        "of being a man, a father, a husband. Openness, honesty, communication are key to "
        "your relationship with your partners, but sometimes, it can be difficult."
    ),
    ("Eddy Kimani", 28): (
        "Once you learn and embrace that especially for men, being weak is normal. It's good "
        "to be weak. We were raised in a society that tells us... be strong. Yes, man, brace "
        "yourself and toughen up. Men are not supposed to talk because even that idea of "
        "toughening up is repeated in different ways. You'll find situations where you want "
        "to share something with your fellow brother but you can't because you think they "
        "will... men don't do that. But that's what society has taught us. Men don't cry. "
        "Men don't show emotions. Men don't fail!"
    ),
    ("Eddy Kimani", 30): (
        "I'm getting also feedback from ladies asking me to help their husbands because they "
        "feel the husband does not want to be helped. Now that one you have to be careful "
        "because again I can't start engaging somebody who may not want or is not ready to "
        "be helped. He may ask, why are you talking to my wife, what business do you have "
        "with her?"
    ),
    ("Eddy Kimani", 44): (
        "So one morning, I walked into a bakery and I sat down, and this guy behind me "
        "greeted me, and he told me his name was Peter and he introduced himself as the "
        "owner of the bakery. He recognized me as that guy of TV. 'Ah, you are that guy "
        "from TV, eh?' I wanted the earth to open, I fall in. This was my Damascus moment. "
        "But that time I had spent almost four hours conversing with Peter, and he went on "
        "his knees and prayed for me. Something that shocked me. He told me I needed to go "
        "back home. I needed to go back to my family. He pushed me every day as we became "
        "friends. But unfortunately, the demons kept on winning this battle all through."
    ),
    ("Eddy Kimani", 62): (
        "People should embrace that aspect of being vulnerable; I am known that way, and "
        "failure is part of the process. It's part of the process, it's actually one of "
        "the key ingredients for success. You just need to make failure your ally. You just "
        "need to make failure, don't hate it. Stand up, dust yourself and move."
    ),
    ("Onyango Otieno (Rixpoet)", 3): (
        "You can't process emotions properly and you can't think rational all right because "
        "your body is really just trying to keep you alive because you are sensing danger "
        "too often sensing danger too often get what I mean. All right this is the epitome "
        "of mental health all right because your daily experiences are connected so your "
        "body moves according to how your environment is. If you feel safe your nerves are "
        "relaxed, right; if you feel in danger your nerves are tight."
    ),
    ("Onyango Otieno (Rixpoet)", 45): (
        "I took it upon myself to call my parents for a meeting, which is very suicidal. "
        "Here I was like, 'Hey, you people, me I'm tired. I'm tired of this skills. I'm "
        "tired of the fights. I just want a peaceful home. And I'm now an adult. I want a "
        "peaceful life.' Now, being a firstborn in the house, I think I was just forced to "
        "grow up so quick. Sometimes it felt like I was actually parenting my own parents."
    ),
    ("Onyango Otieno (Rixpoet)", 65): (
        "I felt cheated. Cheated in the sense that I was 29 years old, and in all my life, "
        "all the education I've been through, nobody had taught me how my mind works. "
        "Nobody had taught me the importance of taking care of my mind. They taught me "
        "algebra, they taught me where Mandrongo not is. They taught me very simple things "
        "about the biology of a body and stuff, things like thorax, things like abdomen, "
        "things I've not even been using. But my mind, the thing that's keeping me alive, "
        "nobody has ever taught me about that."
    ),
    ("Onyango Otieno (Rixpoet)", 85): (
        "What are men doing? What are we doing as peers to say, let's come together and do "
        "something different. But we can't do that because individually we are struggling "
        "with things we don't talk about. So, we can't get together. Women struggle but they "
        "get together. That's how they are able to move. We struggle, we can't get together "
        "because we fear each other. 'Eh, how will Kivu see me? Eh, if I go to Brayo I will "
        "look weak, eh, so we should go drink and let this story end.'"
    ),
    ("Philip Karanja", 27): (
        "Thank you very much for welcoming me here. Mary, thank you so much for welcoming me "
        "here. My name is Philip Karanja, also known as Baba Njeri. And the reason we have "
        "come here is because when I am in Nairobi watching the news, I always see cases of "
        "child defilement and a lot of the time the reports come from this side of Busia."
    ),
    ("Philip Karanja", 28): (
        "Through Equality Now, who agreed to see this vision of mine and partnered with me. "
        "They told me, let us go to Busia. Every tribe has its own culture. I am just here "
        "to find out what's happening on the ground and how we can solve this problem."
    ),
    ("Andrew Kibe", 2): (
        "Especially. Not even. Especially. You see women don't see men the way men think "
        "women see men. Women they see power or puny. That's all they see. Or puny. So if "
        "she doesn't see power, she's seeing a puny guy. Yeah. So imagine you go there to "
        "tell her, oh, I have rent I have not paid. Oh, I don't know, Bro, she's looking at "
        "you like you're a puny because you'd rather be a thief than be a puny. Women fall "
        "in love with people who are in jail. Because those are not punies, those are beasts."
    ),
    ("Andrew Kibe", 3): (
        "That's bullshit. Maybe you see if you want tits and ass, just ask for tits and ass. "
        "Don't go there with your problems. Oh, oh, things are so bad. I'm about to get "
        "fired. I don't know what. Oh, she's looking at you. She's saying, this is a puny "
        "little shit. And she'll give you pity sex because there's still a small spark in you."
    ),
    ("Andrew Kibe", 4): (
        "Bro, you've lost your frame. Every time you do that shit, you do women's things. "
        "Women hate you. Every time you do women's things, she gets tired of you, what is "
        "that? Even if a woman tells you, hey, let's go eat somewhere. Tell her, yeah, let's "
        "go. She gets tired of you. We need to go on holiday. Yes, let's go on holiday. She "
        "gets tired of you because you are a dog. You have no independent brain."
    ),
    ("Andrew Kibe", 6): (
        "So every now and then, they bring somebody to talk to you about sexual harassment. "
        "Like that is a real thing. So I went for one of those trainings. I was like, why "
        "why we're grown-ups? Why are we being taught about sexual harassment? This is our "
        "children sitting down like this to be taught about sexual harassment. I'm a grown-up. "
        "I know what the consequences are to everything. You're not going to police my life."
    ),
    ("Andrew Kibe", 16): (
        "You see, men are not created to be doormats. Yeah. That's such a small problem. For "
        "a man to be solving. Of being a punching bag. Your mother calls you. And you hear "
        "that what she is telling you is nonsense. You can hear it. This is nonsense. And "
        "she is your biological mother. She is telling you about a project. How much money "
        "have have men wasted on stupid ideas coming from their parents?"
    ),
    ("Andrew Kibe", 17): (
        "Telling your mother that your mother is going to be pissed. Why is my mom talking "
        "to me? She has a husband. Go talk to your husband. You are my mother. Don't think "
        "you can dominate me. Go talk to your husband. And I'm not talking about my own mom, "
        "but even though I am, so what? Go talk to your husband. Me, I'm a child. I'm "
        "children. How do you want to dominate me right now? So-called we want to build a "
        "new house. No, your husband. Your mistake is that guy. That is your failure, woman. "
        "Deal with your failures in your life."
    ),
    ("Andrew Kibe", 18): (
        "Totally. Imagine two people raising four kids in a two-bedroom flat with one toilet. "
        "Are those geniuses? All our parents were fools. They were all living in the same "
        "place, the same number of kids. Bro, it was like a pandemic."
    ),
    ("Andrew Kibe", 19): (
        "Yeah, but but as a man you must hold this shit in. You don't express yourself like "
        "a woman. I have friends who are they have to keep apologizing. We cannot be friends. "
        "Because they fuck up all the time. That's part of life. It's fuck up. You understand "
        "me, I understand you. Why do you want me to have to come to down to apologize? For what?"
    ),
    ("Andrew Kibe", 41): (
        "Oh no. No, no, no. It will never happen, Jagero. It will never happen. Never happen. "
        "Unless unless let us say the Prince of Brunei's daughter becomes available and then "
        "calls me. Let us add buses, and then she tells me, I'm going to give you half my "
        "kingdom. Prince. As much as that is a good offer, I will need some time to think "
        "about it. I would tell her, can I talk to you in 20 in 48 hours? Then after 48 "
        "hours, I go to the mountain. I worship God until He tells me, yes, you can marry. "
        "That's when I'll get married again. It's never going to happen."
    ),
    ("Andrew Kibe", 42): (
        "I mean, I don't see why you guys why is it a big deal, Jagero? With three women. "
        "I mean, we're not greedy. We don't have to, you know, you see you married men are "
        "the ones who have to do everything because you're going back to prison. As we are "
        "here in liberty, we can stay even like that without even touching each other. Very "
        "nice. Very just playing around with them."
    ),
    ("Andrew Kibe", 43): (
        "First, even I used to enter the job and go straight to the studio. I do my work. "
        "When I finish in the studio, I leave. I do not greet anyone. Because I'm not there "
        "to make friends. I'm there to pick a bag. I'm just like a mercenary. Come in in the "
        "morning, kill a few people, get the fuck out. I I don't care how much you guys buy "
        "bread for."
    ),
    ("Andrew Kibe", 44): (
        "No, that's not the thing. We'll never stop spending on women. We love these things. "
        "These are our things. These are ours. Like your car, you'll never stop spending on "
        "your car. You love it. So, there is no fear there. It's just that there are too many."
    ),
    ("Andrew Kibe", 45): (
        "Hobbies. So hobbies, you find people playing snooker, whatever, uh, squash, chess. "
        "It is not sitting down just drinking as if you have no work. Imagine how much time "
        "you waste just sitting and drinking. There is no value that is being added to your life."
    ),
    ("Andrew Kibe", 46): (
        "You must escape with an Airbnb and a PS2. So that you can play. Your wife's. You're "
        "about to say your wife's. Your wife's house. I mean, she gives you like once every "
        "month and I can't be there. Jagerro, you've been a good boy this month. You can play "
        "PS2 hours."
    ),
    ("Andrew Kibe", 49): (
        "White women in America, what they do, they identify someone like Wanjala here. A tall "
        "guy who has all these abilities and physique and everything. They seduce him. So they "
        "can get his seed. And supercharge it. So now, if you look at the NBA, all sports in "
        "America now. Look at all sports. All the people are point guards. Because this white "
        "person goes and takes the seed of an older black man. She raises it through the right "
        "system."
    ),
    ("Andrew Kibe", 50): (
        "We as men, we are we judge our women too harshly. A beautiful woman, how can she "
        "resist? That's why I told you, if you're going to marry, marry the ugliest. Get "
        "something you can even send to the shop at 3 a.m. And even the night guards will not "
        "talk to her. Is that the life that you want, Jagero? And if she calls you must pick "
        "up because if she does not pick up you are afraid she is giving it away."
    ),
    ("Andrew Kibe", 52): (
        "And now who is it full of? Women. So where are you looking for women? Where will you "
        "get them from? And the body count started with a Swahili boda rider. God damn it. "
        "The Swahili boda rider was called 14. You are still waiting. Yes, he eats it with "
        "salt and pepper."
    ),
    ("Andrew Kibe", 59): (
        "Me and women have a very good relationship, especially me and prostitutes. We love "
        "each other. I love my prostitutes. I love my hoes. You don't decide who your hoe is "
        "going to be. You let God decide that. So if this evening I go and find myself "
        "somewhere where women are giving themselves away. Whom am I? So should I reject "
        "God's matters? And God is the one who raised this girl until she reached 20 years."
    ),
    ("Andrew Kibe", 60): (
        "Am I stupid? You are not supposed to have a girlfriend. What is that? Am I in class "
        "six? This is my girlfriend. Am I mad? Those are things for children, man. When you "
        "mature, then you realize those those titles mean nothing."
    ),
    ("Andrew Kibe", 63): (
        "This game this game needs beasts. So even when taking a woman, make sure this woman "
        "is a beast like you. Two beasts just existing in the. Even maybe they touch each "
        "other. Eh, all of that is just for the show. When they are done, see you tomorrow. "
        "8:00 in the morning. Yes. We report to work. As wife and husband. It's genius."
    ),
    ("Andrew Kibe", 64): (
        "So imagine now I have to apologize to my father because I impregnated. Why? That "
        "doesn't make any sense. I'm sorry, why? I was not sorry when I was putting my dick."
    ),
    ("Andrew Kibe", 69): (
        "I'll never believe a man who says that. I will never. Bro, the way I have put my "
        "faith in men before. Never put my faith in another man."
    ),
}


def header_map(ws):
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value is not None}


def autosize_summary(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 34)


def main() -> None:
    shutil.copy2(SOURCE, OUTPUT)
    wb = load_workbook(OUTPUT)

    if "Summary Metrics" in wb.sheetnames:
        del wb["Summary Metrics"]

    summary_rows = []
    by_platform = Counter()
    by_content_type = Counter()
    by_audience_type = Counter()
    translated_counts = Counter()
    total_snippets = 0

    for sheet_name in [s for s in wb.sheetnames if s != "Summary Metrics"]:
        ws = wb[sheet_name]
        headers = header_map(ws)
        text_col = headers.get("Text") or headers.get("full_text") or headers.get("text")
        influencer_col = headers.get("Influencer") or headers.get("influencer_name")
        platform_col = headers.get("Platform") or headers.get("platform")
        content_type_col = headers.get("Content Type")

        id_col = headers.get("Segment ID")
        if id_col is None:
            id_col = next(
                (
                    col
                    for col in range(1, ws.max_column + 1)
                    if ws.cell(1, col).value is None
                    and all(ws.cell(row, col).value is None for row in range(2, ws.max_row + 1))
                ),
                None,
            )
        if id_col is None:
            raise ValueError(f"No Segment ID or empty existing column available on {sheet_name}")
        ws.cell(1, id_col, "Segment ID")
        ws.cell(1, id_col).font = Font(bold=True)
        ws.cell(1, id_col).alignment = Alignment(wrap_text=True)

        prefix = PREFIXES.get(sheet_name, sheet_name.upper().replace(" ", "")[:5])
        audience_type = AUDIENCE_TYPES.get(sheet_name, "Unclassified")
        snippet_count = 0
        translated = 0

        for row in range(2, ws.max_row + 1):
            has_text = bool(text_col and ws.cell(row, text_col).value)
            if not has_text:
                continue
            snippet_count += 1
            ws.cell(row, id_col, f"{prefix}-{snippet_count:03d}")

            if (sheet_name, row) in TRANSLATIONS:
                ws.cell(row, text_col, TRANSLATIONS[(sheet_name, row)])
                translated += 1

        influencer = ws.cell(2, influencer_col).value if influencer_col else sheet_name
        platform = ws.cell(2, platform_col).value if platform_col else ""
        content_type = ws.cell(2, content_type_col).value if content_type_col else "X Post"
        total_snippets += snippet_count
        by_platform[str(platform)] += snippet_count
        by_content_type[str(content_type)] += snippet_count
        by_audience_type[audience_type] += snippet_count
        translated_counts[str(influencer)] += translated

        summary_rows.append([influencer, audience_type, platform, content_type, snippet_count, translated, prefix])
        ws.column_dimensions[get_column_letter(id_col)].width = 14

    summary = wb.create_sheet("Summary Metrics")
    title_fill = PatternFill("solid", fgColor="F2F6FA")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    summary["A1"] = "Kenya Content Analysis Snippets Summary"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A1"].fill = title_fill

    overview = [
        ["Metric", "Value"],
        ["Total snippets", total_snippets],
        ["Progressive snippets", by_audience_type["Progressive"]],
        ["Regressive snippets", by_audience_type["Regressive"]],
        ["Influencers", len(summary_rows)],
        ["Translated text snippets", sum(translated_counts.values())],
    ]
    for r, row in enumerate(overview, 3):
        for c, value in enumerate(row, 1):
            summary.cell(r, c, value)

    influencer_table = [
        ["Influencer", "Audience Type", "Platform", "Content Type", "Snippets", "Translated Text Rows", "ID Prefix"],
        *summary_rows,
    ]
    for r, row in enumerate(influencer_table, 9):
        for c, value in enumerate(row, 1):
            summary.cell(r, c, value)

    platform_table = [["Platform", "Snippets"], *by_platform.items()]
    for r, row in enumerate(platform_table, 3):
        for c, value in enumerate(row, 8):
            summary.cell(r, c, value)

    content_table = [["Content Type", "Snippets"], *by_content_type.items()]
    for r, row in enumerate(content_table, 9):
        for c, value in enumerate(row, 8):
            summary.cell(r, c, value)

    audience_table = [["Audience Type", "Snippets"], *by_audience_type.items()]
    for r, row in enumerate(audience_table, 17):
        for c, value in enumerate(row, 8):
            summary.cell(r, c, value)

    for row in [3, 9]:
        for col in range(1, 8):
            cell = summary.cell(row, col)
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in [3, 9]:
        for col in range(8, 10):
            cell = summary.cell(row, col)
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(8, 10):
        cell = summary.cell(17, col)
        if cell.value is not None:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize_summary(summary)
    summary.freeze_panes = "A4"
    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
