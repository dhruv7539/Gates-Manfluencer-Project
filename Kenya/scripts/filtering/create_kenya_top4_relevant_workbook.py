#!/usr/bin/env python3
"""Create a workbook with the most relevant Kenya top-4 filtered comments."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


HERE = Path(__file__).resolve().parent
FILTERING_ROOT = HERE.parent
FILTERED_DIR = FILTERING_ROOT / "outputs" / "Kenya" / "filtered_output"
OUTPUT_PATH = FILTERING_ROOT / "outputs" / "Kenya" / "Kenya_top4_most_relevant_comments.xlsx"

FILES = {
    "Tiktok; Young man who thinks it's a shame not having a car at 35__filtered.xlsx": "Car At 35",
    "Tiktok; Men Are Not Missing. They Are Evolving!.csv__filtered.xlsx": "Men Evolving",
    "Full Tweet Stay away from vulgar women__filtered.xlsx": "Vulgar Women",
    "Tweet_A woman can't love a man. It is a man who loves a woman.__filtered.xlsx": "Love Claim",
}

STRONG_KEYWORDS = {
    "#MasculinitySaturday",
    "Frame",
    "boy child",
    "cunt",
    "feminist",
    "gender",
    "good wife",
    "male",
    "masculine",
    "masculinity",
    "msee",
    "mubaba",
    "mwanaume",
    "obedience",
    "obey",
    "provide",
    "red pill",
    "respectful woman",
    "simp",
    "single mother",
    "soy boy",
    "sponsor",
    "submission",
    "submissive",
    "submit",
    "vulgar women",
    "wanaume",
    "girl child",
}

CAR35_PATTERNS = re.compile(
    r"\b(car|gari|35|ego|growth|arrivalism|pride|arrogance|status|wealth|rentals?|rebuilding|"
    r"sold my car|wife'?s business|pressure)\b",
    re.IGNORECASE,
)
EVOLVING_PATTERNS = re.compile(
    r"\b(men are not missing|masculinity|boy child|girl child|gender|single mother|real man|"
    r"boys? to men|systems? are rigged|wanaume|mwanaume|male disadvantage|men learn)\b",
    re.IGNORECASE,
)
VULGAR_PATTERNS = re.compile(
    r"\b(vulgar|respectful|decency|feminist|soy boy|masculinity|women|woman|simp|cunt|wives?|"
    r"male|patriarch|controlling|strong women)\b",
    re.IGNORECASE,
)
LOVECLAIM_PATTERNS = re.compile(
    r"\b(love|submit|submission|obey|respect|provide|prize|wives?|husband|woman|women|gender|"
    r"red pill|simp|loyal|relationship|support her partner)\b",
    re.IGNORECASE,
)


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        letter = column[0].column_letter
        max_length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 60)


def is_relevant(kind: str, keywords: list[str], text: str) -> bool:
    if any(keyword in STRONG_KEYWORDS for keyword in keywords):
        return True
    if kind == "Car At 35":
        return bool(CAR35_PATTERNS.search(text))
    if kind == "Men Evolving":
        return bool(EVOLVING_PATTERNS.search(text))
    if kind == "Vulgar Women":
        return bool(VULGAR_PATTERNS.search(text))
    if kind == "Love Claim":
        return bool(LOVECLAIM_PATTERNS.search(text))
    return False


def style_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize_columns(worksheet)


def main() -> None:
    out_wb = Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    for filename, tab_name in FILES.items():
        src_wb = load_workbook(FILTERED_DIR / filename, read_only=True, data_only=True)
        src_ws = src_wb["Filtered"]
        headers = list(next(src_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        text_column = "comment" if "comment" in headers else "text"
        text_idx = headers.index(text_column)
        keyword_idx = headers.index("matched_keywords")

        out_ws = out_wb.create_sheet(title=tab_name)
        out_ws.append(headers)

        relevant_count = 0
        for row in src_ws.iter_rows(min_row=2, values_only=True):
            text = "" if row[text_idx] is None else str(row[text_idx])
            keywords = [] if row[keyword_idx] is None else str(row[keyword_idx]).split(" | ")
            if is_relevant(tab_name, keywords, text):
                out_ws.append(list(row))
                relevant_count += 1

        style_worksheet(out_ws)
        print(f"{tab_name}: {relevant_count}")

    out_wb.save(OUTPUT_PATH)
    print(f"Saved workbook to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
